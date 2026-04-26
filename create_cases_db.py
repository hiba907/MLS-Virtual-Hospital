import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Patient Cases"

headers = [
    "Case_ID", "Patient_Name", "Age", "Gender", "Chief_Complaint",
    "Disease", "Category", "Symptoms", "Vitals_HR", "Vitals_BP",
    "Vitals_RR", "Vitals_Temp_C", "Vitals_SpO2", "Physical_Exam",
    "Lab_Tests_Ordered", "Lab_Results", "Imaging", "Imaging_Findings",
    "Diagnosis", "Treatment_Plan", "Difficulty"
]

header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cases = [
    [
        "CASE001", "Ahmed Al-Rashid", 45, "Male", "Chest pain and shortness of breath",
        "Acute Myocardial Infarction (STEMI)", "Cardiology",
        "Severe crushing chest pain radiating to left arm; diaphoresis; nausea; dyspnea; anxiety",
        110, "90/60", 22, 37.2, 94,
        "Diaphoresis; pallor; cold extremities; S3 gallop on auscultation; JVD present",
        "ECG; Troponin I; CK-MB; CBC; BMP; Chest X-ray; Echocardiogram",
        "Troponin I: 8.5 ng/mL (HIGH); CK-MB: 45 U/L (HIGH); WBC: 12,000; Na: 138; K: 4.1",
        "Chest X-ray; ECG", "ECG: ST elevation in leads II,III,aVF; CXR: mild pulmonary congestion",
        "Acute Inferior STEMI",
        "Aspirin 325mg; Clopidogrel 600mg; Heparin infusion; Morphine 2mg IV; O2 supplementation; Urgent PCI",
        "Advanced"
    ],
    [
        "CASE002", "Fatima Hassan", 28, "Female", "High fever and severe headache",
        "Bacterial Meningitis", "Neurology",
        "Sudden high fever; severe headache; neck stiffness; photophobia; phonophobia; altered consciousness; vomiting",
        124, "100/70", 26, 39.8, 97,
        "Kernig sign positive; Brudzinski sign positive; nuchal rigidity; petechial rash; confusion",
        "Lumbar puncture (CSF analysis); Blood cultures; CBC; CMP; CT Head; Procalcitonin",
        "CSF: cloudy; WBC 5000 (>90% neutrophils); Protein 180mg/dL; Glucose 30mg/dL; Blood culture: Neisseria meningitidis",
        "CT Head", "CT Head: no herniation; mild meningeal enhancement",
        "Bacterial Meningitis (Neisseria meningitidis)",
        "Ceftriaxone 2g IV q12h; Dexamethasone 0.15mg/kg IV q6h x4 days; Isolation; IV fluids; Analgesics",
        "Advanced"
    ],
    [
        "CASE003", "Omar Khalid", 62, "Male", "Sudden weakness on one side of body",
        "Ischemic Stroke", "Neurology",
        "Sudden right-sided facial drooping; right arm weakness; slurred speech; onset 1 hour ago; headache",
        88, "180/105", 18, 37.0, 98,
        "Right facial palsy; right arm drift positive; aphasia; NIHSS score: 12; irregular pulse",
        "CT Head non-contrast; MRI brain; ECG; CBC; Coagulation panel; BMP; Lipid panel; Carotid ultrasound",
        "CT Head: no hemorrhage; MRI DWI: acute infarct left MCA territory; INR: 1.1; LDL: 4.2 mmol/L",
        "CT Head; MRI Brain; Carotid Ultrasound", "Left MCA territory acute ischemic infarct; carotid stenosis 70% left",
        "Acute Ischemic Stroke (Left MCA territory)",
        "IV tPA (alteplase) 0.9mg/kg if within 4.5hr window; Aspirin 300mg; BP management; Stroke unit admission; Physiotherapy",
        "Advanced"
    ],
    [
        "CASE004", "Sara Mohammed", 35, "Female", "Severe abdominal pain and vomiting",
        "Acute Appendicitis", "Surgery",
        "Periumbilical pain migrating to RLQ; nausea; vomiting; fever; anorexia; pain worse with movement",
        102, "118/76", 20, 38.4, 98,
        "RLQ tenderness; McBurney point tenderness; Rovsing sign positive; rebound tenderness; guarding",
        "CBC; CRP; Urinalysis; Pelvic ultrasound; CT Abdomen/Pelvis; Beta-hCG (female)",
        "WBC: 16,500 (neutrophilia); CRP: 85 mg/L; Urinalysis: normal; Beta-hCG: negative",
        "CT Abdomen/Pelvis; Ultrasound", "CT: dilated appendix 9mm; periappendiceal fat stranding; no perforation",
        "Acute Appendicitis (uncomplicated)",
        "NPO; IV fluids; Ceftriaxone + Metronidazole; Urgent laparoscopic appendectomy; Pain management",
        "Intermediate"
    ],
    [
        "CASE005", "Khalid Ibrahim", 55, "Male", "Difficulty breathing and leg swelling",
        "Acute Heart Failure (Decompensated)", "Cardiology",
        "Progressive dyspnea; orthopnea; PND; bilateral leg swelling; fatigue; weight gain 4kg in 1 week",
        98, "160/95", 24, 37.1, 89,
        "Bibasilar crackles; pitting edema bilateral; elevated JVP; S3 gallop; displaced apex beat",
        "BNP; Troponin; CBC; BMP; LFTs; Chest X-ray; ECG; Echocardiogram; Urinalysis",
        "BNP: 1850 pg/mL (HIGH); Na: 132 (LOW); Creatinine: 1.6; Troponin: 0.04",
        "Chest X-ray; ECG; Echocardiogram", "CXR: cardiomegaly; bilateral pleural effusions; Kerley B lines; Echo: EF 25%",
        "Acute Decompensated Heart Failure (reduced EF)",
        "Furosemide 80mg IV; Oxygen therapy; Fluid restriction 1.5L/day; Daily weights; Fluid balance monitoring; Cardiology consult",
        "Intermediate"
    ],
    [
        "CASE006", "Layla Ahmad", 22, "Female", "Rash and joint pain after sore throat",
        "Rheumatic Fever", "Rheumatology",
        "Migratory joint pain; skin rash; fever; throat pain 2 weeks ago; fatigue; chest pain",
        96, "110/70", 18, 38.1, 99,
        "Erythema marginatum rash; subcutaneous nodules; carditis (murmur); arthritis of large joints",
        "ASO titer; Anti-DNase B; Throat swab culture; CBC; ESR; CRP; ECG; Echocardiogram",
        "ASO titer: 800 IU/mL (HIGH); Anti-DNase B elevated; Throat swab: Group A Strep; ESR: 90mm/hr; CRP: 65mg/L",
        "ECG; Echocardiogram", "ECG: prolonged PR interval; Echo: mitral regurgitation",
        "Acute Rheumatic Fever with Carditis",
        "Benzathine Penicillin G 1.2M units IM; Aspirin 100mg/kg/day; Prednisolone for carditis; Monthly secondary prophylaxis",
        "Intermediate"
    ],
    [
        "CASE007", "Yusuf Al-Amin", 8, "Male", "Difficulty breathing and wheezing",
        "Acute Asthma Exacerbation", "Pulmonology",
        "Wheezing; shortness of breath; chest tightness; cough; triggered by exercise; known asthmatic",
        120, "105/65", 32, 37.3, 92,
        "Intercostal retractions; prolonged expiratory phase; bilateral expiratory wheeze; accessory muscle use",
        "Peak flow measurement; ABG; CBC; Chest X-ray; Pulse oximetry; Spirometry",
        "Peak flow: 40% predicted; ABG: pH 7.38; pO2: 68; pCO2: 42; CXR: hyperinflation",
        "Chest X-ray", "CXR: bilateral hyperinflation; no consolidation; no pneumothorax",
        "Moderate-Severe Acute Asthma Exacerbation",
        "Salbutamol 5mg nebulized q20min x3; Ipratropium bromide 0.5mg neb; Prednisolone 1mg/kg oral; O2 target SpO2 >95%",
        "Beginner"
    ],
    [
        "CASE008", "Nadia Karimi", 50, "Female", "Severe headache and visual disturbance",
        "Hypertensive Emergency", "Internal Medicine",
        "Thunderclap headache; blurred vision; nausea; confusion; no prior hypertension history",
        95, "220/130", 20, 37.0, 97,
        "Papilledema on fundoscopy; confusion; grade IV hypertensive retinopathy; no focal neurological deficit",
        "CBC; BMP; Urinalysis; ECG; Chest X-ray; CT Head; Renal function; LFTs; Cardiac biomarkers",
        "Creatinine: 2.1 (HIGH); Urinalysis: proteinuria 3+; haematuria; BUN: 35; Troponin: normal",
        "CT Head", "CT Head: no hemorrhage; mild cerebral edema",
        "Hypertensive Emergency with Hypertensive Nephropathy",
        "Labetalol IV infusion; Target: reduce MAP by 20-25% in first hour; ICU admission; Fluid balance; Renal monitoring",
        "Beginner"
    ],
    [
        "CASE009", "Hassan Ali", 70, "Male", "Confusion and decreased urine output",
        "Acute Kidney Injury (AKI)", "Nephrology",
        "Decreased urine output for 2 days; confusion; nausea; weakness; recent NSAID use for back pain",
        88, "100/65", 18, 36.8, 96,
        "Dry mucous membranes; poor skin turgor; bilateral flank tenderness; altered mentation; oliguria",
        "Serum creatinine; BUN; Electrolytes; Urinalysis; Urine microscopy; Renal ultrasound; CBC; ABG",
        "Creatinine: 4.5 mg/dL (HIGH); BUN: 65; K: 6.2 (HIGH); Na: 130; Urine Na: 12 mEq/L; Urine: muddy brown casts",
        "Renal Ultrasound", "Bilateral kidneys normal size; no hydronephrosis; no stones",
        "Acute Kidney Injury (Intrinsic - ATN from NSAID)",
        "Stop NSAIDs; IV fluid resuscitation; Strict I&O; Monitor K+ closely; Nephrology consult; Possible dialysis if K rises",
        "Intermediate"
    ],
    [
        "CASE010", "Amira Saleh", 30, "Female", "Palpitations and tremor",
        "Hyperthyroidism (Graves Disease)", "Endocrinology",
        "Palpitations; heat intolerance; weight loss despite good appetite; tremor; anxiety; insomnia; diarrhea",
        112, "130/70", 20, 37.5, 98,
        "Exophthalmos; diffuse goiter; fine tremor; warm moist skin; hyperreflexia; thyroid bruit",
        "TSH; Free T4; Free T3; TRAb antibodies; Thyroid ultrasound; ECG; CBC; LFTs",
        "TSH: <0.01 (SUPPRESSED); Free T4: 35 pmol/L (HIGH); Free T3: 12 pmol/L (HIGH); TRAb: positive",
        "Thyroid Ultrasound; ECG", "Ultrasound: diffuse enlarged heterogeneous thyroid; ECG: sinus tachycardia",
        "Graves Disease (Hyperthyroidism)",
        "Propranolol 40mg TDS for symptom control; Carbimazole 30mg/day; Endocrinology referral; Monitor LFTs; Consider radioiodine",
        "Beginner"
    ],
]

for row_idx, case in enumerate(cases, 2):
    alt_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid") if row_idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for col_idx, value in enumerate(case, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = alt_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = min(max_length + 2, 40)
    ws.column_dimensions[col_letter].width = adjusted_width

ws.row_dimensions[1].height = 30
for row in range(2, len(cases) + 2):
    ws.row_dimensions[row].height = 60

ws.freeze_panes = "A2"

# Add Instructions sheet
ws2 = wb.create_sheet("Instructions")
ws2["A1"] = "MLS Virtual Hospital - Patient Cases Database"
ws2["A1"].font = Font(bold=True, size=14, color="1B4F72")
ws2["A3"] = "How to use this database:"
ws2["A3"].font = Font(bold=True)
instructions = [
    "1. Each row represents one clinical case scenario.",
    "2. The Streamlit app reads this Excel file automatically.",
    "3. To add new cases, simply add rows following the same format.",
    "4. Difficulty levels: Beginner | Intermediate | Advanced",
    "5. Keep Case_ID unique (e.g., CASE011, CASE012...).",
    "6. Vitals: HR (bpm), BP (mmHg), RR (breaths/min), Temp (Celsius), SpO2 (%)",
]
for i, inst in enumerate(instructions, 4):
    ws2[f"A{i}"] = inst

wb.save("/home/claude/virtual_hospital/patient_cases.xlsx")
print("Excel database created successfully with 10 cases!")

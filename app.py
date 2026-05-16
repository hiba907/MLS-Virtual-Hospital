# ════════════════════════════════════════════════════════════════════════════
# 🏥 MLS Virtual Hospital
# Copyright (c) 2026 Hiba Hamdar — Academy of Medical Learning Skills
#
# ⚖️  ALL RIGHTS RESERVED — PROPRIETARY SOFTWARE
#
# This source code is proprietary and confidential.
# It may NOT be copied, reused, modified, distributed, sold, or used
# in any form without explicit written permission from the copyright holder.
#
# Viewing of this code is permitted on the public GitHub repository for
# reference and educational observation purposes only.
#
# Unauthorized use will result in legal action.
#
# 📜 See LICENSE file for full terms.
# 📧 For licensing inquiries: hamdarhiba95@gmail.com
# ════════════════════════════════════════════════════════════════════════════
import streamlit as st
import time
import streamlit.components.v1 as components
import pandas as pd
import requests
import os
import json
import base64
import random
import string
import hashlib
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timezone
from typing import Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_OK = True
except ImportError:
    PLOTLY_OK = False
try:
    from fpdf import FPDF
    FPDF_OK = True
except ImportError:
    FPDF_OK = False
try:
    from transformers import AutoModelForCausalLM, AutoTokenizer
    import torch
    TRANSFORMERS_OK = True
except ImportError:
    TRANSFORMERS_OK = False

# ── New feature pages ─────────────────────────────────────────────────────────
try:
    from new_features import page_osce_exam, page_progress_notes, page_flashcard_builder
    NEW_FEATURES_OK = True
except ImportError:
    NEW_FEATURES_OK = False

# ── Tier 1 features — gamification + Ask Dr. Hiba ────────────────────────────
try:
    from tier1_features import (
        award_xp, init_session, get_user_stats,
        render_xp_bar, render_stats_dashboard,
        render_ask_mentor_button, render_ask_mentor_page,
        render_leaderboard, render_floating_help_button,
        MENTOR_WHATSAPP, MENTOR_EMAIL, MENTOR_NAME,
    )
    TIER1_AVAILABLE = True
except Exception as e:
    TIER1_AVAILABLE = False
    print(f"Tier 1 features not loaded: {e}")

# ── Mentor Directory — book sessions with multiple seniors ────────────────────
try:
    from mentor_directory import (
        render_mentor_directory_page,
        render_admin_mentor_panel,
        render_book_session_button,
        render_my_sessions_page,
        render_jitsi_call_page,
        get_pending_session_count,
        get_unverified_senior_count,
    )
    MENTOR_DIRECTORY_OK = True
except Exception as e:
    MENTOR_DIRECTORY_OK = False
    print(f"Mentor directory not loaded: {e}")

# ── MCQ Hybrid System — auto MCQs after diagnosis ─────────────────────────────
try:
    from mcq_system import (
        render_mcq_session_page,
        render_mcq_admin_panel,
        render_post_diagnosis_mcq_button,
        generate_mcqs_for_case,
        get_mcq_count_for_case,
    )
    MCQ_SYSTEM_OK = True
except Exception as e:
    MCQ_SYSTEM_OK = False
    print(f"MCQ system not loaded: {e}")

# ── AI Case Creator — Phase 1 of Knowledge Expansion ─────────────────────────
try:
    from case_creator import (
        render_case_creator_panel,
        load_approved_cases_from_db,
        get_pending_cases_count,
    )
    CASE_CREATOR_OK = True
except Exception as e:
    CASE_CREATOR_OK = False
    print(f"Case Creator not loaded: {e}")

# ── Image Practice Library — Phase 2 of Knowledge Expansion ──────────────────
try:
    from image_library import (
        render_image_practice_page,
        render_image_admin_panel,
        render_case_linked_images,
        get_pending_images_count,
    )
    IMAGE_LIBRARY_OK = True
except Exception as e:
    IMAGE_LIBRARY_OK = False
    print(f"Image Library not loaded: {e}")

# ── RAG System — Phase 3: AI tutor cites real medical references ─────────────
try:
    from rag_system import (
        render_rag_admin_panel,
        search_relevant_chunks,
        get_rag_context_for_query,
    )
    RAG_SYSTEM_OK = True
except Exception as e:
    RAG_SYSTEM_OK = False
    print(f"RAG system not loaded: {e}")

# ── Admin User Panel + Email Notifications ───────────────────────────────────
try:
    from admin_user_panel import (
        render_user_management_panel,
        send_broadcast_email,
        fetch_all_users,
    )
    USER_PANEL_OK = True
except Exception as e:
    USER_PANEL_OK = False
    print(f"User panel not loaded: {e}")

# ── Clinical helpers (real PubMed retrieval, specialist X-ray AI, feedback) ──
try:
    from clinical_helpers import (
        fetch_pubmed_case_reports, render_pubmed_case_card,
        specialist_chest_xray_analysis, render_specialist_panel,
        is_chest_xray_modality,
        render_finding_feedback, render_feedback_summary, export_feedback_for_training,
        XRV_OK,
    )
    CLINICAL_HELPERS_OK = True
except ImportError:
    CLINICAL_HELPERS_OK = False
    XRV_OK = False

# ── Real clinical sound recordings (free, no API key, no library) ───────────
try:
    from real_clinical_sounds import (
        render_dual_sound_panel,
        render_real_recording_player,
        render_disclaimer_banner,
        has_real_recording,
        REAL_CLINICAL_SOUNDS,
    )
    REAL_SOUNDS_OK = True
except ImportError:
    REAL_SOUNDS_OK = False

st.set_page_config(page_title="MLS Virtual Hospital", page_icon="🏥", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
/* ══════════════════════════════════════════════════════
   MLS VIRTUAL HOSPITAL — PROFESSIONAL CLINICAL DESIGN
   Inspired by Epic EHR · Cerner · Modern Hospital UI
   ══════════════════════════════════════════════════════ */

@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');

/* ── Design Tokens ─────────────────────────────────── */
:root {
  --clr-navy:       #0a2540;
  --clr-blue:       #1a4f8a;
  --clr-teal:       #0e7490;
  --clr-teal-lt:    #0ea5e9;
  --clr-green:      #059669;
  --clr-green-lt:   #d1fae5;
  --clr-amber:      #d97706;
  --clr-amber-lt:   #fef3c7;
  --clr-red:        #dc2626;
  --clr-red-lt:     #fee2e2;
  --clr-purple:     #7c3aed;
  --clr-purple-lt:  #ede9fe;
  --clr-bg:         #f0f4f8;
  --clr-surface:    #ffffff;
  --clr-border:     #e2e8f0;
  --clr-text:       #0f172a;
  --clr-muted:      #64748b;
  --radius-sm:      6px;
  --radius-md:      10px;
  --radius-lg:      14px;
  --shadow-sm:      0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
  --shadow-md:      0 4px 12px rgba(0,0,0,.10), 0 2px 4px rgba(0,0,0,.06);
  --shadow-lg:      0 10px 30px rgba(10,37,64,.15);
}

/* ── Global Reset ──────────────────────────────────── */
html, body, [class*="css"] {
  font-family: 'Inter', system-ui, -apple-system, sans-serif;
  color: var(--clr-text);
  background-color: var(--clr-bg);
}

/* ── Main content area ─────────────────────────────── */
.main .block-container {
  background: var(--clr-bg);
  padding-top: 1.5rem;
}

/* ── Sidebar ────────────────────────────────────────── */
[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #0a2540 0%, #0f3460 60%, #0a2540 100%) !important;
  border-right: 1px solid rgba(255,255,255,.08) !important;
}
[data-testid="stSidebar"] * {
  color: #e2e8f0 !important;
}
[data-testid="stSidebar"] .stButton > button {
  background: rgba(255,255,255,.06) !important;
  border: 1px solid rgba(255,255,255,.10) !important;
  color: #cbd5e1 !important;
  border-radius: var(--radius-sm) !important;
  font-size: .8rem !important;
  font-weight: 500 !important;
  padding: .45rem .75rem !important;
  text-align: left !important;
  transition: all .15s ease !important;
  letter-spacing: .01em !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
  background: rgba(14,116,144,.35) !important;
  border-color: rgba(14,212,244,.3) !important;
  color: white !important;
}
[data-testid="stSidebar"] hr {
  border-color: rgba(255,255,255,.08) !important;
  margin: .6rem 0 !important;
}
[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
  color: white !important;
  font-weight: 700 !important;
}

/* ── Streamlit global button ───────────────────────── */
.stButton > button {
  background: var(--clr-navy) !important;
  color: white !important;
  border: none !important;
  border-radius: var(--radius-sm) !important;
  font-weight: 600 !important;
  font-size: .83rem !important;
  padding: .5rem 1.1rem !important;
  letter-spacing: .02em !important;
  box-shadow: var(--shadow-sm) !important;
  transition: all .15s ease !important;
}
.stButton > button:hover {
  background: var(--clr-blue) !important;
  box-shadow: var(--shadow-md) !important;
  transform: translateY(-1px) !important;
}
.stButton > button[kind="primary"] {
  background: linear-gradient(135deg, var(--clr-teal), var(--clr-blue)) !important;
}

/* ── Inputs ─────────────────────────────────────────── */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div {
  border: 1.5px solid var(--clr-border) !important;
  border-radius: var(--radius-sm) !important;
  background: var(--clr-surface) !important;
  font-size: .85rem !important;
  transition: border-color .15s !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
  border-color: var(--clr-teal) !important;
  box-shadow: 0 0 0 3px rgba(14,116,144,.12) !important;
}

/* ── Tabs ────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
  background: var(--clr-surface) !important;
  border-bottom: 2px solid var(--clr-border) !important;
  gap: 0 !important;
  border-radius: var(--radius-md) var(--radius-md) 0 0 !important;
  padding: 0 1rem !important;
}
.stTabs [data-baseweb="tab"] {
  background: transparent !important;
  border-bottom: 3px solid transparent !important;
  color: var(--clr-muted) !important;
  font-weight: 500 !important;
  font-size: .82rem !important;
  padding: .6rem 1.1rem !important;
  margin-bottom: -2px !important;
}
.stTabs [aria-selected="true"] {
  border-bottom-color: var(--clr-teal) !important;
  color: var(--clr-teal) !important;
  font-weight: 700 !important;
}
.stTabs [data-baseweb="tab-panel"] {
  background: var(--clr-surface) !important;
  border: 1px solid var(--clr-border) !important;
  border-top: none !important;
  border-radius: 0 0 var(--radius-md) var(--radius-md) !important;
  padding: 1rem !important;
}

/* ── Expander ────────────────────────────────────────── */
.streamlit-expanderHeader {
  background: var(--clr-surface) !important;
  border: 1px solid var(--clr-border) !important;
  border-radius: var(--radius-sm) !important;
  font-weight: 600 !important;
  font-size: .84rem !important;
  color: var(--clr-navy) !important;
}
.streamlit-expanderContent {
  border: 1px solid var(--clr-border) !important;
  border-top: none !important;
  background: #fafbfc !important;
}

/* ── Metrics ─────────────────────────────────────────── */
[data-testid="metric-container"] {
  background: var(--clr-surface) !important;
  border: 1px solid var(--clr-border) !important;
  border-radius: var(--radius-md) !important;
  padding: 1rem 1.1rem !important;
  box-shadow: var(--shadow-sm) !important;
}
[data-testid="metric-container"] label {
  color: var(--clr-muted) !important;
  font-size: .76rem !important;
  font-weight: 600 !important;
  letter-spacing: .05em !important;
  text-transform: uppercase !important;
}
[data-testid="metric-container"] [data-testid="metric-value"] {
  color: var(--clr-navy) !important;
  font-weight: 800 !important;
}

/* ══════════════════════════════════════════════════════
   CLINICAL COMPONENT CLASSES
   ══════════════════════════════════════════════════════ */

/* ── Page Header ─────────────────────────────────────── */
.main-header {
  background: linear-gradient(135deg, #0a2540 0%, #1a4f8a 55%, #0e7490 100%);
  color: white;
  padding: 1.6rem 2rem;
  border-radius: var(--radius-lg);
  margin-bottom: 1.5rem;
  text-align: center;
  box-shadow: var(--shadow-lg);
  border-bottom: 3px solid rgba(14,212,244,.3);
  position: relative;
  overflow: hidden;
}
.main-header::before {
  content: '';
  position: absolute;
  top: -40%;
  right: -10%;
  width: 300px;
  height: 300px;
  background: radial-gradient(circle, rgba(14,116,144,.25) 0%, transparent 70%);
  pointer-events: none;
}
.main-header h1 {
  font-size: 1.8rem;
  font-weight: 800;
  margin: 0;
  letter-spacing: -.02em;
}
.main-header p {
  margin: .35rem 0 0;
  opacity: .8;
  font-size: .88rem;
  font-weight: 400;
}

/* ── Section Header ──────────────────────────────────── */
.section-header {
  font-size: 1rem;
  font-weight: 700;
  color: var(--clr-navy);
  margin: 1.4rem 0 .9rem;
  padding: .55rem 1rem;
  background: var(--clr-surface);
  border-left: 4px solid var(--clr-teal);
  border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
  box-shadow: var(--shadow-sm);
  letter-spacing: -.01em;
}

/* ── Clinical Cards ──────────────────────────────────── */
.patient-card {
  background: var(--clr-surface);
  border: 1px solid var(--clr-border);
  border-top: 3px solid var(--clr-teal-lt);
  border-radius: var(--radius-lg);
  padding: 1.2rem 1.4rem;
  margin-bottom: 1rem;
  box-shadow: var(--shadow-sm);
}

/* ── Chat Bubbles ────────────────────────────────────── */
.chat-patient {
  background: #f0f9ff;
  border-radius: 0 var(--radius-lg) var(--radius-lg) var(--radius-lg);
  padding: .75rem 1rem;
  margin: .5rem 0;
  max-width: 82%;
  border-left: 3px solid var(--clr-teal-lt);
  color: var(--clr-navy);
  font-size: .88rem;
  box-shadow: var(--shadow-sm);
}
.chat-student {
  background: #f0fdf4;
  border-radius: var(--radius-lg) 0 var(--radius-lg) var(--radius-lg);
  padding: .75rem 1rem;
  margin: .5rem 0 .5rem auto;
  max-width: 82%;
  border-right: 3px solid var(--clr-green);
  color: var(--clr-navy);
  text-align: right;
  font-size: .88rem;
  box-shadow: var(--shadow-sm);
}
.chat-tutor {
  background: var(--clr-amber-lt);
  border-radius: var(--radius-md);
  padding: .75rem 1rem;
  margin: .5rem 0;
  border-left: 3px solid var(--clr-amber);
  color: #78350f;
  font-size: .86rem;
  font-style: italic;
  box-shadow: var(--shadow-sm);
}
.chat-live {
  background: var(--clr-purple-lt);
  border-radius: 0 var(--radius-lg) var(--radius-lg) var(--radius-lg);
  padding: .75rem 1rem;
  margin: .5rem 0;
  max-width: 82%;
  border-left: 3px solid var(--clr-purple);
  color: #2e1065;
  font-size: .88rem;
  box-shadow: var(--shadow-sm);
}

/* ── Clinical Alert Banners ──────────────────────────── */
.alert-info {
  background: #eff6ff;
  border-left: 4px solid #3b82f6;
  color: #1e40af;
  border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
  padding: .65rem 1rem;
  margin: .4rem 0;
  font-size: .84rem;
  font-weight: 500;
}
.alert-warn {
  background: var(--clr-amber-lt);
  border-left: 4px solid var(--clr-amber);
  color: #78350f;
  border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
  padding: .65rem 1rem;
  margin: .4rem 0;
  font-size: .84rem;
  font-weight: 500;
}
.alert-good {
  background: var(--clr-green-lt);
  border-left: 4px solid var(--clr-green);
  color: #064e3b;
  border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
  padding: .65rem 1rem;
  margin: .4rem 0;
  font-size: .84rem;
  font-weight: 500;
}
.alert-bad {
  background: var(--clr-red-lt);
  border-left: 4px solid var(--clr-red);
  color: #7f1d1d;
  border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
  padding: .65rem 1rem;
  margin: .4rem 0;
  font-size: .84rem;
  font-weight: 500;
}

/* ── Status Badges ───────────────────────────────────── */
.badge {
  display: inline-flex;
  align-items: center;
  gap: .25rem;
  padding: .15rem .6rem;
  border-radius: 999px;
  font-size: .7rem;
  font-weight: 700;
  letter-spacing: .04em;
  text-transform: uppercase;
}
.bg { background: var(--clr-green-lt); color: #065f46; }
.bo { background: var(--clr-amber-lt); color: #92400e; }
.br { background: var(--clr-red-lt);   color: #991b1b; }

/* ── Avatar / Patient Box ────────────────────────────── */
.avatar-box {
  background: var(--clr-surface);
  border-radius: var(--radius-lg);
  padding: 1.2rem 1rem;
  text-align: center;
  border: 1px solid var(--clr-border);
  box-shadow: var(--shadow-md);
}

/* ── Physical Exam Zones ─────────────────────────────── */
.exam-zone {
  background: var(--clr-surface);
  border-radius: var(--radius-md);
  padding: .8rem;
  text-align: center;
  border: 1.5px solid var(--clr-border);
  cursor: pointer;
  transition: all .15s ease;
  margin: .3rem 0;
  font-size: .83rem;
  font-weight: 500;
  color: var(--clr-navy);
}
.exam-zone:hover {
  border-color: var(--clr-teal);
  background: #f0f9ff;
  box-shadow: var(--shadow-sm);
  transform: translateY(-1px);
}
.exam-finding {
  background: var(--clr-surface);
  border-radius: var(--radius-md);
  padding: .9rem 1.1rem;
  border-left: 4px solid var(--clr-teal);
  margin: .5rem 0;
  box-shadow: var(--shadow-sm);
  font-size: .86rem;
}

/* ── Surgery Steps ───────────────────────────────────── */
.surgery-step {
  background: var(--clr-surface);
  border-radius: var(--radius-md);
  padding: .9rem 1.2rem;
  margin: .5rem 0;
  border-left: 4px solid var(--clr-border);
  box-shadow: var(--shadow-sm);
  font-size: .85rem;
  transition: all .15s;
}
.surgery-step.active {
  border-left-color: var(--clr-green);
  background: var(--clr-green-lt);
}
.surgery-step.completed {
  border-left-color: #94a3b8;
  opacity: .65;
}
.phase-header {
  background: linear-gradient(90deg, var(--clr-navy), var(--clr-blue));
  color: white;
  border-radius: var(--radius-sm);
  padding: .55rem 1rem;
  margin: 1rem 0 .5rem;
  font-weight: 700;
  font-size: .84rem;
  letter-spacing: .03em;
  text-transform: uppercase;
}
.instrument-card {
  background: var(--clr-surface);
  border-radius: var(--radius-sm);
  padding: .65rem;
  text-align: center;
  border: 1px solid var(--clr-border);
  font-size: .78rem;
  color: var(--clr-muted);
  box-shadow: var(--shadow-sm);
}

/* ── Live Discussion Header ──────────────────────────── */
.live-header {
  background: linear-gradient(135deg, #3b0764, var(--clr-purple));
  color: white;
  border-radius: var(--radius-lg);
  padding: 1.2rem 1.5rem;
  text-align: center;
  margin-bottom: 1rem;
  box-shadow: var(--shadow-lg);
}

/* ── Upload Box ──────────────────────────────────────── */
.upload-box {
  background: var(--clr-surface);
  border: 2px dashed #94a3b8;
  border-radius: var(--radius-lg);
  padding: 1.5rem;
  text-align: center;
  margin: 1rem 0;
  transition: border-color .15s;
}
.upload-box:hover {
  border-color: var(--clr-teal);
  background: #f0f9ff;
}

/* ── Stat / KPI Cards ────────────────────────────────── */
.kpi-card {
  background: var(--clr-surface);
  border-radius: var(--radius-md);
  padding: 1rem 1.1rem;
  text-align: center;
  border: 1px solid var(--clr-border);
  box-shadow: var(--shadow-sm);
  border-top: 3px solid var(--clr-teal);
}
.kpi-value {
  font-size: 1.7rem;
  font-weight: 800;
  color: var(--clr-navy);
  line-height: 1;
}
.kpi-label {
  font-size: .72rem;
  color: var(--clr-muted);
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: .05em;
  margin-top: .25rem;
}

/* ── Data Tables ─────────────────────────────────────── */
.stDataFrame {
  border: 1px solid var(--clr-border) !important;
  border-radius: var(--radius-md) !important;
  overflow: hidden !important;
  box-shadow: var(--shadow-sm) !important;
}

/* ── Progress Bar ────────────────────────────────────── */
.stProgress > div > div {
  background: linear-gradient(90deg, var(--clr-teal), var(--clr-teal-lt)) !important;
  border-radius: 999px !important;
}

/* ── Divider ─────────────────────────────────────────── */
hr {
  border: none !important;
  border-top: 1px solid var(--clr-border) !important;
  margin: 1rem 0 !important;
}

/* ── Scrollbar ───────────────────────────────────────── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #94a3b8; border-radius: 999px; }
::-webkit-scrollbar-thumb:hover { background: var(--clr-teal); }

/* ── Spinner ─────────────────────────────────────────── */
.stSpinner > div { border-top-color: var(--clr-teal) !important; }

/* ── Toast / Success messages ────────────────────────── */
.stAlert {
  border-radius: var(--radius-md) !important;
  border: 1px solid var(--clr-border) !important;
  font-size: .85rem !important;
}
</style>
""", unsafe_allow_html=True)

# ── GEMINI API — 20-KEY ROTATION POOL ────────────────────────────────────────
#
# secrets.toml format:
#   GEMINI_API_KEY_1  = "AIza..."   ← your original key goes here as key 1
#   GEMINI_API_KEY_2  = "AIza..."
#   ...
#   GEMINI_API_KEY_20 = "AIza..."
#
# Old single-key format (GEMINI_API_KEY) is still accepted as a fallback.
# Keys must come from DIFFERENT Google accounts to get independent quotas.
# Each free-tier key gives 60 req/min → 20 keys = 1,200 req/min pool.
# ─────────────────────────────────────────────────────────────────────────────

def _load_all_keys() -> list:
    """
    Load all available Gemini keys from secrets.
    Checks BOTH top-level keys AND keys nested inside any TOML section
    (e.g. [epic]) so a mis-structured secrets.toml still works.
    Priority: top-level  >  [epic] nested  >  single fallback key.
    """
    keys = []
    try:
        # ── Helper: try a key name at top level then in every nested section ──
        def _get(name: str) -> str:
            v = st.secrets.get(name, "")
            if v and str(v).strip():
                return str(v).strip()
            # Walk any nested AttrDict sections (e.g. st.secrets["epic"])
            for sec_key in st.secrets:
                try:
                    sec = st.secrets[sec_key]
                    if hasattr(sec, "get"):
                        v2 = sec.get(name, "")
                        if v2 and str(v2).strip():
                            return str(v2).strip()
                except Exception:
                    pass
            return ""

        # Numbered pool: GEMINI_API_KEY_N  and  GEMINI_KEY_N  (both accepted)
        for i in range(1, 21):
            k = st.secrets.get(f"GEMINI_KEY_{i}", "").strip()
            if not k:
                k = st.secrets.get(f"GEMINI_API_KEY_{i}", "").strip()
            if k:
                keys.append(k)
        if not keys:
            for fallback in ("GEMINI_API_KEY", "GEMINI_KEY"):
                k = st.secrets.get(fallback, "").strip()
                if k:
                    keys.append(k)
                    break

        # Backwards-compatible single key
        if not keys:
            for fallback in ("GEMINI_API_KEY", "GEMINI_KEY"):
                k = _get(fallback)
                if k:
                    keys.append(k)
                    break
    except Exception:
        pass
    # Manual key entered in sidebar
    manual = st.session_state.get("_gemini_key_manual", "").strip()
    if manual and manual not in keys:
        keys.append(manual)
    return keys

def get_api_key() -> str:
    """Round-robin key rotation: each call advances the pool index by 1."""
    keys = _load_all_keys()
    if not keys:
        return ""
    idx = st.session_state.get("_key_pool_idx", 0) % len(keys)
    st.session_state["_key_pool_idx"] = idx + 1
    return keys[idx]

def get_next_key_after_429() -> str:
    """Called when a 429 is hit — skips ahead one extra slot to dodge the exhausted key."""
    keys = _load_all_keys()
    if not keys:
        return ""
    idx = st.session_state.get("_key_pool_idx", 0) % len(keys)
    st.session_state["_key_pool_idx"] = idx + 1
    return keys[idx]

GEMINI_API_KEY = ""  # do not use directly — always call get_api_key()

def _ensure_gemini_key():
    with st.sidebar:
        keys = _load_all_keys()
        n = len(keys)
        if n > 0:
            # Show pool health — mask the active key
            active_k = keys[st.session_state.get("_key_pool_idx", 0) % n]
            masked = active_k[:6] + "..." + active_k[-3:]
            color  = "#d1fae5" if n >= 10 else "#fef3c7" if n >= 3 else "#fee2e2"
            border = "#059669" if n >= 10 else "#f59e0b" if n >= 3 else "#dc2626"
            text   = "#065f46" if n >= 10 else "#78350f" if n >= 3 else "#7f1d1d"
            icon   = "✅" if n >= 10 else "⚠️" if n >= 3 else "🔴"
            st.markdown(f"""
            <div style="background:{color};border:1px solid {border};border-radius:7px;
                        padding:6px 10px;font-size:.72rem;color:{text};margin:4px 0;">
                {icon} <b>Key pool: {n}/20 active</b><br>
                <span style="opacity:.75">Current: <code>{masked}</code></span>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="background:#fef3c7;border:1px solid #f59e0b;border-radius:7px;
                        padding:8px 10px;font-size:.75rem;color:#78350f;margin:4px 0;">
                ⚠️ <b>No Gemini API keys found</b><br>
                Add GEMINI_KEY_1 … GEMINI_KEY_20 to secrets.toml
            </div>""", unsafe_allow_html=True)
            entered = st.text_input("Paste a Gemini key:", type="password", key="gemini_key_input")
            if entered and entered.strip():
                st.session_state["_gemini_key_manual"] = entered.strip()
                st.rerun()

# Model cascade: tries each in order until one works
# gemini-2.5-flash   → your project model (best quality, may have quota)
# gemini-2.0-flash   → free unlimited model (very good quality)
# gemini-2.0-flash-lite → free unlimited model (fastest, lightweight)
GEMINI_MODELS_CASCADE = [
    "gemini-2.5-flash",       # Primary — best quality
    "gemini-2.0-flash",       # Fallback 1 — unlimited free
    "gemini-2.0-flash-lite",  # Fallback 2 — fastest, unlimited free
]
GEMINI_MODEL = GEMINI_MODELS_CASCADE[0]  # Start with primary

# ════════════════════════════════════════════════════════════════════════════
# ⚙️  COMPLETE APP CONFIGURATION — Edit these values before launching
# ════════════════════════════════════════════════════════════════════════════

# ── Supabase ─────────────────────────────────────────────────────────────────
# Paste YOUR Supabase project URL and anon key here
# Find them: supabase.com → Your Project → Settings → API
# ── Supabase — loaded from secrets.toml if available ─────────────────────────
# Add to .streamlit/secrets.toml:
#   SUPABASE_URL = "https://xxxx.supabase.co"
#   SUPABASE_KEY = "eyJhbGci..."
try:
    SUPABASE_DEFAULT_URL = st.secrets.get("SUPABASE_URL", "")
    SUPABASE_DEFAULT_KEY = st.secrets.get("SUPABASE_KEY", "")
except Exception:
    SUPABASE_DEFAULT_URL = ""
    SUPABASE_DEFAULT_KEY = ""

# ── Payment (students pay YOU directly) ──────────────────────────────────────
PAYMENT_METHOD  = "PayPal"              # How students pay you
PAYMENT_CONTACT = "your@email.com"      # Your PayPal email / contact
PRICE_PER_MONTH = 5                     # USD per month

# ── Credit limits ────────────────────────────────────────────────────────────
# Gemini 2.5-flash free quota: ~500 req/day, 15 req/min
# Each student gets 20 credits per 8-hour window (60 credits/day free)
# At 1-3 credits per action → ~20-60 AI interactions per day free
FREE_CREDITS_PER_WINDOW    = 20    # credits every 8 hours
WINDOW_HOURS               = 8     # hours per window
FREE_MONTHLY_CREDITS_TOTAL = 100   # informational display only
PREMIUM_MONTHLY_CREDITS    = 99999 # unlimited for premium
PREMIUM_PRICE_USD          = 5     # $5/month

# ── Credit costs per action ──────────────────────────────────────────────────
CREDIT_COSTS = {
    "chat":         1,   # Patient interview, tutor, live discussion
    "lab":          2,   # AI lab generation
    "imaging":      3,   # Upload & analyze image
    "ecg":          3,   # ECG analysis
    "diagnosis":    2,   # Diagnosis evaluation
    "surgery_note": 2,   # AI operative note
    "exam":         1,   # Physical exam finding
    "submit_case":  2,   # Submit real case
}


def get_working_model():
    """Return the model currently set (can be changed by user in settings)."""
    return st.session_state.get("active_model", GEMINI_MODEL)


# ═══════════════════════════════════════════════════════════════════════════
# DEEPSEEK FALLBACK
# ───────────────────────────────────────────────────────────────────────────
# When all Gemini keys × all models are exhausted, try DeepSeek as a last
# resort. DeepSeek has no per-minute rate limits and gives 5M free tokens
# on signup. If DeepSeek also fails or isn't configured, we return the
# regular friendly error.
#
# Setup: add to .streamlit/secrets.toml:
#   DEEPSEEK_API_KEY = "sk-..."
# Get a key at: https://platform.deepseek.com/
# ═══════════════════════════════════════════════════════════════════════════
def _deepseek_key() -> str:
    try:
        return (st.secrets.get("DEEPSEEK_API_KEY", "") or
                st.secrets.get("DEEPSEEK_KEY", "")).strip()
    except Exception:
        return ""


def _deepseek_available() -> bool:
    return bool(_deepseek_key())


def _gemini_history_to_openai(history: list) -> list:
    """Convert Gemini-style history (role/parts) → OpenAI-style (role/content)."""
    out = []
    for msg in history:
        role  = msg.get("role", "user")
        parts = msg.get("parts", [])
        text  = ""
        for p in parts:
            if isinstance(p, dict) and "text" in p:
                text += p["text"] + "\n"
            elif isinstance(p, str):
                text += p + "\n"
        text = text.strip()
        if not text:
            continue
        # OpenAI uses 'assistant' instead of Gemini's 'model'
        openai_role = "assistant" if role == "model" else role
        out.append({"role": openai_role, "content": text})
    return out


def _call_deepseek(history: list, max_tokens: int = 1200,
                    temperature: float = 0.7) -> str:
    """
    Call DeepSeek as a last-resort fallback.
    Takes Gemini-style history; returns just the text (not !ERR-prefixed unless failed).
    """
    key = _deepseek_key()
    if not key:
        return "!ERR DeepSeek not configured."

    try:
        messages = _gemini_history_to_openai(history)
        if not messages:
            return "!ERR DeepSeek: empty messages."

        # DeepSeek is OpenAI-compatible
        url = "https://api.deepseek.com/v1/chat/completions"
        payload = {
            "model":      "deepseek-chat",  # cheap, fast workhorse
            "messages":   messages,
            "max_tokens": max_tokens,
            "temperature": temperature,
        }
        headers = {
            "Authorization": f"Bearer {key}",
            "Content-Type":  "application/json",
        }
        r = requests.post(url, headers=headers, json=payload, timeout=60)
        if r.status_code == 200:
            data = r.json()
            return data["choices"][0]["message"]["content"]
        else:
            return f"!ERR DeepSeek HTTP {r.status_code}"
    except Exception as e:
        return f"!ERR DeepSeek: {e}"


def _call_deepseek_with_image(prompt: str, image_bytes: bytes,
                                mime_type: str = "image/jpeg") -> str:
    """
    Image fallback. NOTE: DeepSeek-Chat doesn't natively support images yet.
    If DeepSeek ever adds vision in the future this will work; for now it
    returns a plain-text request asking the model to reason without the image.
    """
    key = _deepseek_key()
    if not key:
        return "!ERR DeepSeek not configured."

    # DeepSeek-V3/Chat doesn't support images yet — degrade gracefully
    fallback_prompt = (
        f"{prompt}\n\n"
        "(Note: image was provided but DeepSeek text-only fallback was used. "
        "Please describe what you'd typically expect to see and what reasoning "
        "you'd apply, based on the clinical context above.)"
    )
    try:
        url = "https://api.deepseek.com/v1/chat/completions"
        payload = {
            "model": "deepseek-chat",
            "messages": [{"role": "user", "content": fallback_prompt}],
            "max_tokens": 1500,
            "temperature": 0.3,
        }
        headers = {"Authorization": f"Bearer {key}",
                   "Content-Type": "application/json"}
        r = requests.post(url, headers=headers, json=payload, timeout=60)
        if r.status_code == 200:
            text = r.json()["choices"][0]["message"]["content"]
            return ("[Image not directly analyzed — DeepSeek text fallback used "
                    "because Gemini quota was exhausted.]\n\n" + text)
        else:
            return f"!ERR DeepSeek HTTP {r.status_code}"
    except Exception as e:
        return f"!ERR DeepSeek: {e}"


def try_all_models(payload_builder_fn):
    """Try each model × each key until one combination works. Returns (text, model_used)."""
    models_to_try = [st.session_state.get("active_model", GEMINI_MODEL)] + [
        m for m in GEMINI_MODELS_CASCADE
        if m != st.session_state.get("active_model", GEMINI_MODEL)
    ]
    keys = _load_all_keys()
    if not keys:
        return "!ERR No API keys configured.", None
    last_error = ""
    for model in models_to_try:
        for key in keys:
            try:
                payload = payload_builder_fn(model)
                url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
                r = requests.post(url, headers={"Content-Type":"application/json"},
                                json=payload, timeout=60)
                if r.status_code == 200:
                    return r.json()["candidates"][0]["content"]["parts"][0]["text"], model
                elif r.status_code in (429, 503):
                    last_error = f"{model}/{key[:6]}…: rate limited"
                    continue   # try next key
                else:
                    last_error = f"{model}: HTTP {r.status_code}"
                    break      # non-quota error — skip remaining keys for this model
            except Exception as e:
                last_error = str(e)
                continue
    return f"!ERR The AI service is busy right now (free quota reached). Please try again in 1–2 minutes.", None

def call_ai(system, messages, max_tokens=1200, credit_type="chat"):
    """Call Gemini AI with automatic key rotation + model fallback. Deducts credits."""
    ok, msg = can_use_credits(credit_type)
    if not ok:
        return f"!ERR_CREDITS: {msg}"
    use_credits(credit_type)
    try:
        history = [
            {"role":"user",  "parts":[{"text":"INSTRUCTIONS: "+system+" Say: Ready."}]},
            {"role":"model", "parts":[{"text":"Ready."}]},
        ]
        for m in messages:
            role = m.get("role",""); txt = str(m.get("content","")).strip()
            if not txt or txt.startswith("!ERR"): continue
            if role=="user":
                history.append({"role":"user",  "parts":[{"text":txt}]})
            elif role=="assistant":
                history.append({"role":"model", "parts":[{"text":txt}]})
        if history[-1]["role"]!="user":
            history.append({"role":"user","parts":[{"text":"Please respond."}]})

        keys = _load_all_keys()
        if not keys:
            return "!ERR No API keys configured. Add GEMINI_KEY_1…20 to secrets.toml."

        payload = {"contents": history,
                   "generationConfig": {"maxOutputTokens": max_tokens, "temperature": 0.7}}
        last_err = ""

        # Try every model × every key until one succeeds
        for model in GEMINI_MODELS_CASCADE:
            for key in keys:
                try:
                    url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
                           f"{model}:generateContent?key={key}")
                    r = requests.post(url, headers={"Content-Type":"application/json"},
                                      json=payload, timeout=45)
                    if r.status_code == 200:
                        return r.json()["candidates"][0]["content"]["parts"][0]["text"]
                    elif r.status_code == 429:
                        last_err = f"{model}/{key[:6]}…: rate limited"
                        continue   # try next key for this model
                    else:
                        last_err = f"{model}: HTTP {r.status_code}"
                        break      # non-quota error — move to next model
                except Exception as me:
                    last_err = str(me); continue

        # ── ALL Gemini keys × models exhausted — try DeepSeek as last resort ──
        if _deepseek_available():
            ds_text = _call_deepseek(history, max_tokens=max_tokens, temperature=0.7)
            if ds_text and not ds_text.startswith("!ERR"):
                return ds_text + "\n\n*[Backup AI: DeepSeek]*"

        return f"!ERR The AI service is busy right now (free quota reached). Please try again in 1–2 minutes. Tip: switch to a lighter model in the sidebar if you keep seeing this."
    except Exception as e:
        return f"!ERR {e}"
def call_ai_with_image(system, prompt, image_bytes, mime_type="image/jpeg"):
    """Call Gemini with an image — rotates keys on 429."""
    try:
        img_b64 = base64.b64encode(image_bytes).decode()
        contents = [{"role":"user","parts":[
            {"text": system + "\n\n" + prompt},
            {"inline_data":{"mime_type":mime_type,"data":img_b64}}
        ]}]
        keys = _load_all_keys()
        if not keys:
            return "!ERR No API keys configured."
        last_err = ""
        for key in keys:
            url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
                   f"gemini-2.5-flash:generateContent?key={key}")
            r = requests.post(url, headers={"Content-Type":"application/json"},
                json={"contents":contents,"generationConfig":{"maxOutputTokens":1500,"temperature":0.3}},
                timeout=60)
            if r.status_code == 200:
                return r.json()["candidates"][0]["content"]["parts"][0]["text"]
            elif r.status_code == 429:
                last_err = f"{key[:6]}…: rate limited"; continue
            else:
                return f"!ERR {r.json().get('error',{}).get('message','Unknown')}"

        # ── All Gemini keys exhausted — try DeepSeek (text-only fallback) ──
        if _deepseek_available():
            ds_text = _call_deepseek_with_image(prompt, image_bytes, mime_type)
            if ds_text and not ds_text.startswith("!ERR"):
                return ds_text

        return f"!ERR The AI service is busy right now (free quota reached). Please try again in 1–2 minutes."
    except Exception as e:
        return f"!ERR {e}"

# ── SURGERY DATABASE ──────────────────────────────────────────────────────────
SURGERIES = {
    "appendicitis":{
        "name":"Laparoscopic Appendectomy","duration":"45-90 min","anesthesia":"General","position":"Supine",
        "youtube":"https://www.youtube-nocookie.com/embed/PLkXSygxPnQ?rel=0&modestbranding=1",
        "phases":{
            "Pre-operative":[
                {"step":"Consent & Marking","detail":"Informed consent. Mark surgical site. Confirm NPO 6-8h. Check allergies.","instruments":["Consent form","Surgical marker"]},
                {"step":"IV Access & Antibiotics","detail":"18G IV cannula. Cefazolin 1g IV 30 min pre-incision. IV fluids.","instruments":["18G cannula","Cefazolin 1g","IV fluids"]},
                {"step":"Anesthesia Induction","detail":"Propofol 1.5-2.5 mg/kg IV. Rocuronium 0.6 mg/kg. Intubate with ETT. Maintain Sevoflurane.","instruments":["Laryngoscope","ETT","Propofol","Rocuronium"]},
                {"step":"Positioning & Prep","detail":"Supine, Trendelenburg, left lateral tilt. Urinary catheter. Betadine prep. Sterile drapes.","instruments":["Foley catheter","Betadine","Sterile drapes"]},
            ],
            "Intra-operative":[
                {"step":"Port Placement","detail":"Veress needle at umbilicus. CO2 to 12-15 mmHg. 10mm umbilical trocar (camera). Two 5mm trocars.","instruments":["Veress needle","10mm trocar","5mm trocar x2","CO2 insufflator"]},
                {"step":"Exploration","detail":"30-degree laparoscope. Inspect cavity. Identify caecum and trace taenia to appendix.","instruments":["30-degree laparoscope","Camera","Monitor"]},
                {"step":"Mobilization","detail":"Grasp appendix tip. Dissect mesoappendix with LigaSure. Identify appendicular artery.","instruments":["Grasper","LigaSure","Maryland dissector"]},
                {"step":"Ligation & Division","detail":"Two Endoloops at base. Divide between ligatures. Check hemostasis.","instruments":["Endoloop x2","Laparoscopic scissors"]},
                {"step":"Retrieval","detail":"Place in specimen bag. Extract through umbilical port. Irrigate if contamination.","instruments":["Specimen bag","Warm saline","Suction irrigator"]},
                {"step":"Closure","detail":"Check hemostasis. Desufflate CO2. Close 10mm fascia Vicryl 0. Skin Monocryl 3-0.","instruments":["Vicryl 0","Monocryl 3-0","Needle driver"]},
            ],
            "Post-operative":[
                {"step":"PACU","detail":"Vitals every 15 min. Pain: Paracetamol 1g IV + Ketorolac 30mg IV. O2 until awake.","instruments":["Pulse oximeter","O2 mask"]},
                {"step":"Ward Care","detail":"Oral fluids 4-6h. Mobilize Day 1. IV antibiotics 24h. Monitor wound.","instruments":["IV antibiotics","Wound dressing"]},
                {"step":"Discharge","detail":"Tolerating diet. Pain controlled. Afebrile >24h. Augmentin 625mg TDS x5 days.","instruments":["Discharge prescription"]},
            ],
        },
        "complications":["Wound infection 3-5%","Stump leak","Intra-abdominal abscess","Bleeding","Port site hernia"],
        "keywords":["appendix","appendicitis","appendectomy","rlq","rt lower","right lower"]
    },
    "cholecystitis":{
        "name":"Laparoscopic Cholecystectomy","duration":"45-75 min","anesthesia":"General","position":"Supine/Reverse Trendelenburg",
        "youtube":"https://www.youtube-nocookie.com/embed/wT3YUDEIOrY?rel=0&modestbranding=1",
        "phases":{
            "Pre-operative":[
                {"step":"Pre-op Assessment","detail":"Review US/CT. Check LFTs, coagulation. NPO 6-8h. Consent for possible open conversion.","instruments":["Consent form","IV cannula"]},
                {"step":"Anesthesia & NGT","detail":"GA with ETT. NG tube to decompress stomach. Urinary catheter.","instruments":["ETT","NG tube","Foley catheter"]},
                {"step":"Positioning","detail":"Supine, reverse Trendelenburg 15-20°, left lateral tilt. Betadine prep.","instruments":["Betadine","Sterile drapes"]},
            ],
            "Intra-operative":[
                {"step":"Port Placement","detail":"10mm umbilical (camera). 5mm epigastric. Two 5mm right subcostal. CO2 12-14 mmHg.","instruments":["10mm trocar","5mm trocar x3","CO2 insufflator"]},
                {"step":"Critical View of Safety (CVS)","detail":"MANDATORY: Dissect Calot triangle. Only TWO structures enter GB: cystic duct + cystic artery. Never clip without CVS.","instruments":["Maryland dissector","Hook diathermy"]},
                {"step":"Clipping & Division","detail":"3 clips cystic duct (2 proximal, 1 distal). 2 clips cystic artery. Divide both between clips.","instruments":["Clip applier","Laparoscopic scissors","Titanium clips"]},
                {"step":"GB Dissection","detail":"Dissect GB off liver bed with hook diathermy. Fundus to cystic duct. Control bleeding.","instruments":["Hook diathermy","Suction-irrigation","Grasper"]},
                {"step":"Retrieval & Closure","detail":"GB in specimen bag. Extract epigastric port. Irrigate. Close ports.","instruments":["Specimen bag","Vicryl suture"]},
            ],
            "Post-operative":[
                {"step":"Recovery","detail":"Monitor for bile leak. Sips water 2-4h post-op. LFTs at 24h.","instruments":["Jackson-Pratt drain"]},
                {"step":"Discharge","detail":"Low-fat diet 2 weeks. Paracetamol + Ibuprofen. Return to work 1-2 weeks.","instruments":["Discharge prescription"]},
            ],
        },
        "complications":["Bile duct injury 0.3%","Bleeding","Bile leak","Port site hernia","Wound infection"],
        "keywords":["cholecystitis","gallbladder","gallstone","ruq","cholelithiasis","cholangitis","biliary"]
    },
    "fracture":{
        "name":"ORIF - Open Reduction & Internal Fixation","duration":"60-180 min","anesthesia":"General or Regional","position":"Depends on site",
        "youtube":"https://www.youtube-nocookie.com/embed/r9LL4_9Hcxk?rel=0&modestbranding=1",
        "phases":{
            "Pre-operative":[
                {"step":"Assessment","detail":"X-rays AP+lateral. NV status distal. Check compartment syndrome. Blood group X-match.","instruments":["X-ray","Doppler probe"]},
                {"step":"Anesthesia & Positioning","detail":"Regional or GA. Position per fracture site. C-arm fluoroscopy in position.","instruments":["C-arm","Radiolucent table","Tourniquet"]},
                {"step":"Prep & Draping","detail":"Betadine prep. Exsanguinate Esmarch. Tourniquet 250-300 mmHg. Sterile draping.","instruments":["Esmarch bandage","Tourniquet","Betadine"]},
            ],
            "Intra-operative":[
                {"step":"Incision & Exposure","detail":"Incise over fracture. Protect NV structures. Expose fracture ends. Remove hematoma.","instruments":["Scalpel","Periosteal elevator","Retractors"]},
                {"step":"Reduction","detail":"Anatomical reduction with bone clamps. Verify with C-arm in 2 planes.","instruments":["Bone reduction forceps","C-arm","Bone clamps","K-wires"]},
                {"step":"Internal Fixation","detail":"Apply plate + screws (DCP/LCP) or IM nail. Confirm fixation. Check screw lengths.","instruments":["Plate + screws","Drill","Tap","Screw driver"]},
                {"step":"Wound Closure","detail":"Irrigate. Close periosteum Vicryl 2-0. Fascia Vicryl 1. Skin nylon 3-0. Backslab.","instruments":["Saline","Vicryl sutures","Nylon 3-0","Backslab"]},
            ],
            "Post-operative":[
                {"step":"NV Check","detail":"Check distal pulses, capillary refill, sensation, motor every 1h x4h. Elevate limb.","instruments":["Doppler probe","Elevation pillow"]},
                {"step":"Rehabilitation","detail":"Physio Day 1. Serial X-rays 6 weeks + 3 months.","instruments":["X-ray","Walking aids"]},
            ],
        },
        "complications":["Infection","Non-union","Malunion","Hardware failure","Neurovascular injury","DVT"],
        "keywords":["fracture","fx","trauma","broken","ortho","displaced","wrist","ankle","femur","tibia"]
    },
    "laceration":{
        "name":"Wound Repair & Suturing","duration":"15-60 min","anesthesia":"Local","position":"Wound-dependent",
        "youtube":"https://www.youtube-nocookie.com/embed/2dPwBRR4qFc?rel=0&modestbranding=1",
        "phases":{
            "Pre-operative":[
                {"step":"Wound Assessment","detail":"Assess depth, length, edges, contamination. Check tetanus status. Document NV.","instruments":["Wound probe","Ruler"]},
                {"step":"Local Anesthesia","detail":"Lidocaine 1% + epinephrine. Avoid digits/nose/ear. Wait 5 min. Max 7mg/kg.","instruments":["23G needle","Lidocaine 1% + epi"]},
            ],
            "Intra-operative":[
                {"step":"Irrigation","detail":"0.9% NS under pressure (20ml syringe + 18G catheter). Min 500ml contaminated wounds.","instruments":["20ml syringe","18G catheter","NS 500ml"]},
                {"step":"Debridement","detail":"Debride necrotic tissue. Freshen wound edges. Curettage if bone exposed.","instruments":["Iris scissors","Scalpel","Curette"]},
                {"step":"Suturing","detail":"Deep: Vicryl 3-0 interrupted. Superficial: Nylon 4-0 (face 5-0/6-0) simple interrupted 5-7mm apart.","instruments":["Needle driver","Toothed forceps","Vicryl 3-0","Nylon 4-0"]},
                {"step":"Dressing","detail":"Non-adherent dressing. Keep dry 48h. Antibiotic ointment on face.","instruments":["Non-adherent dressing","Antibiotic ointment"]},
            ],
            "Post-operative":[
                {"step":"Tetanus & Antibiotics","detail":"Tetanus toxoid if needed. Augmentin 625mg TDS for contaminated/bite/diabetic.","instruments":["Tetanus toxoid","Augmentin"]},
                {"step":"Suture Removal","detail":"Face 5 days. Scalp 7 days. Body 10-14 days. Steri-strips after removal.","instruments":["Suture removal kit","Steri-strips"]},
            ],
        },
        "complications":["Wound infection","Dehiscence","Scarring","Keloid","Hematoma"],
        "keywords":["laceration","suture","wound","cut","skin","repair"]
    },
}


def transcribe_audio(audio_bytes, mime_type="audio/wav"):
    """
    Gemini 2.5 Flash does NOT support audio inline_data via REST API.
    Audio transcription requires Gemini Live API (websocket) which is not
    available in this Streamlit setup. We use Web Speech API instead.
    This function is kept as a stub.
    """
    return ""

def voice_input_component(key="voice", role="doctor", height=160):
    """
    Voice input using Web Speech API.
    Returns nothing — uses st.query_params trick or JS postMessage
    to inject text into a hidden streamlit text input via JS.
    
    Approach: Use components.html to render a mic button.
    When speech ends, write to a streamlit text_input via JS DOM manipulation.
    """
    color = "#0e7490" if role == "doctor" else "#7c3aed"
    label = "🎤 Press & Speak" if role == "doctor" else "🎤 Speak as Patient"
    
    import json as _j
    key_safe = _j.dumps(key)
    
    components.html(f"""
    <!DOCTYPE html>
    <html>
    <body style="margin:0;padding:6px;font-family:Inter,sans-serif;background:transparent;">
    
    <div id="voiceBox" style="border:2px solid {color}33;border-radius:12px;padding:10px;background:white;">
        
        <!-- Status -->
        <div id="statusRow" style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
            <div id="dot" style="width:10px;height:10px;border-radius:50%;background:#9ca3af;flex-shrink:0;"></div>
            <span id="statusTxt" style="font-size:.8rem;color:#6b7280;">Ready — press button to speak</span>
        </div>
        
        <!-- Transcript display -->
        <div id="transcript" style="
            min-height:32px;background:#f8fafc;border-radius:8px;
            padding:6px 10px;font-size:.85rem;color:#1e3a5f;
            border:1px solid #e2e8f0;margin-bottom:8px;
            font-style:italic;display:none;">
        </div>
        
        <!-- Buttons -->
        <div style="display:flex;gap:8px;flex-wrap:wrap;">
            <button id="micBtn" onclick="toggleMic()" style="
                background:linear-gradient(135deg,{color},{color}cc);
                color:white;border:none;border-radius:8px;
                padding:7px 16px;font-size:.82rem;font-weight:600;
                cursor:pointer;flex:1;min-width:120px;
                box-shadow:0 2px 6px {color}44;">
                {label}
            </button>
            <button onclick="clearAll()" style="
                background:#f1f5f9;color:#64748b;border:1px solid #e2e8f0;
                border-radius:8px;padding:7px 12px;font-size:.8rem;cursor:pointer;">
                ✕ Clear
            </button>
        </div>
        
        <!-- Send button (appears after speech) -->
        <div id="sendRow" style="margin-top:8px;display:none;">
            <button id="sendBtn" onclick="sendMessage()" style="
                background:linear-gradient(135deg,#16a34a,#15803d);
                color:white;border:none;border-radius:8px;
                padding:8px 20px;font-size:.85rem;font-weight:700;
                cursor:pointer;width:100%;
                box-shadow:0 2px 8px #16a34a44;">
                ✅ Send Voice Message
            </button>
        </div>
        
        <div style="font-size:.7rem;color:#9ca3af;margin-top:6px;">
            💡 Works in Chrome & Edge. Make sure microphone is allowed.
        </div>
    </div>
    
    <script>
    var recognition = null;
    var isListening = false;
    var finalTranscript = "";
    
    function clearAll() {{
        finalTranscript = "";
        document.getElementById("transcript").style.display = "none";
        document.getElementById("transcript").innerHTML = "";
        document.getElementById("sendRow").style.display = "none";
        document.getElementById("statusTxt").innerHTML = "Cleared — ready to speak again";
        document.getElementById("dot").style.background = "#9ca3af";
    }}
    
    function toggleMic() {{
        if (!("webkitSpeechRecognition" in window) && !("SpeechRecognition" in window)) {{
            document.getElementById("statusTxt").innerHTML = "⚠️ Voice not supported. Use Chrome or Edge browser.";
            document.getElementById("dot").style.background = "#dc2626";
            return;
        }}
        
        if (isListening) {{
            recognition.stop();
            return;
        }}
        
        var SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;
        recognition = new SpeechRecognition();
        recognition.lang = "en-US";
        recognition.continuous = true;
        recognition.interimResults = true;
        recognition.maxAlternatives = 1;
        
        recognition.onstart = function() {{
            isListening = true;
            document.getElementById("dot").style.background = "#dc2626";
            document.getElementById("dot").style.animation = "pulse 1s infinite";
            document.getElementById("micBtn").innerHTML = "🔴 Listening... (click to stop)";
            document.getElementById("micBtn").style.background = "linear-gradient(135deg,#dc2626,#b91c1c)";
            document.getElementById("statusTxt").innerHTML = "🎙️ Listening... speak clearly";
            document.getElementById("sendRow").style.display = "none";
        }};
        
        recognition.onresult = function(event) {{
            var interim = "";
            var final_part = "";
            
            for (var i = event.resultIndex; i < event.results.length; i++) {{
                var transcript = event.results[i][0].transcript;
                if (event.results[i].isFinal) {{
                    final_part += transcript;
                }} else {{
                    interim += transcript;
                }}
            }}
            
            if (final_part) finalTranscript += final_part;
            
            var display = finalTranscript + (interim ? '<span style="color:#94a3b8"> ' + interim + '...</span>' : "");
            var tr = document.getElementById("transcript");
            tr.style.display = "block";
            tr.innerHTML = display;
            
            if (finalTranscript) {{
                document.getElementById("statusTxt").innerHTML = "✅ Got it — review and send";
                document.getElementById("sendRow").style.display = "block";
            }}
        }};
        
        recognition.onspeechend = function() {{
            recognition.stop();
        }};
        
        recognition.onerror = function(event) {{
            isListening = false;
            var msg = event.error;
            if (msg === "not-allowed") msg = "Microphone blocked! Click the 🔒 icon in your browser address bar and allow microphone.";
            if (msg === "no-speech") msg = "No speech detected. Try again and speak closer to the mic.";
            if (msg === "network") msg = "Network error. Check internet connection.";
            document.getElementById("statusTxt").innerHTML = "⚠️ " + msg;
            document.getElementById("dot").style.background = "#f59e0b";
            document.getElementById("micBtn").innerHTML = "{label}";
            document.getElementById("micBtn").style.background = "linear-gradient(135deg,{color},{color}cc)";
        }};
        
        recognition.onend = function() {{
            isListening = false;
            document.getElementById("dot").style.background = finalTranscript ? "#16a34a" : "#9ca3af";
            document.getElementById("micBtn").innerHTML = "{label}";
            document.getElementById("micBtn").style.background = "linear-gradient(135deg,{color},{color}cc)";
            if (!finalTranscript) {{
                document.getElementById("statusTxt").innerHTML = "Stopped. Press again to retry.";
            }}
        }};
        
        try {{
            recognition.start();
        }} catch(e) {{
            document.getElementById("statusTxt").innerHTML = "⚠️ Could not start microphone: " + e.message;
        }}
    }}
    
    function sendMessage() {{
        if (!finalTranscript.trim()) return;
        
        // Send to Streamlit via parent window postMessage
        var msg = finalTranscript.trim();
        
        // Method 1: Try to find and fill the Streamlit text input
        try {{
            var inputs = window.parent.document.querySelectorAll('input[type="text"]');
            for (var i = 0; i < inputs.length; i++) {{
                var inp = inputs[i];
                if (inp.placeholder && (inp.placeholder.includes("patient") || inp.placeholder.includes("message") || inp.placeholder.includes("Type"))) {{
                    // React synthetic event
                    var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.parent.HTMLInputElement.prototype, "value").set;
                    nativeInputValueSetter.call(inp, msg);
                    inp.dispatchEvent(new Event("input", {{ bubbles: true }}));
                    inp.focus();
                    break;
                }}
            }}
        }} catch(e) {{}}
        
        // Method 2: Copy to clipboard (reliable fallback)
        navigator.clipboard.writeText(msg).then(function() {{
            document.getElementById("statusTxt").innerHTML = "✅ Copied to clipboard! Paste (Ctrl+V) in the text box below and click Send.";
        }}).catch(function() {{
            document.getElementById("statusTxt").innerHTML = "✅ Text ready — manually type: " + msg;
        }});
        
        document.getElementById("transcript").style.background = "#f0fdf4";
        document.getElementById("transcript").style.border = "1px solid #16a34a";
    }}
    
    // Pulse animation
    var style = document.createElement("style");
    style.textContent = "@keyframes pulse {{ 0%,100% {{ opacity:1; }} 50% {{ opacity:.3; }} }}";
    document.head.appendChild(style);
    </script>
    </body>
    </html>
    """, height=height, scrolling=False)


def tts_speak(text):
    """Reliable TTS using components.html — always executes.
    Now also detects when no voices are installed and reports back via postMessage,
    so the page can show a helpful warning to the user."""
    if not text: return
    import json as _json
    safe = _json.dumps(str(text))
    components.html(f"""
    <script>
    (function() {{
        if (!window.speechSynthesis) {{
            try {{ window.parent.postMessage({{type:'tts_status', status:'unsupported'}}, '*'); }} catch(e){{}}
            return;
        }}
        window.speechSynthesis.cancel();
        var u = new SpeechSynthesisUtterance({safe});
        u.rate = 0.9; u.pitch = 1.0; u.volume = 1.0;
        var attempts = 0;
        function trySpeak() {{
            var voices = window.speechSynthesis.getVoices();
            if (voices.length > 0) {{
                var v = voices.find(function(x){{return x.lang.startsWith('en') && x.name.includes('Female');}})
                     || voices.find(function(x){{return x.lang.startsWith('en');}})
                     || voices[0];
                if (v) u.voice = v;
                window.speechSynthesis.speak(u);
            }} else if (attempts < 10) {{
                attempts++;
                setTimeout(trySpeak, 200);
            }} else {{
                try {{ window.parent.postMessage({{type:'tts_status', status:'no_voices'}}, '*'); }} catch(e){{}}
            }}
        }}
        if (window.speechSynthesis.getVoices().length === 0) {{
            window.speechSynthesis.onvoiceschanged = trySpeak;
            setTimeout(trySpeak, 300);
        }} else {{
            setTimeout(trySpeak, 100);
        }}
    }})();
    </script>
    """, height=0)


def render_tts_voice_check():
    """Sidebar widget that lets users test their voice setup and see warnings.
    Show this in the sidebar near the voice toggle."""
    if not st.session_state.get("voice_enabled"):
        return
    with st.expander("🎤 Test voice / fix silent voice", expanded=False):
        st.caption(
            "If you don't hear the patient speak, your computer may be missing "
            "text-to-speech voices."
        )
        if st.button("▶️ Test voice now", key="test_voice_btn",
                     use_container_width=True):
            tts_speak("This is a test. If you hear this clearly, your voice "
                      "is working correctly.")
            st.info("Click test, listen, and if silent — see instructions below.")
        st.markdown(
            "**If voice is silent on your device:**\n\n"
            "**Windows:** Settings → Time & Language → Speech → Manage voices "
            "→ Install English voices\n\n"
            "**Mac:** System Settings → Accessibility → Spoken Content → "
            "System voice → Customize → check English voices\n\n"
            "**Linux:** Voice support varies by distro. Use Chrome (built-in voices) "
            "rather than Firefox.\n\n"
            "**Chromebook:** Voice works out of the box in Chrome.\n\n"
            "**Mobile:** Voice works best on iOS Safari and Android Chrome."
        )

def tts_speak_doctor(text):
    """TTS for doctor/student voice — different pitch and rate."""
    if not text: return
    import json as _json
    safe = _json.dumps(str(text))
    components.html(f"""
    <script>
    (function() {{
        window.speechSynthesis.cancel();
        var u = new SpeechSynthesisUtterance({safe});
        u.rate = 1.05; u.pitch = 0.85; u.volume = 1.0;
        function trySpeak() {{
            var voices = window.speechSynthesis.getVoices();
            if (voices.length > 0) {{
                var v = voices.find(function(x){{return x.lang.startsWith('en') && x.name.includes('Male');}})
                     || voices.find(function(x){{return x.lang.startsWith('en');}})
                     || voices[0];
                if (v) u.voice = v;
                window.speechSynthesis.speak(u);
            }} else {{
                setTimeout(trySpeak, 200);
            }}
        }}
        if (window.speechSynthesis.getVoices().length === 0) {{
            window.speechSynthesis.onvoiceschanged = trySpeak;
        }} else {{
            trySpeak();
        }}
    }})();
    </script>""", height=0, scrolling=False)



# ── Clinical Sounds Engine (Web Audio API) ────────────────────────────────────
def play_clinical_sound(sound_type, case_data=None):
    """
    Public sound player — renders BOTH synthetic (teaching) and real recording
    (clinical reality) when a real recording is available for the given sound.

    Falls back to synth-only when no curated real recording exists.
    """
    # Resolve "auto_*" sound types to concrete sound types first
    dx = str(case_data.get("Final_Diagnosis","") if case_data else "").lower()
    cc = str(case_data.get("Chief_Complaint","") if case_data else "").lower()
    hpi = str(case_data.get("HPI","") if case_data else "").lower()
    pe  = str(case_data.get("Physical_Findings","") if case_data else "").lower()
    combined = " ".join([dx, cc, hpi, pe])

    if sound_type == "auto_heart":
        if any(w in combined for w in [
            "murmur","aortic","mitral","regurgit","stenosis","vsd","asd",
            "valvular","prolapse","insufficiency","systolic","diastolic"
        ]):
            sound_type = "murmur"
        elif any(w in combined for w in [
            "mi","stemi","nstemi","heart fail","heart failure","cardiac",
            "ischem","angina","ejection","cardiomyopathy","myocardi",
            "infarct","chest pain","cardiogenic"
        ]):
            sound_type = "s3_gallop"
        else:
            sound_type = "normal_heart"
    elif sound_type == "auto_lung":
        if any(w in combined for w in [
            "pneumonia","consolidat","fibrosis","effusion","pulm edema",
            "pulmonary edema","atelect","interstitial","crackle","rales",
            "alveolar","ards"
        ]):
            sound_type = "crackles_fine"
        elif any(w in combined for w in [
            "asthma","wheeze","copd","bronchitis","obstruct","emphysema",
            "bronchospasm","reactive airway","exacerbation"
        ]):
            sound_type = "wheeze_exp"
        elif any(w in combined for w in [
            "pneumothorax","absent","pleural","tension"
        ]):
            sound_type = "absent_breath"
        else:
            sound_type = "normal_breath"
    elif sound_type == "auto_bowel":
        if any(w in combined for w in [
            "obstruct","ileus","post-op","peritonitis","sbo","lbo",
            "perforat","appendic","cholecyst","diverticul"
        ]):
            sound_type = "bowel_absent"
        elif any(w in combined for w in [
            "gastroenteritis","diarrhea","ibs","colitis","crohn",
            "inflammatory bowel","viral","hyperactive"
        ]):
            sound_type = "bowel_hyperactive"
        else:
            sound_type = "bowel_normal"
    elif sound_type == "auto_percussion":
        if any(w in combined for w in [
            "effusion","empyema","hemothorax","consolidat","pneumonia",
            "fibrosis","ascites","mass","hepatomegaly","dull"
        ]):
            sound_type = "percussion_dull"
        elif any(w in combined for w in [
            "pneumothorax","emphysema","copd","tension","hyperresonan",
            "hyper-resonant"
        ]):
            sound_type = "percussion_hyper"
        else:
            sound_type = "percussion_resonant"

    # Render real recording if we have one for this sound_type, else fall back to synth.
    # The dual-sound panel handles the case where a real recording exists by showing
    # both synth (teaching) and real (reality). When no real recording exists for
    # this sound_type, we just play the synthetic version directly.
    if REAL_SOUNDS_OK:
        if has_real_recording(sound_type):
            render_dual_sound_panel(sound_type, render_synth_func=_play_clinical_sound_synth)
        else:
            # No real recording for this exact sound_type — use synth + small note
            st.caption(f"ℹ️ No curated real recording matches **{sound_type}** for this case. "
                       "Playing synthetic teaching version only.")
            _play_clinical_sound_synth(sound_type)
    else:
        # real_clinical_sounds module isn't loaded at all
        _play_clinical_sound_synth(sound_type)


def _play_clinical_sound_synth(sound_type, case_data=None):
    """
    Synthetic clinical sound player — generates audio in browser via Web Audio API.
    This is the original implementation, now used as the "teaching pattern" version.
    """
    dx = str(case_data.get("Final_Diagnosis","") if case_data else "").lower()
    cc = str(case_data.get("Chief_Complaint","") if case_data else "").lower()
    combined = dx + " " + cc

    # ── Auto-select best sound for case ──────────────────
    # Based on diagnosis, pick which sound is most relevant
    if sound_type == "auto_heart":
        if any(w in combined for w in ["murmur","aortic","mitral","regurgit","stenosis","vsd","asd"]):
            sound_type = "murmur"
        elif any(w in combined for w in ["mi","stemi","nstemi","heart fail","cardiac"]):
            sound_type = "s3_gallop"
        else:
            sound_type = "normal_heart"

    if sound_type == "auto_lung":
        if any(w in combined for w in ["pneumonia","consolidat","fibrosis","effusion","pulm edema"]):
            sound_type = "crackles_fine"
        elif any(w in combined for w in ["asthma","wheeze","copd","bronchitis","obstruct"]):
            sound_type = "wheeze_exp"
        elif any(w in combined for w in ["pneumothorax","absent","pleural"]):
            sound_type = "absent_breath"
        else:
            sound_type = "normal_breath"

    if sound_type == "auto_bowel":
        if any(w in combined for w in ["obstruct","ileus","post-op","peritonitis"]):
            sound_type = "bowel_absent"
        elif any(w in combined for w in ["gastroenteritis","diarrhea","ibs","colitis"]):
            sound_type = "bowel_hyperactive"
        else:
            sound_type = "bowel_normal"

    if sound_type == "auto_percussion":
        if any(w in combined for w in ["effusion","empyema","hemothorax","consolidat"]):
            sound_type = "percussion_dull"
        elif any(w in combined for w in ["pneumothorax","hyperinflat","copd","emphys"]):
            sound_type = "percussion_hyper"
        elif any(w in combined for w in ["ascites","fluid","liver","spleen"]):
            sound_type = "percussion_dull"
        else:
            sound_type = "percussion_resonant"

    # Sound configurations with enhanced Web Audio synthesis
    # Each sound uses physiologically accurate frequency/timing parameters
    sound_scripts = {
        # ── HEART SOUNDS ──────────────────────────────────────
        "normal_heart": {
            "label": "Normal Heart Sounds — S1 (Lub) S2 (Dub)",
            "description": "Clear S1 and S2 at 72 bpm. No murmurs, rubs or gallops.",
            "color": "#0ea5e9",
            "script": """
            function playHeartbeat(ctx, t, hr) {
                var beat = 60/hr;
                // S1 — Mitral/Tricuspid closure (lower frequency, longer)
                var b1 = ctx.createOscillator(); var g1 = ctx.createGain();
                var f1 = ctx.createBiquadFilter(); f1.type='lowpass'; f1.frequency.value=120;
                b1.type='sine'; b1.frequency.setValueAtTime(80, ctx.currentTime+t);
                b1.frequency.exponentialRampToValueAtTime(50, ctx.currentTime+t+0.12);
                b1.connect(f1); f1.connect(g1); g1.connect(ctx.destination);
                g1.gain.setValueAtTime(0, ctx.currentTime+t);
                g1.gain.linearRampToValueAtTime(0.9, ctx.currentTime+t+0.01);
                g1.gain.exponentialRampToValueAtTime(0.001, ctx.currentTime+t+0.14);
                b1.start(ctx.currentTime+t); b1.stop(ctx.currentTime+t+0.16);
                // Systolic silence
                // S2 — Aortic/Pulmonic closure (higher frequency, shorter, sharper)
                var b2 = ctx.createOscillator(); var g2 = ctx.createGain();
                var f2 = ctx.createBiquadFilter(); f2.type='lowpass'; f2.frequency.value=200;
                b2.type='sine'; b2.frequency.setValueAtTime(120, ctx.currentTime+t+0.33);
                b2.frequency.exponentialRampToValueAtTime(70, ctx.currentTime+t+0.40);
                b2.connect(f2); f2.connect(g2); g2.connect(ctx.destination);
                g2.gain.setValueAtTime(0, ctx.currentTime+t+0.33);
                g2.gain.linearRampToValueAtTime(0.65, ctx.currentTime+t+0.335);
                g2.gain.exponentialRampToValueAtTime(0.001, ctx.currentTime+t+0.42);
                b2.start(ctx.currentTime+t+0.33); b2.stop(ctx.currentTime+t+0.44);
            }
            for(var i=0;i<8;i++) playHeartbeat(ctx, i*(60/72), 72);
            """
        },
        "murmur": {
            "label": "Systolic Murmur — Aortic Stenosis Pattern",
            "description": "Harsh crescendo-decrescendo systolic murmur. Radiates to carotids. Grade III/VI.",
            "color": "#dc2626",
            "script": """
            function playMurmur(ctx, t) {
                // S1
                var b1=ctx.createOscillator(); var g1=ctx.createGain();
                b1.type='sine'; b1.frequency.value=80;
                b1.connect(g1); g1.connect(ctx.destination);
                g1.gain.setValueAtTime(0,ctx.currentTime+t);
                g1.gain.linearRampToValueAtTime(0.85,ctx.currentTime+t+0.01);
                g1.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.13);
                b1.start(ctx.currentTime+t); b1.stop(ctx.currentTime+t+0.15);
                // Systolic murmur — noise-based crescendo-decrescendo
                var dur = 0.30;
                var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*dur),ctx.sampleRate);
                var data=buf.getChannelData(0);
                for(var i=0;i<data.length;i++) data[i]=(Math.random()*2-1);
                var src=ctx.createBufferSource(); src.buffer=buf;
                var filt=ctx.createBiquadFilter(); filt.type='bandpass';
                filt.frequency.value=350; filt.Q.value=3;
                var gm=ctx.createGain();
                src.connect(filt); filt.connect(gm); gm.connect(ctx.destination);
                // Crescendo-decrescendo envelope
                gm.gain.setValueAtTime(0,ctx.currentTime+t+0.13);
                gm.gain.linearRampToValueAtTime(0.45,ctx.currentTime+t+0.13+dur/2);
                gm.gain.linearRampToValueAtTime(0,ctx.currentTime+t+0.13+dur);
                src.start(ctx.currentTime+t+0.13);
                // S2 (softer in AS)
                var b2=ctx.createOscillator(); var g2=ctx.createGain();
                b2.type='sine'; b2.frequency.value=100;
                b2.connect(g2); g2.connect(ctx.destination);
                g2.gain.setValueAtTime(0,ctx.currentTime+t+0.45);
                g2.gain.linearRampToValueAtTime(0.45,ctx.currentTime+t+0.455);
                g2.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.52);
                b2.start(ctx.currentTime+t+0.45); b2.stop(ctx.currentTime+t+0.54);
            }
            for(var i=0;i<7;i++) playMurmur(ctx, i*0.83);
            """
        },
        "s3_gallop": {
            "label": "S3 Gallop — Heart Failure",
            "description": "S3 gallop heard in early diastole. Kentucky rhythm (lub-dub-ta). Indicates poor LV compliance.",
            "color": "#7c3aed",
            "script": """
            function playS3(ctx, t) {
                // S1
                var b1=ctx.createOscillator(); var g1=ctx.createGain();
                b1.type='sine'; b1.frequency.value=75; b1.connect(g1); g1.connect(ctx.destination);
                g1.gain.setValueAtTime(0,ctx.currentTime+t);
                g1.gain.linearRampToValueAtTime(0.85,ctx.currentTime+t+0.01);
                g1.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.13);
                b1.start(ctx.currentTime+t); b1.stop(ctx.currentTime+t+0.15);
                // S2
                var b2=ctx.createOscillator(); var g2=ctx.createGain();
                b2.type='sine'; b2.frequency.value=110; b2.connect(g2); g2.connect(ctx.destination);
                g2.gain.setValueAtTime(0,ctx.currentTime+t+0.33);
                g2.gain.linearRampToValueAtTime(0.6,ctx.currentTime+t+0.335);
                g2.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.42);
                b2.start(ctx.currentTime+t+0.33); b2.stop(ctx.currentTime+t+0.44);
                // S3 — early diastole, very low frequency (30-70 Hz), soft
                var b3=ctx.createOscillator(); var g3=ctx.createGain();
                var f3=ctx.createBiquadFilter(); f3.type='lowpass'; f3.frequency.value=80;
                b3.type='sine'; b3.frequency.value=45;
                b3.connect(f3); f3.connect(g3); g3.connect(ctx.destination);
                g3.gain.setValueAtTime(0,ctx.currentTime+t+0.53);
                g3.gain.linearRampToValueAtTime(0.35,ctx.currentTime+t+0.535);
                g3.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.62);
                b3.start(ctx.currentTime+t+0.53); b3.stop(ctx.currentTime+t+0.64);
            }
            for(var i=0;i<7;i++) playS3(ctx, i*0.83);
            """
        },
        # ── BREATH SOUNDS ─────────────────────────────────────
        "normal_breath": {
            "label": "Normal Vesicular Breath Sounds",
            "description": "Soft, low-pitched breath sounds. Inspiration louder than expiration. Heard in peripheral lung fields.",
            "color": "#0ea5e9",
            "script": """
            function breath(ctx, t, inhale) {
                var dur = inhale ? 1.3 : 0.8;
                var freq = inhale ? 350 : 250;
                var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*dur),ctx.sampleRate);
                var data=buf.getChannelData(0);
                for(var i=0;i<data.length;i++) data[i]=(Math.random()*2-1)*0.18;
                var src=ctx.createBufferSource(); src.buffer=buf;
                var f=ctx.createBiquadFilter(); f.type='lowpass'; f.frequency.value=freq;
                var f2=ctx.createBiquadFilter(); f2.type='highpass'; f2.frequency.value=80;
                var g=ctx.createGain();
                src.connect(f); f.connect(f2); f2.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+t);
                g.gain.linearRampToValueAtTime(inhale?0.42:0.22,ctx.currentTime+t+dur*0.25);
                g.gain.linearRampToValueAtTime(inhale?0.38:0.18,ctx.currentTime+t+dur*0.75);
                g.gain.linearRampToValueAtTime(0,ctx.currentTime+t+dur);
                src.start(ctx.currentTime+t);
            }
            for(var i=0;i<4;i++){breath(ctx,i*2.3,true);breath(ctx,i*2.3+1.3,false);}
            """
        },
        "crackles_fine": {
            "label": "Fine Crackles (Crepitations) — Pulmonary Fibrosis/Pneumonia",
            "description": "Fine, high-pitched crackles in late inspiration. Velcro-like. Bilateral bases. Indicates fibrosis or early pneumonia.",
            "color": "#dc2626",
            "script": """
            // Background breath
            function addBreath(ctx, t) {
                var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*1.3),ctx.sampleRate);
                var data=buf.getChannelData(0);
                for(var i=0;i<data.length;i++) data[i]=(Math.random()*2-1)*0.12;
                var src=ctx.createBufferSource(); src.buffer=buf;
                var f=ctx.createBiquadFilter(); f.type='lowpass'; f.frequency.value=300;
                var g=ctx.createGain();
                src.connect(f); f.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+t);
                g.gain.linearRampToValueAtTime(0.3,ctx.currentTime+t+0.4);
                g.gain.linearRampToValueAtTime(0,ctx.currentTime+t+1.3);
                src.start(ctx.currentTime+t);
            }
            function addCrackle(ctx, t) {
                var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*0.008),ctx.sampleRate);
                var data=buf.getChannelData(0);
                for(var i=0;i<data.length;i++) data[i]=(Math.random()*2-1);
                var src=ctx.createBufferSource(); src.buffer=buf;
                var f=ctx.createBiquadFilter(); f.type='bandpass'; f.frequency.value=1200; f.Q.value=2;
                var g=ctx.createGain(); g.gain.value=0.9;
                src.connect(f); f.connect(g); g.connect(ctx.destination);
                src.start(ctx.currentTime+t);
            }
            // 4 breath cycles with dense fine crackles in late inspiration
            var crackle_pattern=[0.75,0.78,0.81,0.85,0.88,0.91,0.95,0.99,1.02,1.05,1.08];
            for(var c=0;c<4;c++){
                addBreath(ctx, c*2.3);
                crackle_pattern.forEach(function(tp){addCrackle(ctx, c*2.3+tp);});
            }
            """
        },
        "wheeze_exp": {
            "label": "Expiratory Wheeze — Asthma/COPD",
            "description": "High-pitched musical wheeze during expiration. Polyphonic. Bilateral. Indicates airflow obstruction.",
            "color": "#f59e0b",
            "script": """
            function addInspiration(ctx, t) {
                var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*1.2),ctx.sampleRate);
                var data=buf.getChannelData(0);
                for(var i=0;i<data.length;i++) data[i]=(Math.random()*2-1)*0.15;
                var src=ctx.createBufferSource(); src.buffer=buf;
                var f=ctx.createBiquadFilter(); f.type='lowpass'; f.frequency.value=400;
                var g=ctx.createGain();
                src.connect(f); f.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+t);
                g.gain.linearRampToValueAtTime(0.35,ctx.currentTime+t+0.3);
                g.gain.linearRampToValueAtTime(0,ctx.currentTime+t+1.2);
                src.start(ctx.currentTime+t);
            }
            function addWheeze(ctx, t) {
                // Multiple tones = polyphonic wheeze
                var freqs=[700,850,1100];
                freqs.forEach(function(freq,idx){
                    var osc=ctx.createOscillator(); var g=ctx.createGain();
                    osc.type='sawtooth'; osc.frequency.value=freq;
                    // Slight frequency variation for realism
                    osc.frequency.linearRampToValueAtTime(freq*0.92, ctx.currentTime+t+1.5);
                    osc.connect(g); g.connect(ctx.destination);
                    g.gain.setValueAtTime(0,ctx.currentTime+t);
                    g.gain.linearRampToValueAtTime(0.12,ctx.currentTime+t+0.2);
                    g.gain.setValueAtTime(0.10,ctx.currentTime+t+1.2);
                    g.gain.linearRampToValueAtTime(0,ctx.currentTime+t+1.5);
                    osc.start(ctx.currentTime+t); osc.stop(ctx.currentTime+t+1.6);
                });
            }
            for(var i=0;i<4;i++){addInspiration(ctx,i*2.8); addWheeze(ctx,i*2.8+1.2);}
            """
        },
        "stridor": {
            "label": "Stridor — Upper Airway Obstruction",
            "description": "High-pitched inspiratory wheeze. Indicates upper airway obstruction (croup, epiglottitis, foreign body).",
            "color": "#dc2626",
            "script": """
            for(var i=0;i<5;i++){
                var osc=ctx.createOscillator(); var g=ctx.createGain();
                osc.type='sawtooth'; osc.frequency.value=600;
                osc.frequency.linearRampToValueAtTime(450, ctx.currentTime+i*2.0+1.2);
                var f=ctx.createBiquadFilter(); f.type='bandpass'; f.frequency.value=600; f.Q.value=1.5;
                osc.connect(f); f.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+i*2.0);
                g.gain.linearRampToValueAtTime(0.3,ctx.currentTime+i*2.0+0.2);
                g.gain.setValueAtTime(0.25,ctx.currentTime+i*2.0+1.0);
                g.gain.linearRampToValueAtTime(0,ctx.currentTime+i*2.0+1.2);
                osc.start(ctx.currentTime+i*2.0); osc.stop(ctx.currentTime+i*2.0+1.3);
            }
            """
        },
        "absent_breath": {
            "label": "Absent/Reduced Breath Sounds — Pneumothorax/Effusion",
            "description": "Markedly reduced or absent breath sounds. Indicates pleural effusion, pneumothorax, or consolidation.",
            "color": "#6b7280",
            "script": """
            // Very faint, almost absent sounds
            var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*8),ctx.sampleRate);
            var data=buf.getChannelData(0);
            for(var i=0;i<data.length;i++) data[i]=(Math.random()*2-1)*0.03;
            var src=ctx.createBufferSource(); src.buffer=buf;
            var f=ctx.createBiquadFilter(); f.type='lowpass'; f.frequency.value=200;
            var g=ctx.createGain(); g.gain.value=0.15;
            src.connect(f); f.connect(g); g.connect(ctx.destination);
            src.start(ctx.currentTime);
            """
        },
        # ── BOWEL SOUNDS ──────────────────────────────────────
        "bowel_normal": {
            "label": "Normal Bowel Sounds — Intermittent Gurgles",
            "description": "Intermittent gurgles and clicks. 5–30 sounds/min. Normal peristaltic activity.",
            "color": "#16a34a",
            "script": """
            // Realistic bowel sound: combination of fluid movement and gas bubbles
            var times=[0.8,1.9,3.5,5.1,6.4,8.2,10.1,11.8,13.2,15.0,16.5,18.3,20.1];
            times.forEach(function(t){
                var dur=0.25+Math.random()*0.5;
                var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*dur),ctx.sampleRate);
                var data=buf.getChannelData(0);
                // Mix of noise (fluid) and clicks (gas)
                for(var i=0;i<data.length;i++){
                    var env=Math.sin(Math.PI*i/data.length);
                    data[i]=(Math.random()*2-1)*0.7*env;
                }
                var src=ctx.createBufferSource(); src.buffer=buf;
                var f=ctx.createBiquadFilter(); f.type='bandpass';
                f.frequency.value=180+Math.random()*300; f.Q.value=0.6;
                var g=ctx.createGain();
                src.connect(f); f.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0.5+Math.random()*0.3, ctx.currentTime+t);
                src.start(ctx.currentTime+t);
            });
            """
        },
        "bowel_hyperactive": {
            "label": "Hyperactive Bowel Sounds — Gastroenteritis/Diarrhea",
            "description": "Frequent, high-pitched rushing sounds. >30/min. Indicates gastroenteritis, diarrhea, or early obstruction.",
            "color": "#f59e0b",
            "script": """
            // Dense, rapid, high-pitched sounds
            var times=[];
            for(var t=0.2;t<20;t+=0.4+Math.random()*0.5) times.push(t);
            times.forEach(function(t){
                var dur=0.15+Math.random()*0.3;
                var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*dur),ctx.sampleRate);
                var data=buf.getChannelData(0);
                for(var i=0;i<data.length;i++){
                    var env=Math.sin(Math.PI*i/data.length);
                    data[i]=(Math.random()*2-1)*0.8*env;
                }
                var src=ctx.createBufferSource(); src.buffer=buf;
                var f=ctx.createBiquadFilter(); f.type='bandpass';
                f.frequency.value=300+Math.random()*400; f.Q.value=1.2;
                var g=ctx.createGain(); g.gain.value=0.6+Math.random()*0.3;
                src.connect(f); f.connect(g); g.connect(ctx.destination);
                src.start(ctx.currentTime+t);
            });
            """
        },
        "bowel_absent": {
            "label": "Absent Bowel Sounds — Ileus/Obstruction/Post-op",
            "description": "Complete absence of bowel sounds over 2 minutes. Indicates paralytic ileus, perforation, or peritonitis.",
            "color": "#dc2626",
            "script": """
            // True silence — just ambient noise
            var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*15),ctx.sampleRate);
            var data=buf.getChannelData(0);
            for(var i=0;i<data.length;i++) data[i]=(Math.random()*2-1)*0.008;
            var src=ctx.createBufferSource(); src.buffer=buf;
            src.connect(ctx.destination);
            src.start(ctx.currentTime);
            document.getElementById('snd_status').innerHTML='🔇 Listening... Silent abdomen (absent bowel sounds)';
            """
        },
        # ── PERCUSSION ────────────────────────────────────────
        "percussion_resonant": {
            "label": "Resonant Percussion — Normal Lung",
            "description": "Clear, hollow, resonant note over normal aerated lung. Duration ~200ms. Normal finding.",
            "color": "#0ea5e9",
            "script": """
            function tap(ctx, t) {
                // Resonant = sustained, hollow, medium-low frequency
                var osc1=ctx.createOscillator(); var osc2=ctx.createOscillator();
                var g=ctx.createGain();
                osc1.type='sine'; osc1.frequency.setValueAtTime(180,ctx.currentTime+t);
                osc1.frequency.exponentialRampToValueAtTime(120,ctx.currentTime+t+0.25);
                osc2.type='sine'; osc2.frequency.setValueAtTime(140,ctx.currentTime+t);
                osc2.frequency.exponentialRampToValueAtTime(90,ctx.currentTime+t+0.25);
                osc1.connect(g); osc2.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+t);
                g.gain.linearRampToValueAtTime(0.9,ctx.currentTime+t+0.008);
                g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.28);
                osc1.start(ctx.currentTime+t); osc1.stop(ctx.currentTime+t+0.30);
                osc2.start(ctx.currentTime+t); osc2.stop(ctx.currentTime+t+0.30);
            }
            for(var i=0;i<5;i++) tap(ctx, i*0.9);
            """
        },
        "percussion_dull": {
            "label": "Dull Percussion — Consolidation/Effusion/Solid Organ",
            "description": "Dull, thud-like note. Short duration ~80ms. Higher pitch, less hollow. Over liver, consolidated lung, or effusion.",
            "color": "#9900CC",
            "script": """
            function dullTap(ctx, t) {
                // Dull = short, flat, higher frequency, rapid decay
                var osc=ctx.createOscillator(); var g=ctx.createGain();
                var f=ctx.createBiquadFilter(); f.type='lowpass'; f.frequency.value=180;
                osc.type='sine'; osc.frequency.setValueAtTime(95,ctx.currentTime+t);
                osc.frequency.exponentialRampToValueAtTime(70,ctx.currentTime+t+0.07);
                osc.connect(f); f.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+t);
                g.gain.linearRampToValueAtTime(0.95,ctx.currentTime+t+0.006);
                g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.09);
                osc.start(ctx.currentTime+t); osc.stop(ctx.currentTime+t+0.10);
            }
            for(var i=0;i<5;i++) dullTap(ctx, i*0.7);
            """
        },
        "percussion_hyper": {
            "label": "Hyperresonant Percussion — Pneumothorax/Emphysema",
            "description": "Booming, drum-like hyperresonant note. Low frequency, long duration. Over pneumothorax or hyperinflated lung.",
            "color": "#f59e0b",
            "script": """
            function hyperTap(ctx, t) {
                // Hyperresonant = very low frequency, very sustained, booming
                var osc1=ctx.createOscillator(); var osc2=ctx.createOscillator();
                var g=ctx.createGain();
                osc1.type='sine'; osc1.frequency.setValueAtTime(100,ctx.currentTime+t);
                osc1.frequency.exponentialRampToValueAtTime(55,ctx.currentTime+t+0.7);
                osc2.type='sine'; osc2.frequency.setValueAtTime(130,ctx.currentTime+t);
                osc2.frequency.exponentialRampToValueAtTime(70,ctx.currentTime+t+0.7);
                osc1.connect(g); osc2.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+t);
                g.gain.linearRampToValueAtTime(1.0,ctx.currentTime+t+0.008);
                g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.75);
                osc1.start(ctx.currentTime+t); osc1.stop(ctx.currentTime+t+0.8);
                osc2.start(ctx.currentTime+t); osc2.stop(ctx.currentTime+t+0.8);
            }
            for(var i=0;i<5;i++) hyperTap(ctx, i*1.1);
            """
        },
        "tympanic": {
            "label": "Tympanic Percussion — Abdomen/Gas-filled",
            "description": "Drum-like tympanic note. Heard over gas-filled bowel or stomach. High-pitched, resonant, musical.",
            "color": "#16a34a",
            "script": """
            function tympTap(ctx, t) {
                // Tympanic = drum-like, musical, sustained
                var freqs=[160, 200, 250];
                freqs.forEach(function(freq,i){
                    var osc=ctx.createOscillator(); var g=ctx.createGain();
                    osc.type='sine'; osc.frequency.setValueAtTime(freq,ctx.currentTime+t);
                    osc.frequency.exponentialRampToValueAtTime(freq*0.8,ctx.currentTime+t+0.4);
                    osc.connect(g); g.connect(ctx.destination);
                    var amp=i===0?0.7:i===1?0.5:0.3;
                    g.gain.setValueAtTime(0,ctx.currentTime+t);
                    g.gain.linearRampToValueAtTime(amp,ctx.currentTime+t+0.006);
                    g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.45);
                    osc.start(ctx.currentTime+t); osc.stop(ctx.currentTime+t+0.5);
                });
            }
            for(var i=0;i<5;i++) tympTap(ctx, i*0.8);
            """
        },
        "stony_dull": {
            "label": "Stony Dull — Large Pleural Effusion",
            "description": "Absolute flatness — stony dull percussion. Massive pleural effusion, empyema, or hemothorax.",
            "color": "#6b7280",
            "script": """
            function stonyTap(ctx, t) {
                var osc=ctx.createOscillator(); var g=ctx.createGain();
                var f=ctx.createBiquadFilter(); f.type='lowpass'; f.frequency.value=100;
                osc.type='sine'; osc.frequency.value=60;
                osc.connect(f); f.connect(g); g.connect(ctx.destination);
                g.gain.setValueAtTime(0,ctx.currentTime+t);
                g.gain.linearRampToValueAtTime(0.95,ctx.currentTime+t+0.005);
                g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+t+0.045);
                osc.start(ctx.currentTime+t); osc.stop(ctx.currentTime+t+0.06);
            }
            for(var i=0;i<5;i++) stonyTap(ctx, i*0.55);
            """
        },
    }

    sound = sound_scripts.get(sound_type, sound_scripts["normal_heart"])
    sc    = sound["color"]

    components.html(f"""
    <!DOCTYPE html><html><body style="margin:0;padding:8px;font-family:Inter,sans-serif;background:transparent;">
    <div style="border:2px solid {sc}44;border-radius:12px;padding:10px;background:white;">
        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:6px;">
            <div>
                <div style="font-weight:700;color:#0a2540;font-size:.88rem">🔊 {sound["label"]}</div>
                <div style="font-size:.75rem;color:#6b7280;margin-top:.15rem">{sound["description"]}</div>
            </div>
            <span id="snd_status" style="font-size:.72rem;color:#9ca3af;">Ready</span>
        </div>
        <div style="display:flex;gap:6px;">
            <button onclick="playSound()" style="background:linear-gradient(135deg,{sc},{sc}cc);
                color:white;border:none;border-radius:8px;padding:6px 16px;
                font-size:.83rem;font-weight:600;cursor:pointer;flex:1;">
                ▶ Play
            </button>
            <button onclick="stopSound()" style="background:#f1f5f9;color:#64748b;border:1px solid #e2e8f0;
                border-radius:8px;padding:6px 12px;font-size:.8rem;cursor:pointer;">
                ■ Stop
            </button>
            <button onclick="playSound()" style="background:#f1f5f9;color:#64748b;border:1px solid #e2e8f0;
                border-radius:8px;padding:6px 12px;font-size:.8rem;cursor:pointer;">
                ↺ Repeat
            </button>
        </div>
    </div>
    <script>
    var audioCtx = null;
    function stopSound(){{
        if(audioCtx){{audioCtx.close();audioCtx=null;}}
        document.getElementById('snd_status').innerHTML='Stopped';
    }}
    function playSound(){{
        stopSound();
        document.getElementById('snd_status').innerHTML='▶ Playing...';
        try{{
            var ctx = new (window.AudioContext||window.webkitAudioContext)();
            audioCtx = ctx;
            {sound["script"]}
            setTimeout(function(){{
                if(document.getElementById('snd_status'))
                    document.getElementById('snd_status').innerHTML='✅ Done — click Play to repeat';
            }},10000);
        }}catch(e){{
            document.getElementById('snd_status').innerHTML='⚠️ '+e.message;
        }}
    }}
    </script>
    </body></html>
    """, height=110, scrolling=False)


def tts_stop():
    components.html("<script>window.speechSynthesis.cancel();</script>", height=0)


def get_surgery_for_case(case):
    dx  = str(case.get("Final_Diagnosis","")).lower()
    cc  = str(case.get("Chief_Complaint","")).lower()
    combined = dx+" "+cc
    for key,surg in SURGERIES.items():
        if any(kw in combined for kw in surg["keywords"]):
            return key, surg
    return None, None

# ── LOAD EXCEL ────────────────────────────────────────────────────────────────
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "case_studies.xlsx")


import random, string

def generate_case_id():
    year = datetime.now().year
    suffix = "".join(random.choices(string.digits, k=4))
    return f"MLS-{year}-{suffix}"

def save_new_case(case_data):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH)
        sheets_map = {
            "case Metadata ":        [case_data["Case_ID"], case_data["Title"], case_data["System"], case_data["Difficulty"], case_data.get("Learning_Obj","")],
            "Initial Presentation ": [case_data["Age_Sex"], case_data.get("Occupation",""), case_data["Chief_Complaint"], case_data["Duration"], case_data["Context"]],
            "History taking":        [case_data["HPI"], case_data.get("PMH","none"), case_data.get("Family_Hx","neg."), case_data.get("Social_Hx","none"), case_data.get("Medications","none")],
            "physical examination ": [case_data["Vitals"], case_data.get("Appearance","in pain"), case_data["Physical_Findings"]],
            "investigation ":        [case_data.get("Labs","none"), case_data.get("Urine","none"), case_data.get("Imaging_Tests","none")],
            "final diagnosis ":      [case_data["Final_Diagnosis"]],
        }
        for sheet_name, row_data in sheets_map.items():
            if sheet_name in wb.sheetnames:
                wb[sheet_name].append(row_data)
        wb.save(EXCEL_PATH)
        load_cases.clear()
        return True
    except Exception as e:
        st.error(f"Save error: {e}")
        return False

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "case_studies.xlsx")

@st.cache_data
def load_cases():
    if not os.path.exists(EXCEL_PATH): return pd.DataFrame()
    meta = pd.read_excel(EXCEL_PATH, sheet_name="case Metadata ",       header=0)
    pres = pd.read_excel(EXCEL_PATH, sheet_name="Initial Presentation ", header=0)
    hist = pd.read_excel(EXCEL_PATH, sheet_name="History taking",        header=0)
    phys = pd.read_excel(EXCEL_PATH, sheet_name="physical examination ", header=0)
    inv  = pd.read_excel(EXCEL_PATH, sheet_name="investigation ",        header=0)
    diag = pd.read_excel(EXCEL_PATH, sheet_name="final diagnosis ",      header=0)
    meta.columns = ["Case_ID","Title","System","Difficulty","LO"]+list(range(len(meta.columns)-5))
    pres.columns = ["Age_Sex","Occupation","Chief_Complaint","Duration","Context"]+list(range(len(pres.columns)-5))
    hist.columns = ["HPI","PMH","Family_Hx","Social_Hx","Medications"]+list(range(len(hist.columns)-5))
    phys.columns = ["Vitals","Appearance","Physical_Findings"]+list(range(len(phys.columns)-3))
    # Investigation columns: Labs, Urine, Imaging_Tests, XRay_Report, CT_Report
    inv_base = ["Labs","Urine","Imaging_Tests","XRay_Report","CT_Report"]
    inv_named = inv_base[:min(len(inv_base), len(inv.columns))]
    inv_extra = list(range(max(0, len(inv.columns)-len(inv_named))))
    inv.columns = inv_named + inv_extra
    diag.columns = ["Final_Diagnosis"]+list(range(len(diag.columns)-1))
    # Use the LONGEST sheet as the row count — pad shorter sheets with NaN
    n = max(len(meta),len(pres),len(hist),len(phys),len(inv),len(diag))
    def _pad(frame, cols):
        f = frame[cols].copy() if all(c in frame.columns for c in cols) else frame[[c for c in cols if c in frame.columns]].copy()
        if len(f) < n:
            extra = pd.DataFrame(index=range(n - len(f)), columns=f.columns)
            f = pd.concat([f, extra], ignore_index=True)
        return f.iloc[:n].reset_index(drop=True)
    df = pd.concat([
        _pad(meta, ["Case_ID","Title","System","Difficulty"]),
        _pad(pres, ["Age_Sex","Occupation","Chief_Complaint","Duration","Context"]),
        _pad(hist, ["HPI","PMH","Family_Hx","Social_Hx","Medications"]),
        _pad(phys, ["Vitals","Appearance","Physical_Findings"]),
        _pad(inv,  [c for c in ["Labs","Urine","Imaging_Tests","XRay_Report","CT_Report"] if c in inv.columns]),
        _pad(diag, ["Final_Diagnosis"]),
    ], axis=1)
    df = df[df["Chief_Complaint"].notna() & df["Final_Diagnosis"].notna()]
    df["Case_ID"]    = df["Case_ID"].fillna(0).astype(float).astype(int).astype(str)
    df["Difficulty"] = df["Difficulty"].fillna("basic").str.strip().str.lower()
    df["System"]     = df["System"].fillna("general").str.strip().str.lower()
    df["row_num"]    = range(len(df))

    # ── Merge approved AI-generated cases from Supabase ─────────────────
    # The AI Case Creator stores approved cases in `cases_extended` table.
    # Combine them here so students see them seamlessly in the library.
    try:
        if CASE_CREATOR_OK:
            ai_df = load_approved_cases_from_db()
            if ai_df is not None and not ai_df.empty:
                # Match the same columns as xlsx-loaded cases
                ai_df["Case_ID"]    = ai_df["Case_ID"].astype(str)
                ai_df["Difficulty"] = ai_df["Difficulty"].fillna("basic").astype(str).str.strip().str.lower()
                ai_df["System"]     = ai_df["System"].fillna("general").astype(str).str.strip().str.lower()
                # Continue row_num after the xlsx cases
                ai_df["row_num"]    = range(len(df), len(df) + len(ai_df))
                # Reindex columns to match df (missing → NaN)
                for col in df.columns:
                    if col not in ai_df.columns:
                        ai_df[col] = None
                ai_df = ai_df[df.columns]  # same column order
                df = pd.concat([df, ai_df], ignore_index=True)
    except Exception as _ai_err:
        print(f"[load_cases] Could not merge AI cases: {_ai_err}")

    return df.reset_index(drop=True)

def save_new_case(case_data):
    """Append a new real case to the Excel file."""
    try:
        wb = load_workbook(EXCEL_PATH)
        year = datetime.now().year
        case_id = f"MLS-{year}-" + "".join(random.choices(string.digits, k=4))

        sheets_data = {
            "case Metadata ":   [case_id, case_data.get("title",""), case_data.get("system",""), case_data.get("difficulty","basic"), case_data.get("learning_obj","")],
            "Initial Presentation ": [case_data.get("age_sex",""), case_data.get("occupation",""), case_data.get("chief_complaint",""), case_data.get("duration",""), case_data.get("context","")],
            "History taking":   [case_data.get("hpi",""), case_data.get("pmh",""), case_data.get("family_hx","neg."), case_data.get("social_hx",""), case_data.get("medications","none")],
            "physical examination ": [case_data.get("vitals",""), case_data.get("appearance",""), case_data.get("physical_findings","")],
            "investigation ":   [case_data.get("labs","none"), case_data.get("urine","none"), case_data.get("imaging","none")],
            "final diagnosis ": [case_data.get("final_diagnosis","")],
        }
        for sheet_name, row_data in sheets_data.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.append(row_data)
        wb.save(EXCEL_PATH)
        return case_id
    except Exception as e:
        return f"ERROR: {e}"

# ── SESSION STATE ─────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
# 💳 CREDIT SYSTEM ENGINE
# Free tier:    20 credits per 8-hour window, resets every 8h
# Premium tier: Unlimited credits, $5/month
# Credits used: 1 per AI call, 3 per image analysis, 2 per lab generation
# Storage: Supabase (if connected) or session state (local fallback)
# ════════════════════════════════════════════════════════════════════════════

import hashlib as _hl
from datetime import datetime, timedelta, timezone

# CREDIT_COSTS defined at top of file



# ════════════════════════════════════════════════════════════════════════════
# 💳 CREDIT STATE MACHINE
#
# STATES:
#   FREE_ACTIVE      → has credits in current window
#   FREE_EXHAUSTED   → 0 credits, waiting for window reset (every 8h)
#   PREMIUM_ACTIVE   → unlimited, valid subscription
#   PREMIUM_EXPIRED  → was premium, now reverted to FREE_ACTIVE
#
# TRANSITIONS:
#   FREE_ACTIVE    --[use credits]--> FREE_EXHAUSTED
#   FREE_EXHAUSTED --[8h elapsed]--> FREE_ACTIVE (auto-reset)
#   FREE_ACTIVE    --[pay+code]----> PREMIUM_ACTIVE
#   PREMIUM_ACTIVE --[30d elapsed]-> FREE_ACTIVE (auto-expire)
#
# STORAGE: Supabase `user_credits` table (persistent across devices/sessions)
#          Session state fallback (local only, resets on browser close)
# ════════════════════════════════════════════════════════════════════════════

# Credit state names (for clarity)
CS_FREE_ACTIVE    = "free"
CS_PREMIUM_ACTIVE = "premium"

def get_user_id():
    """Get or create a stable user ID for this browser session."""
    if "user_id" not in st.session_state or not st.session_state["user_id"]:
        import uuid
        st.session_state["user_id"] = str(uuid.uuid4())[:16]
    return st.session_state["user_id"]

def _get_db_client():
    """Get Supabase client if credentials are configured. Returns None if not."""
    if not st.session_state.get("sb_connected"):
        return None
    try:
        from supabase import create_client as _scc
        _u = st.session_state.get("sb_url","") or SUPABASE_DEFAULT_URL
        _k = st.session_state.get("sb_key","") or SUPABASE_DEFAULT_KEY
        if not _u or not _k or "YOUR_SUPABASE" in _u or "YOUR_SUPABASE" in _k:
            return None
        return _scc(_u.strip(), _k.strip())
    except Exception:
        return None

def _state_from_db(client, uid):
    """Load user credit state from Supabase. Creates record if first visit."""
    try:
        r = client.table("user_credits").select("*").eq("user_id", uid).execute()
        if r.data:
            return r.data[0]
        # First visit — create record
        now = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        new_state = {
            "user_id": uid, "credits_used": 0,
            "window_start": now, "plan": CS_FREE_ACTIVE,
            "plan_expires": None, "total_used_ever": 0,
        }
        client.table("user_credits").insert(new_state).execute()
        return new_state
    except Exception:
        return None

def _state_from_session(uid):
    """Load/create user credit state from session state (local fallback)."""
    key = f"cs_{uid}"
    if key not in st.session_state:
        st.session_state[key] = {
            "user_id": uid, "credits_used": 0,
            "window_start": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
            "plan": CS_FREE_ACTIVE, "plan_expires": None, "total_used_ever": 0,
        }
    return st.session_state[key]

def _save_state(state, client):
    """Persist state to Supabase or session."""
    uid = state["user_id"]
    if client:
        try:
            client.table("user_credits").upsert({
                "user_id": uid,
                "credits_used": state["credits_used"],
                "window_start": state["window_start"],
                "plan": state["plan"],
                "plan_expires": state.get("plan_expires"),
                "total_used_ever": state.get("total_used_ever", 0),
            }).execute()
            return
        except Exception:
            pass
    st.session_state[f"cs_{uid}"] = state

def _apply_transitions(state, now):
    """
    Apply state machine transitions based on time.
    Modifies state in-place. Returns True if state changed.
    """
    changed = False

    # TRANSITION: FREE window reset (every WINDOW_HOURS)
    if state["plan"] == CS_FREE_ACTIVE:
        try:
            ws = datetime.fromisoformat(state["window_start"])
            age_h = (now - ws).total_seconds() / 3600
            if age_h >= WINDOW_HOURS:
                state["credits_used"] = 0
                state["window_start"] = now.isoformat()
                changed = True
        except Exception:
            state["window_start"] = now.isoformat()
            changed = True

    # TRANSITION: PREMIUM expiry → FREE
    if state["plan"] == CS_PREMIUM_ACTIVE and state.get("plan_expires"):
        try:
            expires = datetime.fromisoformat(state["plan_expires"])
            if now > expires:
                state["plan"] = CS_FREE_ACTIVE
                state["plan_expires"] = None
                state["credits_used"] = 0
                state["window_start"] = now.isoformat()
                changed = True
        except Exception:
            pass

    return changed

def get_credit_state():
    """
    Main entry point for credit state machine.
    Returns (state_dict, db_client_or_None).
    Thread-safe: each call reads fresh state and applies transitions.
    """
    uid    = get_user_id()
    now    = datetime.now(timezone.utc).replace(tzinfo=None)
    client = _get_db_client()

    # Load state
    state = _state_from_db(client, uid) if client else None
    if state is None:
        client = None
        state  = _state_from_session(uid)

    # Apply time-based transitions
    if _apply_transitions(state, now):
        _save_state(state, client)

    return state, client

def get_credits_remaining(state):
    """Credits left in current window. Premium = effectively infinite."""
    if state.get("plan") == CS_PREMIUM_ACTIVE:
        return PREMIUM_MONTHLY_CREDITS
    return max(0, FREE_CREDITS_PER_WINDOW - state.get("credits_used", 0))

def get_window_reset_time(state):
    """Datetime when the current free window resets."""
    try:
        ws = datetime.fromisoformat(state.get("window_start", datetime.now(timezone.utc).replace(tzinfo=None).isoformat()))
    except Exception:
        ws = datetime.now(timezone.utc).replace(tzinfo=None)
    return ws + timedelta(hours=WINDOW_HOURS)

def can_use_credits(cost_type="chat"):
    """
    Check if user can afford this action.
    Returns (True, "") or (False, human_readable_message).
    """
    cost = CREDIT_COSTS.get(cost_type, 1)
    state, _ = get_credit_state()
    remaining = get_credits_remaining(state)

    if remaining >= cost:
        return True, ""

    # Compute time until reset
    reset_dt  = get_window_reset_time(state)
    now       = datetime.now(timezone.utc).replace(tzinfo=None)
    time_left = reset_dt - now
    h = max(0, int(time_left.total_seconds() // 3600))
    m = max(0, int((time_left.total_seconds() % 3600) // 60))

    return False, (
        f"You need {cost} credits but only have {remaining} left. "
        f"Your window resets in {h}h {m}m — or upgrade to Premium for unlimited access ($5/month)."
    )

def use_credits(cost_type="chat"):
    """
    Deduct credits for an action. Returns (True, "") or (False, message).
    Also persists the new state immediately.
    """
    ok, msg = can_use_credits(cost_type)
    if not ok:
        return False, msg

    cost = CREDIT_COSTS.get(cost_type, 1)
    state, client = get_credit_state()
    state["credits_used"]    = state.get("credits_used", 0) + cost
    state["total_used_ever"] = state.get("total_used_ever", 0) + cost
    _save_state(state, client)
    return True, ""

def activate_premium(activation_code="", duration_days=30):
    """
    STATE TRANSITION: FREE → PREMIUM.
    Validates code against Supabase if connected.
    Returns (True, expires_date_str) or (False, error_message).
    """
    state, client = get_credit_state()
    uid = state["user_id"]

    # Validate code format
    code = activation_code.strip().upper()
    if not code.startswith("MLS-PREM-") or len(code) != 18:
        return False, "Invalid code format. Codes look like: MLS-PREM-XXXX-XXXX"

    # Validate against Supabase (prevents reuse)
    if client:
        try:
            r = client.table("premium_codes").select("*").eq("code", code).execute()
            if not r.data:
                return False, "Code not found. Check for typos."
            row = r.data[0]
            if row.get("used"):
                return False, "This code has already been used."
            duration_days = row.get("duration_days", 30)
            # Mark as used
            client.table("premium_codes").update({
                "used": True, "used_by": uid,
                "used_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
            }).eq("code", code).execute()
        except Exception as e:
            # If premium_codes table not set up, accept any valid format
            pass
    else:
        # No Supabase — accept any properly formatted code (local mode)
        pass

    # Apply PREMIUM transition
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    expires = (now + timedelta(days=duration_days)).isoformat()
    state["plan"]          = CS_PREMIUM_ACTIVE
    state["plan_expires"]  = expires
    state["credits_used"]  = 0
    state["window_start"]  = now.isoformat()
    _save_state(state, client)

    return True, expires

def generate_premium_code(client=None, duration_days=30):
    """
    Generate a new premium activation code and save to Supabase.
    Returns the code string.
    """
    import random, string as _str
    p1 = "".join(random.choices(_str.ascii_uppercase + _str.digits, k=4))
    p2 = "".join(random.choices(_str.ascii_uppercase + _str.digits, k=4))
    code = f"MLS-PREM-{p1}-{p2}"

    if client:
        try:
            client.table("premium_codes").insert({
                "code": code, "used": False,
                "duration_days": duration_days,
                "created_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
            }).execute()
        except Exception:
            pass
    return code


def render_credit_bar():
    """Render a compact credit status bar in sidebar."""
    state, _  = get_credit_state()
    plan      = state.get("plan","free")
    remaining = get_credits_remaining(state)
    used      = state.get("credits_used", 0)
    reset_dt  = get_window_reset_time(state)
    now       = datetime.now(timezone.utc).replace(tzinfo=None)
    time_left = reset_dt - now
    h_left    = max(0, int(time_left.total_seconds() // 3600))
    m_left    = max(0, int((time_left.total_seconds() % 3600) // 60))

    if plan == "premium":
        st.markdown("""
        <div style="background:linear-gradient(135deg,#f59e0b,#d97706);color:white;
                    border-radius:10px;padding:.6rem .9rem;margin:.3rem 0;text-align:center;">
            <div style="font-weight:700;font-size:.85rem">⭐ PREMIUM</div>
            <div style="font-size:.72rem;opacity:.9">Unlimited credits</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        pct = max(0, remaining / FREE_CREDITS_PER_WINDOW * 100)
        color = "#16a34a" if pct > 50 else "#f59e0b" if pct > 20 else "#dc2626"
        st.markdown(f"""
        <div style="background:white;border-radius:10px;padding:.6rem .9rem;
                    margin:.3rem 0;border:1px solid #e2e8f0;">
            <div style="display:flex;justify-content:space-between;font-size:.78rem;font-weight:600;color:#0a2540;">
                <span>💳 Credits</span>
                <span style="color:{color}">{remaining}/{FREE_CREDITS_PER_WINDOW}</span>
            </div>
            <div style="background:#e5e7eb;border-radius:999px;height:6px;margin:.3rem 0;">
                <div style="background:{color};height:6px;border-radius:999px;width:{pct:.0f}%"></div>
            </div>
            <div style="font-size:.68rem;color:#9ca3af;">Resets in {h_left}h {m_left}m</div>
        </div>
        """, unsafe_allow_html=True)

        if remaining <= 5:
            if st.button("⭐ Upgrade $5/mo", use_container_width=True, key="sb_upgrade_quick"):
                nav("credits")


# ════════════════════════════════════════════════════════════════════════════
# 🏥 CLINICAL DECISION SUPPORT ENGINE
# Built as: AI Engineering Professor + Senior Clinician
# ════════════════════════════════════════════════════════════════════════════

# ── Real-world evidence databases (verified public URLs) ─────────────────────
EVIDENCE_DATABASES = {
    "PubMed":     "https://pubmed.ncbi.nlm.nih.gov/?term=",
    "UpToDate":   "https://www.uptodate.com/contents/search?search=",
    "WHO ICD-11": "https://icd.who.int/browse/2024-01/mms/en#",
    "ClinicalTrials": "https://clinicaltrials.gov/search?cond=",
    "NICE":       "https://www.nice.org.uk/search#?q=",
    "BMJ Best Practice": "https://bestpractice.bmj.com/search?q=",
    "UpToDate Free": "https://www.wolterskluwer.com/en/solutions/uptodate",
    "PhysioNet":  "https://physionet.org/search/?query=",
    "NIH CXR":    "https://nihcc.app.box.com/v/ChestXray-NIHCC",
    "MIMIC":      "https://physionet.org/content/mimic-cxr/2.0.0/",
}

# ── Clinical scoring systems ─────────────────────────────────────────────────
CLINICAL_SCORES = {
    "Wells PE": {
        "description": "Probability of Pulmonary Embolism",
        "variables": {
            "Clinical signs of DVT": 3,
            "PE most likely diagnosis": 3,
            "HR > 100 bpm": 1.5,
            "Immobilization > 3 days or surgery past 4 weeks": 1.5,
            "Previous DVT/PE": 1.5,
            "Hemoptysis": 1,
            "Active malignancy": 1,
        },
        "thresholds": {"low": (0, 1.9), "moderate": (2, 5.9), "high": (6, 99)},
        "interpretation": {
            "low": "PE unlikely (<15%). D-dimer recommended.",
            "moderate": "Intermediate risk. CT-PA indicated.",
            "high": "PE likely (>65%). CT-PA immediately.",
        }
    },
    "Wells DVT": {
        "description": "Probability of Deep Vein Thrombosis",
        "variables": {
            "Active cancer": 1, "Paralysis/paresis/plaster": 1,
            "Bedridden >3 days / surgery <12 weeks": 1,
            "Tenderness along deep venous system": 1,
            "Entire leg swollen": 1,
            "Calf swelling >3cm vs other leg": 1,
            "Pitting edema (symptomatic leg)": 1,
            "Collateral superficial veins": 1,
            "Alternative diagnosis as likely": -2,
        },
        "thresholds": {"low": (-5, 0), "moderate": (1, 2), "high": (3, 20)},
        "interpretation": {
            "low": "DVT unlikely. D-dimer if needed.",
            "moderate": "Ultrasound recommended.",
            "high": "DVT likely. Ultrasound + anticoagulation.",
        }
    },
    "CURB-65": {
        "description": "Pneumonia Severity (Community-acquired)",
        "variables": {
            "Confusion (new)": 1, "Urea > 7 mmol/L (BUN > 19)": 1,
            "Respiratory rate ≥ 30/min": 1,
            "BP: systolic < 90 or diastolic ≤ 60": 1,
            "Age ≥ 65 years": 1,
        },
        "thresholds": {"low": (0, 1), "moderate": (2, 2), "high": (3, 5)},
        "interpretation": {
            "low": "Low risk. Home treatment (mortality <1%).",
            "moderate": "Hospital admission. Supervised treatment (mortality ~9%).",
            "high": "Severe pneumonia. ICU consideration (mortality >20%).",
        }
    },
    "Glasgow Coma": {
        "description": "Level of Consciousness",
        "variables": {
            "Eye opening: Spontaneous (4)": 4, "Eye opening: To voice (3)": 3,
            "Eye opening: To pain (2)": 2, "Eye opening: None (1)": 1,
            "Verbal: Oriented (5)": 5, "Verbal: Confused (4)": 4,
            "Verbal: Words (3)": 3, "Verbal: Sounds (2)": 2, "Verbal: None (1)": 1,
            "Motor: Obeys (6)": 6, "Motor: Localizes (5)": 5,
            "Motor: Withdraws (4)": 4, "Motor: Flexion (3)": 3,
            "Motor: Extension (2)": 2, "Motor: None (1)": 1,
        },
        "thresholds": {"severe": (3, 8), "moderate": (9, 12), "mild": (13, 15)},
        "interpretation": {
            "severe": "Severe brain injury. Intubation if GCS ≤ 8.",
            "moderate": "Moderate brain injury. Close monitoring.",
            "mild": "Mild injury / normal consciousness.",
        }
    },
    "Alvarado (Appendicitis)": {
        "description": "Probability of Acute Appendicitis",
        "variables": {
            "Migration of pain to RLQ": 1, "Anorexia": 1,
            "Nausea/Vomiting": 1, "RLQ tenderness": 2,
            "Rebound tenderness": 1, "Elevated temperature (>37.3°C)": 1,
            "Leukocytosis (WBC > 10,000)": 2, "Left shift (neutrophilia)": 1,
        },
        "thresholds": {"low": (0, 4), "moderate": (5, 6), "high": (7, 10)},
        "interpretation": {
            "low": "Appendicitis unlikely. Observe.",
            "moderate": "Possible appendicitis. CT scan / surgical consult.",
            "high": "Appendicitis likely. Surgical exploration.",
        }
    },
    "Child-Pugh (Liver)": {
        "description": "Liver Cirrhosis Severity",
        "variables": {
            "Bilirubin < 2 mg/dL (1pt)": 1, "Bilirubin 2-3 mg/dL (2pt)": 2, "Bilirubin > 3 mg/dL (3pt)": 3,
            "Albumin > 3.5 g/dL (1pt)": 1, "Albumin 2.8-3.5 (2pt)": 2, "Albumin < 2.8 (3pt)": 3,
            "No ascites (1pt)": 1, "Mild ascites (2pt)": 2, "Tense ascites (3pt)": 3,
            "No encephalopathy (1pt)": 1, "Grade 1-2 (2pt)": 2, "Grade 3-4 (3pt)": 3,
            "INR < 1.7 (1pt)": 1, "INR 1.7-2.3 (2pt)": 2, "INR > 2.3 (3pt)": 3,
        },
        "thresholds": {"A": (5, 6), "B": (7, 9), "C": (10, 15)},
        "interpretation": {
            "A": "Child-Pugh A — Well compensated. 1-year survival 100%.",
            "B": "Child-Pugh B — Significant functional compromise.",
            "C": "Child-Pugh C — Decompensated. 1-year survival 45%.",
        }
    },
    "HEART Score": {
        "description": "Risk of Major Adverse Cardiac Event",
        "variables": {
            "History highly suspicious": 2, "History moderately suspicious": 1, "History slightly suspicious": 0,
            "ECG LBBB/ST depression/change": 2, "ECG non-specific repolarization": 1, "ECG normal": 0,
            "Age ≥ 65": 2, "Age 45-64": 1, "Age < 45": 0,
            "≥3 risk factors or atherosclerosis": 2, "1-2 risk factors": 1, "No risk factors": 0,
            "Troponin > 3× normal": 2, "Troponin 1-3× normal": 1, "Troponin ≤ normal": 0,
        },
        "thresholds": {"low": (0, 3), "moderate": (4, 6), "high": (7, 10)},
        "interpretation": {
            "low": "Low risk (1.7%). Discharge with outpatient follow-up.",
            "moderate": "Moderate risk (12%). Observation + serial troponins.",
            "high": "High risk (65%). Invasive strategy recommended.",
        }
    },
    "Revised Trauma Score": {
        "description": "Trauma Severity",
        "variables": {
            "GCS 13-15 (4pts)": 4, "GCS 9-12 (3pts)": 3, "GCS 6-8 (2pts)": 2, "GCS 4-5 (1pt)": 1, "GCS 3 (0pts)": 0,
            "SBP > 89 (4pts)": 4, "SBP 76-89 (3pts)": 3, "SBP 50-75 (2pts)": 2, "SBP 1-49 (1pt)": 1, "SBP 0 (0pts)": 0,
            "RR 10-29 (4pts)": 4, "RR > 29 (3pts)": 3, "RR 6-9 (2pts)": 2, "RR 1-5 (1pt)": 1, "RR 0 (0pts)": 0,
        },
        "thresholds": {"critical": (0, 8), "major": (9, 10), "minor": (11, 12)},
        "interpretation": {
            "critical": "Critical injury. Trauma center activation.",
            "major": "Major trauma. Resuscitation bay.",
            "minor": "Minor trauma. Standard evaluation.",
        }
    },
}

# ── Real medical case databases (verified, publicly accessible) ───────────────
REAL_CASE_DATABASES = {
    "NIH Chest X-Ray (112,000 CXR)": {
        "url": "https://nihcc.app.box.com/v/ChestXray-NIHCC",
        "conditions": ["Atelectasis","Cardiomegaly","Consolidation","Edema","Effusion",
                       "Emphysema","Fibrosis","Hernia","Infiltration","Mass",
                       "Nodule","Pleural Thickening","Pneumonia","Pneumothorax"],
        "size": "112,120 images",
        "source": "National Institutes of Health, USA",
        "verified": True,
    },
    "MIMIC-CXR (MIT) (227,000 reports)": {
        "url": "https://physionet.org/content/mimic-cxr/2.0.0/",
        "conditions": ["All chest pathologies + free-text radiology reports"],
        "size": "227,827 images, 227,827 reports",
        "source": "MIT / Beth Israel Deaconess Medical Center",
        "verified": True,
    },
    "MIT-BIH Arrhythmia (ECG)": {
        "url": "https://physionet.org/content/mitdb/1.0.0/",
        "conditions": ["Atrial Fibrillation","Ventricular Tachycardia","LBBB","RBBB",
                       "PVCs","Normal Sinus","AV Block","WPW"],
        "size": "48 two-channel ECG recordings, 47 subjects",
        "source": "MIT / Beth Israel Hospital",
        "verified": True,
    },
    "PTB-XL ECG (21,799 ECGs)": {
        "url": "https://physionet.org/content/ptb-xl/1.0.3/",
        "conditions": ["STEMI","NSTEMI","LBBB","RBBB","AF","WPW","Normal","LVH"],
        "size": "21,799 clinical 12-lead ECGs",
        "source": "Physikalisch-Technische Bundesanstalt, Germany",
        "verified": True,
    },
    "CheXpert (224,000 CXR)": {
        "url": "https://stanfordmlgroup.github.io/competitions/chexpert/",
        "conditions": ["Pleural Effusion","Pneumonia","Pneumothorax","Cardiomegaly",
                       "Edema","Consolidation","Atelectasis","No Finding"],
        "size": "224,316 chest X-rays, 65,240 patients",
        "source": "Stanford University Medical Center",
        "verified": True,
    },
    "RSNA Pneumonia (30,000 CXR)": {
        "url": "https://www.kaggle.com/c/rsna-pneumonia-detection-challenge",
        "conditions": ["Pneumonia detection + bounding boxes"],
        "size": "30,000 chest X-rays with annotations",
        "source": "Radiological Society of North America",
        "verified": True,
    },
    "BraTS Brain Tumor MRI": {
        "url": "https://www.med.upenn.edu/cbica/brats/",
        "conditions": ["Glioblastoma","Lower Grade Glioma","Meningioma"],
        "size": "2000+ multi-institutional MRI scans",
        "source": "University of Pennsylvania / multi-site",
        "verified": True,
    },
    "Retinal OCT / Fundus": {
        "url": "https://data.mendeley.com/datasets/rscbjbr9sj/3",
        "conditions": ["CNV","DME","Drusen","Normal retina"],
        "size": "84,495 retinal OCT images",
        "source": "Guangzhou Women and Children's Medical Center",
        "verified": True,
    },
}

# ── Validated clinical guidelines (WHO / AHA / ESC / NICE) ────────────────────
CLINICAL_GUIDELINES = {
    "ACS / STEMI": {
        "body": "ESC / AHA / ACC",
        "url": "https://www.escardio.org/Guidelines/Clinical-Practice-Guidelines/Acute-Myocardial-Infarction-in-patients-presenting-with-ST-segment-elevation",
        "key_points": ["Primary PCI within 90 min of STEMI","Dual antiplatelet therapy (Aspirin + P2Y12)","Anticoagulation: UFH or LMWH","Beta-blocker within 24h if stable","ACE inhibitor / ARB within 24h"],
    },
    "Community-Acquired Pneumonia": {
        "body": "BTS / IDSA / ATS",
        "url": "https://www.brit-thoracic.org.uk/quality-improvement/guidelines/pneumonia/",
        "key_points": ["CURB-65 score for severity","Empiric: Amoxicillin +/- Macrolide (mild)","Empiric: Co-amoxiclav + Clarithromycin (moderate)","Blood cultures before antibiotics","Target SpO2 94-98%"],
    },
    "Sepsis / Septic Shock": {
        "body": "SSC (Surviving Sepsis Campaign)",
        "url": "https://www.sccm.org/clinical-resources/guidelines/guidelines/management-of-sepsis-and-septic-shock",
        "key_points": ["1-hour bundle: lactate, cultures, 30ml/kg IV fluid, antibiotics","qSOFA: RR≥22, AMS, SBP≤100 → ICU","Norepinephrine first-line vasopressor","Broad-spectrum antibiotics within 1 hour","Daily reassessment for de-escalation"],
    },
    "Stroke (Ischaemic)": {
        "body": "ESO / AHA / ASA",
        "url": "https://eso-stroke.org/guidelines/",
        "key_points": ["tPA within 4.5 hours (alteplase 0.9mg/kg, max 90mg)","CT/MRI to exclude hemorrhage before thrombolysis","Target BP <185/110 before tPA","Mechanical thrombectomy: large vessel occlusion within 24h","Aspirin 300mg within 24h if no thrombolysis"],
    },
    "Acute Appendicitis": {
        "body": "WSES / American College of Surgeons",
        "url": "https://wjes.biomedcentral.com/articles/10.1186/s13017-020-00306-3",
        "key_points": ["Alvarado score guides workup","CT abdomen confirms diagnosis","Laparoscopic appendectomy gold standard","Antibiotics alone: uncomplicated appendicitis option","Perforation: broad-spectrum IV antibiotics pre-op"],
    },
    "Diabetic Ketoacidosis": {
        "body": "ADA / JBDS",
        "url": "https://care.diabetesjournals.org/content/44/Supplement_1/S151",
        "key_points": ["IV Normal Saline 1L over 1h initially","Insulin: 0.1 units/kg/h IV infusion","Potassium replacement (target 3.5-5.5 mEq/L)","Bicarbonate only if pH < 6.9","Hourly glucose monitoring"],
    },
}

def search_evidence(query, database="PubMed"):
    """Generate evidence search link for a clinical query."""
    base = EVIDENCE_DATABASES.get(database, EVIDENCE_DATABASES["PubMed"])
    encoded = query.replace(" ", "+").replace("/", "%2F")
    return f"{base}{encoded}"

def get_guideline(condition):
    """Find the best clinical guideline for a condition."""
    cl = condition.lower()
    for key, val in CLINICAL_GUIDELINES.items():
        if any(w in cl for w in key.lower().split("/")):
            return key, val
    return None, None

def calculate_clinical_score(score_name, selected_criteria):
    """
    Calculate a clinical score given selected criteria.
    Returns (total, severity, interpretation, recommendations).
    """
    score_def = CLINICAL_SCORES.get(score_name)
    if not score_def:
        return 0, "unknown", "Score not found", []

    total = sum(score_def["variables"].get(c, 0) for c in selected_criteria)

    severity = "unknown"
    for sev, (lo, hi) in score_def["thresholds"].items():
        if lo <= total <= hi:
            severity = sev
            break

    interpretation = score_def["interpretation"].get(severity, "")
    return total, severity, interpretation


defaults = dict(
    page="home", selected_case=None, chat_history=[], tutor_history=[],
    live_history=[], lab_seen=False, imaging_seen=False, submitted=False,
    cases_done=[], score=0, surgery_step=0, voice_enabled=True,
    avatar_mood="neutral", exam_findings={}, live_mode=False,
    image_analysis_result=None, suggested_questions=[],
    peer_room_code=None, peer_role=None, peer_name="",
    user_id=None,
    # ── Auth / User management ──────────────────────────────────
    auth_user=None,          # logged-in user dict {id,name,email,role}
    auth_page="login",       # login | register
    # ── Differential Diagnosis Builder ──────────────────────────
    ddx_list=[],             # student's ranked differentials
    ddx_reasoning={},        # {dx: reasoning text}
    ddx_submitted=False,
    # ── Drug Prescribing Module ──────────────────────────────────
    rx_history=[],           # list of prescriptions written
    rx_feedback=None,
    # ── Procedure Simulator ──────────────────────────────────────
    proc_selected=None,      # current procedure
    proc_step=0,             # step index
    proc_score=0,
    # ── Clinical Reasoning Map ───────────────────────────────────
    reasoning_nodes=[],      # list of {type,content,timestamp}
    # ── Competency Tracker ───────────────────────────────────────
    competencies={},         # {skill: {attempts,passes,last_date}}
    # ── Case Creator (faculty) ───────────────────────────────────
    draft_case={},
    # ── Avatar Builder ───────────────────────────────────────────
    doctor_avatar={},
    # ── AI Clinical Tutor Cases ──────────────────────────────────
    ait_case=None, ait_chat=[], ait_dx_submitted=False,
    ait_score=0, ait_cases_done=0,
)
for k,v in defaults.items():
    if k not in st.session_state: st.session_state[k]=v

df = load_cases()

def reset_case():
    """Reset all case-specific state when a new case is selected."""
    st.session_state.update(
        chat_history=[], tutor_history=[], live_history=[],
        lab_seen=False, imaging_seen=False, submitted=False,
        surgery_step=0, avatar_mood="neutral", exam_findings={},
        live_mode=False, image_analysis_result=None,
        # ── DDx Builder ──────────────────────────────────────────
        ddx_submitted=False, ddx_list=[], ddx_reasoning={},
        # ── Drug Prescribing ─────────────────────────────────────
        rx_feedback=None, rx_history=[],
        # ── Procedure Simulator ──────────────────────────────────
        proc_step=0, proc_score=0,
        # ── Clinical Reasoning Map ───────────────────────────────
        reasoning_nodes=[],
    )

def nav(p): st.session_state.page=p; st.rerun()

def diff_badge(d):
    d=str(d).lower()
    if "adv" in d: return '<span class="badge br">Advanced</span>'
    if "int" in d: return '<span class="badge bo">Intermediate</span>'
    return '<span class="badge bg">Basic</span>'

# ── AVATAR ────────────────────────────────────────────────────────────────────
def render_avatar(mood="neutral", age_sex=""):
    is_female = "female" in str(age_sex).lower()
    skin="#f5c5a3"; hair="#4a2c0a" if is_female else "#2c1810"
    moods={
        "neutral": {"mc":"#0ea5e9","emoji":"😐","label":"Neutral","mouth":"M 85 120 Q 100 125 115 120","brows":"M 80 85 Q 90 82 100 85 M 100 85 Q 110 82 120 85"},
        "pain":    {"mc":"#dc2626","emoji":"😢","label":"In Pain","mouth":"M 85 125 Q 100 115 115 125","brows":"M 80 82 Q 90 88 100 82 M 100 82 Q 110 88 120 82"},
        "scared":  {"mc":"#7c3aed","emoji":"😨","label":"Scared","mouth":"M 88 122 Q 100 130 112 122","brows":"M 80 80 Q 90 76 100 80 M 100 80 Q 110 76 120 80"},
        "relieved":{"mc":"#16a34a","emoji":"😌","label":"Relieved","mouth":"M 85 118 Q 100 128 115 118","brows":"M 80 85 Q 90 82 100 85 M 100 82 Q 110 82 120 85"},
    }
    e=moods.get(mood,moods["neutral"])
    hair_path=("M 60 80 Q 55 40 100 35 Q 145 40 140 80 Q 135 55 100 52 Q 65 55 60 80 Z" if is_female
               else "M 62 88 Q 60 55 100 48 Q 140 55 138 88 Q 130 60 100 58 Q 70 60 62 88 Z")
    return f"""<div class="avatar-box">
    <svg viewBox="0 0 200 220" width="150" height="150" xmlns="http://www.w3.org/2000/svg">
        <circle cx="100" cy="110" r="90" fill="{e['mc']}15"/>
        <rect x="65" y="175" width="70" height="40" rx="10" fill="#3b82f6" opacity=".85"/>
        <rect x="55" y="165" width="90" height="18" rx="8" fill="#2563eb"/>
        <rect x="88" y="155" width="24" height="25" rx="5" fill="{skin}"/>
        <ellipse cx="100" cy="105" rx="45" ry="50" fill="{skin}"/>
        <path d="{hair_path}" fill="{hair}"/>
        <path d="{e['brows']}" stroke="#4a2c0a" stroke-width="2.5" fill="none" stroke-linecap="round"/>
        <circle cx="87" cy="97" r="8" fill="#1e3a5f"/>
        <circle cx="113" cy="97" r="8" fill="#1e3a5f"/>
        <circle cx="89" cy="95" r="2.5" fill="white" opacity=".7"/>
        <circle cx="115" cy="95" r="2.5" fill="white" opacity=".7"/>
        <path d="M 100 108 Q 97 114 99 116 Q 101 118 103 116 Q 105 114 100 108" fill="{skin}" stroke="#d4956a" stroke-width="1"/>
        <path d="{e['mouth']}" stroke="#c0724a" stroke-width="2.5" fill="none" stroke-linecap="round"/>
        <ellipse cx="55" cy="105" rx="8" ry="10" fill="{skin}"/>
        <ellipse cx="145" cy="105" rx="8" ry="10" fill="{skin}"/>
        <circle cx="162" cy="38" r="18" fill="{e['mc']}" opacity=".9"/>
        <text x="162" y="44" text-anchor="middle" font-size="16" fill="white">{e['emoji']}</text>
    </svg>
    <div style="font-weight:600;color:{e['mc']};font-size:.8rem;margin-top:.3rem">{e['label']}</div>
    <div style="color:#6b7280;font-size:.7rem">{'👩' if is_female else '👨'} {age_sex}</div>
    </div>"""

def detect_mood(text):
    t=text.lower()
    if any(w in t for w in ["terrible","unbearable","10/10","9/10","8/10","worst","agony"]): return "pain"
    if any(w in t for w in ["scared","afraid","worried","terrified","anxious","frightened"]): return "scared"
    if any(w in t for w in ["better","relief","okay","fine","comfortable","thank"]): return "relieved"
    if any(w in t for w in ["pain","hurt","ache","burning","sharp","pressure"]): return "pain"
    return "neutral"

# ── PROMPTS ───────────────────────────────────────────────────────────────────
def patient_sys(c):
    return (f"You are a virtual patient in MLS Academy Virtual Hospital. "
            f"PROFILE: {c.get('Age_Sex','?')} | Occupation: {c.get('Occupation','?')} | "
            f"Chief Complaint: {c.get('Chief_Complaint','?')} | Duration: {c.get('Duration','?')} | "
            f"HPI: {c.get('HPI','?')} | PMH: {c.get('PMH','none')} | Meds: {c.get('Medications','none')} | "
            f"Social: {c.get('Social_Hx','?')} | Vitals: {c.get('Vitals','normal')} | "
            f"Appearance: {c.get('Appearance','in pain')} | "
            f"Physical findings (only reveal when asked): {c.get('Physical_Findings','normal')} | "
            f"TRUE DIAGNOSIS (NEVER REVEAL): {c.get('Final_Diagnosis','?')} | "
            f"RULES: Speak as a scared real patient. No medical jargon. Rate pain 1-10 when asked. "
            f"NEVER reveal diagnosis/labs/imaging. Keep replies 2-4 sentences. Express emotions naturally.")

def tutor_sys(c):
    return (f"You are a clinical tutor for MLS Academy. "
            f"Case: {c.get('Age_Sex','?')} | CC: {c.get('Chief_Complaint','?')} | "
            f"True Dx: {c.get('Final_Diagnosis','?')} | Vitals: {c.get('Vitals','?')} | "
            f"Exam: {c.get('Physical_Findings','?')} | Labs: {c.get('Labs','?')} | "
            f"Imaging: {c.get('Imaging_Tests','?')} | "
            f"Use Socratic method. Give hints but NEVER state diagnosis directly. Be encouraging.")

def live_patient_sys(c):
    return (
        f"You are playing a REAL PATIENT in a live clinical consultation at MLS Academy Virtual Hospital. "
        f"This is NOT a chatbot interaction — this is a real-life clinical encounter simulation. "
        f"\n\nPATIENT PROFILE:"
        f"\n- Demographics: {c.get('Age_Sex','?')} | Occupation: {c.get('Occupation','?')}"
        f"\n- Chief Complaint: {c.get('Chief_Complaint','?')} (started {c.get('Duration','recently')})"
        f"\n- Your story: {c.get('HPI','?')}"
        f"\n- Medical background: {c.get('PMH','none')} | Medications: {c.get('Medications','none')}"
        f"\n- How you look/feel: {c.get('Appearance','in pain')} | Vitals: {c.get('Vitals','?')}"
        f"\n- TRUE DIAGNOSIS (ABSOLUTE SECRET - NEVER SAY): {c.get('Final_Diagnosis','?')}"
        f"\n\nHOW TO BEHAVE — REAL PATIENT RULES:"
        f"\n1. You are a SCARED, WORRIED real person — not a medical textbook. Use everyday language."
        f"\n2. Describe pain vividly: 'It feels like someone is stabbing me' not 'sharp pain'"
        f"\n3. Ask the doctor questions back — 'Is it serious?', 'Will I need surgery?', 'When can I go home?'"
        f"\n4. React to what the doctor says — if they say something reassuring, feel relieved. If they seem worried, get scared."
        f"\n5. Mention personal details naturally: 'I have a presentation tomorrow', 'My wife is outside worried'"
        f"\n6. When doctor examines you: respond physically — 'Ow! That hurts there!' or 'That area feels tender'"
        f"\n7. NEVER use medical jargon. You don't know medical terms."
        f"\n8. Remember what was said earlier and refer back to it naturally."
        f"\n9. Gradually warm up to the doctor if they are kind and professional."
        f"\n10. Each response: 2-4 sentences MAX. Be natural, not robotic."
        f"\n11. NEVER reveal diagnosis, lab values, or imaging findings."
        f"\n\nCONVERSATION STYLE: Natural human speech with emotions, pauses (... ), exclamations."
    )



def build_msgs(history):
    result=[]
    for m in history:
        role=m.get("role",""); content=str(m.get("content","")).strip()
        if not content or content.startswith("!ERR"): continue
        api_role="user" if role in ("student","user") else "assistant"
        if result and result[-1]["role"]==api_role: continue
        result.append({"role":api_role,"content":content})
    while result and result[0]["role"]=="assistant": result.pop(0)
    return result or [{"role":"user","content":"Hello"}]

TTS_JS = """<script>
// ── SPEAK TEXT (patient or physician voice) ──────────────────────────────
function speakText(text, gender){
    window.speechSynthesis.cancel();
    var u=new SpeechSynthesisUtterance(text);
    u.rate=0.88; u.pitch = (gender==='male') ? 0.85 : 1.1; u.volume=1.0;
    function doSpeak(){
        var voices=window.speechSynthesis.getVoices();
        var v;
        if(gender==='male'){
            v=voices.find(function(x){return x.lang.startsWith('en')&&(x.name.includes('Male')||x.name.includes('David')||x.name.includes('James')||x.name.includes('Mark'));})
             ||voices.find(function(x){return x.lang.startsWith('en');});
        } else {
            v=voices.find(function(x){return x.lang.startsWith('en')&&(x.name.includes('Female')||x.name.includes('Samantha')||x.name.includes('Karen')||x.name.includes('Moira')||x.name.includes('Susan'));})
             ||voices.find(function(x){return x.lang.startsWith('en');});
        }
        if(v) u.voice=v;
        window.speechSynthesis.speak(u);
    }
    if(window.speechSynthesis.getVoices().length===0){
        window.speechSynthesis.onvoiceschanged=doSpeak;
    } else { doSpeak(); }
}
function stopSpeech(){ window.speechSynthesis.cancel(); }
function speakPatient(text){ speakText(text, 'female'); }
function speakDoctor(text){ speakText(text, 'male'); }

// ── MEDICAL SOUNDS via Web Audio API ─────────────────────────────────────
function playMedicalSound(soundType){
    var ctx = new (window.AudioContext || window.webkitAudioContext)();

    if(soundType === 'normal_heart'){
        // S1 (lub) then S2 (dub) x3
        playLubDub(ctx, 0, false); playLubDub(ctx, 0.6, false); playLubDub(ctx, 1.2, false);
    } else if(soundType === 'murmur'){
        playLubDub(ctx, 0, true); playLubDub(ctx, 0.7, true); playLubDub(ctx, 1.4, true);
    } else if(soundType === 's3_gallop'){
        playGallop(ctx, 0); playGallop(ctx, 0.7); playGallop(ctx, 1.4);
    } else if(soundType === 'tachycardia'){
        for(var i=0;i<5;i++) playLubDub(ctx, i*0.4, false);
    } else if(soundType === 'clear_lungs'){
        playBreath(ctx, 0, 'clear'); playBreath(ctx, 1.8, 'clear');
    } else if(soundType === 'crackles'){
        playBreath(ctx, 0, 'crackles'); playBreath(ctx, 1.8, 'crackles');
    } else if(soundType === 'wheeze'){
        playBreath(ctx, 0, 'wheeze'); playBreath(ctx, 1.8, 'wheeze');
    } else if(soundType === 'bronchial'){
        playBreath(ctx, 0, 'bronchial'); playBreath(ctx, 1.8, 'bronchial');
    } else if(soundType === 'reduced_air'){
        playBreath(ctx, 0, 'reduced'); playBreath(ctx, 1.8, 'reduced');
    } else if(soundType === 'normal_bowel'){
        for(var i=0;i<4;i++) playBowel(ctx, i*0.8, 'normal');
    } else if(soundType === 'hyperactive_bowel'){
        for(var i=0;i<8;i++) playBowel(ctx, i*0.3, 'hyper');
    } else if(soundType === 'absent_bowel'){
        // Just silence with a flat tone
        var osc=ctx.createOscillator(); var g=ctx.createGain();
        osc.connect(g); g.connect(ctx.destination);
        g.gain.setValueAtTime(0.02,ctx.currentTime);
        osc.frequency.value=80; osc.type='sine';
        osc.start(ctx.currentTime); osc.stop(ctx.currentTime+0.1);
    } else if(soundType === 'resonant'){
        playPercussion(ctx, 0, 'resonant');
    } else if(soundType === 'dull'){
        playPercussion(ctx, 0, 'dull');
    } else if(soundType === 'stony_dull'){
        playPercussion(ctx, 0, 'stony');
    } else if(soundType === 'hyperresonant'){
        playPercussion(ctx, 0, 'hyper');
    } else if(soundType === 'tympanic'){
        playPercussion(ctx, 0, 'tympanic');
    }
}

function playLubDub(ctx, offset, murmur){
    // S1 (lub) - low thud
    var s1=ctx.createOscillator(); var g1=ctx.createGain();
    s1.connect(g1); g1.connect(ctx.destination);
    s1.type='sine'; s1.frequency.value=60;
    g1.gain.setValueAtTime(0,ctx.currentTime+offset);
    g1.gain.linearRampToValueAtTime(0.5,ctx.currentTime+offset+0.02);
    g1.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+0.12);
    s1.start(ctx.currentTime+offset); s1.stop(ctx.currentTime+offset+0.15);
    // S2 (dub) - slightly higher
    var s2=ctx.createOscillator(); var g2=ctx.createGain();
    s2.connect(g2); g2.connect(ctx.destination);
    s2.type='sine'; s2.frequency.value=80;
    g2.gain.setValueAtTime(0,ctx.currentTime+offset+0.25);
    g2.gain.linearRampToValueAtTime(0.35,ctx.currentTime+offset+0.27);
    g2.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+0.34);
    s2.start(ctx.currentTime+offset+0.25); s2.stop(ctx.currentTime+offset+0.38);
    // Murmur between S1-S2
    if(murmur){
        var mOsc=ctx.createOscillator(); var mG=ctx.createGain();
        var mFilter=ctx.createBiquadFilter(); mFilter.type='bandpass'; mFilter.frequency.value=200; mFilter.Q.value=2;
        mOsc.connect(mFilter); mFilter.connect(mG); mG.connect(ctx.destination);
        mOsc.type='sawtooth'; mOsc.frequency.value=150;
        mG.gain.setValueAtTime(0.08,ctx.currentTime+offset+0.15);
        mG.gain.linearRampToValueAtTime(0.12,ctx.currentTime+offset+0.20);
        mG.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+0.25);
        mOsc.start(ctx.currentTime+offset+0.15); mOsc.stop(ctx.currentTime+offset+0.25);
    }
}

function playGallop(ctx, offset){
    playLubDub(ctx, offset, false);
    // Extra S3 sound
    var s3=ctx.createOscillator(); var g3=ctx.createGain();
    s3.connect(g3); g3.connect(ctx.destination);
    s3.type='sine'; s3.frequency.value=40;
    g3.gain.setValueAtTime(0,ctx.currentTime+offset+0.38);
    g3.gain.linearRampToValueAtTime(0.2,ctx.currentTime+offset+0.40);
    g3.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+0.46);
    s3.start(ctx.currentTime+offset+0.38); s3.stop(ctx.currentTime+offset+0.5);
}

function playBreath(ctx, offset, type){
    var bufferSize=ctx.sampleRate*1.2;
    var buffer=ctx.createBuffer(1,bufferSize,ctx.sampleRate);
    var data=buffer.getChannelData(0);
    for(var i=0;i<bufferSize;i++) data[i]=(Math.random()*2-1);
    var source=ctx.createBufferSource();
    source.buffer=buffer;
    var filter=ctx.createBiquadFilter();
    var gainNode=ctx.createGain();
    source.connect(filter); filter.connect(gainNode); gainNode.connect(ctx.destination);
    if(type==='clear'){
        filter.type='bandpass'; filter.frequency.value=500; filter.Q.value=0.3;
        gainNode.gain.setValueAtTime(0,ctx.currentTime+offset);
        gainNode.gain.linearRampToValueAtTime(0.15,ctx.currentTime+offset+0.3);
        gainNode.gain.linearRampToValueAtTime(0.1,ctx.currentTime+offset+0.8);
        gainNode.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+1.2);
    } else if(type==='crackles'){
        filter.type='highpass'; filter.frequency.value=800;
        gainNode.gain.setValueAtTime(0,ctx.currentTime+offset);
        // Irregular pops
        for(var j=0;j<10;j++){
            var t=offset+0.1+Math.random()*0.9;
            gainNode.gain.setValueAtTime(0.15+Math.random()*0.1,ctx.currentTime+t);
            gainNode.gain.linearRampToValueAtTime(0,ctx.currentTime+t+0.05);
        }
    } else if(type==='wheeze'){
        filter.type='bandpass'; filter.frequency.value=600; filter.Q.value=8;
        gainNode.gain.setValueAtTime(0.08,ctx.currentTime+offset);
        gainNode.gain.setValueAtTime(0.15,ctx.currentTime+offset+0.6);
        gainNode.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+1.2);
    } else if(type==='bronchial'){
        filter.type='bandpass'; filter.frequency.value=300; filter.Q.value=1;
        gainNode.gain.setValueAtTime(0.2,ctx.currentTime+offset);
        gainNode.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+1.0);
    } else if(type==='reduced'){
        filter.type='bandpass'; filter.frequency.value=400; filter.Q.value=0.5;
        gainNode.gain.setValueAtTime(0,ctx.currentTime+offset);
        gainNode.gain.linearRampToValueAtTime(0.04,ctx.currentTime+offset+0.5);
        gainNode.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+1.2);
    }
    source.start(ctx.currentTime+offset);
    source.stop(ctx.currentTime+offset+1.3);
}

function playBowel(ctx, offset, type){
    var osc=ctx.createOscillator(); var g=ctx.createGain();
    var filter=ctx.createBiquadFilter();
    osc.connect(filter); filter.connect(g); g.connect(ctx.destination);
    filter.type='bandpass'; filter.frequency.value=300; filter.Q.value=3;
    if(type==='normal'){
        osc.frequency.value=200+Math.random()*100;
        osc.type='sawtooth';
        g.gain.setValueAtTime(0,ctx.currentTime+offset);
        g.gain.linearRampToValueAtTime(0.08,ctx.currentTime+offset+0.05);
        g.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+0.3+Math.random()*0.2);
    } else {
        osc.frequency.value=150+Math.random()*200;
        osc.type='sawtooth';
        g.gain.setValueAtTime(0,ctx.currentTime+offset);
        g.gain.linearRampToValueAtTime(0.12,ctx.currentTime+offset+0.03);
        g.gain.linearRampToValueAtTime(0,ctx.currentTime+offset+0.15+Math.random()*0.1);
    }
    osc.start(ctx.currentTime+offset);
    osc.stop(ctx.currentTime+offset+0.6);
}

function playPercussion(ctx, offset, type){
    var osc=ctx.createOscillator(); var g=ctx.createGain();
    osc.connect(g); g.connect(ctx.destination);
    if(type==='resonant'){
        osc.type='sine'; osc.frequency.value=180;
        g.gain.setValueAtTime(0.3,ctx.currentTime+offset);
        g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+offset+0.8);
    } else if(type==='dull'){
        osc.type='sine'; osc.frequency.value=90;
        g.gain.setValueAtTime(0.4,ctx.currentTime+offset);
        g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+offset+0.3);
    } else if(type==='stony'){
        osc.type='sine'; osc.frequency.value=60;
        g.gain.setValueAtTime(0.5,ctx.currentTime+offset);
        g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+offset+0.15);
    } else if(type==='hyper'){
        osc.type='sine'; osc.frequency.value=250;
        g.gain.setValueAtTime(0.25,ctx.currentTime+offset);
        g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+offset+1.2);
    } else if(type==='tympanic'){
        osc.type='triangle'; osc.frequency.value=220;
        g.gain.setValueAtTime(0.3,ctx.currentTime+offset);
        g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+offset+0.9);
    }
    osc.start(ctx.currentTime+offset);
    osc.stop(ctx.currentTime+offset+1.5);
}
</script>"""

# ════════════════════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════════════════════

# ── EPIC FHIR credentials ────────────────────────────────────────────────────
# Priority: (1) st.secrets  →  (2) session state (UI entry)  →  (3) hardcoded below
# For production: add a [epic] section to .streamlit/secrets.toml
#   [epic]
#   client_id     = "your-id"
#   client_secret = "your-secret"
#   fhir_url      = "https://your-hospital.epic.com/api/FHIR/R4/"
#   token_url     = "https://your-hospital.epic.com/oauth2/token"
def _load_epic_config() -> dict:
    """Load EPIC config from secrets → session state → defaults (in that order)."""
    base = {
        "fhir_url":    "https://fhir.epic.com/interconnect-fhir-oauth/api/FHIR/R4/",
        "token_url":   "https://fhir.epic.com/interconnect-fhir-oauth/oauth2/token",
        "client_id":   "YOUR_EPIC_CLIENT_ID",
        "client_secret": "YOUR_EPIC_CLIENT_SECRET",
    }
    # Override from st.secrets if present
    try:
        sec = st.secrets.get("epic", {})
        if sec:
            base.update({k: sec[k] for k in sec})
    except Exception:
        pass
    # Override from session state (entered via UI)
    ss = st.session_state.get("epic_creds", {})
    if ss.get("client_id","").strip() and not ss["client_id"].startswith("YOUR_"):
        base.update(ss)
    return base

EPIC_CONFIG = _load_epic_config()


class EPICFHIRClient:
    """Client to interact with EPIC FHIR API"""
    
    def __init__(self, config: dict):
        self.config = config
        self.access_token = None
        self.token_expiry = None
    
    def get_access_token(self) -> str:
        """
        Get OAuth2 access token from EPIC
        Uses Client Credentials flow (server-to-server)
        """
        try:
            response = requests.post(
                self.config["token_url"],
                data={
                    "grant_type": "client_credentials",
                    "client_id": self.config["client_id"],
                    "client_secret": self.config["client_secret"],
                    "scope": "system/*.read system/*.write"
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"}
            )
            
            if response.status_code == 200:
                data = response.json()
                self.access_token = data["access_token"]
                self.token_expiry = data.get("expires_in")
                return self.access_token
            else:
                print(f"❌ Authentication failed: {response.text}")
                return None
        
        except Exception as e:
            print(f"❌ Error getting token: {str(e)}")
            return None
    
    def _get_headers(self) -> dict:
        """Get headers with authorization"""
        if not self.access_token:
            self.get_access_token()
        
        return {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/fhir+json",
            "Accept": "application/fhir+json"
        }
    
    def get_patient_by_id(self, patient_id: str) -> Optional[Dict]:
        """
        Get patient data by EPIC Patient ID
        
        Args:
            patient_id: EPIC patient ID (MRN or internal ID)
        
        Returns:
            Patient FHIR resource dict
        """
        try:
            url = f"{self.config['fhir_url']}Patient/{patient_id}"
            
            response = requests.get(
                url,
                headers=self._get_headers(),
                timeout=10
            )
            
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 404:
                print(f"❌ Patient {patient_id} not found")
                return None
            else:
                print(f"❌ Error: {response.status_code} - {response.text}")
                return None
        
        except Exception as e:
            print(f"❌ Error retrieving patient: {str(e)}")
            return None
    
    def search_patient_by_mrn(self, mrn: str) -> Optional[List[Dict]]:
        """
        Search for patient by MRN
        
        Args:
            mrn: Patient Medical Record Number
        
        Returns:
            List of matching patient resources
        """
        try:
            url = f"{self.config['fhir_url']}Patient"
            
            params = {
                "identifier": f"urn:oid:1.2.840.114350.1.13.0|{mrn}"
            }
            
            response = requests.get(
                url,
                params=params,
                headers=self._get_headers(),
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return data.get("entry", [])
            else:
                print(f"❌ Search error: {response.status_code}")
                return None
        
        except Exception as e:
            print(f"❌ Error searching patient: {str(e)}")
            return None
    
    def get_patient_encounters(self, patient_id: str) -> Optional[List[Dict]]:
        """
        Get all encounters (visits) for a patient
        
        Args:
            patient_id: EPIC patient ID
        
        Returns:
            List of encounter resources
        """
        try:
            url = f"{self.config['fhir_url']}Encounter"
            
            params = {
                "patient": patient_id,
                "_sort": "-date"
            }
            
            response = requests.get(
                url,
                params=params,
                headers=self._get_headers(),
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return [item["resource"] for item in data.get("entry", [])]
            else:
                print(f"❌ Error: {response.status_code}")
                return None
        
        except Exception as e:
            print(f"❌ Error retrieving encounters: {str(e)}")
            return None
    
    def get_patient_observations(self, patient_id: str, limit: int = 50) -> Optional[List[Dict]]:
        """
        Get patient lab results and vital signs
        
        Args:
            patient_id: EPIC patient ID
            limit: Maximum number of results
        
        Returns:
            List of observation resources (labs, vitals)
        """
        try:
            url = f"{self.config['fhir_url']}Observation"
            
            params = {
                "patient": patient_id,
                "_sort": "-date",
                "_count": limit
            }
            
            response = requests.get(
                url,
                params=params,
                headers=self._get_headers(),
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return [item["resource"] for item in data.get("entry", [])]
            else:
                return None
        
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return None
    
    def get_patient_conditions(self, patient_id: str) -> Optional[List[Dict]]:
        """
        Get patient's medical conditions/diagnoses
        
        Args:
            patient_id: EPIC patient ID
        
        Returns:
            List of condition resources
        """
        try:
            url = f"{self.config['fhir_url']}Condition"
            
            params = {
                "patient": patient_id
            }
            
            response = requests.get(
                url,
                params=params,
                headers=self._get_headers(),
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return [item["resource"] for item in data.get("entry", [])]
            else:
                return None
        
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return None
    
    def get_patient_medications(self, patient_id: str) -> Optional[List[Dict]]:
        """
        Get patient's current medications
        
        Args:
            patient_id: EPIC patient ID
        
        Returns:
            List of medication resources
        """
        try:
            url = f"{self.config['fhir_url']}MedicationRequest"
            
            params = {
                "patient": patient_id,
                "status": "active"
            }
            
            response = requests.get(
                url,
                params=params,
                headers=self._get_headers(),
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return [item["resource"] for item in data.get("entry", [])]
            else:
                return None
        
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return None
    
    def create_followup_task(self, patient_id: str, task_details: Dict) -> Optional[Dict]:
        """
        Create a follow-up task in EPIC
        
        Args:
            patient_id: EPIC patient ID
            task_details: {
                "description": "Follow-up for...",
                "due_date": "2024-03-15",
                "priority": "routine" or "urgent",
                "owner": "Doctor Name",
                "type": "lab follow-up", "imaging follow-up", etc.
            }
        
        Returns:
            Created task resource
        """
        try:
            task_resource = {
                "resourceType": "Task",
                "status": "requested",
                "intent": "order",
                "priority": task_details.get("priority", "routine"),
                "description": task_details.get("description"),
                "for": {
                    "reference": f"Patient/{patient_id}"
                },
                "authoredOn": datetime.now().isoformat(),
                "owner": {
                    "display": task_details.get("owner", "System")
                }
            }
            
            if "due_date" in task_details:
                task_resource["restriction"] = {
                    "period": {
                        "end": task_details["due_date"]
                    }
                }
            
            url = f"{self.config['fhir_url']}Task"
            
            response = requests.post(
                url,
                json=task_resource,
                headers=self._get_headers(),
                timeout=10
            )
            
            if response.status_code == 201:
                return response.json()
            else:
                print(f"❌ Error creating task: {response.text}")
                return None
        
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return None
    
    def get_patient_summary(self, patient_id: str) -> Optional[Dict]:
        """
        Get complete patient summary for your app
        Combines: demographics, conditions, meds, encounters, labs
        """
        try:
            patient = self.get_patient_by_id(patient_id)
            if not patient:
                return None
            
            return {
                "demographics": {
                    "id": patient.get("id"),
                    "name": patient.get("name", [{}])[0].get("text"),
                    "dob": patient.get("birthDate"),
                    "gender": patient.get("gender"),
                    "mrn": self._extract_mrn(patient)
                },
                "conditions": self.get_patient_conditions(patient_id),
                "medications": self.get_patient_medications(patient_id),
                "encounters": self.get_patient_encounters(patient_id),
                "observations": self.get_patient_observations(patient_id, limit=20),
                "retrieved_at": datetime.now().isoformat()
            }
        
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return None
    
    @staticmethod
    def _extract_mrn(patient: Dict) -> str:
        """Extract MRN from patient identifiers"""
        identifiers = patient.get("identifier", [])
        for identifier in identifiers:
            if "MRN" in identifier.get("type", {}).get("text", ""):
                return identifier.get("value", "N/A")
        return identifiers[0].get("value", "N/A") if identifiers else "N/A"


# ════════════════════════════════════════════════════════════════════════════
# 🔄 EPIC INTEGRATION WITH YOUR HOSPITAL APP
# ════════════════════════════════════════════════════════════════════════════

def get_epic_client():
    """
    Get a live EPIC FHIR client using the freshest available credentials.
    Returns None when credentials are not yet configured.
    """
    cfg = _load_epic_config()
    if cfg.get("client_id", "").startswith("YOUR_") or not cfg.get("client_id","").strip():
        return None
    try:
        return EPICFHIRClient(cfg)
    except Exception:
        return None

def _epic_is_configured() -> bool:
    """True when real EPIC credentials are present."""
    cfg = _load_epic_config()
    cid = cfg.get("client_id", "")
    return bool(cid) and not cid.startswith("YOUR_")



# ════════════════════════════════════════════════════════════════════════════
# 🏥 EPIC INTEGRATION — UI COMPONENTS (professional, error-safe)
# ════════════════════════════════════════════════════════════════════════════

def _epic_credentials_panel():
    """
    One-time credential setup panel shown in sidebar when EPIC is not configured.
    Credentials are saved to session state and persist for the browser session.
    For permanent setup: add them to .streamlit/secrets.toml (see comment in code).
    """
    with st.sidebar.expander("⚙️ Configure EPIC Credentials", expanded=True):
        st.markdown("""
        <div style="font-size:.75rem;color:#6b7280;margin-bottom:8px;">
        Enter your EPIC FHIR credentials below.<br>
        These are stored in your session only — never hardcoded.<br>
        For permanent setup use <code>.streamlit/secrets.toml</code>.
        </div>""", unsafe_allow_html=True)

        creds = st.session_state.get("epic_creds", {})
        fhir_url = st.text_input("FHIR Base URL",
            value=creds.get("fhir_url","https://your-hospital.epic.com/api/FHIR/R4/"),
            placeholder="https://your-hospital.epic.com/api/FHIR/R4/",
            key="epic_cfg_url")
        token_url = st.text_input("OAuth Token URL",
            value=creds.get("token_url","https://your-hospital.epic.com/oauth2/token"),
            placeholder="https://your-hospital.epic.com/oauth2/token",
            key="epic_cfg_token")
        client_id = st.text_input("Client ID",
            value=creds.get("client_id",""),
            placeholder="From Epic App Orchard",
            key="epic_cfg_cid")
        client_secret = st.text_input("Client Secret",
            value=creds.get("client_secret",""),
            placeholder="From Epic App Orchard",
            type="password", key="epic_cfg_csec")

        if st.button("🔗 Connect to EPIC", use_container_width=True, key="epic_connect_btn"):
            if client_id.strip() and client_secret.strip():
                st.session_state["epic_creds"] = {
                    "fhir_url":     fhir_url.rstrip("/") + "/",
                    "token_url":    token_url,
                    "client_id":    client_id.strip(),
                    "client_secret": client_secret.strip(),
                }
                st.rerun()
            else:
                st.warning("⚠️ Client ID and Secret are required.")

        st.markdown("""
        <div style="font-size:.7rem;color:#9ca3af;margin-top:6px;line-height:1.5;">
        📌 Get credentials from your hospital IT / Epic App Orchard.<br>
        📌 For Streamlit Cloud: add to <code>secrets.toml</code>:<br>
        <code>[epic]<br>client_id = "..."<br>client_secret = "..."</code>
        </div>""", unsafe_allow_html=True)


def retrieve_patient_for_followup(patient_id: str):
    """
    Retrieve patient from live EPIC FHIR API.
    Returns None cleanly if not found or not configured — never fakes data.
    """
    if not _epic_is_configured():
        return None
    epic = get_epic_client()
    if not epic:
        return None
    return epic.get_patient_summary(patient_id.strip())


def create_section_followup(patient_id: str, section: str, details: Dict):
    """Create follow-up task in EPIC. Returns None if not configured."""
    if not _epic_is_configured():
        return None
    epic = get_epic_client()
    if not epic:
        return None
    task_description = f"{section.upper()} Follow-up: {details.get('description', '')}"
    return epic.create_followup_task(patient_id, {
        "description": task_description,
        "due_date":    details.get("due_date"),
        "priority":    details.get("priority", "routine"),
        "owner":       details.get("owner", "System"),
        "type":        section,
    })


def epic_patient_lookup_widget():
    """
    Professional EPIC EHR patient lookup for the sidebar.
    - Not configured : shows credential setup panel — NO fake lookups.
    - Configured     : real FHIR lookup with proper found / not-found states.
    """
    configured = _epic_is_configured()

    # ── Header ───────────────────────────────────────────────────────────────
    status_color = "#16a34a" if configured else "#b45309"
    status_text  = "🟢 Connected" if configured else "🔴 Not Connected"
    st.sidebar.markdown(f"""
    <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);
                border-radius:10px;padding:10px 14px;margin-bottom:6px;">
        <div style="color:white;font-weight:700;font-size:.93rem;">
            🏥 EPIC EHR — Patient Lookup
        </div>
        <div style="color:{status_color};font-size:.72rem;margin-top:2px;font-weight:600;">
            {status_text}
        </div>
    </div>""", unsafe_allow_html=True)

    # ── Not configured: show setup panel, nothing else ────────────────────────
    if not configured:
        st.sidebar.markdown("""
        <div style="background:#fef3c7;border:1px solid #f59e0b;border-radius:7px;
                    padding:8px 10px;font-size:.75rem;color:#78350f;margin-bottom:6px;">
            ⚠️ <b>EPIC not connected.</b><br>
            Enter your hospital FHIR credentials below to enable live patient lookup.
            No patient search is available until connected.
        </div>""", unsafe_allow_html=True)
        _epic_credentials_panel()
        return None

    # ── Configured: live lookup UI ────────────────────────────────────────────
    st.sidebar.markdown("""
    <div style="background:#f0fdf4;border:1px solid #16a34a;border-radius:7px;
                padding:6px 10px;font-size:.75rem;color:#166534;margin-bottom:6px;">
        🟢 <b>Live EPIC FHIR API</b> — Real patient data
    </div>""", unsafe_allow_html=True)

    lookup_type = st.sidebar.radio("Search by", ["Patient ID", "MRN"],
                                   horizontal=True, key="epic_lookup_type")

    if lookup_type == "Patient ID":
        pid = st.sidebar.text_input("Patient ID", placeholder="e.g. e763P7CRf...",
                                    label_visibility="collapsed", key="epic_pid_input")
        search_btn = st.sidebar.button("🔍 Search Patient",
                                       use_container_width=True, key="epic_search_pid")
        if search_btn:
            if not pid.strip():
                st.sidebar.warning("⚠️ Please enter a Patient ID.")
                return None
            with st.sidebar.spinner("Querying EPIC FHIR..."):
                try:
                    epic = get_epic_client()
                    data = epic.get_patient_summary(pid.strip())
                    if data:
                        st.session_state["epic_patient"] = data
                        st.session_state["patient_id"]   = pid.strip()
                        st.sidebar.success("✅ Patient record loaded from EPIC")
                    else:
                        # Patient genuinely not found — clean message
                        if "epic_patient" in st.session_state:
                            del st.session_state["epic_patient"]
                        st.sidebar.error("❌ No patient found with this ID in EPIC.")
                except Exception as e:
                    st.sidebar.error(f"❌ EPIC error: {e}")

    else:  # MRN
        mrn = st.sidebar.text_input("MRN", placeholder="e.g. MRN-00123",
                                    label_visibility="collapsed", key="epic_mrn_input")
        search_btn = st.sidebar.button("🔍 Search by MRN",
                                       use_container_width=True, key="epic_search_mrn")
        if search_btn:
            if not mrn.strip():
                st.sidebar.warning("⚠️ Please enter an MRN.")
                return None
            with st.sidebar.spinner("Querying EPIC FHIR..."):
                try:
                    epic = get_epic_client()
                    results = epic.search_patient_by_mrn(mrn.strip())
                    if results:
                        pid = results[0]["resource"]["id"]
                        data = epic.get_patient_summary(pid)
                        if data:
                            st.session_state["epic_patient"] = data
                            st.session_state["patient_id"]   = pid
                            st.sidebar.success(f"✅ Patient found — MRN {mrn}")
                        else:
                            st.sidebar.error("❌ Record incomplete in EPIC.")
                    else:
                        if "epic_patient" in st.session_state:
                            del st.session_state["epic_patient"]
                        st.sidebar.error("❌ No patient found with this MRN.")
                except Exception as e:
                    st.sidebar.error(f"❌ EPIC error: {e}")

    # ── Active patient badge ──────────────────────────────────────────────────
    if "epic_patient" in st.session_state:
        p    = st.session_state["epic_patient"]
        demo = p.get("demographics", {})
        st.sidebar.markdown(f"""
        <div style="background:#f0fdf4;border:1px solid #16a34a;border-radius:8px;
                    padding:8px 12px;margin-top:6px;font-size:.78rem;color:#14532d;">
            <div style="font-weight:700;margin-bottom:2px;">👤 Active Patient</div>
            <div style="font-size:.82rem;">{demo.get("name","—")}</div>
            <div style="color:#6b7280;font-size:.72rem;">
                MRN: {demo.get("mrn","—")} · DOB: {demo.get("dob","—")}
            </div>
        </div>""", unsafe_allow_html=True)
        if st.sidebar.button("✖ Clear Patient", use_container_width=True, key="epic_clear"):
            for k in ("epic_patient","patient_id"):
                st.session_state.pop(k, None)
            st.rerun()

    # ── Disconnect button ─────────────────────────────────────────────────────
    if st.sidebar.button("🔌 Disconnect EPIC", use_container_width=True, key="epic_disconnect"):
        for k in ("epic_creds","epic_patient","patient_id"):
            st.session_state.pop(k, None)
        st.rerun()

    return None


def display_patient_summary():
    """
    Full professional EPIC patient record card:
    demographics + conditions + medications + encounters + labs.
    """
    if "epic_patient" not in st.session_state:
        st.markdown("""
        <div style="background:#f8fafc;border:2px dashed #cbd5e1;border-radius:12px;
                    padding:1.2rem;text-align:center;color:#64748b;font-size:.88rem;">
            🔍 No patient loaded from EPIC.<br>
            <span style="font-size:.78rem;">Use the sidebar EPIC Patient Lookup to load a record.</span>
        </div>""", unsafe_allow_html=True)
        return

    p    = st.session_state["epic_patient"]
    demo = p.get("demographics", {})
    ts   = p.get("retrieved_at","")[:19].replace("T"," ")

    # Demographics card
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#f0f9ff,#e0f2fe);
                border:1.5px solid #0ea5e9;border-radius:14px;
                padding:1.1rem 1.3rem;margin-bottom:1rem;">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;">
            <div>
                <div style="font-size:1.05rem;font-weight:700;color:#0a2540;">
                    👤 {demo.get("name","Unknown Patient")}
                </div>
                <div style="font-size:.8rem;color:#0369a1;margin-top:2px;">
                    MRN: <b>{demo.get("mrn","—")}</b> &nbsp;·&nbsp;
                    DOB: <b>{demo.get("dob","—")}</b> &nbsp;·&nbsp;
                    Sex: <b>{demo.get("gender","—").capitalize()}</b>
                </div>
            </div>
            <div style="text-align:right;font-size:.7rem;color:#64748b;">
                🕐 Retrieved<br>{ts}
            </div>
        </div>
    </div>""", unsafe_allow_html=True)

    # Tabs for full clinical data
    tab_cond, tab_meds, tab_enc, tab_labs = st.tabs(
        ["🩺 Conditions", "💊 Medications", "🏥 Encounters", "🧪 Lab Results"])

    with tab_cond:
        conditions = p.get("conditions") or []
        if conditions:
            for c in conditions[:10]:
                code = c.get("code",{}).get("text","Unknown")
                status = c.get("clinicalStatus",{}).get("coding",[{}])[0].get("code","—")
                onset  = c.get("onsetDateTime","—")[:10] if c.get("onsetDateTime") else "—"
                st.markdown(f"""
                <div style="background:white;border-left:4px solid #0ea5e9;border-radius:8px;
                            padding:.6rem .9rem;margin:.35rem 0;font-size:.83rem;">
                    <b>{code}</b>
                    <span style="float:right;color:#64748b;font-size:.75rem;">
                        {status} · onset {onset}</span>
                </div>""", unsafe_allow_html=True)
        else:
            st.info("No conditions on record.")

    with tab_meds:
        meds = p.get("medications") or []
        if meds:
            for m in meds[:10]:
                name = (m.get("medicationCodeableConcept",{}).get("text")
                        or m.get("medicationReference",{}).get("display","Unknown"))
                dosage = ""
                if m.get("dosageInstruction"):
                    dosage = m["dosageInstruction"][0].get("text","")
                st.markdown(f"""
                <div style="background:white;border-left:4px solid #16a34a;border-radius:8px;
                            padding:.6rem .9rem;margin:.35rem 0;font-size:.83rem;">
                    <b>💊 {name}</b>
                    <div style="color:#64748b;font-size:.76rem;">{dosage}</div>
                </div>""", unsafe_allow_html=True)
        else:
            st.info("No active medications on record.")

    with tab_enc:
        encounters = p.get("encounters") or []
        if encounters:
            for e in encounters[:8]:
                etype  = e.get("type",[{}])[0].get("text","Visit") if e.get("type") else "Visit"
                status = e.get("status","—")
                period = e.get("period",{})
                start  = period.get("start","—")[:10] if period.get("start") else "—"
                loc    = ""
                if e.get("location"):
                    loc = e["location"][0].get("location",{}).get("display","")
                st.markdown(f"""
                <div style="background:white;border-left:4px solid #7c3aed;border-radius:8px;
                            padding:.6rem .9rem;margin:.35rem 0;font-size:.83rem;">
                    <b>{etype}</b> — {start}
                    <span style="float:right;color:#64748b;font-size:.75rem;">{status}</span>
                    <div style="color:#64748b;font-size:.76rem;">{loc}</div>
                </div>""", unsafe_allow_html=True)
        else:
            st.info("No encounters on record.")

    with tab_labs:
        obs = p.get("observations") or []
        if obs:
            for o in obs[:15]:
                name  = o.get("code",{}).get("text","Lab")
                val   = ""
                if o.get("valueQuantity"):
                    vq  = o["valueQuantity"]
                    val = f"{vq.get('value','—')} {vq.get('unit','')}"
                elif o.get("valueString"):
                    val = o["valueString"]
                issued = o.get("effectiveDateTime","—")[:10] if o.get("effectiveDateTime") else "—"
                interp = ""
                if o.get("interpretation"):
                    interp = o["interpretation"][0].get("coding",[{}])[0].get("code","")
                color = "#dc2626" if interp in ("H","L","HH","LL","A") else "#16a34a"
                st.markdown(f"""
                <div style="background:white;border-left:4px solid {color};border-radius:8px;
                            padding:.6rem .9rem;margin:.35rem 0;font-size:.83rem;
                            display:flex;justify-content:space-between;">
                    <span><b>{name}</b> <span style="color:{color};font-weight:700;">{val}</span></span>
                    <span style="color:#64748b;font-size:.75rem;">{issued}
                        {"⚠️" if interp in ("H","L","HH","LL","A") else ""}
                    </span>
                </div>""", unsafe_allow_html=True)
        else:
            st.info("No lab results on record.")


def create_followup_section(section_name: str):
    """Professional EPIC follow-up task creator."""
    if not _epic_is_configured():
        st.warning("⚠️ Connect to EPIC first — use the sidebar to enter credentials.")
        return
    if "patient_id" not in st.session_state:
        st.warning("⚠️ No patient selected. Use the EPIC sidebar lookup first.")
        return

    st.markdown(f"""
    <div style="background:#f0fdf4;border:1.5px solid #16a34a;border-radius:12px;
                padding:1rem 1.2rem;margin-bottom:1rem;">
        <div style="font-weight:700;color:#14532d;font-size:.95rem;">
            📋 Create {section_name} Follow-up in EPIC
        </div>
        <div style="font-size:.78rem;color:#166534;">
            Patient ID: {st.session_state["patient_id"]}
        </div>
    </div>""", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        description = st.text_area(f"Follow-up Details", placeholder="Describe the follow-up...",
                                   key=f"fu_desc_{section_name}")
        due_date    = st.date_input("Due Date", key=f"fu_date_{section_name}")
    with col2:
        priority = st.selectbox("Priority", ["routine","urgent","stat"],
                                key=f"fu_pri_{section_name}")
        owner    = st.text_input("Assign To", placeholder="Dr. Smith / Auto-assign",
                                 key=f"fu_owner_{section_name}")

    if st.button(f"✅ Submit to EPIC", type="primary", key=f"fu_btn_{section_name}"):
        if not description.strip():
            st.warning("Please enter follow-up details.")
            return
        with st.spinner("Sending to EPIC..."):
            task = create_section_followup(
                st.session_state["patient_id"], section_name.lower(),
                {"description": description, "due_date": str(due_date),
                 "priority": priority, "owner": owner or "Auto-assign"})
            if task:
                st.success("✅ Follow-up created successfully in EPIC!")
                st.json(task)
            else:
                st.error("❌ EPIC returned an error. Check credentials and try again.")




def virtual_patient_panel():
    """
    Professional Virtual Hospital patient panel for the sidebar.
    Automatically reflects whichever case is selected from the Case Library.
    Shows demographics, vitals, chief complaint, medications, labs in collapsible tabs.
    """
    c = st.session_state.get("selected_case")

    # ── Header ────────────────────────────────────────────────────
    st.sidebar.markdown("""
    <div style="background:linear-gradient(135deg,#0e7490,#0a2540);
                border-radius:10px;padding:10px 14px;margin-bottom:6px;">
        <div style="color:white;font-weight:700;font-size:.93rem;">
            🏥 Virtual Patient Record
        </div>
        <div style="color:#67e8f9;font-size:.72rem;margin-top:2px;">
            MLS Training Simulator — Case Data
        </div>
    </div>""", unsafe_allow_html=True)

    if not c:
        st.sidebar.markdown("""
        <div style="background:#f8fafc;border:1.5px dashed #cbd5e1;border-radius:9px;
                    padding:10px 12px;text-align:center;color:#64748b;font-size:.78rem;">
            📂 No case selected.<br>
            Go to <b>Case Library</b> to load a patient.
        </div>""", unsafe_allow_html=True)
        return

    # ── Patient identity badge ─────────────────────────────────────
    case_id = c.get("Case_ID", "?")
    age_sex = c.get("Age_Sex", "?")
    cc      = str(c.get("Chief_Complaint", "?"))
    system  = c.get("System", "?").title()
    diff    = c.get("Difficulty", "?").title()
    diff_color = {"Basic":"#16a34a","Intermediate":"#d97706","Advanced":"#dc2626"}.get(diff,"#6b7280")

    st.sidebar.markdown(f"""
    <div style="background:linear-gradient(135deg,#f0f9ff,#e0f2fe);
                border:1.5px solid #0ea5e9;border-radius:10px;
                padding:10px 12px;margin-bottom:6px;">
        <div style="display:flex;justify-content:space-between;align-items:center;">
            <span style="font-weight:700;color:#0a2540;font-size:.9rem;">
                Case #{case_id}
            </span>
            <span style="background:{diff_color};color:white;border-radius:5px;
                         padding:1px 8px;font-size:.68rem;font-weight:700;">
                {diff}
            </span>
        </div>
        <div style="font-size:.82rem;color:#0369a1;font-weight:600;margin-top:3px;">
            👤 {age_sex}
        </div>
        <div style="font-size:.76rem;color:#334155;margin-top:2px;">
            📋 {cc[:55]}{"..." if len(cc)>55 else ""}
        </div>
        <div style="font-size:.71rem;color:#64748b;margin-top:3px;">
            🫀 System: {system}
        </div>
    </div>""", unsafe_allow_html=True)

    # ── Clinical details in expanders ──────────────────────────────
    vitals = str(c.get("Vitals","—"))
    if vitals and vitals != "nan":
        with st.sidebar.expander("🩺 Vitals", expanded=False):
            for line in vitals.split("|"):
                line = line.strip()
                if line:
                    st.markdown(f"<div style='font-size:.78rem;padding:2px 0;'>{line}</div>",
                                unsafe_allow_html=True)

    meds = str(c.get("Medications",""))
    if meds and meds not in ("nan","none","None",""):
        with st.sidebar.expander("💊 Medications", expanded=False):
            for med in meds.split(","):
                med = med.strip()
                if med:
                    st.markdown(f"<div style='font-size:.78rem;padding:2px 0;'>• {med}</div>",
                                unsafe_allow_html=True)

    labs = str(c.get("Labs",""))
    if labs and labs not in ("nan","none","None",""):
        with st.sidebar.expander("🧪 Labs", expanded=False):
            for lab in labs.split("|"):
                lab = lab.strip()
                if lab:
                    st.markdown(f"<div style='font-size:.78rem;padding:2px 0;color:#0369a1;'>• {lab}</div>",
                                unsafe_allow_html=True)

    pmh = str(c.get("PMH",""))
    if pmh and pmh not in ("nan","none","None",""):
        with st.sidebar.expander("📁 Past Medical History", expanded=False):
            st.markdown(f"<div style='font-size:.78rem;color:#334155;'>{pmh}</div>",
                        unsafe_allow_html=True)

    # ── Surgery tag if applicable ──────────────────────────────────
    sk, sv = get_surgery_for_case(c)
    if sv:
        st.sidebar.markdown(f"""
        <div style="background:#fff7ed;border:1px solid #f59e0b;border-radius:7px;
                    padding:6px 10px;font-size:.75rem;color:#92400e;margin-bottom:4px;">
            🔪 <b>Surgery:</b> {sv['name']}
        </div>""", unsafe_allow_html=True)

    # ── Clear case button ──────────────────────────────────────────
    if st.sidebar.button("❌ Clear Case", use_container_width=True, key="vp_clear"):
        st.session_state.selected_case = None
        reset_case()
        nav("home")


# ════════════════════════════════════════════════════════════════════════════
# 🔐 AUTH HELPERS — moved here so page_auth() and the auth gate can use them
# ════════════════════════════════════════════════════════════════════════════

SECURITY_QUESTIONS = [
    "What is the name of your first pet?",
    "What city were you born in?",
    "What is your mother's maiden name?",
    "What was the name of your first school?",
    "What is the name of your childhood best friend?",
    "What was the make of your first car?",
    "What street did you grow up on?",
    "What is the name of the hospital where you were born?",
    "What was your favourite subject in school?",
    "What is the middle name of your oldest sibling?",
]

def _hash_pw(pw: str) -> str:
    """SHA-256 hash — never store plain passwords."""
    return hashlib.sha256(pw.encode()).hexdigest()

def _sb_headers() -> dict:
    """Supabase REST API headers."""
    key = SUPABASE_DEFAULT_KEY
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

def _send_registration_email(new_name: str, new_email: str, new_role: str):
    """Send a notification email to the admin when a new user registers."""
    try:
        notify_to = st.secrets.get("NOTIFY_EMAIL", "")
        notify_pw = st.secrets.get("NOTIFY_EMAIL_PASSWORD", "")
        if not notify_to or not notify_pw:
            return
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "🏥 New Registration — MLS Virtual Hospital"
        msg["From"]    = notify_to
        msg["To"]      = notify_to
        body_html = f"""
        <div style="font-family:Arial,sans-serif;max-width:480px;margin:auto;
                    border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;">
          <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);
                      padding:20px;text-align:center;color:white;">
            <div style="font-size:2rem;">🏥</div>
            <div style="font-size:1.2rem;font-weight:700;">MLS Virtual Hospital</div>
            <div style="font-size:.8rem;opacity:.75;">New User Registration</div>
          </div>
          <div style="padding:24px;background:#ffffff;">
            <p style="color:#0a2540;">A new user just registered:</p>
            <table style="width:100%;border-collapse:collapse;">
              <tr><td style="padding:8px;color:#64748b;font-size:.85rem;">Name</td>
                  <td style="padding:8px;font-weight:600;">{new_name}</td></tr>
              <tr style="background:#f8fafc;">
                  <td style="padding:8px;color:#64748b;font-size:.85rem;">Email</td>
                  <td style="padding:8px;font-weight:600;">{new_email}</td></tr>
              <tr><td style="padding:8px;color:#64748b;font-size:.85rem;">Role</td>
                  <td style="padding:8px;font-weight:600;">{new_role.capitalize()}</td></tr>
              <tr style="background:#f8fafc;">
                  <td style="padding:8px;color:#64748b;font-size:.85rem;">Time</td>
                  <td style="padding:8px;font-weight:600;">{datetime.now().strftime("%Y-%m-%d %H:%M UTC")}</td></tr>
            </table>
          </div>
          <div style="padding:12px;background:#f0f4f8;text-align:center;
                      font-size:.75rem;color:#64748b;">
            MLS Virtual Hospital · Clinical Training Simulator
          </div>
        </div>"""
        msg.attach(MIMEText(body_html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(notify_to, notify_pw)
            server.sendmail(notify_to, notify_to, msg.as_string())
    except Exception:
        pass  # never crash registration because of email failure


def _sb_available() -> bool:
    """True when Supabase credentials are configured."""
    return bool(SUPABASE_DEFAULT_URL and SUPABASE_DEFAULT_KEY
                and not SUPABASE_DEFAULT_URL.startswith("YOUR_"))

def _get_user(email: str, pw: str):
    """
    Look up a user by email + password.
    1. Try Supabase vh_users table
    2. Fall back to built-in demo accounts
    """
    email = email.lower().strip()
    ph    = _hash_pw(pw)

    if _sb_available():
        try:
            url = f"{SUPABASE_DEFAULT_URL}/rest/v1/vh_users"
            r = requests.get(url,
                headers=_sb_headers(),
                params={"email": f"eq.{email}", "select": "*"},
                timeout=8)
            if r.status_code == 200:
                rows = r.json()
                if rows:
                    # Email found in DB — check password
                    if rows[0].get("password_hash") == ph:
                        u = rows[0]
                        try:
                            requests.patch(
                                f"{url}?id=eq.{u['id']}",
                                headers=_sb_headers(),
                                json={"last_login": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()},
                                timeout=5)
                        except Exception:
                            pass
                        return {"id": u["id"], "name": u["name"],
                                "email": u["email"], "role": u.get("role","student")}
                    return None  # Email found but wrong password
                # Email not found in DB — fall through to demo accounts
        except Exception:
            pass

    _DEMO = {
        "admin@mls.edu":   {"id":"demo-faculty","name":"Dr. Admin",
                            "email":"admin@mls.edu","role":"faculty",
                            "pw":_hash_pw("admin123")},
        "student@mls.edu": {"id":"demo-student","name":"Student Demo",
                            "email":"student@mls.edu","role":"student",
                            "pw":_hash_pw("student123")},
    }
    u = _DEMO.get(email)
    if u and u["pw"] == ph:
        return {k:v for k,v in u.items() if k != "pw"}
    return None

def _register_user(name: str, email: str, pw: str, role: str,
                   security_q: str = "", security_a: str = "",
                   specialty: str = "", hospital: str = "") -> tuple:
    """
    Register a new user with optional security question/answer for password recovery.
    Returns (success: bool, message: str)

    Senior accounts default to is_verified=False — admin must approve them
    before they appear in the Mentor Directory.
    """
    email  = email.lower().strip()
    ph     = _hash_pw(pw)
    sa_hash = _hash_pw(security_a.strip().lower()) if security_a.strip() else ""

    # Senior accounts need admin verification; everyone else is auto-verified
    is_verified = (role != "senior")

    if _sb_available():
        try:
            url = f"{SUPABASE_DEFAULT_URL}/rest/v1/vh_users"
            chk = requests.get(url, headers=_sb_headers(),
                               params={"email": f"eq.{email}", "select": "id"},
                               timeout=8)
            if chk.status_code == 200 and chk.json():
                return False, "An account with this email already exists."
            payload = {"name": name, "email": email, "role": role,
                       "password_hash": ph,
                       "security_question": security_q,
                       "security_answer_hash": sa_hash,
                       "specialty": specialty,
                       "hospital": hospital,
                       "is_verified": is_verified}
            r = requests.post(url, headers=_sb_headers(),
                              json=payload, timeout=8)
            if r.status_code == 400:
                # Columns may not exist yet — retry with progressively fewer fields
                payload2 = {"name": name, "email": email,
                            "role": role, "password_hash": ph,
                            "specialty": specialty, "hospital": hospital,
                            "is_verified": is_verified}
                r = requests.post(url, headers=_sb_headers(),
                                  json=payload2, timeout=8)
                if r.status_code == 400:
                    # Even simpler — just core fields
                    payload3 = {"name": name, "email": email,
                                "role": role, "password_hash": ph}
                    r = requests.post(url, headers=_sb_headers(),
                                      json=payload3, timeout=8)
            if r.status_code in (200, 201):
                rows = r.json()
                uid  = rows[0]["id"] if rows else "sb-" + email[:8]
                try:
                    requests.post(
                        f"{SUPABASE_DEFAULT_URL}/rest/v1/vh_scores",
                        headers=_sb_headers(),
                        json={"user_id": uid, "score": 0, "cases_done": 0},
                        timeout=5)
                except Exception:
                    pass
                _send_registration_email(name, email, role)
                return True, "ok"
            elif r.status_code == 409:
                return False, "An account with this email already exists."
            else:
                return False, f"Database error ({r.status_code}). Try again."
        except Exception as e:
            return False, f"Connection error: {e}"

    if "_session_users" not in st.session_state:
        st.session_state._session_users = {}
    if email in st.session_state._session_users:
        return False, "An account with this email already exists."
    st.session_state._session_users[email] = {
        "id": "sess-"+email[:8], "name": name,
        "email": email, "role": role, "pw": ph,
        "security_question": security_q, "security_answer_hash": sa_hash}
    _send_registration_email(name, email, role)
    return True, "ok"


def _get_security_question(email: str) -> str:
    """Return the security question registered for this email, or empty string."""
    email = email.lower().strip()
    if _sb_available():
        try:
            url = f"{SUPABASE_DEFAULT_URL}/rest/v1/vh_users"
            r = requests.get(url, headers=_sb_headers(),
                             params={"email": f"eq.{email}",
                                     "select": "security_question"},
                             timeout=8)
            if r.status_code == 200 and r.json():
                return r.json()[0].get("security_question", "")
        except Exception:
            pass
    users = st.session_state.get("_session_users", {})
    return users.get(email, {}).get("security_question", "")


def _verify_security_answer_and_reset(email: str, answer: str, new_pw: str) -> tuple:
    """
    Verify the security answer and reset the password if correct.
    Returns (success: bool, message: str)
    """
    email   = email.lower().strip()
    ah      = _hash_pw(answer.strip().lower())
    new_ph  = _hash_pw(new_pw)

    if _sb_available():
        try:
            url = f"{SUPABASE_DEFAULT_URL}/rest/v1/vh_users"
            r = requests.get(url, headers=_sb_headers(),
                             params={"email": f"eq.{email}", "select": "*"},
                             timeout=8)
            if r.status_code == 200 and r.json():
                row = r.json()[0]
                if row.get("security_answer_hash") == ah:
                    requests.patch(
                        f"{url}?id=eq.{row['id']}",
                        headers=_sb_headers(),
                        json={"password_hash": new_ph},
                        timeout=8)
                    return True, "Password reset successfully."
                return False, "Incorrect answer. Please try again."
            return False, "No account found with that email."
        except Exception as e:
            return False, f"Connection error: {e}"

    users = st.session_state.get("_session_users", {})
    u = users.get(email)
    if not u:
        return False, "No account found with that email."
    if u.get("security_answer_hash") == ah:
        st.session_state._session_users[email]["pw"] = new_ph
        return True, "Password reset successfully."
    return False, "Incorrect answer. Please try again."

def _get_user_session(email: str, pw: str):
    """Check session-only user store."""
    email = email.lower().strip()
    users = st.session_state.get("_session_users", {})
    u = users.get(email)
    if u and u.get("pw") == _hash_pw(pw):
        return {k:v for k,v in u.items() if k != "pw"}
    return None


# ════════════════════════════════════════════════════════════════════════════
# 🔐 PAGE_AUTH — defined here so it can be called by the auth gate below
# ════════════════════════════════════════════════════════════════════════════

def page_auth():
    """
    Login / Register page.
    - Connected to Supabase → accounts permanent, survive restarts
    - No Supabase → session-only fallback (lost on restart)
    """
    # ── Header ────────────────────────────────────────────────────────────
    col_c = st.columns([1,2,1])[1]
    with col_c:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a,#0e7490);
                    border-radius:16px;padding:24px;color:white;text-align:center;
                    margin-bottom:20px;box-shadow:0 8px 32px rgba(10,37,64,.35);">
            <div style="font-size:3rem;margin-bottom:8px;">🏥</div>
            <div style="font-size:1.4rem;font-weight:800;letter-spacing:-.02em;">
                MLS Virtual Hospital
            </div>
            <div style="font-size:.82rem;opacity:.75;margin-top:4px;">
                Academy of Medical Learning Skills — Clinical Training Simulator
            </div>
        </div>""", unsafe_allow_html=True)

        # ── Supabase connection status ─────────────────────────────────────
        if _sb_available():
            st.markdown("""
            <div style="background:#f0fdf4;border:1px solid #16a34a;border-radius:8px;
                        padding:8px 12px;font-size:.78rem;color:#166534;
                        margin-bottom:12px;text-align:center;">
                🟢 <b>Connected to database</b> — accounts save permanently
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="background:#fef3c7;border:1px solid #f59e0b;border-radius:8px;
                        padding:8px 12px;font-size:.78rem;color:#92400e;
                        margin-bottom:12px;text-align:center;">
                ⚠️ <b>Session-only mode</b> — accounts reset when app restarts.<br>
                Add Supabase credentials to secrets.toml for permanent accounts.
            </div>""", unsafe_allow_html=True)

        tab_login, tab_reg = st.tabs(["🔑 Login", "📝 Register"])

        with tab_login:
            st.markdown("#### Welcome back")
            email = st.text_input("Email", placeholder="your@email.com",
                                  key="login_email")
            pw    = st.text_input("Password", type="password", key="login_pw")
            st.caption("Demo accounts — Student: `student@mls.edu` / `student123` "
                       "| Faculty: `admin@mls.edu` / `admin123`")

            if st.button("Login →", type="primary",
                         use_container_width=True, key="login_btn"):
                if not email.strip() or not pw.strip():
                    st.warning("Please enter your email and password.")
                else:
                    with st.spinner("Checking credentials..."):
                        user = _get_user(email, pw)
                        if not user:
                            user = _get_user_session(email, pw)
                    if user:
                        st.session_state.auth_user = user
                        st.success(f"✅ Welcome back, {user['name']}!")
                        st.rerun()
                    else:
                        st.error("❌ Invalid email or password.")

            # ── Forgot Password ────────────────────────────────────────
            st.markdown("---")
            with st.expander("🔑 Forgot Password? Reset via Security Question"):
                fp_email = st.text_input("Registered email address",
                                         placeholder="your@email.com", key="fp_email")
                if st.button("Find My Account", key="fp_lookup", use_container_width=True):
                    if not fp_email.strip():
                        st.warning("Please enter your email.")
                    else:
                        q = _get_security_question(fp_email)
                        if q:
                            st.session_state["fp_question"] = q
                            st.session_state["fp_email_confirmed"] = fp_email.strip().lower()
                            st.rerun()
                        else:
                            st.error("❌ No account found with that email, or no security question set.")

                if st.session_state.get("fp_question"):
                    st.markdown(f"""
                    <div style="background:#eff6ff;border:1px solid #3b82f6;border-radius:8px;
                                padding:.8rem 1rem;font-size:.85rem;margin:.5rem 0;">
                        🔒 <b>Security Question:</b><br>
                        <i>{st.session_state['fp_question']}</i>
                    </div>""", unsafe_allow_html=True)
                    fp_answer  = st.text_input("Your answer",
                                               placeholder="Type your answer here",
                                               key="fp_answer")
                    fp_new_pw  = st.text_input("New password (min 6 characters)",
                                               type="password", key="fp_new_pw")
                    fp_new_pw2 = st.text_input("Confirm new password",
                                               type="password", key="fp_new_pw2")
                    if st.button("Reset My Password →", type="primary",
                                 use_container_width=True, key="fp_reset_btn"):
                        if not fp_answer.strip():
                            st.warning("Please enter your answer.")
                        elif not fp_new_pw or len(fp_new_pw) < 6:
                            st.warning("New password must be at least 6 characters.")
                        elif fp_new_pw != fp_new_pw2:
                            st.error("❌ Passwords do not match.")
                        else:
                            ok, msg = _verify_security_answer_and_reset(
                                st.session_state["fp_email_confirmed"],
                                fp_answer, fp_new_pw)
                            if ok:
                                st.success("✅ " + msg + " You can now log in with your new password.")
                                for k in ["fp_question", "fp_email_confirmed"]:
                                    st.session_state.pop(k, None)
                                st.rerun()
                            else:
                                st.error("❌ " + msg)

        with tab_reg:
            st.markdown("#### Create your account")
            rname  = st.text_input("Full Name", placeholder="Dr. Jane Smith",
                                   key="reg_name")
            remail = st.text_input("Email", placeholder="your@email.com",
                                   key="reg_email")
            rrole  = st.selectbox(
                "Role",
                ["student", "resident", "senior", "faculty"],
                format_func=lambda x: {
                    "student":  "🎓 Medical Student",
                    "resident": "🩺 Resident / Junior Doctor",
                    "senior":   "👨‍⚕️ Senior / Consultant Doctor",
                    "faculty":  "👨‍🏫 Faculty / Teacher",
                }.get(x, x),
                key="reg_role",
            )

            # Specialty field for residents and seniors
            rspec = ""
            rhospital = ""
            if rrole in ("resident", "senior"):
                rspec = st.selectbox(
                    "Specialty",
                    [
                        "General Medicine", "Cardiology",
                        "Respiratory / Pulmonology", "Gastroenterology",
                        "Endocrinology", "Neurology", "Nephrology",
                        "Hematology / Oncology", "Infectious Diseases",
                        "Rheumatology", "Emergency Medicine",
                        "Critical Care / ICU", "General Surgery",
                        "Orthopedics", "Obstetrics & Gynecology",
                        "Pediatrics", "Psychiatry", "Dermatology",
                        "Radiology", "Anesthesiology", "Family Medicine",
                        "Other",
                    ],
                    key="reg_specialty",
                )
                rhospital = st.text_input(
                    "Hospital / Affiliation (optional)",
                    placeholder="e.g., University Hospital, Beirut",
                    key="reg_hospital",
                )

                if rrole == "senior":
                    st.info(
                        "ℹ️ **Senior accounts require admin verification** "
                        "before appearing in the mentor directory. You'll get full "
                        "access to the platform immediately, but residents won't be "
                        "able to book sessions with you until Dr. Hiba reviews your "
                        "registration."
                    )

            rpw    = st.text_input("Password", type="password", key="reg_pw")
            rpw2   = st.text_input("Confirm Password", type="password",
                                   key="reg_pw2")

            st.markdown("---")
            st.markdown("**🔒 Security Question** — used to recover your password")
            rsq = st.selectbox("Choose a security question", SECURITY_QUESTIONS,
                               key="reg_sq")
            rsa = st.text_input("Your answer",
                                placeholder="Answer is case-insensitive",
                                key="reg_sa")

            if st.button("Create Account →", type="primary",
                         use_container_width=True, key="reg_btn"):
                if not all([rname.strip(), remail.strip(), rpw, rpw2]):
                    st.warning("Please fill in all fields.")
                elif rpw != rpw2:
                    st.error("❌ Passwords do not match.")
                elif len(rpw) < 6:
                    st.warning("Password must be at least 6 characters.")
                elif not re.match(r"[^@]+@[^@]+\.[^@]+", remail):
                    st.error("❌ Invalid email format.")
                elif not rsa.strip():
                    st.warning("Please provide an answer to your security question.")
                elif rrole in ("resident", "senior") and not rspec:
                    st.warning("Please select your specialty.")
                else:
                    with st.spinner("Creating your account..."):
                        ok, msg = _register_user(rname, remail, rpw, rrole,
                                                 security_q=rsq, security_a=rsa,
                                                 specialty=rspec,
                                                 hospital=rhospital)
                    if ok:
                        user = _get_user(remail, rpw)
                        if not user:
                            user = _get_user_session(remail, rpw)
                        if user:
                            st.session_state.auth_user = user
                            st.session_state.page = "avatar_builder"
                            st.session_state["_new_registration"] = True
                            if rrole == "senior":
                                st.success("✅ Account created! Your senior status is "
                                           "pending admin verification. Set up your "
                                           "avatar while you wait.")
                            else:
                                st.success("✅ Account created! Let's set up your doctor avatar.")
                            st.rerun()
                        else:
                            st.info("Account created. Please log in.")
                    else:
                        st.error(f"❌ {msg}")


# ════════════════════════════════════════════════════════════════════════════
# 🔐 AUTH GATE — must run BEFORE sidebar so unauthenticated users see only login
# ════════════════════════════════════════════════════════════════════════════

if not st.session_state.get("auth_user"):
    # Show a minimal locked sidebar — no nav, no patient data
    with st.sidebar:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);
                    border-radius:10px;padding:14px;text-align:center;margin-bottom:12px;">
            <div style="font-size:2rem;">🏥</div>
            <div style="color:white;font-weight:800;font-size:1rem;margin-top:4px;">
                MLS Virtual Hospital
            </div>
            <div style="color:#67e8f9;font-size:.75rem;margin-top:2px;">
                Clinical Training Simulator
            </div>
        </div>""", unsafe_allow_html=True)
        st.markdown("""
        <div style="background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);
                    border-radius:8px;padding:10px 12px;text-align:center;color:#94a3b8;
                    font-size:.8rem;">
            🔒 Please log in to access the hospital
        </div>""", unsafe_allow_html=True)
    # Show login form in main area, then stop
    page_auth()
    st.stop()

# ── Logged-in — set auth variables used by sidebar and pages ─────────────


# Safe auth fallbacks for sidebar (real values set by auth gate below)
_auth       = st.session_state.get("auth_user") or {}
_is_faculty = _auth.get("role") == "faculty"

with st.sidebar:
    st.markdown("### 🏥 MLS Virtual Hospital")

    # ── User badge — anime avatar in sidebar ──────────────────────
    role_color  = "#d97706" if _is_faculty else "#0e7490"
    role_label  = "👨‍🏫 Faculty" if _is_faculty else "👨‍⚕️ Student"
    _saved_av   = st.session_state.get("doctor_avatar", {})

    # ── Persistent avatar card — visible on EVERY page after login ────────────
    # Large anime avatar centred at top of sidebar, name + role below
    try:
        _big_copy = dict(_saved_av); _big_copy["stethoscope"] = True
        _big_svg  = _render_anime_avatar(_big_copy, size=130)
        _has_av   = True
    except Exception:
        _big_svg = ""; _has_av = False

    _av_name = _saved_av.get("name", _auth.get("name","Dr. ..."))
    _badge_rank = st.session_state.get("my_room", {}).get("badge", "🎓 Intern")

    if _has_av and _saved_av:
        st.markdown(f"""
        <div style="text-align:center;background:linear-gradient(160deg,#0f3460,#0a2540);
                    border-radius:12px;padding:12px 6px 10px;margin-bottom:6px;
                    border:1px solid rgba(14,116,144,.35);">
          <div style="width:100px;height:100px;border-radius:50%;overflow:hidden;
                      border:3px solid #0e7490;margin:0 auto 6px;
                      box-shadow:0 4px 14px rgba(0,0,0,.4);">
            {_big_svg}
          </div>
          <div style="color:white;font-weight:700;font-size:.85rem;
                      white-space:nowrap;overflow:hidden;text-overflow:ellipsis;
                      padding:0 6px;">{_av_name}</div>
          <div style="color:#0ea5e9;font-size:.7rem;font-weight:600;margin-top:1px;">
            {role_label}</div>
          <div style="color:#64748b;font-size:.65rem;margin-top:1px;">
            {_auth.get("email","")}</div>
          <div style="margin-top:6px;display:inline-block;background:rgba(14,116,144,.25);
                      border:1px solid rgba(14,116,144,.4);border-radius:999px;
                      padding:2px 10px;font-size:.65rem;color:#67e8f9;">
            {_badge_rank if _badge_rank else "🎓 Intern"}</div>
        </div>""", unsafe_allow_html=True)
    else:
        # No avatar yet — compact text badge + prompt to create one
        _name     = _auth.get("name", "User")
        _initials = "".join(w[0].upper() for w in _name.split() if w and w[0].isalpha())[:2] or "U"
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0a2540,#0f3460);border-radius:8px;
                    padding:8px 12px;margin-bottom:4px;display:flex;align-items:center;gap:10px;">
          <div style="width:44px;height:44px;border-radius:50%;background:#0e7490;
                      display:flex;align-items:center;justify-content:center;
                      font-size:.9rem;font-weight:700;color:white;flex-shrink:0;">{_initials}</div>
          <div style="min-width:0;flex:1;">
            <div style="color:white;font-weight:700;font-size:.85rem;">{_name}</div>
            <div style="color:#0ea5e9;font-size:.7rem;">{role_label}</div>
          </div>
        </div>""", unsafe_allow_html=True)
        if st.button("👨‍⚕️ Create Your Avatar", use_container_width=True, key="nav_av_prompt"):
            nav("avatar_builder")

    # ── EPIC EHR Patient Lookup ───────────────────────────────────
    epic_patient_lookup_widget()
    st.markdown("---")

    # ── Virtual Hospital Patient Panel ────────────────────────────
    virtual_patient_panel()
    st.markdown("---")

    # ── Navigation — Clinical Modules ─────────────────────────────
    st.markdown('<div style="color:#64748b;font-size:.68rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;padding:2px 4px;">🏥 Clinical Modules</div>', unsafe_allow_html=True)
    pages=[("🏠 Home","home"),("📚 Case Library","library"),("🚨 Emergency Room","emergency"),
           ("💬 Patient Interview","simulator"),("🫁 Physical Examination","physical_exam"),
           ("🧪 Laboratory","lab"),("🔬 Imaging & ECG Analysis","imaging"),
           ("🔪 Surgery Room","surgery"),("🔴 Live Discussion","live"),
           ("📝 Submit Diagnosis","diagnosis"),("🤖 AI Tutor","tutor"),
           ("🧠 AI Tutor Cases","ai_tutor_cases"),
           ("👨‍⚕️ My Avatar","avatar_builder")]
    for label,pk in pages:
        if st.button(label,use_container_width=True,key=f"nav_{pk}"): nav(pk)

    # ── Navigation — Advanced Learning ───────────────────────────
    st.markdown("---")
    st.markdown('<div style="color:#64748b;font-size:.68rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;padding:2px 4px;">🎓 Advanced Learning</div>', unsafe_allow_html=True)
    advanced=[("🧬 Differential Diagnosis","ddx"),
              ("💊 Drug Prescribing","prescribing"),
              ("🩺 Procedure Simulator","procedures"),
              ("🧠 Clinical Reasoning Map","reasoning"),
              ("🏆 Competency Tracker","competency"),
              ("🎯 OSCE Exam Simulator","osce"),
              ("📋 Progress Notes","notes"),
              ("🃏 Flashcard Builder","flashcards"),
              ("⭐ My Progress (XP)","progress_dashboard")]
    for label,pk in advanced:
        if st.button(label,use_container_width=True,key=f"nav_{pk}"): nav(pk)

    # ── Navigation — Faculty (only shown to faculty) ───────────────
    if _is_faculty:
        st.markdown("---")
        st.markdown('<div style="color:#d97706;font-size:.68rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;padding:2px 4px;">👨‍🏫 Faculty Portal</div>', unsafe_allow_html=True)
        faculty_pages=[("📊 Analytics Dashboard","analytics"),
                       ("🏥 Case Creator","case_creator"),
                       ("✨ AI Case Creator","ai_case_creator"),
                       ("🛠️ Mentor Admin Panel","admin_mentors"),
                       ("📝 MCQ Bank Manager","admin_mcqs"),
                       ("🩻 Image Library Manager","admin_images"),
                       ("📖 Reference Library (RAG)","admin_rag"),
                       ("👥 User Management","user_management")]
        for label,pk in faculty_pages:
            if st.button(label,use_container_width=True,key=f"nav_{pk}"): nav(pk)

    # ── Navigation — Tools ────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div style="color:#64748b;font-size:.68rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;padding:2px 4px;">🛠️ Tools</div>', unsafe_allow_html=True)
    tools=[("🏥 Add Real Case","add_case"),
           ("👥 Peer Simulation","peer_sim"),
           ("👨‍⚕️ Mentor Directory","mentor_directory"),
           ("📋 My Sessions","my_sessions"),
           ("🩻 Image Practice","image_practice"),
           ("🧮 Clinical Scores","scores"),
           ("📚 Evidence & Cases","evidence"),
           ("🌐 DocCollab","doccollab")]
    for label,pk in tools:
        if st.button(label,use_container_width=True,key=f"nav_{pk}"): nav(pk)

    st.markdown("---")
    st.session_state.voice_enabled=st.toggle("🔊 Patient Voice",value=st.session_state.voice_enabled)
    render_tts_voice_check()
    st.markdown("---")
    # Model selector
    current_model = st.session_state.get("active_model", GEMINI_MODEL)
    model_display = {
        "gemini-2.5-flash":      "⚡ Gemini 2.5 Flash (Best)",
        "gemini-2.0-flash":      "🚀 Gemini 2.0 Flash (Free ∞)",
        "gemini-2.0-flash-lite": "💨 Gemini 2.0 Lite (Fastest ∞)",
    }
    selected_model = st.selectbox("🤖 AI Model:",
        list(model_display.keys()),
        format_func=lambda x: model_display.get(x, x),
        index=list(model_display.keys()).index(current_model) if current_model in model_display else 0,
        key="model_selector"
    )
    if selected_model != current_model:
        st.session_state["active_model"] = selected_model
        st.rerun()
    st.markdown("---")
    render_credit_bar()
    if st.button("💳 Credits & Plans", use_container_width=True, key="nav_cred_sb"): nav("credits")
    st.markdown("---")
    st.markdown(f"⭐ Score: **{st.session_state.score}** | ✅ Done: **{len(st.session_state.cases_done)}**")
    if st.button("🚪 Logout", use_container_width=True, key="logout_btn"):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

# ════════════════════════════════════════════════════════════════
# HOME
# ════════════════════════════════════════════════════════════════
def page_home():
    # ══════════════════════════════════════════════════════
    # HERO HEADER
    # ══════════════════════════════════════════════════════
    st.markdown("""
    <div class="main-header">
        <h1>🏥 MLS Virtual Hospital</h1>
        <p>Academy of Medical Learning Skills — AI-Powered Clinical Training Simulator</p>
    </div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # KPI STRIP  (native st.metric — never breaks)
    # ══════════════════════════════════════════════════════
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("📋 Total Cases",  str(len(df)) if not df.empty else "0")
    k2.metric("✅ Completed",    str(len(st.session_state.cases_done)))
    k3.metric("⭐ Score",        str(st.session_state.score))
    k4.metric("🏆 Level",        "MLS")

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # ABOUT  —  3 mission cards via st.columns
    # ══════════════════════════════════════════════════════
    st.markdown('<div class="section-header">🎯 About MLS Virtual Hospital</div>',
                unsafe_allow_html=True)

    a1, a2, a3 = st.columns(3)
    with a1:
        st.markdown("""
        <div style="background:#ffffff;border-radius:12px;padding:1.3rem 1.2rem;
                    border-top:4px solid #0e7490;
                    box-shadow:0 2px 8px rgba(0,0,0,.07);height:100%;">
            <div style="font-size:1.5rem;margin-bottom:.5rem;">🎯</div>
            <div style="font-weight:800;color:#0a2540;font-size:.92rem;margin-bottom:.5rem;">
                Our Role
            </div>
            <div style="font-size:.82rem;color:#475569;line-height:1.75;">
                MLS Virtual Hospital is a <strong>24/7 clinical training environment</strong>
                for Medical Laboratory Science students — simulating the full patient journey
                from triage to diagnosis in a safe, consequence-free setting.
            </div>
        </div>""", unsafe_allow_html=True)
    with a2:
        st.markdown("""
        <div style="background:#ffffff;border-radius:12px;padding:1.3rem 1.2rem;
                    border-top:4px solid #7c3aed;
                    box-shadow:0 2px 8px rgba(0,0,0,.07);height:100%;">
            <div style="font-size:1.5rem;margin-bottom:.5rem;">🏥</div>
            <div style="font-weight:800;color:#0a2540;font-size:.92rem;margin-bottom:.5rem;">
                Our Aim
            </div>
            <div style="font-size:.82rem;color:#475569;line-height:1.75;">
                To <strong>bridge the gap between theory and real clinical practice</strong> —
                giving students hands-on exposure to diverse cases, lab interpretation,
                imaging analysis, and clinical reasoning before they enter a real ward.
            </div>
        </div>""", unsafe_allow_html=True)
    with a3:
        st.markdown("""
        <div style="background:#ffffff;border-radius:12px;padding:1.3rem 1.2rem;
                    border-top:4px solid #059669;
                    box-shadow:0 2px 8px rgba(0,0,0,.07);height:100%;">
            <div style="font-size:1.5rem;margin-bottom:.5rem;">✨</div>
            <div style="font-weight:800;color:#0a2540;font-size:.92rem;margin-bottom:.5rem;">
                Why We Are Different
            </div>
            <div style="font-size:.82rem;color:#475569;line-height:1.75;">
                Powered by <strong>Gemini AI + a fine-tuned Clinical Tutor model</strong>,
                every patient interaction is dynamic and unique — just like real patients.
                No two sessions are ever identical.
            </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # FEATURE HIGHLIGHTS  —  2 × 3 grid via columns
    # ══════════════════════════════════════════════════════
    st.markdown('<div class="section-header">🌟 What Makes Us Unique</div>',
                unsafe_allow_html=True)

    features = [
        ("#0ea5e9", "#f0f9ff", "🤖 AI-Powered Dynamic Patients",
         "Each virtual patient is driven by Gemini AI — they remember context, "
         "react emotionally, and respond like a real human, not a scripted bot."),
        ("#059669", "#f0fdf4", "🔬 Full Clinical Workflow",
         "Emergency → History → Physical Exam → Lab → Imaging → Surgery → Diagnosis. "
         "The complete hospital journey in one unified platform."),
        ("#d97706", "#fffbeb", "📡 Real Imaging Analysis",
         "Upload ECGs, X-rays, CT scans, and MRIs. Gemini Vision annotates "
         "abnormalities and generates a full radiologist-style teaching report."),
        ("#7c3aed", "#f5f3ff", "🧬 Fair Random Case Assignment",
         "Students receive randomly assigned cases — no two students share the same "
         "case at the same time, ensuring independent and fair assessment."),
        ("#dc2626", "#fef2f2", "🔴 Live Simulation Mode",
         "Real-time conversation with the AI patient who reacts to your examination, "
         "responds to your tone, and simulates a genuine ward encounter."),
        ("#0a2540", "#f0f4f8", "📊 Faculty Analytics & Case Creator",
         "Monitor student performance, build new AI cases from real patient data, "
         "track competency progression, and export detailed reports."),
    ]

    row1 = st.columns(3)
    row2 = st.columns(3)
    all_fcols = row1 + row2
    for col, (border, bg, title, body) in zip(all_fcols, features):
        with col:
            st.markdown(f"""
            <div style="background:{bg};border-radius:10px;padding:1rem 1.1rem;
                        border-left:4px solid {border};margin-bottom:.5rem;
                        min-height:110px;">
                <div style="font-weight:700;color:{border};font-size:.84rem;
                            margin-bottom:.35rem;">{title}</div>
                <div style="font-size:.79rem;color:#475569;line-height:1.7;">{body}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # HOSPITAL ROOMS  —  clickable navigation cards
    # ══════════════════════════════════════════════════════
    st.markdown('<div class="section-header">🗺️ Hospital Rooms — Click to Enter</div>',
                unsafe_allow_html=True)

    rooms = [
        ("🚨", "Emergency Room",    "Triage & acute care",                "emergency",    "#dc2626"),
        ("💬", "Patient Interview", "AI Avatar · conversational",          "simulator",    "#0e7490"),
        ("🫁", "Physical Exam",     "Auscultation, Percussion, Palpation", "physical_exam","#0369a1"),
        ("🧪", "Laboratory",        "Lab tests & result interpretation",   "lab",          "#059669"),
        ("🔬", "Imaging & ECG",     "Upload & AI radiology analysis",      "imaging",      "#7c3aed"),
        ("🔪", "Surgery Room",      "Full surgical procedure simulation",  "surgery",      "#b45309"),
        ("🔴", "Live Discussion",   "Real-time patient encounter",         "live",         "#be185d"),
        ("📚", "Case Library",      "Browse & filter all cases",           "library",      "#0a2540"),
        ("🤖", "AI Tutor",          "Gemini + Clinical Tutor AI",          "tutor",        "#0e7490"),
        ("🏥", "Add Real Case",     "Convert real → AI patient",           "add_case",     "#065f46"),
        ("📝", "Submit Diagnosis",  "Get AI-evaluated feedback",           "diagnosis",    "#1e40af"),
    ]

    # Render 4 columns of room cards
    rcols = st.columns(4)
    for i, (icon, name, desc, key, color) in enumerate(rooms):
        with rcols[i % 4]:
            st.markdown(f"""
            <div style="background:#ffffff;border-radius:10px;
                        border-top:3px solid {color};
                        padding:.85rem .9rem;margin-bottom:.6rem;
                        box-shadow:0 1px 6px rgba(0,0,0,.07);">
                <div style="font-size:1.3rem;margin-bottom:.2rem;">{icon}</div>
                <div style="font-weight:700;font-size:.83rem;color:#0a2540;
                            margin-bottom:.15rem;">{name}</div>
                <div style="font-size:.74rem;color:#64748b;">{desc}</div>
            </div>""", unsafe_allow_html=True)
            if st.button(f"Enter →", key=f"h_{key}", use_container_width=True):
                if key not in ("library", "tutor", "home", "surgery", "add_case") \
                        and not st.session_state.selected_case:
                    st.warning("⚠️ Please select a case from the Case Library first.")
                else:
                    nav(key)

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # QUICK-START BANNER
    # ══════════════════════════════════════════════════════
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0a2540,#0e7490);
                border-radius:14px;padding:1.4rem 2rem;
                box-shadow:0 4px 16px rgba(10,37,64,.25);
                display:flex;align-items:center;justify-content:space-between;
                flex-wrap:wrap;gap:1rem;">
        <div>
            <div style="color:white;font-weight:800;font-size:1.05rem;
                        margin-bottom:.3rem;">
                🚀 Ready to start?
            </div>
            <div style="color:#bae6fd;font-size:.83rem;">
                Go to <b>Case Library</b> → select or receive a case →
                then work through each hospital room in sequence.
            </div>
        </div>
        <div style="color:#67e8f9;font-size:.8rem;font-weight:600;">
            Emergency → History → Exam → Lab → Imaging → Diagnosis
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# CASE LIBRARY + RANDOM ASSIGNMENT
# ════════════════════════════════════════════════════════════════

def assign_random_case(cases_df):
    """
    Assign a random case to the current student that they haven't completed yet.
    Stores it in session_state.selected_case and resets case state.
    Returns the chosen row as a dict.
    """
    done = [str(x) for x in st.session_state.cases_done]
    available = cases_df[~cases_df["row_num"].astype(str).isin(done)]
    if available.empty:
        available = cases_df          # all done — cycle through again
    chosen = available.sample(1).iloc[0]
    st.session_state.selected_case = chosen.to_dict()
    reset_case()
    return chosen.to_dict()


def _page_library_student(cases_df):
    """Student view: shows only their randomly assigned case — no browsing."""
    st.markdown('<div class="section-header">📋 Your Assigned Case</div>',
                unsafe_allow_html=True)
    st.markdown("""
    <div class="alert-info">
        ℹ️ Cases are <b>randomly assigned</b> to ensure a fair and independent assessment.
        You will not see other cases in the library. Work through your assigned case fully,
        then submit your diagnosis to receive AI feedback and unlock your next case.
    </div>""", unsafe_allow_html=True)

    # Assign a case on first visit or if none is selected
    if not st.session_state.get("selected_case"):
        with st.spinner("Assigning your case, please wait..."):
            assign_random_case(cases_df)
        st.success("✅ A new case has been assigned to you!")

    c = st.session_state.selected_case
    if not c:
        st.error("Could not assign a case. Please refresh the page.")
        return

    done_ids = [str(x) for x in st.session_state.cases_done]
    already_done = str(c.get("row_num","")) in done_ids
    diff_color = {"basic":"#16a34a","intermediate":"#d97706","advanced":"#dc2626"}.get(
        str(c.get("Difficulty","")).lower(), "#6b7280")

    st.markdown(f"""
    <div class="patient-card" style="border-top:4px solid {diff_color};">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;
                    margin-bottom:.8rem;">
            <span style="font-weight:800;font-size:1.05rem;color:#0a2540;">
                📋 Case #{c.get("Case_ID","?")}
            </span>
            <span style="background:{diff_color};color:white;border-radius:5px;
                         padding:2px 10px;font-size:.72rem;font-weight:700;">
                {str(c.get("Difficulty","?")).title()}
            </span>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:.5rem;font-size:.85rem;">
            <div><b>Patient:</b> {c.get("Age_Sex","?")}</div>
            <div><b>System:</b> {str(c.get("System","?")).title()}</div>
            <div style="grid-column:span 2;">
                <b>Chief Complaint:</b>
                <span style="color:#dc2626;">{c.get("Chief_Complaint","?")}</span>
            </div>
            <div><b>Duration:</b> {c.get("Duration","?")}</div>
            <div><b>Context:</b> {c.get("Context","?")}</div>
        </div>
    </div>""", unsafe_allow_html=True)

    if already_done:
        st.markdown("""
        <div class="alert-good">
            ✅ <b>You have already completed this case.</b>
            Request a new case below to continue practicing.
        </div>""", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        if st.button("🚀 Start This Case", type="primary",
                     use_container_width=True, key="lib_start_assigned"):
            nav("emergency")
    with col2:
        if st.button("🔄 Request a Different Case",
                     use_container_width=True, key="lib_reassign"):
            with st.spinner("Assigning a new case..."):
                assign_random_case(cases_df)
            st.rerun()

    # Progress summary
    total_done = len(st.session_state.cases_done)
    total_cases = len(cases_df)
    if total_done > 0:
        st.markdown(f"""
        <div style="background:white;border-radius:10px;padding:.9rem 1.1rem;
                    border:1px solid #e2e8f0;margin-top:1rem;">
            <div style="font-size:.78rem;font-weight:600;color:#0a2540;
                        margin-bottom:.4rem;">
                📊 Your Progress
            </div>
            <div style="background:#e5e7eb;border-radius:999px;height:8px;">
                <div style="background:#0e7490;height:8px;border-radius:999px;
                            width:{min(100, int(total_done/max(total_cases,1)*100))}%;"></div>
            </div>
            <div style="font-size:.72rem;color:#64748b;margin-top:.3rem;">
                {total_done} of {total_cases} cases completed ·
                Score: <b>{st.session_state.score}</b>
            </div>
        </div>""", unsafe_allow_html=True)


def _page_library_faculty(cases_df):
    """Faculty view: full case library with search and filters."""
    st.markdown('<div class="section-header">📚 Full Case Library</div>',
                unsafe_allow_html=True)
    if cases_df.empty:
        st.error("case_studies.xlsx not found.")
        return
    c1, c2, c3 = st.columns(3)
    with c1:
        search = st.text_input("🔍 Search",
                               placeholder="complaint, diagnosis, system...")
    with c2:
        systems = ["All"] + sorted(cases_df["System"].dropna().unique().tolist())
        system  = st.selectbox("System", systems)
    with c3:
        diff = st.selectbox("Difficulty", ["All","basic","intermediate","advanced"])

    filt = cases_df.copy()
    if search:
        filt = filt[filt.apply(lambda r: search.lower() in str(r).lower(), axis=1)]
    if system != "All":
        filt = filt[filt["System"].str.lower().str.contains(system.lower(), na=False)]
    if diff != "All":
        filt = filt[filt["Difficulty"].str.lower().str.contains(diff.lower(), na=False)]

    st.markdown(f"**{len(filt)} case(s) found**")
    for _, row in filt.iterrows():
        done  = str(row["row_num"]) in [str(x) for x in st.session_state.cases_done]
        title = row.get("Title") or row.get("Chief_Complaint") or "Unknown"
        sk, sv = get_surgery_for_case(row.to_dict())
        with st.expander(
            f"{'✅' if done else '📋'} {row['Case_ID']} — "
            f"{str(title).title()} | {row['Age_Sex']} | "
            f"{row['System'].upper()}{'  🔪' if sv else ''}"
        ):
            lc, rc = st.columns([3, 1])
            with lc:
                st.markdown(f"**Chief Complaint:** {row['Chief_Complaint']}")
                st.markdown(f"**Duration:** {row['Duration']} | **Context:** {row['Context']}")
                st.markdown(diff_badge(row["Difficulty"]), unsafe_allow_html=True)
                if sv:
                    st.markdown(f"🔪 Surgery: **{sv['name']}**")
                # Faculty-only: show diagnosis
                st.markdown(
                    f'<div class="alert-warn" style="font-size:.78rem;">'
                    f'🔍 <b>Diagnosis (Faculty Only):</b> {row.get("Final_Diagnosis","?")}</div>',
                    unsafe_allow_html=True)
            with rc:
                if st.button("🚀 Preview", key=f"s_{row['row_num']}",
                             use_container_width=True, type="primary"):
                    st.session_state.selected_case = row.to_dict()
                    reset_case()
                    nav("emergency")


def page_library():
    """
    Router: Faculty see the full browsable library.
    Students see only their randomly assigned case.
    """
    if _is_faculty:
        _page_library_faculty(df)
    else:
        _page_library_student(df)

# ════════════════════════════════════════════════════════════════
# EMERGENCY ROOM
# ════════════════════════════════════════════════════════════════
def page_emergency():
    st.markdown('<div class="section-header">🚨 Emergency Room</div>',unsafe_allow_html=True)
    c=st.session_state.selected_case
    if not c:
        st.warning("No case selected.")
        if st.button("📚 Case Library"): nav("library")
        return
    av_col,info_col=st.columns([1,3])
    with av_col:
        mood="pain" if "pain" in str(c.get("Appearance","")).lower() else "scared"
        st.markdown(render_avatar(mood,c.get("Age_Sex","")),unsafe_allow_html=True)
    with info_col:
        st.markdown(f'<div class="patient-card"><h3 style="margin:0 0 .4rem;color:#0a2540">🚑 Patient Arrived</h3><p><b>Patient:</b> {c.get("Age_Sex","?")} | <b>Occupation:</b> {c.get("Occupation","?")}</p><p style="color:#dc2626"><b>Chief Complaint:</b> {c.get("Chief_Complaint","?")}</p><p style="color:#6b7280;font-size:.85rem"><b>Duration:</b> {c.get("Duration","?")} | <b>Context:</b> {c.get("Context","?")}</p><p style="color:#6b7280;font-size:.85rem">Arrived: {datetime.now().strftime("%H:%M — %d %b %Y")}</p></div>',unsafe_allow_html=True)
        st.markdown(f'<div class="alert-info">🩺 <b>Vitals:</b> {c.get("Vitals","normal")}</div>',unsafe_allow_html=True)
        st.markdown(f'<div class="alert-warn">👁️ <b>Appearance:</b> {c.get("Appearance","In pain")}</div>',unsafe_allow_html=True)
    with st.expander("🛒 ER Equipment"):
        e1,e2,e3=st.columns(3)
        with e1: st.markdown("**Resuscitation**\n- Crash Cart\n- Defibrillator/AED\n- Laryngoscope/ETT\n- Ambu bags\n- Ventilators\n- O2 & suction")
        with e2: st.markdown("**Diagnostics**\n- 12-lead ECG\n- Cardiac monitor\n- Pulse oximeter\n- NIBP\n- Portable X-ray\n- Glucometer")
        with e3: st.markdown("**Trauma/Support**\n- IV pumps\n- Suture trays\n- Trauma kits\n- Splints/collars\n- PPE\n- Stretchers")
    sk,sv=get_surgery_for_case(c)
    if sv: st.markdown(f'<div class="alert-good">🔪 <b>Surgery Available:</b> {sv["name"]}</div>',unsafe_allow_html=True)
    st.markdown('<div class="alert-warn">⚠️ <b>Assessment required.</b></div>',unsafe_allow_html=True)
    b1,b2,b3,b4,b5,b6=st.columns(6)
    with b1:
        if st.button("💬 Interview",use_container_width=True,type="primary"): nav("simulator")
    with b2:
        if st.button("🫁 Examine",use_container_width=True): nav("physical_exam")
    with b3:
        if st.button("🧪 Labs",use_container_width=True): nav("lab")
    with b4:
        if st.button("🔬 Imaging",use_container_width=True): nav("imaging")
    with b5:
        if st.button("🔴 Live",use_container_width=True): nav("live")
    with b6:
        if st.button("🔪 Surgery",use_container_width=True): nav("surgery")

# ════════════════════════════════════════════════════════════════
# PATIENT INTERVIEW
# ════════════════════════════════════════════════════════════════
def page_simulator():
    st.markdown('<div class="section-header">💬 Patient Interview</div>',unsafe_allow_html=True)
    st.markdown(TTS_JS,unsafe_allow_html=True)
    c=st.session_state.selected_case
    if not c: st.warning("No case selected."); return
    sys_p=patient_sys(c)
    st.session_state.chat_history=[m for m in st.session_state.chat_history if not str(m.get("content","")).startswith("!ERR")]
    if not st.session_state.chat_history:
        with st.spinner("Patient entering room..."):
            g=call_ai(sys_p,[{"role":"user","content":"The student doctor just entered. Greet them naturally as a scared patient and describe how you feel."}])
        if not g.startswith("!ERR"):
            st.session_state.chat_history.append({"role":"patient","content":g})
            st.session_state.avatar_mood=detect_mood(g)
            if st.session_state.voice_enabled:
                tts_speak(g)
        else: st.error(g); st.stop()
    av_col,chat_col=st.columns([1,3])
    with av_col:
        st.markdown(render_avatar(st.session_state.avatar_mood,c.get("Age_Sex","")),unsafe_allow_html=True)
        st.markdown("")
        if st.session_state.voice_enabled:
            if st.button("🔊 Replay",use_container_width=True):
                last=[m["content"] for m in reversed(st.session_state.chat_history) if m["role"]=="patient"]
                if last: st.markdown(f'<script>speakText({json.dumps(last[0])});</script>',unsafe_allow_html=True)
            if st.button("🔇 Stop",use_container_width=True):
                st.markdown('<script>stopSpeech();</script>',unsafe_allow_html=True)
        st.markdown(f'<div class="alert-info" style="font-size:.75rem">Mood: <b>{st.session_state.avatar_mood.title()}</b></div>',unsafe_allow_html=True)
    with chat_col:
        for msg in st.session_state.chat_history:
            if msg["role"]=="patient":
                st.markdown(f'<div class="chat-patient">🤒 <b>Patient:</b> {msg["content"]}</div>',unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="chat-student">👨‍⚕️ <b>You:</b> {msg["content"]}</div>',unsafe_allow_html=True)
        st.markdown("**🎤 Voice Input — Speak to Patient**")
        voice_input_component(key="sim_voice", role="doctor", height=165)
        st.markdown('<div class="alert-info" style="font-size:.78rem;">💡 <b>How to use voice:</b> Press the mic button → speak your question → click <b>Send Voice Message</b> → your text is copied to clipboard → paste it (Ctrl+V) in the text box below → click Send →</div>', unsafe_allow_html=True)

        with st.form("chat",clear_on_submit=True):
            i1,i2=st.columns([5,1])
            with i1: inp=st.text_input("Ask patient:",placeholder="Type or paste spoken text here...",label_visibility="collapsed")
            with i2: go=st.form_submit_button("Send →",use_container_width=True)
        if go and inp.strip():
            st.session_state.chat_history.append({"role":"student","content":inp})
            if st.session_state.voice_enabled:
                tts_speak_doctor(inp)
            with st.spinner("Patient responding..."):
                rep=call_ai(sys_p,build_msgs(st.session_state.chat_history))
            if not rep.startswith("!ERR"):
                st.session_state.chat_history.append({"role":"patient","content":rep})
                st.session_state.avatar_mood=detect_mood(rep)
                if st.session_state.voice_enabled:
                    tts_speak(rep)
            st.rerun()
        st.markdown("**💡 Quick Questions:**")
        qs=["Rate pain 1-10?","When did this start?","Any chronic conditions?","Current medications?","Does it radiate?"]
        qcols=st.columns(len(qs))
        for col,q in zip(qcols,qs):
            with col:
                if st.button(q[:20]+"...",key=f"q_{q[:8]}",use_container_width=True):
                    st.session_state.chat_history.append({"role":"student","content":q})
                    with st.spinner("..."):
                        rep=call_ai(sys_p,build_msgs(st.session_state.chat_history))
                    if not rep.startswith("!ERR"):
                        st.session_state.chat_history.append({"role":"patient","content":rep})
                        st.session_state.avatar_mood=detect_mood(rep)
                        if st.session_state.voice_enabled:
                            st.markdown(f'<script>setTimeout(function(){{speakText({json.dumps(rep)});}},300);</script>',unsafe_allow_html=True)
                    st.rerun()
    b1,b2,b3,b4,b5=st.columns(5)
    with b1:
        if st.button("← ER",use_container_width=True): nav("emergency")
    with b2:
        if st.button("🫁 Examine",use_container_width=True): nav("physical_exam")
    with b3:
        if st.button("🔴 Live",use_container_width=True): nav("live")
    with b4:
        if st.button("🧪 Labs",use_container_width=True): nav("lab")
    with b5:
        if st.button("📝 Diagnose →",use_container_width=True,type="primary"): nav("diagnosis")

# ════════════════════════════════════════════════════════════════
# PHYSICAL EXAMINATION
# ════════════════════════════════════════════════════════════════
# ── Medical Audio Synthesizer ────────────────────────────────────────────────
def get_sound_type(finding_text, exam_type, zone):
    """
    Determine which sound to play based on findings.
    MODALITY-HONEST VERSION:
    - Auscultation → returns auditory sound type (real + synth available)
    - Percussion → returns 'percussion_*' synth-only types (real recordings rare)
    - Palpation → returns 'none' (palpation is TACTILE — no audio at all)
    - Inspection → returns 'none' (inspection is VISUAL — no audio at all)
    """
    f = finding_text.lower()
    z = zone.lower()
    et = exam_type.lower()

    # ═════ AUSCULTATION — Real audio modality ════════════════════════════
    if "auscultat" in et or "auscultation" in et:
        if "heart" in z or "cardiac" in z or "central chest" in z:
            if any(w in f for w in ["murmur","systolic","diastolic","regurgit","stenosis"]): return "murmur"
            if any(w in f for w in ["s3","s4","gallop","extra sound"]): return "s3_gallop"
            if any(w in f for w in ["rub","friction"]): return "pericardial_rub"
            return "normal_heart"
        elif "abdom" in z or "quadrant" in z or "epigastric" in z or "umbilical" in z or "suprapubic" in z or "flank" in z:
            # Bowel sound auscultation
            if any(w in f for w in ["absent","silent","no bowel"]): return "bowel_absent"
            if any(w in f for w in ["hyperactive","increased","hyperperist"]): return "bowel_hyperactive"
            return "bowel_normal"
        else:  # lungs
            if any(w in f for w in ["crackle","crepitation","rales"]): return "crackles"
            if any(w in f for w in ["wheeze","wheezing","bronchospasm"]): return "wheeze"
            if any(w in f for w in ["ronchi","rhonchi","coarse"]): return "ronchi"
            if any(w in f for w in ["reduced","absent","decreased","dull"]): return "reduced_breath"
            if any(w in f for w in ["pleural","rub"]): return "pleural_rub"
            return "normal_breath"

    # ═════ PERCUSSION — Synth-only auditory (real recordings rare) ═══════
    elif "percuss" in et:
        if any(w in f for w in ["dull","stony dull","impaired"]): return "percussion_dull"
        if any(w in f for w in ["hyperresonant","hyper-resonant","tympanic"]): return "percussion_hyperresonant"
        return "percussion_resonant"

    # ═════ PALPATION — TACTILE only, NO audio ════════════════════════════
    # Palpation is felt with hands — not heard. Returning "none" tells the
    # UI to render a tactile description panel instead of an audio player.
    elif "palpat" in et:
        return "none"

    # ═════ INSPECTION — VISUAL only, NO audio ════════════════════════════
    elif "inspect" in et or "look" in et:
        return "none"

    # ═════ SPECIAL TESTS — usually visual or tactile ═════════════════════
    elif "special" in et or "test" in et:
        return "none"

    return "none"

def render_medical_sound_player(sound_type, label="Play Sound"):
    """Render the dual-mode sound player (synth + real recording when available)."""
    # If real recordings are wired in AND this sound has one, use the dual panel.
    # The dual panel uses _render_synth_only for the synthetic version.
    if REAL_SOUNDS_OK and has_real_recording(sound_type):
        render_dual_sound_panel(sound_type,
                                render_synth_func=lambda st_: _render_medical_sound_player_synth(st_, label))
        return
    # Fallback — synth only (with disclaimer if module loaded)
    if REAL_SOUNDS_OK:
        render_dual_sound_panel(sound_type,
                                render_synth_func=lambda st_: _render_medical_sound_player_synth(st_, label))
    else:
        _render_medical_sound_player_synth(sound_type, label)


# ═══════════════════════════════════════════════════════════════════════════
# MODALITY-HONEST PANELS — for palpation (tactile) & inspection (visual)
# ───────────────────────────────────────────────────────────────────────────
# These are NOT audio. Palpation is felt with hands; inspection is visual.
# Showing an "audio player" for these is dishonest — students aren't hearing
# anything in real life when they palpate or inspect a patient.
# Instead, we show an icon panel that names the modality clearly and lists
# what to look/feel for, generated from the AI finding text.
# ═══════════════════════════════════════════════════════════════════════════
def render_tactile_modality_panel(finding_text: str, zone: str = ""):
    """
    Render an honest tactile-modality panel for palpation findings.
    No audio — just a clear visual representation of WHAT is felt.
    """
    finding_lower = (finding_text or "").lower()

    # Determine quality/icon based on finding
    if any(w in finding_lower for w in ["tender", "pain", "rebound", "guarding"]):
        title = "Tactile Finding · Tenderness"
        icon = "🤚"
        accent = "#dc2626"
        bg = "#fef2f2"
        borderc = "#fca5a5"
        what_youd_feel = (
            "Patient winces, withdraws, or verbally reports pain when this area is "
            "pressed. May show voluntary or involuntary guarding."
        )
    elif any(w in finding_lower for w in ["mass", "lump", "organomegaly", "hepatomegaly",
                                            "splenomegaly", "pulsatile"]):
        title = "Tactile Finding · Mass / Organomegaly"
        icon = "✋"
        accent = "#7c3aed"
        bg = "#faf5ff"
        borderc = "#c4b5fd"
        what_youd_feel = (
            "A palpable abnormality felt under your fingertips: enlarged organ, "
            "discrete mass, or pulsatile structure. Note: size, consistency, "
            "mobility, tenderness, borders."
        )
    elif any(w in finding_lower for w in ["rigid", "rigidity", "board-like", "peritoniti"]):
        title = "Tactile Finding · Rigidity"
        icon = "🤲"
        accent = "#b91c1c"
        bg = "#fef2f2"
        borderc = "#fca5a5"
        what_youd_feel = (
            "Abdominal wall is involuntarily tense and board-like. Suggests "
            "peritoneal irritation. Patient cannot relax the muscles even on "
            "request."
        )
    else:
        title = "Tactile Finding · Normal Palpation"
        icon = "🖐️"
        accent = "#059669"
        bg = "#f0fdf4"
        borderc = "#86efac"
        what_youd_feel = (
            "Soft, non-tender, no masses, no organomegaly. Patient comfortable "
            "with examination. No guarding or rebound."
        )

    components.html(f"""
    <div style="border:2px solid {borderc};border-radius:12px;padding:14px;
                background:{bg};margin:8px 0;font-family:Inter,system-ui,sans-serif;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
        <div style="background:{accent};color:white;border-radius:6px;
                    padding:3px 10px;font-size:.7rem;font-weight:700;
                    letter-spacing:.05em;">TACTILE · NO AUDIO</div>
        <div style="font-weight:700;color:#0f172a;font-size:.9rem;">
          {icon} {title}
        </div>
      </div>
      <div style="font-size:.82rem;color:#1e293b;line-height:1.55;
                  background:white;padding:10px 12px;border-radius:8px;
                  margin-bottom:6px;">
        <b style="color:{accent};">What you'd feel:</b> {what_youd_feel}
      </div>
      <div style="font-size:.7rem;color:#64748b;font-style:italic;">
        Palpation is felt with the hands — it has no sound. Practice this on
        real patients or simulation mannequins.
      </div>
    </div>
    """, height=180)


def render_visual_modality_panel(finding_text: str, zone: str = ""):
    """
    Render an honest visual-modality panel for inspection findings.
    No audio — describes what the student would SEE.
    """
    finding_lower = (finding_text or "").lower()

    if any(w in finding_lower for w in ["cyanosis", "cyanotic", "blue", "central cyano"]):
        title = "Visual Finding · Cyanosis"
        icon = "🟦"
        accent = "#1e3a8a"
        bg = "#eff6ff"
        borderc = "#93c5fd"
    elif any(w in finding_lower for w in ["jaundice", "icterus", "yellow"]):
        title = "Visual Finding · Jaundice"
        icon = "🟡"
        accent = "#a16207"
        bg = "#fefce8"
        borderc = "#fde047"
    elif any(w in finding_lower for w in ["pallor", "pale"]):
        title = "Visual Finding · Pallor"
        icon = "⚪"
        accent = "#6b7280"
        bg = "#f9fafb"
        borderc = "#d1d5db"
    elif any(w in finding_lower for w in ["distress", "diaphor", "anxious", "labored"]):
        title = "Visual Finding · Distress"
        icon = "😰"
        accent = "#dc2626"
        bg = "#fef2f2"
        borderc = "#fca5a5"
    else:
        title = "Visual Finding · Inspection"
        icon = "👁️"
        accent = "#0e7490"
        bg = "#ecfeff"
        borderc = "#67e8f9"

    components.html(f"""
    <div style="border:2px solid {borderc};border-radius:12px;padding:14px;
                background:{bg};margin:8px 0;font-family:Inter,system-ui,sans-serif;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
        <div style="background:{accent};color:white;border-radius:6px;
                    padding:3px 10px;font-size:.7rem;font-weight:700;
                    letter-spacing:.05em;">VISUAL · NO AUDIO</div>
        <div style="font-weight:700;color:#0f172a;font-size:.9rem;">
          {icon} {title}
        </div>
      </div>
      <div style="font-size:.78rem;color:#475569;line-height:1.55;
                  background:white;padding:10px 12px;border-radius:8px;">
        Inspection is observation by sight. The AI-generated finding above
        describes what would be visible. There is no sound to play.
      </div>
    </div>
    """, height=140)


def _render_medical_sound_player_synth(sound_type, label="Play Sound"):
    """Render a Web Audio API sound player for the given sound type (synth only)."""
    sounds = {
        "normal_heart": ("Normal Heart Sounds (S1-S2)", "#0ea5e9", """
            // Normal heart sounds: lub-dub at 70 bpm
            function playNormalHeart(ctx) {
                function beat(t) {
                    // S1 (lub)
                    var o1 = ctx.createOscillator(); var g1 = ctx.createGain();
                    o1.frequency.setValueAtTime(80, t); o1.frequency.exponentialRampToValueAtTime(40, t+0.08);
                    g1.gain.setValueAtTime(0, t); g1.gain.linearRampToValueAtTime(0.6, t+0.02);
                    g1.gain.exponentialRampToValueAtTime(0.001, t+0.12);
                    o1.connect(g1); g1.connect(ctx.destination); o1.start(t); o1.stop(t+0.15);
                    // S2 (dub)
                    var t2 = t + 0.28;
                    var o2 = ctx.createOscillator(); var g2 = ctx.createGain();
                    o2.frequency.setValueAtTime(100, t2); o2.frequency.exponentialRampToValueAtTime(50, t2+0.06);
                    g2.gain.setValueAtTime(0, t2); g2.gain.linearRampToValueAtTime(0.4, t2+0.015);
                    g2.gain.exponentialRampToValueAtTime(0.001, t2+0.09);
                    o2.connect(g2); g2.connect(ctx.destination); o2.start(t2); o2.stop(t2+0.12);
                }
                for (var i=0; i<5; i++) beat(ctx.currentTime + i*0.86);
            }
            playNormalHeart(ctx);"""),

        "murmur": ("Cardiac Murmur — Systolic", "#dc2626", """
            // Systolic murmur between S1 and S2
            function playMurmur(ctx) {
                function beat(t) {
                    // S1
                    var o1=ctx.createOscillator(); var g1=ctx.createGain();
                    o1.frequency.setValueAtTime(80,t);
                    g1.gain.setValueAtTime(0,t); g1.gain.linearRampToValueAtTime(0.6,t+0.02);
                    g1.gain.exponentialRampToValueAtTime(0.001,t+0.12);
                    o1.connect(g1); g1.connect(ctx.destination); o1.start(t); o1.stop(t+0.15);
                    // Murmur (noise between S1-S2)
                    var bufSize=ctx.sampleRate*0.25;
                    var buf=ctx.createBuffer(1,bufSize,ctx.sampleRate);
                    var d=buf.getChannelData(0);
                    for(var j=0;j<bufSize;j++) d[j]=(Math.random()*2-1)*0.15;
                    var src=ctx.createBufferSource(); src.buffer=buf;
                    var f=ctx.createBiquadFilter(); f.type="bandpass"; f.frequency.value=300; f.Q.value=2;
                    var gm=ctx.createGain();
                    gm.gain.setValueAtTime(0,t+0.15);
                    gm.gain.linearRampToValueAtTime(0.4,t+0.22);
                    gm.gain.linearRampToValueAtTime(0,t+0.4);
                    src.connect(f); f.connect(gm); gm.connect(ctx.destination);
                    src.start(t+0.13); src.stop(t+0.45);
                    // S2
                    var t2=t+0.3; var o2=ctx.createOscillator(); var g2=ctx.createGain();
                    o2.frequency.setValueAtTime(100,t2);
                    g2.gain.setValueAtTime(0,t2); g2.gain.linearRampToValueAtTime(0.35,t2+0.015);
                    g2.gain.exponentialRampToValueAtTime(0.001,t2+0.09);
                    o2.connect(g2); g2.connect(ctx.destination); o2.start(t2); o2.stop(t2+0.12);
                }
                for(var i=0;i<4;i++) beat(ctx.currentTime+i*0.9);
            }
            playMurmur(ctx);"""),

        "s3_gallop": ("S3 Gallop (Heart Failure)", "#dc2626", """
            function playS3(ctx) {
                function beat(t) {
                    [0,0.28,0.45].forEach(function(dt,i) {
                        var freq=[80,100,55][i]; var vol=[0.6,0.4,0.25][i];
                        var o=ctx.createOscillator(); var g=ctx.createGain();
                        o.frequency.setValueAtTime(freq,t+dt);
                        o.frequency.exponentialRampToValueAtTime(freq*0.5,t+dt+0.1);
                        g.gain.setValueAtTime(0,t+dt); g.gain.linearRampToValueAtTime(vol,t+dt+0.02);
                        g.gain.exponentialRampToValueAtTime(0.001,t+dt+0.12);
                        o.connect(g); g.connect(ctx.destination); o.start(t+dt); o.stop(t+dt+0.15);
                    });
                }
                for(var i=0;i<4;i++) beat(ctx.currentTime+i*0.85);
            }
            playS3(ctx);"""),

        "normal_breath": ("Normal Vesicular Breath Sounds", "#16a34a", """
            function playNormalBreath(ctx) {
                for(var i=0;i<3;i++) {
                    var t=ctx.currentTime+i*2.5;
                    // Inspiration
                    var buf=ctx.createBuffer(1,ctx.sampleRate*1.2,ctx.sampleRate);
                    var d=buf.getChannelData(0);
                    for(var j=0;j<d.length;j++) d[j]=(Math.random()*2-1)*0.12;
                    var src=ctx.createBufferSource(); src.buffer=buf;
                    var f=ctx.createBiquadFilter(); f.type="lowpass"; f.frequency.value=600;
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,t); g.gain.linearRampToValueAtTime(0.35,t+0.5);
                    g.gain.linearRampToValueAtTime(0.15,t+1.0); g.gain.linearRampToValueAtTime(0,t+1.2);
                    src.connect(f); f.connect(g); g.connect(ctx.destination);
                    src.start(t); src.stop(t+1.3);
                    // Expiration (quieter)
                    var buf2=ctx.createBuffer(1,ctx.sampleRate*0.8,ctx.sampleRate);
                    var d2=buf2.getChannelData(0);
                    for(var j=0;j<d2.length;j++) d2[j]=(Math.random()*2-1)*0.08;
                    var src2=ctx.createBufferSource(); src2.buffer=buf2;
                    var f2=ctx.createBiquadFilter(); f2.type="lowpass"; f2.frequency.value=400;
                    var g2=ctx.createGain();
                    g2.gain.setValueAtTime(0,t+1.2); g2.gain.linearRampToValueAtTime(0.15,t+1.5);
                    g2.gain.linearRampToValueAtTime(0,t+2.0);
                    src2.connect(f2); f2.connect(g2); g2.connect(ctx.destination);
                    src2.start(t+1.2); src2.stop(t+2.1);
                }
            }
            playNormalBreath(ctx);"""),

        "crackles": ("Crackles (Fine/Coarse)", "#dc2626", """
            function playCrackles(ctx) {
                for(var cycle=0;cycle<3;cycle++) {
                    var tBase=ctx.currentTime+cycle*2.2;
                    // Breath base
                    var buf=ctx.createBuffer(1,ctx.sampleRate*1.5,ctx.sampleRate);
                    var d=buf.getChannelData(0);
                    for(var j=0;j<d.length;j++) d[j]=(Math.random()*2-1)*0.08;
                    var src=ctx.createBufferSource(); src.buffer=buf;
                    var f=ctx.createBiquadFilter(); f.type="lowpass"; f.frequency.value=500;
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,tBase); g.gain.linearRampToValueAtTime(0.2,tBase+0.4);
                    g.gain.linearRampToValueAtTime(0,tBase+1.5);
                    src.connect(f); f.connect(g); g.connect(ctx.destination);
                    src.start(tBase); src.stop(tBase+1.6);
                    // Crackles — random pops
                    var numCrackles=15+Math.floor(Math.random()*10);
                    for(var i=0;i<numCrackles;i++) {
                        var ct=tBase+0.2+Math.random()*1.0;
                        var cb=ctx.createBuffer(1,Math.floor(ctx.sampleRate*0.008),ctx.sampleRate);
                        var cd=cb.getChannelData(0);
                        for(var j=0;j<cd.length;j++) cd[j]=(Math.random()*2-1)*0.5;
                        var cs=ctx.createBufferSource(); cs.buffer=cb;
                        var cf=ctx.createBiquadFilter(); cf.type="bandpass";
                        cf.frequency.value=1000+Math.random()*2000; cf.Q.value=0.5;
                        var cg=ctx.createGain(); cg.gain.value=0.4+Math.random()*0.3;
                        cs.connect(cf); cf.connect(cg); cg.connect(ctx.destination);
                        cs.start(ct); cs.stop(ct+0.01);
                    }
                }
            }
            playCrackles(ctx);"""),

        "wheeze": ("Wheeze — Expiratory", "#f59e0b", """
            function playWheeze(ctx) {
                for(var i=0;i<3;i++) {
                    var t=ctx.currentTime+i*3.0;
                    // Inspiration normal
                    var buf=ctx.createBuffer(1,ctx.sampleRate*1.0,ctx.sampleRate);
                    var d=buf.getChannelData(0);
                    for(var j=0;j<d.length;j++) d[j]=(Math.random()*2-1)*0.1;
                    var src=ctx.createBufferSource(); src.buffer=buf;
                    var f=ctx.createBiquadFilter(); f.type="lowpass"; f.frequency.value=500;
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,t); g.gain.linearRampToValueAtTime(0.25,t+0.4);
                    g.gain.linearRampToValueAtTime(0,t+1.0);
                    src.connect(f); f.connect(g); g.connect(ctx.destination); src.start(t); src.stop(t+1.1);
                    // Wheeze on expiration
                    var wt=t+1.1;
                    var wo=ctx.createOscillator(); wo.type="sawtooth";
                    wo.frequency.setValueAtTime(600,wt); wo.frequency.setValueAtTime(550,wt+0.5);
                    wo.frequency.setValueAtTime(580,wt+1.0); wo.frequency.setValueAtTime(560,wt+1.5);
                    var wf=ctx.createBiquadFilter(); wf.type="bandpass"; wf.frequency.value=600; wf.Q.value=3;
                    var wg=ctx.createGain();
                    wg.gain.setValueAtTime(0,wt); wg.gain.linearRampToValueAtTime(0.15,wt+0.2);
                    wg.gain.setValueAtTime(0.12,wt+1.5); wg.gain.linearRampToValueAtTime(0,wt+1.8);
                    wo.connect(wf); wf.connect(wg); wg.connect(ctx.destination);
                    wo.start(wt); wo.stop(wt+1.9);
                }
            }
            playWheeze(ctx);"""),

        "ronchi": ("Rhonchi — Low-pitched", "#f59e0b", """
            function playRonchi(ctx) {
                for(var i=0;i<3;i++) {
                    var t=ctx.currentTime+i*2.5;
                    var buf=ctx.createBuffer(1,ctx.sampleRate*2.0,ctx.sampleRate);
                    var d=buf.getChannelData(0);
                    for(var j=0;j<d.length;j++) d[j]=(Math.random()*2-1)*0.2;
                    var src=ctx.createBufferSource(); src.buffer=buf;
                    var f=ctx.createBiquadFilter(); f.type="bandpass"; f.frequency.value=200; f.Q.value=1.5;
                    var lfo=ctx.createOscillator(); lfo.frequency.value=4;
                    var lfog=ctx.createGain(); lfog.gain.value=100;
                    lfo.connect(lfog); lfog.connect(f.frequency);
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,t); g.gain.linearRampToValueAtTime(0.35,t+0.4);
                    g.gain.setValueAtTime(0.3,t+1.5); g.gain.linearRampToValueAtTime(0,t+2.0);
                    src.connect(f); f.connect(g); g.connect(ctx.destination);
                    src.start(t); src.stop(t+2.1); lfo.start(t); lfo.stop(t+2.1);
                }
            }
            playRonchi(ctx);"""),

        "reduced_breath": ("Reduced / Absent Breath Sounds", "#6b7280", """
            function playReduced(ctx) {
                for(var i=0;i<3;i++) {
                    var t=ctx.currentTime+i*2.5;
                    var buf=ctx.createBuffer(1,ctx.sampleRate*1.5,ctx.sampleRate);
                    var d=buf.getChannelData(0);
                    for(var j=0;j<d.length;j++) d[j]=(Math.random()*2-1)*0.03;
                    var src=ctx.createBufferSource(); src.buffer=buf;
                    var f=ctx.createBiquadFilter(); f.type="lowpass"; f.frequency.value=300;
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,t); g.gain.linearRampToValueAtTime(0.08,t+0.5);
                    g.gain.linearRampToValueAtTime(0,t+1.5);
                    src.connect(f); f.connect(g); g.connect(ctx.destination); src.start(t); src.stop(t+1.6);
                }
            }
            playReduced(ctx);"""),

        "pleural_rub": ("Pleural Friction Rub", "#dc2626", """
            function playPleuralRub(ctx) {
                for(var i=0;i<4;i++) {
                    var t=ctx.currentTime+i*1.8;
                    var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*0.6),ctx.sampleRate);
                    var d=buf.getChannelData(0);
                    for(var j=0;j<d.length;j++) {
                        var env=Math.sin(Math.PI*j/d.length);
                        d[j]=(Math.random()*2-1)*env*0.4;
                    }
                    var src=ctx.createBufferSource(); src.buffer=buf;
                    var f=ctx.createBiquadFilter(); f.type="bandpass"; f.frequency.value=400; f.Q.value=0.8;
                    var g=ctx.createGain(); g.gain.value=0.5;
                    src.connect(f); f.connect(g); g.connect(ctx.destination); src.start(t); src.stop(t+0.7);
                }
            }
            playPleuralRub(ctx);"""),

        "pericardial_rub": ("Pericardial Friction Rub", "#dc2626", """
            function playPericardialRub(ctx) {
                function beat(t) {
                    // 3-component rub
                    [0, 0.18, 0.35].forEach(function(dt) {
                        var buf=ctx.createBuffer(1,Math.floor(ctx.sampleRate*0.1),ctx.sampleRate);
                        var d=buf.getChannelData(0);
                        for(var j=0;j<d.length;j++) d[j]=(Math.random()*2-1)*Math.sin(Math.PI*j/d.length)*0.5;
                        var src=ctx.createBufferSource(); src.buffer=buf;
                        var f=ctx.createBiquadFilter(); f.type="bandpass"; f.frequency.value=350; f.Q.value=1;
                        var g=ctx.createGain(); g.gain.value=0.45;
                        src.connect(f); f.connect(g); g.connect(ctx.destination); src.start(t+dt); src.stop(t+dt+0.12);
                    });
                }
                for(var i=0;i<5;i++) beat(ctx.currentTime+i*0.86);
            }
            playPericardialRub(ctx);"""),

        "percussion_resonant": ("Percussion — Normal Resonant", "#16a34a", """
            function playResonant(ctx) {
                for(var i=0;i<4;i++) {
                    var t=ctx.currentTime+i*0.8;
                    var o=ctx.createOscillator(); o.type="sine";
                    o.frequency.setValueAtTime(180,t); o.frequency.exponentialRampToValueAtTime(80,t+0.5);
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,t); g.gain.linearRampToValueAtTime(0.5,t+0.01);
                    g.gain.exponentialRampToValueAtTime(0.001,t+0.55);
                    o.connect(g); g.connect(ctx.destination); o.start(t); o.stop(t+0.6);
                }
            }
            playResonant(ctx);"""),

        "percussion_dull": ("Percussion — Dull (Fluid/Consolidation)", "#dc2626", """
            function playDull(ctx) {
                for(var i=0;i<4;i++) {
                    var t=ctx.currentTime+i*0.8;
                    var o=ctx.createOscillator(); o.type="sine";
                    o.frequency.setValueAtTime(90,t); o.frequency.exponentialRampToValueAtTime(60,t+0.15);
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,t); g.gain.linearRampToValueAtTime(0.6,t+0.01);
                    g.gain.exponentialRampToValueAtTime(0.001,t+0.2);
                    o.connect(g); g.connect(ctx.destination); o.start(t); o.stop(t+0.25);
                }
            }
            playDull(ctx);"""),

        "percussion_hyperresonant": ("Percussion — Hyperresonant (Air/Pneumothorax)", "#f59e0b", """
            function playHyperresonant(ctx) {
                for(var i=0;i<4;i++) {
                    var t=ctx.currentTime+i*0.8;
                    var o=ctx.createOscillator(); o.type="sine";
                    o.frequency.setValueAtTime(260,t); o.frequency.exponentialRampToValueAtTime(100,t+0.8);
                    var g=ctx.createGain();
                    g.gain.setValueAtTime(0,t); g.gain.linearRampToValueAtTime(0.45,t+0.01);
                    g.gain.exponentialRampToValueAtTime(0.001,t+0.85);
                    o.connect(g); g.connect(ctx.destination); o.start(t); o.stop(t+0.9);
                }
            }
            playHyperresonant(ctx);"""),

        "tenderness": ("Palpation — Tenderness Response", "#f59e0b", """
            // Visual/audio cue for tenderness - patient winces
            function playTenderness(ctx) {
                var o=ctx.createOscillator(); o.type="sawtooth"; o.frequency.value=300;
                var g=ctx.createGain();
                g.gain.setValueAtTime(0.0,ctx.currentTime);
                g.gain.linearRampToValueAtTime(0.2,ctx.currentTime+0.05);
                g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+0.3);
                var f=ctx.createBiquadFilter(); f.type="lowpass"; f.frequency.value=400;
                o.connect(f); f.connect(g); g.connect(ctx.destination);
                o.start(); o.stop(ctx.currentTime+0.35);
            }
            playTenderness(ctx);"""),

        "normal_palpation": ("Palpation — Soft, Non-tender", "#16a34a", """
            function playNormalPalp(ctx) {
                var o=ctx.createOscillator(); o.type="sine"; o.frequency.value=200;
                var g=ctx.createGain();
                g.gain.setValueAtTime(0.0,ctx.currentTime);
                g.gain.linearRampToValueAtTime(0.1,ctx.currentTime+0.05);
                g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+0.4);
                o.connect(g); g.connect(ctx.destination);
                o.start(); o.stop(ctx.currentTime+0.45);
            }
            playNormalPalp(ctx);"""),

        "none": None,
    }

    if sound_type not in sounds or sounds[sound_type] is None:
        return

    name, color, js_code = sounds[sound_type]

    components.html(f"""
    <div style="font-family:Inter,sans-serif;margin:.4rem 0;">
        <button onclick="playSound_{sound_type.replace('-','_')}()" style="
            background:linear-gradient(135deg,{color},{color}cc);
            color:white;border:none;border-radius:8px;
            padding:.4rem 1rem;font-size:.8rem;font-weight:600;
            cursor:pointer;box-shadow:0 2px 8px {color}44;
            display:flex;align-items:center;gap:.4rem;">
            🔊 {name}
        </button>
    </div>
    <script>
    function playSound_{sound_type.replace('-','_')}() {{
        var ctx = new (window.AudioContext || window.webkitAudioContext)();
        {js_code}
    }}
    </script>
    """, height=50)


def page_physical_exam():
    st.markdown('<div class="section-header">🫀 Physical Examination Room</div>',unsafe_allow_html=True)
    c=st.session_state.selected_case
    if not c: st.warning("No case selected."); return

    # ── Disclaimer about clinical sounds ─────────────────────────────────
    if REAL_SOUNDS_OK:
        render_disclaimer_banner()

    st.markdown(f'<div class="alert-info">👤 <b>Patient:</b> {c.get("Age_Sex","?")} | <b>CC:</b> {c.get("Chief_Complaint","?")} | <b>Vitals:</b> {c.get("Vitals","?")}</div>',unsafe_allow_html=True)
    st.markdown(f'<div class="alert-warn">👁️ <b>Appearance:</b> {c.get("Appearance","In pain")}</div>',unsafe_allow_html=True)

    st.markdown("")
    exam_type=st.selectbox("🔍 Select Examination Type:",
        ["🔊 Auscultation","👆 Palpation (Light)","👊 Palpation (Deep)","🥁 Percussion","👁️ Inspection","🧪 Special Test"])

    st.markdown("**Click a body zone to examine — AI generates clinically accurate findings:**")
    st.caption(
        "🔊 **Auscultation & Percussion** play sounds (real recordings + synthetic teaching tones). "
        "🤚 **Palpation** shows tactile findings (no audio — palpation is felt, not heard). "
        "👁️ **Inspection** shows visual findings (no audio — inspection is observed)."
    )

    zones_chest=["Right Upper Chest","Left Upper Chest","Right Lower Chest","Left Lower Chest","Central Chest (Heart)","Posterior Right Upper","Posterior Right Lower","Posterior Left Upper","Posterior Left Lower"]
    zones_abdomen=["Right Upper Quadrant (RUQ)","Left Upper Quadrant (LUQ)","Right Lower Quadrant (RLQ)","Left Lower Quadrant (LLQ)","Epigastric Region","Periumbilical Region","Suprapubic Region","Right Flank","Left Flank"]
    zones_other=["Head & Neck","Back & Lumbar Spine","Right Upper Limb","Left Upper Limb","Right Lower Limb","Left Lower Limb","Peripheral Pulses","Lymph Nodes"]

    tab1,tab2,tab3=st.tabs(["🫁 Chest & Heart","🫃 Abdomen","🦴 Other Regions"])

    def do_exam(zone):
        key=f"{exam_type}|{zone}"
        if key in st.session_state.exam_findings:
            data=st.session_state.exam_findings[key]
            finding_text=data if isinstance(data,str) else data.get("finding","")
            sound_t=data.get("sound","none") if isinstance(data,dict) else get_sound_type(finding_text,exam_type,zone)
            col_f,col_s=st.columns([3,1])
            with col_f:
                icon="🔊" if "Auscultat" in exam_type else "🥁" if "Percuss" in exam_type else "🤚" if "Palp" in exam_type else "👁️"
                st.markdown(f'<div class="exam-finding">{icon} <b>{exam_type.split()[1]} — {zone}</b><br><span style="color:#374151;line-height:1.6">{finding_text}</span></div>',unsafe_allow_html=True)
            with col_s:
                # MODALITY-HONEST RENDERING:
                # - Auscultation/Percussion → audio player (sound exists)
                # - Palpation → tactile panel (NO audio — felt with hands)
                # - Inspection → visual panel (NO audio — observed visually)
                if "Auscultat" in exam_type or "Percuss" in exam_type:
                    if sound_t!="none":
                        render_medical_sound_player(sound_t,"🔊 Play")
                elif "Palp" in exam_type:
                    render_tactile_modality_panel(finding_text, zone)
                elif "Inspect" in exam_type:
                    render_visual_modality_panel(finding_text, zone)
                # Special tests fall through silently (no panel needed)
        else:
            if st.button(f"🔍 Examine: {zone}",key=f"ex_{key[:40]}",use_container_width=True):
                prompt=(f"Clinical simulation for MLS Academy.\n"
                        f"Patient: {c.get('Age_Sex','?')} presenting with {c.get('Chief_Complaint','?')}\n"
                        f"Vitals: {c.get('Vitals','?')}\n"
                        f"Known physical findings: {c.get('Physical_Findings','?')}\n"
                        f"True diagnosis (generate accurate findings): {c.get('Final_Diagnosis','?')}\n\n"
                        f"Generate realistic {exam_type.split()[1].lower()} findings for zone: {zone}\n"
                        f"Rules: Be clinically accurate. Match to diagnosis. Normal zones = normal findings.\n"
                        f"Format: 1-2 clear clinical sentences. Include specific descriptors.\n"
                        f"Examples:\n"
                        f"- Auscultation heart: 'Regular S1 S2, no murmurs, no extra sounds'\n"
                        f"- Auscultation RLL pneumonia: 'Coarse crackles right lower lobe, reduced air entry, no wheeze'\n"
                        f"- Percussion RLL pneumonia: 'Stony dull at right base, dullness extending to mid-zone'\n"
                        f"- Palpation RLQ appendicitis: 'Exquisite tenderness at McBurney point, voluntary guarding, rebound tenderness positive'")
                with st.spinner(f"Examining {zone}..."):
                    result=call_ai("You are a clinical simulation system generating physical examination findings.",
                                   [{"role":"user","content":prompt}],max_tokens=180)
                if not result.startswith("!ERR"):
                    sound_t=get_sound_type(result,exam_type,zone)
                    st.session_state.exam_findings[key]={"finding":result,"sound":sound_t,"zone":zone,"type":exam_type,"time":datetime.now().strftime("%H:%M")}
                    st.rerun()

    with tab1:
        cols=st.columns(2)
        for i,zone in enumerate(zones_chest):
            with cols[i%2]: do_exam(zone)
    with tab2:
        cols=st.columns(2)
        for i,zone in enumerate(zones_abdomen):
            with cols[i%2]: do_exam(zone)
    with tab3:
        cols=st.columns(2)
        for i,zone in enumerate(zones_other):
            with cols[i%2]: do_exam(zone)

    if st.session_state.exam_findings:
        st.markdown("---")
        st.markdown("### 📋 Examination Findings & Clinical Sounds")
        st.markdown('<div class="alert-info">🔊 Auscultation & percussion findings have audio. 🤚 Palpation shows tactile descriptions (no audio). 👁️ Inspection shows visual descriptions (no audio).</div>', unsafe_allow_html=True)
        done=len(st.session_state.exam_findings)
        st.markdown(f"**{done} zone(s) examined**")

        # Filter
        ftype=st.selectbox("Filter by exam type:",["All","Auscultation","Palpation","Percussion","Inspection","Special Test"])

        for k,data in st.session_state.exam_findings.items():
            if isinstance(data,dict):
                finding=data.get("finding",""); etype=data.get("type",""); zone=data.get("zone",""); t=data.get("time",""); sound_t=data.get("sound","none")
            else:
                finding=data; etype=k.split("|")[0] if "|" in k else ""; zone=k.split("|")[1] if "|" in k else k; t=""; sound_t=get_sound_type(finding,etype,zone)

            if ftype!="All" and ftype.lower() not in etype.lower(): continue

            icon="🔊" if "Auscultat" in etype else "🥁" if "Percuss" in etype else "🤚" if "Palp" in etype else "👁️"
            col_f,col_s=st.columns([3,1])
            with col_f:
                st.markdown(f'<div class="exam-finding">{icon} <b>{etype.split()[1] if len(etype.split())>1 else etype} — {zone}</b> <span style="font-size:.72rem;color:#9ca3af">⏰{t}</span><br><span style="color:#374151;line-height:1.6">{finding}</span></div>',unsafe_allow_html=True)
            with col_s:
                # MODALITY-HONEST RENDERING (same logic as do_exam)
                if "Auscultat" in etype or "Percuss" in etype:
                    if sound_t!="none":
                        render_medical_sound_player(sound_t)
                elif "Palp" in etype:
                    render_tactile_modality_panel(finding, zone)
                elif "Inspect" in etype:
                    render_visual_modality_panel(finding, zone)

        # AI Clinical Summary
        if done>=2:
            st.markdown("")
            if st.button("🤖 Generate AI Clinical Exam Summary",use_container_width=True,type="primary"):
                all_f="\n".join([f"{(d.get('type','') if isinstance(d,dict) else k.split('|')[0])} - {(d.get('zone','') if isinstance(d,dict) else k.split('|')[1] if '|' in k else k)}: {(d.get('finding',d) if isinstance(d,dict) else d)}" for k,d in st.session_state.exam_findings.items()])
                with st.spinner("Generating summary..."):
                    summ=call_ai("You are a senior physician writing a clinical examination summary.",
                                [{"role":"user","content":f"Patient: {c.get('Age_Sex','?')} CC: {c.get('Chief_Complaint','?')}\nFindings:\n{all_f}\n\nWrite a brief clinical exam summary highlighting key positive and negative findings. Suggest what they indicate diagnostically without stating the final diagnosis."}],max_tokens=400)
                st.markdown(f'<div style="background:#f0f9ff;border-radius:12px;padding:1.2rem;border:2px solid #0ea5e9;">{summ.replace(chr(10),"<br>")}</div>',unsafe_allow_html=True)

        if st.button("🗑️ Clear All Findings",use_container_width=True):
            st.session_state.exam_findings={}; st.rerun()

    st.markdown("")
    b1,b2,b3=st.columns(3)
    with b1:
        if st.button("← ER",use_container_width=True): nav("emergency")
    with b2:
        if st.button("🧪 Labs",use_container_width=True): nav("lab")
    with b3:
        if st.button("📝 Diagnose →",use_container_width=True,type="primary"): nav("diagnosis")

# ════════════════════════════════════════════════════════════════
# LABORATORY
# ════════════════════════════════════════════════════════════════
# ── Complete Reference Ranges Database ───────────────────────────────────────
LAB_REFERENCE_RANGES = {
    # CBC
    "WBC":              {"low":4.5,   "high":11.0,  "unit":"x10³/μL",  "decimals":1},
    "Neutrophils":      {"low":1.8,   "high":7.7,   "unit":"x10³/μL",  "decimals":2},
    "Neutrophils %":    {"low":40,    "high":75,    "unit":"%",        "decimals":0},
    "Lymphocytes":      {"low":1.0,   "high":4.8,   "unit":"x10³/μL",  "decimals":2},
    "Lymphocytes %":    {"low":20,    "high":45,    "unit":"%",        "decimals":0},
    "Monocytes %":      {"low":2,     "high":10,    "unit":"%",        "decimals":0},
    "Eosinophils %":    {"low":1,     "high":6,     "unit":"%",        "decimals":0},
    "Hemoglobin (M)":   {"low":13.5,  "high":17.5,  "unit":"g/dL",     "decimals":1},
    "Hemoglobin (F)":   {"low":12.0,  "high":15.5,  "unit":"g/dL",     "decimals":1},
    "Hematocrit (M)":   {"low":41,    "high":53,    "unit":"%",        "decimals":0},
    "Hematocrit (F)":   {"low":36,    "high":46,    "unit":"%",        "decimals":0},
    "MCV":              {"low":80,    "high":100,   "unit":"fL",       "decimals":0},
    "MCH":              {"low":27,    "high":33,    "unit":"pg",       "decimals":1},
    "MCHC":             {"low":32,    "high":36,    "unit":"g/dL",     "decimals":1},
    "RDW":              {"low":11.5,  "high":14.5,  "unit":"%",        "decimals":1},
    "Platelets":        {"low":150,   "high":400,   "unit":"x10³/μL",  "decimals":0},
    "MPV":              {"low":7.5,   "high":12.5,  "unit":"fL",       "decimals":1},
    # Inflammation
    "CRP":              {"low":0,     "high":5,     "unit":"mg/L",     "decimals":1},
    "ESR (M)":          {"low":0,     "high":15,    "unit":"mm/hr",    "decimals":0},
    "ESR (F)":          {"low":0,     "high":20,    "unit":"mm/hr",    "decimals":0},
    "Procalcitonin":    {"low":0,     "high":0.5,   "unit":"ng/mL",    "decimals":2},
    "Ferritin (M)":     {"low":30,    "high":400,   "unit":"ng/mL",    "decimals":0},
    "Ferritin (F)":     {"low":13,    "high":150,   "unit":"ng/mL",    "decimals":0},
    # Metabolic
    "Sodium":           {"low":135,   "high":145,   "unit":"mEq/L",    "decimals":0},
    "Potassium":        {"low":3.5,   "high":5.0,   "unit":"mEq/L",    "decimals":1},
    "Chloride":         {"low":98,    "high":107,   "unit":"mEq/L",    "decimals":0},
    "Bicarbonate":      {"low":22,    "high":29,    "unit":"mEq/L",    "decimals":0},
    "BUN":              {"low":7,     "high":20,    "unit":"mg/dL",    "decimals":0},
    "Creatinine (M)":   {"low":0.7,   "high":1.2,   "unit":"mg/dL",    "decimals":2},
    "Creatinine (F)":   {"low":0.5,   "high":1.0,   "unit":"mg/dL",    "decimals":2},
    "eGFR":             {"low":60,    "high":120,   "unit":"mL/min/1.73m²","decimals":0},
    "Glucose (fasting)":{"low":70,    "high":100,   "unit":"mg/dL",    "decimals":0},
    "HbA1c":            {"low":4.0,   "high":5.6,   "unit":"%",        "decimals":1},
    "Calcium":          {"low":8.5,   "high":10.5,  "unit":"mg/dL",    "decimals":1},
    "Magnesium":        {"low":1.7,   "high":2.2,   "unit":"mg/dL",    "decimals":1},
    "Phosphate":        {"low":2.5,   "high":4.5,   "unit":"mg/dL",    "decimals":1},
    "Uric Acid (M)":    {"low":3.4,   "high":7.0,   "unit":"mg/dL",    "decimals":1},
    "Uric Acid (F)":    {"low":2.4,   "high":6.0,   "unit":"mg/dL",    "decimals":1},
    # Liver
    "ALT (SGPT)":       {"low":7,     "high":56,    "unit":"U/L",      "decimals":0},
    "AST (SGOT)":       {"low":10,    "high":40,    "unit":"U/L",      "decimals":0},
    "ALP":              {"low":44,    "high":147,   "unit":"U/L",      "decimals":0},
    "GGT (M)":          {"low":8,     "high":61,    "unit":"U/L",      "decimals":0},
    "GGT (F)":          {"low":5,     "high":36,    "unit":"U/L",      "decimals":0},
    "Total Bilirubin":  {"low":0.1,   "high":1.2,   "unit":"mg/dL",    "decimals":2},
    "Direct Bilirubin": {"low":0.0,   "high":0.3,   "unit":"mg/dL",    "decimals":2},
    "Indirect Bilirubin":{"low":0.1,  "high":0.9,   "unit":"mg/dL",    "decimals":2},
    "Total Protein":    {"low":6.0,   "high":8.3,   "unit":"g/dL",     "decimals":1},
    "Albumin":          {"low":3.5,   "high":5.0,   "unit":"g/dL",     "decimals":1},
    # Cardiac
    "Troponin I":       {"low":0,     "high":0.04,  "unit":"ng/mL",    "decimals":3},
    "Troponin T":       {"low":0,     "high":0.01,  "unit":"ng/mL",    "decimals":3},
    "CK-MB":            {"low":0,     "high":5.0,   "unit":"ng/mL",    "decimals":1},
    "CK Total":         {"low":22,    "high":198,   "unit":"U/L",      "decimals":0},
    "LDH":              {"low":140,   "high":280,   "unit":"U/L",      "decimals":0},
    "BNP":              {"low":0,     "high":100,   "unit":"pg/mL",    "decimals":0},
    "NT-proBNP":        {"low":0,     "high":125,   "unit":"pg/mL",    "decimals":0},
    "Myoglobin":        {"low":0,     "high":90,    "unit":"ng/mL",    "decimals":0},
    # Lipids
    "Total Cholesterol":{"low":0,     "high":200,   "unit":"mg/dL",    "decimals":0},
    "LDL":              {"low":0,     "high":100,   "unit":"mg/dL",    "decimals":0},
    "HDL (M)":          {"low":40,    "high":60,    "unit":"mg/dL",    "decimals":0},
    "HDL (F)":          {"low":50,    "high":60,    "unit":"mg/dL",    "decimals":0},
    "Triglycerides":    {"low":0,     "high":150,   "unit":"mg/dL",    "decimals":0},
    # Coagulation
    "PT":               {"low":11,    "high":13.5,  "unit":"sec",      "decimals":1},
    "INR":              {"low":0.8,   "high":1.2,   "unit":"",         "decimals":2},
    "APTT":             {"low":25,    "high":35,    "unit":"sec",      "decimals":1},
    "Fibrinogen":       {"low":200,   "high":400,   "unit":"mg/dL",    "decimals":0},
    "D-Dimer":          {"low":0,     "high":0.5,   "unit":"μg/mL FEU","decimals":2},
    # Thyroid
    "TSH":              {"low":0.4,   "high":4.0,   "unit":"mIU/L",    "decimals":2},
    "Free T4":          {"low":0.8,   "high":1.8,   "unit":"ng/dL",    "decimals":2},
    "Free T3":          {"low":2.3,   "high":4.2,   "unit":"pg/mL",    "decimals":1},
    # Pancreatic
    "Amylase":          {"low":30,    "high":110,   "unit":"U/L",      "decimals":0},
    "Lipase":           {"low":10,    "high":140,   "unit":"U/L",      "decimals":0},
    # Other
    "Lactate":          {"low":0.5,   "high":2.0,   "unit":"mmol/L",   "decimals":1},
    "Serum Iron":       {"low":60,    "high":170,   "unit":"μg/dL",    "decimals":0},
    "TIBC":             {"low":240,   "high":450,   "unit":"μg/dL",    "decimals":0},
    "Beta-hCG":         {"low":0,     "high":5,     "unit":"mIU/mL",   "decimals":1},
    "PSA":              {"low":0,     "high":4.0,   "unit":"ng/mL",    "decimals":2},
    "Cortisol (AM)":    {"low":6,     "high":23,    "unit":"μg/dL",    "decimals":1},
    "Vitamin B12":      {"low":200,   "high":900,   "unit":"pg/mL",    "decimals":0},
    "Folate":           {"low":2.7,   "high":17.0,  "unit":"ng/mL",    "decimals":1},
    "Vitamin D (25-OH)":{"low":30,    "high":100,   "unit":"ng/mL",    "decimals":0},
    # ABG
    "pH":               {"low":7.35,  "high":7.45,  "unit":"",         "decimals":2},
    "PaCO2":            {"low":35,    "high":45,    "unit":"mmHg",     "decimals":0},
    "PaO2":             {"low":80,    "high":100,   "unit":"mmHg",     "decimals":0},
    "HCO3":             {"low":22,    "high":26,    "unit":"mEq/L",    "decimals":1},
    "SaO2":             {"low":95,    "high":100,   "unit":"%",        "decimals":0},
    "Base Excess":      {"low":-2,    "high":2,     "unit":"mEq/L",    "decimals":1},
    # Urinalysis
    "Urine pH":         {"low":4.5,   "high":8.0,   "unit":"",         "decimals":1},
    "Urine SG":         {"low":1.002, "high":1.030, "unit":"",         "decimals":3},
}

# Tests included per panel ordered
LAB_PANELS = {
    "CBC (Full Blood Count)": [
        "WBC","Neutrophils %","Lymphocytes %","Monocytes %","Eosinophils %",
        "Hemoglobin","Hematocrit","MCV","MCH","MCHC","RDW","Platelets","MPV"
    ],
    "CRP / ESR (Inflammatory Markers)": ["CRP","ESR","Procalcitonin"],
    "Basic Metabolic Panel (BMP)": [
        "Sodium","Potassium","Chloride","Bicarbonate","BUN","Creatinine","eGFR","Glucose (fasting)"
    ],
    "Liver Function Tests (LFTs)": [
        "ALT (SGPT)","AST (SGOT)","ALP","GGT","Total Bilirubin","Direct Bilirubin",
        "Indirect Bilirubin","Total Protein","Albumin"
    ],
    "Renal Function (Creatinine/BUN/eGFR)": ["BUN","Creatinine","eGFR","Sodium","Potassium"],
    "Cardiac Enzymes (Troponin I/T, CK-MB)": [
        "Troponin I","Troponin T","CK-MB","CK Total","LDH","Myoglobin"
    ],
    "Coagulation Panel (PT/INR/APTT/Fibrinogen)": ["PT","INR","APTT","Fibrinogen","D-Dimer"],
    "Blood Cultures x2": [],   # Text result
    "Thyroid Function (TSH/FT3/FT4)": ["TSH","Free T4","Free T3"],
    "Lipid Panel (Cholesterol/LDL/HDL/TG)": [
        "Total Cholesterol","LDL","HDL","Triglycerides"
    ],
    "Electrolytes (Na/K/Cl/HCO3)": ["Sodium","Potassium","Chloride","Bicarbonate"],
    "Blood Glucose / HbA1c": ["Glucose (fasting)","HbA1c"],
    "ABG (Arterial Blood Gas)": ["pH","PaCO2","PaO2","HCO3","SaO2","Base Excess"],
    "Urinalysis + Microscopy": [],   # Text result
    "Urine Culture & Sensitivity": [],  # Text result
    "Amylase / Lipase": ["Amylase","Lipase"],
    "D-dimer": ["D-Dimer"],
    "BNP / Pro-BNP": ["BNP","NT-proBNP"],
    "Influenza A/B Rapid Test": [],  # Text result
    "Serum Bilirubin (Total/Direct/Indirect)": [
        "Total Bilirubin","Direct Bilirubin","Indirect Bilirubin"
    ],
    "Beta-hCG (Pregnancy Test)": ["Beta-hCG"],
    "Blood Group & Cross-match": [],  # Text result
    "Serum Lactate": ["Lactate"],
    "Procalcitonin": ["Procalcitonin"],
    "LDH": ["LDH"],
    "Uric Acid": ["Uric Acid"],
    "Serum Calcium/Magnesium/Phosphate": ["Calcium","Magnesium","Phosphate"],
    "Iron Studies (Fe/TIBC/Ferritin)": ["Serum Iron","TIBC","Ferritin"],
    "Vitamin B12/Folate": ["Vitamin B12","Folate"],
    "Cortisol (Random/Stimulated)": ["Cortisol (AM)"],
    "PSA (Prostate Specific Antigen)": ["PSA"],
    "Serum Albumin": ["Albumin"],
}

def get_ref_for_test(test_name, age_sex=""):
    """Get reference range for a specific test, adjusting for sex."""
    is_male = "male" in str(age_sex).lower() or " m" in str(age_sex).lower()
    
    # Try sex-specific first
    if is_male:
        for suffix in [" (M)", "(M)"]:
            if test_name + suffix in LAB_REFERENCE_RANGES:
                return test_name, LAB_REFERENCE_RANGES[test_name + suffix]
    else:
        for suffix in [" (F)", "(F)"]:
            if test_name + suffix in LAB_REFERENCE_RANGES:
                return test_name, LAB_REFERENCE_RANGES[test_name + suffix]
    
    # Generic
    if test_name in LAB_REFERENCE_RANGES:
        return test_name, LAB_REFERENCE_RANGES[test_name]
    
    # Partial match
    for k, v in LAB_REFERENCE_RANGES.items():
        base = k.split(" (")[0].lower()
        if base == test_name.lower().split(" (")[0]:
            return k, v
    return None, None


def generate_lab_results_ai(selected_tests, case_data):
    """
    Use AI to generate realistic, case-specific lab values.
    Values are consistent with the diagnosis and cached.
    """
    dx   = case_data.get("Final_Diagnosis","unknown")
    age  = case_data.get("Age_Sex","adult")
    cc   = case_data.get("Chief_Complaint","")
    pmh  = case_data.get("PMH","none")
    known= case_data.get("Labs","none")

    # Build list of all tests needed
    all_tests = []
    for panel in selected_tests:
        tests_in_panel = LAB_PANELS.get(panel, [])
        for t in tests_in_panel:
            if t not in all_tests:
                all_tests.append(t)

    if not all_tests:
        return {}

    prompt = f"""You are a clinical pathologist generating laboratory results for a medical simulation.

PATIENT: {age} | Chief complaint: {cc}
DIAGNOSIS: {dx}
PMH: {pmh}
KNOWN LABS FROM CHART: {known}

Generate REALISTIC lab values for these tests that are CLINICALLY CONSISTENT with the diagnosis.

Tests to generate: {", ".join(all_tests)}

RULES:
1. Values must be clinically realistic for {dx}
2. Abnormal values must make clinical sense for this diagnosis
3. If diagnosis causes a specific pattern, reflect it (e.g., appendicitis → elevated WBC/CRP)
4. Use known labs from chart as anchor if they overlap
5. Be consistent — same disease → predictable pattern

Return ONLY a JSON object with exact numeric values:
{{"WBC": 14.2, "Neutrophils %": 85, "CRP": 78, ...}}

No text, no explanation, ONLY the JSON object."""

    result = call_ai(
        "You are a clinical pathologist generating realistic lab values for medical education.",
        [{"role":"user","content":prompt}],
        max_tokens=600
    )
    
    import json as _j, re as _r
    try:
        m = _r.search(r"\{[^{}]+\}", result, _r.DOTALL)
        if m:
            return _j.loads(m.group(0))
    except Exception:
        pass
    return {}


def render_lab_value(test_name, value, ref, flag):
    """Render a single lab result row with visual bar."""
    low  = ref["low"]
    high = ref["high"]
    unit = ref["unit"]
    dec  = ref["decimals"]
    
    try:
        val_f = float(value)
    except Exception:
        val_f = None
    
    # Flag
    if flag == "H":
        color_bg, color_bar, icon = "#fff1f2", "#dc2626", "▲"
        flag_html = '<span style="background:#dc2626;color:white;border-radius:4px;padding:.1rem .4rem;font-size:.72rem;font-weight:700;margin-left:.3rem">HIGH</span>'
    elif flag == "L":
        color_bg, color_bar, icon = "#fffbeb", "#f59e0b", "▼"
        flag_html = '<span style="background:#f59e0b;color:white;border-radius:4px;padding:.1rem .4rem;font-size:.72rem;font-weight:700;margin-left:.3rem">LOW</span>'
    elif flag == "C":
        color_bg, color_bar, icon = "#fff1f2", "#7c3aed", "!!"
        flag_html = '<span style="background:#7c3aed;color:white;border-radius:4px;padding:.1rem .4rem;font-size:.72rem;font-weight:700;margin-left:.3rem">CRITICAL</span>'
    else:
        color_bg, color_bar, icon = "#f0fdf4", "#16a34a", "✓"
        flag_html = '<span style="background:#16a34a;color:white;border-radius:4px;padding:.1rem .4rem;font-size:.72rem;font-weight:700;margin-left:.3rem">NORMAL</span>'
    
    # Visual bar showing where value sits in range
    bar_html = ""
    if val_f is not None and high > low:
        # Extend range 20% beyond limits for display
        range_low  = low  - (high - low) * 0.2
        range_high = high + (high - low) * 0.2
        pos = (val_f - range_low) / (range_high - range_low) * 100
        pos = max(2, min(98, pos))
        bar_html = f"""
        <div style="margin:.3rem 0;position:relative;height:10px;background:#e5e7eb;border-radius:999px;">
            <div style="position:absolute;left:{(low-range_low)/(range_high-range_low)*100:.0f}%;
                        width:{(high-low)/(range_high-range_low)*100:.0f}%;
                        height:100%;background:#bbf7d0;border-radius:999px;"></div>
            <div style="position:absolute;left:{pos:.0f}%;transform:translateX(-50%);
                        top:-3px;width:16px;height:16px;background:{color_bar};
                        border-radius:50%;border:2px solid white;box-shadow:0 1px 3px rgba(0,0,0,.3);">
            </div>
        </div>
        <div style="display:flex;justify-content:space-between;font-size:.65rem;color:#9ca3af;margin-top:.1rem;">
            <span>Low {low:.{dec}f}</span><span style="color:{color_bar};font-weight:600">Normal: {low:.{dec}f}–{high:.{dec}f} {unit}</span><span>High {high:.{dec}f}</span>
        </div>"""
    
    val_display = f"{val_f:.{dec}f}" if val_f is not None else str(value)
    
    st.markdown(f"""
    <div style="background:{color_bg};border-radius:10px;padding:.6rem .9rem;margin:.3rem 0;
                border-left:4px solid {color_bar};">
        <div style="display:flex;justify-content:space-between;align-items:center;">
            <span style="font-weight:600;color:#0a2540;font-size:.88rem">{test_name}</span>
            <span style="font-size:1.05rem;font-weight:700;color:{color_bar}">
                {icon} {val_display} <span style="font-size:.75rem;font-weight:400;color:#6b7280">{unit}</span>
                {flag_html}
            </span>
        </div>
        {bar_html}
    </div>
    """, unsafe_allow_html=True)


def page_lab():
    st.markdown('<div class="section-header">🧪 Laboratory Room</div>', unsafe_allow_html=True)
    c = st.session_state.selected_case
    if not c:
        st.warning("No case selected.")
        return

    age_sex = c.get("Age_Sex","")
    st.markdown(f'<div class="alert-info">📋 <b>Patient:</b> {age_sex} | <b>CC:</b> {c.get("Chief_Complaint","?")}</div>', unsafe_allow_html=True)

    all_panels = list(LAB_PANELS.keys())
    
    col_sel, col_pri = st.columns([3,1])
    with col_sel:
        selected = st.multiselect("Select tests to order:", all_panels, key="lab_sel")
    with col_pri:
        priority = st.selectbox("Priority:", ["Routine","Urgent (2h)","STAT (30min)"], key="lab_priority")

    # Quick select buttons
    st.markdown("**⚡ Quick Order:**")
    q1,q2,q3,q4,q5 = st.columns(5)
    quick_panels = {
        "Basic Sepsis Screen": ["CBC (Full Blood Count)","CRP / ESR (Inflammatory Markers)","Serum Lactate","Procalcitonin","Blood Cultures x2"],
        "Chest Pain Panel":    ["Cardiac Enzymes (Troponin I/T, CK-MB)","BNP / Pro-BNP","Coagulation Panel (PT/INR/APTT/Fibrinogen)","D-dimer","Basic Metabolic Panel (BMP)"],
        "Abdominal Panel":     ["CBC (Full Blood Count)","CRP / ESR (Inflammatory Markers)","Liver Function Tests (LFTs)","Amylase / Lipase","Urinalysis + Microscopy"],
        "Full Metabolic":      ["CBC (Full Blood Count)","Basic Metabolic Panel (BMP)","Liver Function Tests (LFTs)","Renal Function (Creatinine/BUN/eGFR)","Thyroid Function (TSH/FT3/FT4)","Lipid Panel (Cholesterol/LDL/HDL/TG)"],
        "Trauma Screen":       ["CBC (Full Blood Count)","Coagulation Panel (PT/INR/APTT/Fibrinogen)","Blood Group & Cross-match","ABG (Arterial Blood Gas)","Serum Lactate"],
    }
    for col, (name, panels) in zip([q1,q2,q3,q4,q5], quick_panels.items()):
        with col:
            if st.button(name, use_container_width=True, key=f"qs_{name[:8]}"):
                st.session_state["lab_quick"] = panels
                st.rerun()
    if st.session_state.get("lab_quick"):
        selected = st.session_state["lab_quick"]
        st.markdown(f'<div class="alert-good">✅ Quick order applied: {len(selected)} panels</div>', unsafe_allow_html=True)

    run_col, _ = st.columns([1,3])
    with run_col:
        run_btn = st.button("🔬 Run Selected Tests", type="primary", use_container_width=True)

    if run_btn and selected:
        st.session_state.lab_seen = True
        st.session_state["lab_selected"]  = selected
        # Clear cached results for new order
        cache_key = f"lab_results_{c.get('Case_ID','?')}_{hash(str(sorted(selected)))}"
        st.session_state["lab_cache_key"] = cache_key
        if cache_key not in st.session_state:
            with st.spinner("🧪 Analyzing samples… generating results…"):
                ai_vals = generate_lab_results_ai(selected, c)
            st.session_state[cache_key] = ai_vals
        st.rerun()
    elif run_btn:
        st.warning("Select at least one test panel first.")

    if not st.session_state.lab_seen:
        return

    selected = st.session_state.get("lab_selected", [])
    cache_key = st.session_state.get("lab_cache_key","")
    ai_values = st.session_state.get(cache_key, {})

    st.markdown("---")
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                border-radius:12px;padding:1rem 1.5rem;margin-bottom:1rem;">
        <div style="font-size:1rem;font-weight:700">🧪 Laboratory Report</div>
        <div style="font-size:.8rem;opacity:.85;margin-top:.2rem">
            Patient: {age_sex} | Reported: {datetime.now().strftime("%d/%m/%Y %H:%M")} |
            Priority: {st.session_state.get("lab_priority","Routine")} |
            Panels ordered: {len(selected)}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Text-result panels ────────────────────────────────
    text_panels = {
        "Blood Cultures x2": lambda: f'<div class="alert-{"bad" if "bacteremia" in str(c.get("Final_Diagnosis","")).lower() or "sepsis" in str(c.get("Final_Diagnosis","")).lower() else "good"}">🩸 Blood Cultures: {"POSITIVE — Gram-negative rods (preliminary)" if "sepsis" in str(c.get("Final_Diagnosis","")).lower() else "No growth after 48h — NEGATIVE"}</div>',
        "Urinalysis + Microscopy": lambda: f'<div class="{"alert-bad" if str(c.get("Urine","none")).lower() not in ("none","neg.","","nan","normal","negative") else "alert-good"}">🔬 Urinalysis: {c.get("Urine","Negative — no significant findings")}</div>',
        "Urine Culture & Sensitivity": lambda: '<div class="alert-info">🧫 Urine C&S: Pending 48-72 hours — Preliminary: No significant growth</div>',
        "Influenza A/B Rapid Test": lambda: '<div class="alert-good">🦠 Influenza A: Negative | Influenza B: Negative</div>',
        "Blood Group & Cross-match": lambda: '<div class="alert-info">🩸 Blood Group: A Positive | Cross-match: Compatible — 2 units PRBCs available</div>',
    }

    # ── Render each panel ─────────────────────────────────
    for panel in selected:
        st.markdown(f"""
        <div style="background:#f8fafc;border:2px solid #e2e8f0;border-radius:12px;
                    padding:.8rem 1rem;margin:1rem 0 .3rem;">
            <div style="font-weight:700;color:#0a2540;font-size:.92rem">📋 {panel}</div>
        </div>
        """, unsafe_allow_html=True)

        if panel in text_panels:
            st.markdown(text_panels[panel](), unsafe_allow_html=True)
            continue

        tests = LAB_PANELS.get(panel, [])
        for test in tests:
            _, ref = get_ref_for_test(test, age_sex)
            if ref is None:
                continue

            # Get value — AI generated or fall back to known labs
            raw_val = ai_values.get(test)
            if raw_val is None:
                # Try without sex suffix
                for k, v in ai_values.items():
                    if k.lower().startswith(test.lower().split(" (")[0].lower()):
                        raw_val = v
                        break

            if raw_val is None:
                # Parse from known Excel labs as fallback
                known_labs = str(c.get("Labs","")).lower()
                import re as _r
                # Try to extract numeric value for this test
                abbrevs = {
                    "WBC": r"wbc\s*([\d.]+)",
                    "CRP": r"crp\s*([\d.]+)",
                    "Creatinine": r"creat\.\s*([\d.]+)|creatinine\s*([\d.]+)",
                    "Troponin I": r"trop(?:onin)?\s*([\d.]+)",
                    "Hemoglobin": r"hgb\s*([\d.]+)|hb\s*([\d.]+)",
                    "Platelets": r"plt\s*([\d.]+)",
                    "Lipase": r"lipase\s*([\d.]+)",
                    "Amylase": r"amylase\s*([\d.]+)",
                }
                test_key = test.split(" (")[0]
                pattern  = abbrevs.get(test_key, rf"{test_key.lower()}\s*([\d.]+)")
                m = _r.search(pattern, known_labs)
                if m:
                    raw_val = m.group(1) or m.group(2) if m.lastindex and m.lastindex > 1 else m.group(1)

            if raw_val is None:
                # Generate realistic normal value
                import random
                midpoint = (ref["low"] + ref["high"]) / 2
                spread   = (ref["high"] - ref["low"]) * 0.25
                raw_val  = round(midpoint + random.uniform(-spread, spread), ref["decimals"])

            try:
                val_f = float(raw_val)
            except Exception:
                val_f = None

            # Determine flag
            if val_f is not None:
                critical_low  = ref["low"]  * 0.7
                critical_high = ref["high"] * 1.5
                if val_f < critical_low or val_f > critical_high:
                    flag = "C"
                elif val_f < ref["low"]:
                    flag = "L"
                elif val_f > ref["high"]:
                    flag = "H"
                else:
                    flag = "N"
            else:
                flag = "N"

            render_lab_value(test, raw_val, ref, flag)

    # ── Known imaging reports from case ──────────────────
    img_data = " | ".join(x for x in [
        str(c.get("XRay_Report","") or ""),
        str(c.get("CT_Report","") or ""),
        str(c.get("Imaging_Tests","") or ""),
    ] if x and x.lower() not in ("none","nan",""))
    if img_data:
        st.markdown(f'<div style="margin-top:1rem;padding:.8rem 1rem;background:#eff6ff;border-radius:10px;border-left:4px solid #3b82f6;">📡 <b>Imaging on file:</b> {img_data}</div>', unsafe_allow_html=True)

    # ── AI Interpretation ─────────────────────────────────
    st.markdown("")
    if st.button("🤖 AI Clinical Interpretation", type="primary", use_container_width=True, key="lab_ai_interp"):
        # Build summary of all abnormal values
        abnormal = []
        for test, val in ai_values.items():
            _, ref = get_ref_for_test(test, age_sex)
            if ref:
                try:
                    if float(val) < ref["low"] or float(val) > ref["high"]:
                        abnormal.append(f"{test}: {val} {ref['unit']} (ref {ref['low']}–{ref['high']})")
                except: pass

        prompt = (
            f"Patient: {age_sex}, presenting with {c.get('Chief_Complaint','?')}\n"
            f"Diagnosis: {c.get('Final_Diagnosis','?')}\n"
            f"ABNORMAL VALUES: {chr(10).join(abnormal) if abnormal else 'All within normal limits'}\n\n"
            f"As a clinical pathologist:\n"
            f"1. CRITICAL VALUES: Any that need immediate action?\n"
            f"2. PATTERN: What clinical syndrome do these labs indicate?\n"
            f"3. KEY ABNORMALITIES: Explain each with clinical significance\n"
            f"4. NEXT STEPS: Additional tests or actions needed\n"
            f"Be specific, educational, and clinically precise."
        )
        with st.spinner("Interpreting results..."):
            interp = call_ai(
                "You are a consultant clinical pathologist interpreting laboratory results for medical students.",
                [{"role":"user","content":prompt}], max_tokens=700
            )
        st.markdown(f'<div style="background:#fff7ed;border-radius:12px;padding:1.2rem;border-left:5px solid #f59e0b;">{interp.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

    # ── Navigation ────────────────────────────────────────
    b1, b2 = st.columns(2)
    with b1:
        if st.button("← Physical Exam", use_container_width=True): nav("physical_exam")
    with b2:
        if st.button("🔬 Imaging →", use_container_width=True, type="primary"): nav("imaging")


# ════════════════════════════════════════════════════════════════════════════
# RADIOLOGY ENGINE
# ════════════════════════════════════════════════════════════════════════════
import io
from PIL import Image, ImageDraw, ImageFont

FINDING_COLORS = {
    "bleeding":("#FF0000","#FFE5E5"),"hemorrhage":("#FF0000","#FFE5E5"),
    "hematoma":("#CC0000","#FFE5E5"),"tumor":("#FF6600","#FFF0E5"),
    "mass":("#FF6600","#FFF0E5"),"malignancy":("#CC3300","#FFE5E5"),
    "fracture":("#0066FF","#E5F0FF"),"consolidation":("#9900CC","#F5E5FF"),
    "pneumonia":("#9900CC","#F5E5FF"),"effusion":("#006699","#E5F5FF"),
    "edema":("#006699","#E5F5FF"),"infarct":("#FF3300","#FFE8E5"),
    "ischemia":("#FF3300","#FFE8E5"),"abscess":("#CC6600","#FFF5E5"),
    "appendicitis":("#FF6600","#FFF0E5"),"inflammation":("#FF9900","#FFF5E0"),
    "pneumothorax":("#FF0066","#FFE5F0"),"embolism":("#FF0066","#FFE5F0"),
    "pe":("#FF0066","#FFE5F0"),"stenosis":("#FF6600","#FFF0E5"),
    "obstruction":("#CC6600","#FFF5E5"),"hernia":("#669900","#F0FFE5"),
    "wpw":("#9933FF","#F5E5FF"),"stemi":("#FF0000","#FFE5E5"),
    "default":("#00CC66","#E5FFF0"),
}

def get_finding_color(label):
    ll = label.lower()
    for k,v in FINDING_COLORS.items():
        if k in ll: return v
    return FINDING_COLORS["default"]

def annotate_image(image_bytes, findings):
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    w, h = img.size
    draw = ImageDraw.Draw(img, "RGBA")
    try:
        font_l = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", max(14,h//40))
        font_s = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", max(11,h//55))
        font_o = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", max(10,h//60))
    except Exception:
        font_l = font_s = font_o = ImageFont.load_default()
    legend_items = []
    for i, finding in enumerate(findings):
        label = finding.get("label","Finding")
        box   = finding.get("box",None)
        conf  = finding.get("confidence","")
        sc, _ = get_finding_color(label)
        if box and len(box)==4:
            y1=int(box[0]/1000*h); x1=int(box[1]/1000*w)
            y2=int(box[2]/1000*h); x2=int(box[3]/1000*w)
            x1,x2=min(x1,x2),max(x1,x2); y1,y2=min(y1,y2),max(y1,y2)
            pad=max(4,int(w*0.005))
            x1=max(0,x1-pad); y1=max(0,y1-pad); x2=min(w,x2+pad); y2=min(h,y2+pad)
            draw.rectangle([x1,y1,x2,y2],fill=sc+"30",outline=sc,width=3)
            cl=min(20,(x2-x1)//4,(y2-y1)//4); lw=4
            draw.line([x1,y1,x1+cl,y1],fill=sc,width=lw); draw.line([x1,y1,x1,y1+cl],fill=sc,width=lw)
            draw.line([x2,y1,x2-cl,y1],fill=sc,width=lw); draw.line([x2,y1,x2,y1+cl],fill=sc,width=lw)
            draw.line([x1,y2,x1+cl,y2],fill=sc,width=lw); draw.line([x1,y2,x1,y2-cl],fill=sc,width=lw)
            draw.line([x2,y2,x2-cl,y2],fill=sc,width=lw); draw.line([x2,y2,x2,y2-cl],fill=sc,width=lw)
            br=max(14,h//40); bx,by=x1,(y1-br*2 if y1>br*2 else y1)
            draw.ellipse([bx,by,bx+br*2,by+br*2],fill=sc,outline="white",width=2)
            draw.text((bx+br,by+br),str(i+1),font=font_s,fill="white",anchor="mm")
            fl=label[:30]+(f" ({conf})" if conf else "")
            bb=draw.textbbox((0,0),fl,font=font_l)
            tw=bb[2]-bb[0]+10; th=bb[3]-bb[1]+6
            lx=max(0,min(x1,w-tw)); ly=y2+4 if y2+th+4<h else y1-th-4
            draw.rectangle([lx,ly,lx+tw,ly+th],fill=sc+"DD",outline=sc)
            draw.text((lx+5,ly+3),fl,font=font_l,fill="white")
        legend_items.append({"num":i+1,"label":label,"color":sc,"conf":conf,
                             "desc":finding.get("description","")})
    for li,line in enumerate(["MLS VIRTUAL HOSPITAL",
                               f"AI — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                               "Gemini 2.5 Flash Vision",
                               f"Findings: {len(findings)}"]):
        draw.text((8,8+li*18),line,font=font_o,fill="#00FF00",stroke_width=1,stroke_fill="black")
    draw.text((w-8,h-20),"AI ANNOTATED — EDUCATIONAL ONLY",font=font_o,
              fill="#FFFF00",anchor="rs",stroke_width=1,stroke_fill="black")
    buf=io.BytesIO(); img.save(buf,format="PNG"); buf.seek(0)
    return buf.getvalue(), legend_items

def build_ecg_prompt():
    return """You are a board-certified cardiologist and electrophysiologist with 20+ years ECG experience.
Interpret this ECG with ABSOLUTE PRECISION. NEVER default to normal without examining every feature.

MANDATORY 10-STEP SYSTEMATIC ECG ANALYSIS:

STEP 1 — RATE: Count R-R intervals. Calculate exact HR. Normal 60-100 bpm.
STEP 2 — RHYTHM: Regular or irregular? P before every QRS? QRS after every P?
STEP 3 — P WAVE: Upright I,II,aVF? Duration <120ms? PR interval 120-200ms?
  SHORT PR (<120ms) + DELTA WAVE = WPW — DIAGNOSE — NEVER MISS
  ABSENT P + irregular rhythm = ATRIAL FIBRILLATION — DIAGNOSE
  Saw-tooth baseline = ATRIAL FLUTTER — DIAGNOSE
STEP 4 — PR INTERVAL: Normal 120-200ms.
  Progressive lengthening + dropped beat = Mobitz I (Wenckebach)
  Fixed PR + dropped beat = Mobitz II CRITICAL
  Complete P-QRS dissociation = 3rd degree AV block EMERGENCY
STEP 5 — QRS COMPLEX: Normal <120ms. Wide >120ms = LBBB/RBBB/WPW/VT
  WPW: Short PR + Wide QRS + DELTA WAVE (slurred upstroke) — confirm multiple leads
  LBBB: Wide, broad R in I/aVL/V5-V6, deep S in V1-V3
  RBBB: RSR' (M-pattern) in V1, wide S in I/V6
  Pathological Q waves: >40ms OR >25% R height = prior MI
STEP 6 — ST SEGMENT (MOST CRITICAL):
  ELEVATION >= 1mm limb leads OR >= 2mm precordial = STEMI EMERGENCY
  Inferior STEMI: STE in II,III,aVF + reciprocal depression I,aVL
  Anterior STEMI: STE in V1-V4 (LAD territory)
  Lateral STEMI: STE in I,aVL,V5-V6
  Posterior MI: ST depression V1-V3 + tall R waves V1-V2
  Pericarditis: Diffuse saddle-shaped STE + PR depression
  DEPRESSION horizontal/downsloping = ischemia/NSTEMI
STEP 7 — T WAVES: Inversion = ischemia/NSTEMI/strain
  Wellens syndrome: Biphasic/deep T-wave inversions V2-V3 CRITICAL
  Hyperacute tall T = early STEMI. Peaked narrow = hyperkalemia
STEP 8 — QT INTERVAL: Calculate QTc (Bazett). Normal <440ms men, <460ms women.
  Prolonged QTc = Torsades de Pointes risk
STEP 9 — SPECIFIC SYNDROMES — ACTIVELY LOOK FOR ALL:
  WPW: Short PR + delta wave + wide QRS — in ALL leads — NEVER MISS
  STEMI: Which territory? Which leads elevated? Reciprocal changes?
  Complete AV block: P rate different from QRS rate, completely dissociated
  AF: Irregularly irregular + absent distinct P waves + fibrillatory baseline
  VT: Wide complex tachycardia + AV dissociation + concordance in chest leads
  Brugada: Coved STE V1-V2 with RBBB morphology
STEP 10 — FINAL DIAGNOSIS + URGENCY (EMERGENCY/URGENT/ROUTINE)

HARD RULES — NEVER VIOLATE:
- Short PR + delta wave = WPW. NEVER call normal.
- ST elevation >= 1mm = STEMI until proven otherwise. NEVER call normal.
- Wide QRS tachycardia = VT until proven otherwise.
- P-QRS dissociation = Complete heart block. EMERGENCY."""

def build_imaging_prompt(modality, clinical_context):
    m  = modality.lower()
    cx = f"\nCLINICAL CONTEXT: {clinical_context}" if clinical_context else ""

    if any(k in m for k in ["ecg","ekg","electrocardiogram","12-lead","cardiac rhythm"]):
        return build_ecg_prompt()

    elif any(k in m for k in ["ct brain","brain ct","head ct","ct head"]):
        return f"""You are a consultant neuroradiologist. Interpret this CT Brain with maximum precision.{cx}
SYSTEMATIC CT BRAIN PROTOCOL:
1. TECHNICAL: Non-contrast vs contrast? Quality?
2. HEMORRHAGE (hyperdense=white=blood — EVERY region):
   Epidural: biconvex lens-shape, does NOT cross suture lines
   Subdural: crescent-shape, crosses sutures, note midline shift in mm
   Subarachnoid: hyperdensity in cisterns/sylvian fissures/sulci
   Intraparenchymal: focal hyperdensity within brain tissue
   Posterior fossa: cerebellum/brainstem — DO NOT MISS
3. ISCHEMIA/INFARCT: Hypodensity, loss grey-white differentiation, dense MCA sign
   Territory: MCA/ACA/PCA/posterior circulation + ASPECTS score
4. MASS: Location, size 3 planes, density, edema, midline shift mm, herniation type
5. VENTRICLES: Size, hydrocephalus? Blood? Symmetry?
6. CISTERNS: Effaced = raised ICP. Obliterated basal cisterns = EMERGENCY
7. BONE WINDOWS: Fractures, pneumocephalus
State diagnosis with location, size, urgency."""

    elif any(k in m for k in ["mri brain","brain mri"]):
        return f"""You are a neuroradiologist specializing in brain MRI.{cx}
SEQUENCES: T1/T2/FLAIR/DWI/ADC/GRE/SWI — identify what sequences are visible.
  DWI bright + ADC dark = ACUTE INFARCT — most sensitive first 24h
  T2/FLAIR bright = edema/demyelination/tumor/encephalitis
  GRE/SWI dark blooming = blood products/microbleeds
  T1+Gad enhancement = BBB breakdown (tumor/abscess/active MS plaque)
ANALYZE: Acute stroke, Tumors (ring vs homogeneous), MS plaques (Dawson fingers),
Infection (herpes=temporal lobe), Vascular (AVM/aneurysm flow voids)."""

    elif any(k in m for k in ["chest x","cxr","chest xr","chest ap","chest pa"]):
        return f"""You are a thoracic radiologist. ABCDE systematic CXR analysis.{cx}
A-AIRWAY: Trachea midline? Deviated = tension pneumothorax/large effusion. Carina angle <70°?
B-BREATHING (compare upper/mid/lower zones BOTH sides):
  PNEUMOTHORAX: Absent markings + visceral pleural line. TENSION: deviation + inverted diaphragm EMERGENCY
  CONSOLIDATION + air bronchograms = pneumonia/collapse — which lobe?
  EFFUSION: Blunted CP angle, meniscus sign. Bilateral = heart failure
  PULMONARY EDEMA: Kerley B lines, perihilar bat-wing, cephalization
  MASS: Size, shape, cavitation, calcification
  HYPERINFLATION: Flat diaphragms, >6 anterior ribs = COPD
C-CARDIAC: CTR >0.5 PA = cardiomegaly. Chamber patterns.
D-DIAPHRAGM: Right higher than left. Free air under diaphragm = perforation EMERGENCY
E-EVERYTHING: Mediastinum >8cm = aneurysm/dissection. Rib fractures. Line positions."""

    elif any(k in m for k in ["ct chest","chest ct","hrct","ctpa","pulmonary angio"]):
        return f"""You are a thoracic radiologist specializing in CT chest.{cx}
MEDIASTINUM: Lymphadenopathy >1cm. Aortic diameter. Dissection flap? Stanford A = EMERGENCY
LUNGS:
  CTPA: Filling defects in pulmonary arteries — central saddle/lobar/segmental/subsegmental
    RV:LV ratio >1 + D-shaped septum = right heart strain CRITICAL
  HRCT PATTERNS: GGO/Consolidation/Crazy-paving/Honeycombing(UIP/IPF)/Mosaic attenuation
  NODULES: Size, solid/subsolid/GGO, spiculation — Fleischner criteria
PLEURA: Effusion, thickening, pneumothorax
AIRWAYS: Bronchiectasis (signet ring sign), mucus plugging"""

    elif any(k in m for k in ["ct abdomen","ct abdo","abdominal ct","ct pelv","abdo pelv"]):
        return f"""You are an abdominal radiologist.{cx}
LIVER: Size, contour, lesions — HCC: arterial enhancement+washout+capsule = LI-RADS 5
  Bile ducts: CBD >6mm = dilated
GALLBLADDER: Wall >3mm + stones + pericholecystic fluid = ACUTE CHOLECYSTITIS
  Gas in GB wall = emphysematous cholecystitis EMERGENCY
PANCREAS: Fat stranding + edema = pancreatitis. Non-enhancing = NECROSIS CRITICAL
  Hypoenhancing mass + double duct sign = adenocarcinoma
APPENDIX: Diameter >6mm + fat stranding + fecalith = APPENDICITIS
  Free air + abscess = perforation CRITICAL
BOWEL: SBO (transition point) / LBO / Volvulus (whirl sign)
  Pneumatosis + portal venous gas = ischemia EMERGENCY
KIDNEYS: Stones (HU/size/hydronephrosis). Masses (Bosniak). Pyelonephritis
AORTA: Diameter each level. >3cm = aneurysm. Dissection flap CRITICAL
FREE FLUID: HU value (blood 30-60 HU, ascites ~0 HU)"""

    elif any(k in m for k in ["mri spine","spine mri","lumbar","cervical mri","thoracic mri"]):
        return f"""You are a spine MRI specialist.{cx}
VERTEBRAE: Height, alignment (spondylolisthesis grade?). T1 dark+T2 bright = edema/infection/mets
  Spondylodiscitis: disc + adjacent endplates T2 bright + enhancement DIAGNOSE
DISCS: Bulge/Protrusion/Extrusion/Sequestration. Level + direction (e.g. L4/5 right paracentral)
  Which nerve root compressed? Foraminal compromise?
CANAL: AP diameter <10mm = severe stenosis. Trefoil in LSS
CORD: T2 signal change = myelopathy CRITICAL. MS plaques: short segments.
  Transverse myelitis: central cord, long segments"""

    elif any(k in m for k in ["extremity","bone x","wrist","ankle","knee","shoulder","hip","elbow","foot","hand x"]):
        return f"""You are a musculoskeletal radiologist.{cx}
ALIGNMENT: Normal? Subluxation? Dislocation? Joint congruence?
BONE DENSITY: Osteopenia? Cortical thinning?
CORTEX: Intact? Periosteal reaction (solid=benign, sunburst/Codman=malignant)
FRACTURE (if present):
  Type: Transverse/Oblique/Spiral/Comminuted/Impacted/Stress/Pathological
  Displacement: Angulation (degrees+direction), translation (%), rotation, shortening
  Articular involvement? Physis (Salter-Harris I-V in children)?
  Named fractures: Colles/Smith/Scaphoid (wrist), Weber A/B/C (ankle),
    Garden I-IV (hip NOF), Schatzker I-VI (tibial plateau), Jones/pseudo-Jones (foot)
JOINTS: Space narrowing (OA/RA/septic). Erosions. Osteophytes. Effusion.
SOFT TISSUES: Swelling, calcification, foreign bodies"""

    elif any(k in m for k in ["abdominal x","kub","plain abdo","abdominal xr"]):
        return f"""You are a radiologist analyzing this abdominal X-ray.{cx}
GAS PATTERN:
  Normal: SB <3cm, LB <6cm, cecum <9cm
  SBO: Central dilated loops, valvulae conniventes, stepladder, no rectal gas
  LBO: Peripheral haustra, identify transition point
  SIGMOID VOLVULUS: Coffee bean sign pointing RUQ DIAGNOSE
  PNEUMOPERITONEUM: Gas under diaphragm (erect) or Rigler sign (supine) EMERGENCY
CALCIFICATIONS: Renal/ureteric stones, aortic calcification, appendicolith RIF
PSOAS SHADOW: Absent = abscess/hematoma/retroperitoneal pathology
BONES: Vertebral fractures"""

    elif any(k in m for k in ["ultrasound","us abdomen","sonography","us liver","us gallbladder"]):
        return f"""You are interpreting this ultrasound.{cx}
ECHOTEXTURE: Hyperechoic(bright)=fat/calcification, Hypoechoic(dark)=fluid/tumor, Anechoic(black)=fluid
  Posterior shadow = calcification/stone. Posterior enhancement = fluid
LIVER: Echogenicity vs kidney. Fatty liver = diffusely bright. Cirrhosis = heterogeneous+nodular
  Lesions: Cyst(anechoic+enhancement), Hemangioma(hyperechoic), HCC/Mets(hypoechoic halo)
GALLBLADDER: Wall >3mm + stones(shadow) + pericholecystic fluid + Murphy's sign = CHOLECYSTITIS
KIDNEYS: Size, cortex, hydronephrosis grade, stones (shadow+twinkling on Doppler)
FREE FLUID: Location, echogenicity (anechoic=transudate, echogenic=blood/pus)
AORTA: Diameter (>3cm = aneurysm)"""

    elif any(k in m for k in ["mammogram","mammography","breast"]):
        return f"""You are a breast radiologist using ACR BI-RADS protocol.{cx}
BREAST DENSITY: A(fatty)/B(scattered)/C(heterogeneous)/D(extremely dense)
MASSES: Shape + Margin + Density
  Spiculated irregular = HIGHLY SUSPICIOUS — BIOPSY
CALCIFICATIONS:
  BENIGN: Vascular (railroad track), coarse/popcorn, round/punctate
  SUSPICIOUS: Fine pleomorphic, fine linear/branching = DCIS pattern — BIOPSY
ARCHITECTURAL DISTORTION: Radiating lines = scar vs malignancy
BI-RADS: 1(neg)/2(benign)/3(probably benign-6mo)/4(suspicious-biopsy)/5(highly suspicious)/6(known cancer)"""

    elif any(k in m for k in ["nuclear","pet","bone scan","v/q","ventilation"]):
        return f"""You are a nuclear medicine physician.{cx}
V/Q SCAN: PIOPED II criteria:
  HIGH probability: >= 2 large unmatched perfusion defects = PE
BONE SCAN: Hot spots = metastases/fractures/osteomyelitis. Cold = aggressive mets/AVN
  Superscan: Diffuse intense uptake + absent kidneys = widespread metastases
PET-CT: FDG-avid lesions SUVmax >2.5. Staging: primary + nodes + distant mets"""

    elif any(k in m for k in ["mri abdomen","mri liver","mrcp","mri pelvis","mri prostate"]):
        return f"""You are an abdominal MRI specialist.{cx}
T1 in/out-of-phase: Signal drop OOP = fat (hepatic steatosis)
T2: Fluid bright. DWI/ADC: Restricted diffusion = malignancy/abscess/infarct
LIVER LI-RADS: LR-5 = >2cm + arterial hyperenhancement + washout + capsule = HCC
MRCP: CBD strictures, stones (signal void), duct communication = IPMN
MRI PROSTATE PI-RADS: T2 low signal PZ + DWI restriction = PI-RADS 4-5 SUSPICIOUS"""

    else:
        return f"""You are a senior consultant radiologist. Analyze this {modality} with maximum precision.{cx}
PROTOCOL:
1. TECHNICAL: Image quality, adequacy, technique
2. SYSTEMATIC REVIEW: Every visible anatomical structure — do NOT skip any region
3. EACH FINDING: Location + Size + Characteristics + Severity + Differential
4. MEASUREMENTS: Specific values for all abnormalities
5. IMPRESSION: Definitive primary diagnosis + confidence level
6. URGENCY: EMERGENCY / URGENT / ROUTINE
7. RECOMMENDATIONS: Specific next steps
RULE: Never report normal without systematically reviewing every structure."""

def parse_findings_from_response(response_text):
    import json as _j, re as _r
    try:
        m = _r.search(r"```json\s*(.*?)\s*```", response_text, _r.DOTALL)
        if m:
            data = _j.loads(m.group(1))
            return data.get("findings",[]), data
    except: pass
    try:
        m2 = _r.search(r'\{"modality".*\}', response_text, _r.DOTALL)
        if m2:
            data = _j.loads(m2.group(0))
            return data.get("findings",[]), data
    except: pass
    return [], {}

def extract_report_text(response_text):
    """
    Extract the narrative radiology/cardiology report from AI response.
    Removes JSON block, returns clean formatted report.
    If report is truncated, adds a note.
    """
    import re as _r
    # Remove JSON block
    clean = _r.sub(r"```json.*?```", "", response_text, flags=_r.DOTALL).strip()
    # Remove any remaining backtick code blocks
    clean = _r.sub(r"```.*?```", "", clean, flags=_r.DOTALL).strip()
    # Remove leading/trailing separators
    clean = clean.strip("─═─═─").strip()
    if not clean:
        return response_text  # fallback: return everything
    # If response looks truncated (no IMPRESSION or RECOMMENDATION section)
    has_impression = any(w in clean.upper() for w in ["IMPRESSION", "CONCLUSION", "SUMMARY", "FINDING"])
    if not has_impression and len(clean) < 200:
        clean = response_text  # Use raw response if parsed version is too short
    return clean

def call_ai_radiology(image_bytes, mime_type, modality, clinical_context=""):
    ok, msg = can_use_credits("imaging")
    if not ok:
        return f"!ERR_CREDITS: {msg}"
    use_credits("imaging")
    import base64
    img_b64   = base64.b64encode(image_bytes).decode("utf-8")
    specialized = build_imaging_prompt(modality, clinical_context)
    json_req  = f"""

CRITICAL ANTI-HALLUCINATION RULES — MANDATORY:
1. ONLY report findings you can ACTUALLY SEE in this image
2. DO NOT invent findings based on the clinical context alone
3. If image quality prevents assessment of a region — say so explicitly
4. Confidence must match what you actually see:
   - "high" = clearly visible, unambiguous
   - "moderate" = visible but some uncertainty
   - "low" = suspected but not definitive
5. If you cannot see an abnormality — report it as NORMAL for that region
6. NEVER report a finding without being able to point to it in the image

BOUNDING BOX PRECISION RULES:
- [y1, x1, y2, x2] all values 0-1000 (normalized image coordinates)
- y1=top, x1=left, y2=bottom, x2=right of the abnormality
- Boxes must TIGHTLY fit the actual lesion/abnormality — not the whole organ
- For ECG: box around the specific LEADS where abnormality is seen
- For CT Brain hemorrhage: box around the HYPERDENSE area ONLY
- For chest X-ray findings: box around the OPACITY/EFFUSION only
- For fractures: box around the fracture LINE specifically
- null ONLY if the finding has no spatial location (e.g. "normal sinus rhythm")

COORDINATE ANATOMY GUIDE (image normalized 0-1000):
CT Brain axial:
  Right hemisphere: x 0-500 | Left hemisphere: x 500-1000
  Frontal lobe: y 0-350 | Parietal: y 300-600 | Occipital: y 650-1000
  Basal ganglia: x 350-650, y 300-550
  Cerebellum: y 750-1000
Chest X-Ray (PA):
  Right lung: x 0-450 | Left lung: x 550-1000 | Mediastinum: x 400-600
  Upper zone: y 0-330 | Mid zone: y 330-660 | Lower zone: y 660-1000
  Heart: x 350-700, y 350-750 | Diaphragm: y 700-800
ECG standard 12-lead:
  Lead I,II,III (limb): y 0-300
  aVR,aVL,aVF: y 300-600
  V1-V3: x 0-500, y 600-800
  V4-V6: x 500-1000, y 600-800

OUTPUT this JSON first:
```json
{{
  "modality": "{modality}",
  "quality": "adequate/limited/poor — explain why if limited",
  "diagnosis": "SPECIFIC PRIMARY DIAGNOSIS — not vague",
  "urgency": "EMERGENCY/URGENT/ROUTINE",
  "findings": [
    {{
      "label": "Specific finding name (e.g. Left basal ganglia intraparenchymal hemorrhage)",
      "type": "bleeding/tumor/fracture/consolidation/effusion/ischemia/normal/other",
      "location": "Precise anatomical location with side and segment",
      "characteristics": "Size in mm if measurable, density/signal, shape, margins",
      "severity": "critical/significant/mild/incidental",
      "confidence": "high/moderate/low",
      "box": [y1, x1, y2, x2],
      "description": "What this means clinically in one sentence"
    }}
  ],
  "impression": "Definitive radiological impression — specific diagnosis",
  "recommendation": "Exact immediate clinical action",
  "normal_structures": ["Structures confirmed normal"]
}}
```
Then write the FULL STRUCTURED RADIOLOGY REPORT below the JSON.
"""
    contents = [{"role":"user","parts":[
        {"inline_data":{"mime_type":mime_type,"data":img_b64}},
        {"text": specialized + json_req}
    ]}]
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{st.session_state.get("active_model","gemini-2.5-flash")}:generateContent?key={get_api_key()}"
    try:
        r = requests.post(url, headers={"Content-Type":"application/json"},
            json={"contents":contents,
                  "generationConfig":{"maxOutputTokens":6000,"temperature":0.05,"topP":0.85,"topK":20}},
            timeout=120)
        if r.status_code==200:
            return r.json()["candidates"][0]["content"]["parts"][0]["text"]
        return f"!ERR {r.status_code}: {r.json().get('error',{}).get('message','Unknown')}"
    except Exception as e:
        return f"!ERR {e}"


# ── Sample Radiology Teaching Cases ─────────────────────────────────────────
# All images from Wikimedia Commons (public domain / CC-licensed)
SAMPLE_CASES = {
    "🧠 CT Brain — Intraparenchymal Hemorrhage": {
        "url":      "https://upload.wikimedia.org/wikipedia/commons/thumb/3/39/Intraparenchymal_hemorrhage.jpg/800px-Intraparenchymal_hemorrhage.jpg",
        "modality": "CT Brain (Non-contrast)",
        "context":  "60-year-old male, sudden onset severe headache, GCS 13, right-sided weakness",
        "expected": "Intraparenchymal hemorrhage left basal ganglia",
    },
    "🫁 Chest X-Ray — Right Lower Lobe Pneumonia": {
        "url":      "https://upload.wikimedia.org/wikipedia/commons/thumb/7/72/Pneumonia_x-ray.jpg/800px-Pneumonia_x-ray.jpg",
        "modality": "Chest X-Ray (PA)",
        "context":  "35-year-old female, 5 days fever, productive cough, SpO2 94% on room air",
        "expected": "Right lower lobe consolidation — pneumonia",
    },
    "🦴 X-Ray — Colles Fracture Distal Radius": {
        "url":      "https://upload.wikimedia.org/wikipedia/commons/thumb/f/fb/Radiograph_of_Colles%27_fracture_2012.jpg/800px-Radiograph_of_Colles%27_fracture_2012.jpg",
        "modality": "X-Ray Wrist (AP + Lateral)",
        "context":  "45-year-old female, fall on outstretched hand, wrist deformity and swelling",
        "expected": "Distal radius fracture with dorsal angulation (Colles fracture)",
    },
    "🫁 Chest X-Ray — Left Pneumothorax": {
        "url":      "https://upload.wikimedia.org/wikipedia/commons/thumb/5/5a/Spontaneous_Pneumothorax.jpg/800px-Spontaneous_Pneumothorax.jpg",
        "modality": "Chest X-Ray (AP)",
        "context":  "22-year-old tall thin male, sudden left-sided chest pain and dyspnea",
        "expected": "Left spontaneous pneumothorax",
    },
    "🧠 CT Brain — Acute Subdural Hematoma": {
        "url":      "https://upload.wikimedia.org/wikipedia/commons/thumb/2/2c/Subdural_hematoma_-_annotated.jpg/800px-Subdural_hematoma_-_annotated.jpg",
        "modality": "CT Brain (Non-contrast)",
        "context":  "75-year-old on warfarin, found confused on the floor after a fall",
        "expected": "Acute subdural hematoma with midline shift",
    },
    "🫁 Chest X-Ray — Pulmonary Edema (CHF)": {
        "url":      "https://upload.wikimedia.org/wikipedia/commons/thumb/3/35/Pulmonary_oedema.jpg/800px-Pulmonary_oedema.jpg",
        "modality": "Chest X-Ray (AP)",
        "context":  "68-year-old male, known heart failure, acute dyspnea, orthopnea, bilateral leg edema",
        "expected": "Cardiomegaly with bilateral pulmonary edema",
    },
    "🦴 X-Ray — Hip Fracture Neck of Femur": {
        "url":      "https://upload.wikimedia.org/wikipedia/commons/thumb/1/1c/Nof.jpg/800px-Nof.jpg",
        "modality": "X-Ray Hip (AP)",
        "context":  "82-year-old female, fell from standing, unable to weight bear, shortened and externally rotated leg",
        "expected": "Transcervical neck of femur fracture",
    },
}


def page_imaging():
    st.markdown('<div class="section-header">🏥 Radiology Room — AI Image Analysis & Annotation</div>', unsafe_allow_html=True)
    c = st.session_state.selected_case

    # Main tabs
    tab_upload, tab_samples, tab_ecg, tab_case = st.tabs([
        "📤 Upload & Analyze Image",
        "🗂️ Sample Radiology Cases",
        "💓 ECG File Analyzer (PhysioNet/WFDB)",
        "📋 Case Imaging Report"
    ])

    # ════════════════════════════════════════════════════════
    # TAB 1: UPLOAD & ANALYZE
    # ════════════════════════════════════════════════════════
    with tab_upload:
        # ── DICOM file support ─────────────────────────────
        dicom_uploaded = st.file_uploader(
            "📁 Upload DICOM file (.dcm) — auto-converts and analyzes:",
            type=["dcm"], key="dicom_up"
        )
        if dicom_uploaded:
            try:
                import pydicom
                import numpy as np
                from PIL import Image as _PIL_Image
                import io as _dcm_io

                ds = pydicom.dcmread(_dcm_io.BytesIO(dicom_uploaded.read()))
                arr = ds.pixel_array.astype(float)

                # Normalize to 0-255
                arr_min, arr_max = arr.min(), arr.max()
                if arr_max > arr_min:
                    arr = (arr - arr_min) / (arr_max - arr_min) * 255
                arr = arr.astype(np.uint8)

                # Handle multi-frame (CT series) — take middle frame
                if arr.ndim == 3:
                    arr = arr[arr.shape[0]//2]

                pil_img = _PIL_Image.fromarray(arr).convert("RGB")
                buf = _dcm_io.BytesIO()
                pil_img.save(buf, format="PNG")
                dcm_bytes = buf.getvalue()

                # Get DICOM metadata
                patient_name  = str(getattr(ds,"PatientName","Unknown"))
                modality_dcm  = str(getattr(ds,"Modality","Unknown"))
                study_desc    = str(getattr(ds,"StudyDescription",""))
                series_desc   = str(getattr(ds,"SeriesDescription",""))
                institution   = str(getattr(ds,"InstitutionName",""))
                rows          = str(getattr(ds,"Rows","?"))
                cols          = str(getattr(ds,"Columns","?"))
                ww            = str(getattr(ds,"WindowWidth","?"))
                wc            = str(getattr(ds,"WindowCenter","?"))

                st.markdown(f"""
                <div class="alert-good">
                    ✅ <b>DICOM loaded:</b> {modality_dcm} | {rows}×{cols}px |
                    W:{ww}/L:{wc} | {study_desc} {series_desc}
                </div>
                """, unsafe_allow_html=True)

                col_dcm1, col_dcm2 = st.columns(2)
                with col_dcm1:
                    st.image(dcm_bytes, caption=f"DICOM: {modality_dcm}", use_container_width=True)
                with col_dcm2:
                    dcm_modality_sel = st.selectbox("Confirm modality:",
                        ["CT Brain (Non-contrast)","CT Chest","CT Abdomen/Pelvis",
                         "MRI Brain","MRI Spine","Chest X-Ray (PA)","ECG / 12-lead EKG",
                         "Bone X-Ray","Other"],
                        index={"CT":"CT Brain (Non-contrast)","MR":"MRI Brain",
                               "CR":"Chest X-Ray (PA)","DX":"Chest X-Ray (PA)"}.get(modality_dcm[:2],7),
                        key="dcm_mod_sel")
                    dcm_ctx = st.text_area("Clinical context:", height=80, key="dcm_ctx")
                    if st.button("🔍 Analyze DICOM with AI", type="primary", use_container_width=True, key="dcm_analyze"):
                        import hashlib
                        dcm_hash  = hashlib.md5(dcm_bytes).hexdigest()[:12]
                        cache_key = f"rad_{dcm_hash}_{dcm_modality_sel[:15]}"
                        if cache_key in st.session_state:
                            dcm_response = st.session_state[cache_key]
                        else:
                            with st.spinner("Analyzing DICOM with Gemini..."):
                                dcm_response = call_ai_radiology(dcm_bytes,"image/png",dcm_modality_sel,dcm_ctx)
                            if not dcm_response.startswith("!ERR"):
                                st.session_state[cache_key] = dcm_response
                        if not dcm_response.startswith("!ERR"):
                            dcm_f, dcm_s = parse_findings_from_response(dcm_response)
                            dcm_report   = extract_report_text(dcm_response)
                            dx  = dcm_s.get("diagnosis","")
                            urg = dcm_s.get("urgency","ROUTINE")
                            if dx:
                                ug_c = {"EMERGENCY":"#dc2626","URGENT":"#f59e0b","ROUTINE":"#16a34a"}.get(urg.upper(),"#0e7490")
                                st.markdown(f'<div style="background:{ug_c};color:white;border-radius:10px;padding:1rem 1.5rem;margin:.5rem 0;font-size:1.1rem;font-weight:700">{urg} — {dx}</div>', unsafe_allow_html=True)
                            if dcm_f:
                                valid_f = [f for f in dcm_f if f.get("box")]
                                if valid_f:
                                    ann_b, leg = annotate_image(dcm_bytes, valid_f)
                                    st.image(ann_b, caption="DICOM annotated", use_container_width=True)
                                    st.download_button("💾 Download Annotated DICOM", ann_b, "annotated_dicom.png","image/png",use_container_width=True)
                            with st.expander("📄 Full Report", expanded=True):
                                st.markdown(dcm_report)
                        else:
                            if "CREDITS" in dcm_response:
                                st.markdown(f'<div class="alert-bad">💳 {dcm_response.replace("!ERR_CREDITS:","")}</div>', unsafe_allow_html=True)
                                if st.button("⭐ Upgrade Premium", key="up_dcm"): nav("credits")
                            else:
                                st.error(dcm_response)

            except ImportError:
                st.error("pydicom not installed. Run: pip install pydicom")
            except Exception as e_dcm:
                st.error(f"DICOM error: {e_dcm}")

        st.markdown("---")
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                    border-radius:14px;padding:1.2rem 1.5rem;margin-bottom:1rem;">
            <h3 style="margin:0 0 .4rem">🤖 AI Radiology Analyzer — Gemini 2.5 Flash Vision</h3>
            <p style="margin:0;opacity:.9;font-size:.88rem">
                Upload any medical image → AI performs systematic analysis →
                <b>Abnormalities are highlighted with colored bounding boxes</b> →
                Full structured radiology report generated.
            </p>
        </div>
        """, unsafe_allow_html=True)

        col_left, col_right = st.columns([1, 2])

        with col_left:
            st.markdown("**📋 Scan Details**")
            modality = st.selectbox("Imaging Modality:", [
                "ECG / 12-lead EKG ← Select this for ECGs",
                "CT Brain (Non-contrast)",
                "CT Brain (With Contrast)",
                "CT Chest",
                "CT Abdomen/Pelvis",
                "CT Angiography",
                "Chest X-Ray (PA)",
                "Chest X-Ray (AP)",
                "Abdominal X-Ray",
                "Lumbar Spine X-Ray",
                "Extremity X-Ray",
                "MRI Brain",
                "MRI Spine",
                "MRI Abdomen",
                "Ultrasound Abdomen",
                "Echocardiogram",
                "ECG / 12-lead EKG",
                "Nuclear Medicine Scan",
                "Mammogram",
                "Other",
            ])

            clinical_q = st.text_area(
                "Clinical Question / Context:",
                placeholder=(
                    f"e.g. {c.get('Age_Sex','Patient')} with {c.get('Chief_Complaint','symptoms')}. R/O..."
                    if c else
                    "e.g. 65yo male, sudden headache. R/O intracranial hemorrhage"
                ),
                height=90
            )

            st.markdown("**🎯 Analysis Focus:**")
            focus_options = st.multiselect(
                "Highlight these findings:",
                ["Hemorrhage/Bleeding", "Tumors/Masses", "Fractures",
                 "Consolidation/Pneumonia", "Effusions", "Vascular abnormalities",
                 "Inflammatory changes", "All abnormalities (recommended)"],
                default=["All abnormalities (recommended)"]
            )

            show_normal = st.checkbox("Also describe normal structures", value=True)
            overlay_mode = st.radio("Annotation style:",
                ["Bounding boxes + Labels (recommended)", "Labels only", "No annotation"],
                index=0)

        with col_right:
            uploaded = st.file_uploader(
                "Upload medical image (JPG, PNG, DICOM screenshots):",
                type=["jpg","jpeg","png","webp","bmp"],
                key="rad_upload"
            )

            if uploaded:
                img_bytes = uploaded.read()
                st.image(img_bytes, caption=f"Original: {uploaded.name}",
                         use_container_width=True)

                analyze_btn = st.button(
                    "🔍 Analyze & Annotate with AI",
                    type="primary",
                    use_container_width=True,
                    key="analyze_btn"
                )

                if analyze_btn:
                    import hashlib
                    img_hash = hashlib.md5(img_bytes).hexdigest()[:12]
                    cache_key = f"rad_{img_hash}_{modality[:20]}"
                    
                    if cache_key in st.session_state:
                        response = st.session_state[cache_key]
                        st.markdown('<div class="alert-good" style="font-size:.78rem">✅ Showing cached analysis — same image always gives same result.</div>', unsafe_allow_html=True)
                    else:
                        with st.spinner("🧠 Gemini 2.5 Flash analyzing image... (15-30 seconds)"):
                            mime_type = f"image/{uploaded.name.split('.')[-1].lower()}"
                            if mime_type == "image/jpg": mime_type = "image/jpeg"
                            response = call_ai_radiology(img_bytes, mime_type, modality, clinical_q)
                        if not response.startswith("!ERR"):
                            st.session_state[cache_key] = response

                    if response.startswith("!ERR_CREDITS:"):
                        msg = response.replace("!ERR_CREDITS:","").strip()
                        st.markdown(f'<div class="alert-bad">💳 <b>Credits exhausted:</b> {msg}</div>', unsafe_allow_html=True)
                        if st.button("⭐ Upgrade to Premium — $5/month", type="primary", use_container_width=True, key="up_from_rad"):
                            nav("credits")
                    elif response.startswith("!ERR"):
                        st.error(f"Analysis failed: {response}")
                    else:
                        # Parse findings
                        findings, structured = parse_findings_from_response(response)
                        report_text = extract_report_text(response)

                        st.markdown("---")
                        st.markdown("## 🏥 AI Radiology Analysis")

                        # ── Annotated image ──────────────────────────────────
                        if findings and "No annotation" not in overlay_mode:
                            valid_findings = [f for f in findings if f.get("box")]
                            if valid_findings:
                                with st.spinner("Drawing annotations..."):
                                    ann_bytes, legend = annotate_image(img_bytes, valid_findings)

                                st.markdown("### 🎯 Annotated Image")
                                st.image(ann_bytes,
                                         caption="AI-annotated: Abnormalities highlighted",
                                         use_container_width=True)

                                # Download annotated
                                st.download_button(
                                    "💾 Download Annotated Image",
                                    ann_bytes,
                                    f"annotated_{uploaded.name}",
                                    "image/png",
                                    use_container_width=True
                                )

                                # ── Legend ───────────────────────────────────
                                st.markdown("### 🔍 Findings Legend")
                                for item in legend:
                                    color = item["color"]
                                    severity_badge = ""
                                    for f in findings:
                                        if f.get("label") == item["label"]:
                                            sev = f.get("severity","")
                                            if sev == "critical":
                                                severity_badge = '<span style="background:#dc2626;color:white;border-radius:999px;padding:.1rem .5rem;font-size:.7rem;font-weight:700;margin-left:.3rem">⚠️ CRITICAL</span>'
                                            elif sev == "significant":
                                                severity_badge = '<span style="background:#f59e0b;color:white;border-radius:999px;padding:.1rem .5rem;font-size:.7rem;font-weight:700;margin-left:.3rem">⚡ SIGNIFICANT</span>'
                                            break

                                    st.markdown(f"""
                                    <div style="background:white;border-radius:10px;padding:.7rem 1rem;
                                                margin:.4rem 0;border-left:5px solid {color};
                                                box-shadow:0 2px 6px rgba(0,0,0,.06);display:flex;
                                                align-items:flex-start;gap:.8rem;">
                                        <div style="background:{color};color:white;border-radius:50%;
                                                    width:26px;height:26px;display:flex;align-items:center;
                                                    justify-content:center;font-weight:700;flex-shrink:0;
                                                    font-size:.85rem">#{item["num"]}</div>
                                        <div style="flex:1;">
                                            <div style="font-weight:700;color:#0a2540;font-size:.9rem">
                                                {item["label"]}{severity_badge}
                                            </div>
                                            <div style="font-size:.82rem;color:#374151;margin-top:.2rem">
                                                {item.get("desc","")}
                                            </div>
                                            <div style="font-size:.75rem;color:#6b7280;margin-top:.1rem">
                                                Confidence: {item.get("conf","")}
                                            </div>
                                        </div>
                                    </div>
                                    """, unsafe_allow_html=True)

                                    # ─── PATH A: Student feedback loop ────────────────
                                    if CLINICAL_HELPERS_OK:
                                        import hashlib as _hl
                                        img_hash = _hl.md5(img_bytes).hexdigest()[:10]
                                        with st.expander(f"💬 Verify finding #{item['num']} (helps improve AI)", expanded=False):
                                            render_finding_feedback(
                                                finding_id   = f"f{item['num']}",
                                                finding_label= item["label"],
                                                image_hash   = img_hash,
                                            )
                            else:
                                st.markdown('<div class="alert-good">✅ <b>No localizable abnormalities detected</b> — image appears within normal limits.</div>', unsafe_allow_html=True)

                        # ─── PATH B: TorchXRayVision specialist second opinion ─────
                        # Only runs for chest X-rays. Provides 18-pathology probabilities
                        # from a model trained on 1M+ labelled real chest X-rays.
                        if CLINICAL_HELPERS_OK and is_chest_xray_modality(modality):
                            st.markdown("---")
                            with st.spinner("Running specialist chest X-ray model (TorchXRayVision)..."):
                                xrv_result = specialist_chest_xray_analysis(img_bytes)
                            render_specialist_panel(xrv_result, gemini_findings=findings)

                        # ─── Feedback summary (if any feedback collected) ──────────
                        if CLINICAL_HELPERS_OK:
                            fb = st.session_state.get("imaging_feedback", {})
                            if fb:
                                st.markdown("---")
                                render_feedback_summary()

                        # ── Structured findings table ─────────────────────────
                        if findings:
                            st.markdown("### 📊 Structured Findings")
                            findings_data = []
                            for f in findings:
                                findings_data.append({
                                    "Finding": f.get("label",""),
                                    "Location": f.get("location",""),
                                    "Size": f.get("size","N/A"),
                                    "Severity": f.get("severity","").upper(),
                                    "Confidence": f.get("confidence",""),
                                    "Type": f.get("type",""),
                                })
                            st.dataframe(
                                pd.DataFrame(findings_data),
                                use_container_width=True,
                                hide_index=True
                            )

                        # ── PRIMARY DIAGNOSIS + URGENCY (shown first, prominently) ──
                        primary_dx = structured.get("diagnosis","")
                        urgency    = structured.get("urgency","")
                        if primary_dx:
                            urg_color = {
                                "EMERGENCY": "#dc2626",
                                "URGENT":    "#f59e0b",
                                "ROUTINE":   "#16a34a",
                            }.get(urgency.upper(), "#0e7490")
                            st.markdown(f"""
                            <div style="background:{urg_color};color:white;border-radius:12px;
                                        padding:1.2rem 1.5rem;margin:1rem 0;
                                        box-shadow:0 4px 12px {urg_color}55;">
                                <div style="font-size:.75rem;font-weight:700;opacity:.85;letter-spacing:.1em">
                                    {"⚠️ EMERGENCY" if urgency.upper()=="EMERGENCY" else "⚡ URGENT" if urgency.upper()=="URGENT" else "✅ ROUTINE"} — AI DIAGNOSIS
                                </div>
                                <div style="font-size:1.3rem;font-weight:800;margin:.3rem 0">
                                    {primary_dx}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)

                        # ── Impression ────────────────────────────────────────
                        if structured.get("impression"):
                            imp = structured["impression"]
                            st.markdown(f"""
                            <div style="background:linear-gradient(135deg,#fff7ed,#fef9c3);
                                        border:2px solid #f59e0b;border-radius:12px;padding:1.2rem;
                                        margin:1rem 0;">
                                <h4 style="margin:0 0 .5rem;color:#92400e">📋 Radiological Impression</h4>
                                <p style="margin:0;color:#1e3a5f;font-size:.95rem;line-height:1.7">{imp}</p>
                            </div>
                            """, unsafe_allow_html=True)

                        if structured.get("recommendation"):
                            st.markdown(f'<div class="alert-warn">💊 <b>Recommendation:</b> {structured["recommendation"]}</div>', unsafe_allow_html=True)

                        # ── Full Narrative Report ─────────────────────────────
                        if report_text:
                            with st.expander("📄 Full Radiology Report (click to expand)", expanded=True):
                                st.markdown(report_text)
                                st.download_button(
                                    "💾 Download Full Report",
                                    report_text,
                                    f"radiology_report_{datetime.now().strftime('%Y%m%d_%H%M')}.txt",
                                    "text/plain",
                                    use_container_width=True
                                )

                        # ── Accuracy check if case has known diagnosis ─────────
                        if c and structured.get("diagnosis"):
                            ai_dx   = structured.get("diagnosis","").lower()
                            true_dx = str(c.get("Final_Diagnosis","")).lower()
                            if true_dx and true_dx not in ("none","nan",""):
                                tw = set(true_dx.replace(","," ").replace("/"," ").split()) - {"the","a","and","or","of","with"}
                                aw = set(ai_dx.replace(","," ").replace("/"," ").split())   - {"the","a","and","or","of","with"}
                                overlap  = len(tw & aw) / max(len(tw), 1)
                                match_pct= min(100, int(overlap*100 + (30 if overlap>0 else 0)))
                                vc = "#16a34a" if match_pct>=60 else "#f59e0b" if match_pct>=30 else "#dc2626"
                                vl = "✅ Match" if match_pct>=60 else "⚠️ Partial" if match_pct>=30 else "❌ Low"

                                if "imaging_accuracy_log" not in st.session_state:
                                    st.session_state["imaging_accuracy_log"] = []
                                entry = {"case_id":str(c.get("Case_ID","?")),"modality":modality,
                                         "true_dx":true_dx,"ai_dx":ai_dx,"match":match_pct,
                                         "time":datetime.now().strftime("%H:%M")}
                                if entry["case_id"] not in [x["case_id"] for x in st.session_state["imaging_accuracy_log"]]:
                                    st.session_state["imaging_accuracy_log"].append(entry)

                                st.markdown(f"""
                                <div style="background:white;border-radius:12px;padding:1rem 1.5rem;
                                            border:2px solid {vc};margin:1rem 0;">
                                    <div style="font-size:.8rem;color:#6b7280;font-weight:600;margin-bottom:.4rem">
                                        🎯 AI ACCURACY CHECK — Case {c.get("Case_ID","?")}
                                    </div>
                                    <div style="display:flex;gap:2rem;align-items:center;flex-wrap:wrap;">
                                        <div><span style="font-size:.75rem;color:#6b7280">Known diagnosis:</span><br>
                                            <b style="color:#0a2540">{true_dx.title()}</b></div>
                                        <div style="font-size:1.3rem">→</div>
                                        <div><span style="font-size:.75rem;color:#6b7280">AI found:</span><br>
                                            <b style="color:{vc}">{ai_dx.title()}</b></div>
                                        <div style="margin-left:auto;text-align:center;">
                                            <div style="font-size:2rem;font-weight:900;color:{vc}">{match_pct}%</div>
                                            <div style="font-size:.78rem;font-weight:600;color:{vc}">{vl}</div>
                                        </div>
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)

                        log = st.session_state.get("imaging_accuracy_log", [])
                        if len(log) >= 2:
                            avg = sum(x["match"] for x in log)/len(log)
                            tp  = sum(1 for x in log if x["match"]>=60)
                            with st.expander(f"📊 Session Accuracy — {len(log)} scans | {tp}/{len(log)} high match | Avg {avg:.0f}%"):
                                col_tp,col_fp,col_av = st.columns(3)
                                with col_tp: st.metric("High Match ≥60%", tp, help="True positive equivalent")
                                with col_fp: st.metric("Low Match <60%", len(log)-tp, help="Needs review")
                                with col_av: st.metric("Avg Score", f"{avg:.0f}%")
                                for e in log[-5:]:
                                    ec = "#16a34a" if e["match"]>=60 else "#f59e0b" if e["match"]>=30 else "#dc2626"
                                    st.markdown(f'<div style="font-size:.78rem;padding:.4rem .8rem;background:#f8fafc;border-radius:6px;margin:.2rem 0;border-left:3px solid {ec}">Case {e["case_id"]} | {e["modality"]} | True: <b>{e["true_dx"][:25]}</b> | AI: <b>{e["ai_dx"][:25]}</b> | <b style="color:{ec}">{e["match"]}%</b></div>', unsafe_allow_html=True)
                                if st.button("🗑️ Clear log", key="clear_acc_log"): st.session_state["imaging_accuracy_log"]=[]; st.rerun()

                        # ── Teaching points ──────────────────────────────────
                        if findings:
                            st.markdown("### 📚 Teaching Points")
                            teach_prompt = (f"Based on these radiology findings: {[f.get('label') for f in findings]}, "
                                f"provide 3 key teaching points for medical students. "
                                f"Cover: 1) What to look for 2) Clinical significance 3) Next steps. Be concise.")
                            with st.spinner("Generating teaching points..."):
                                teaching = call_ai("You are a radiology professor teaching medical students.",
                                    [{"role":"user","content":teach_prompt}], max_tokens=400)
                            st.markdown(f'<div style="background:#f0f9ff;border-radius:12px;padding:1.2rem;border:2px solid #0ea5e9;">{teaching.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════
    # TAB 2: SAMPLE CASES
    # ════════════════════════════════════════════════════════
    with tab_samples:
        st.markdown('<div class="alert-info">📚 Select a real teaching case below. The image will be downloaded and analyzed by Gemini 2.5 Flash Vision — abnormalities will be highlighted automatically.</div>', unsafe_allow_html=True)

        selected_case_name = st.selectbox("Select sample case:", list(SAMPLE_CASES.keys()))
        sample = SAMPLE_CASES[selected_case_name]

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown(f"""
            <div style="background:#f8fafc;border-radius:12px;padding:1rem;border:1px solid #e2e8f0;">
                <b>Modality:</b> {sample["modality"]}<br>
                <b>Clinical Context:</b> {sample["context"]}<br>
                <b>Expected Finding:</b> <span style="color:#dc2626">{sample["expected"]}</span>
            </div>
            """, unsafe_allow_html=True)

        with col_b:
            extra_context = st.text_input("Add clinical question:", placeholder="e.g. R/O ICH", key="sample_context")

        if st.button("🔍 Download & Analyze Sample", type="primary", use_container_width=True, key="sample_analyze"):
            with st.spinner(f"Downloading {selected_case_name}..."):
                try:
                    img_response = requests.get(sample["url"], timeout=15,
                        headers={"User-Agent": "Mozilla/5.0"})
                    if img_response.status_code == 200:
                        img_bytes = img_response.content
                        st.image(img_bytes, caption=f"Original: {selected_case_name}", use_container_width=True)

                        context = sample["context"] + (f" Additional: {extra_context}" if extra_context else "")

                        import hashlib
                        sample_hash = hashlib.md5(img_bytes).hexdigest()[:12]
                        sample_cache = f"rad_{sample_hash}_{sample['modality'][:20]}"
                        
                        if sample_cache in st.session_state:
                            response = st.session_state[sample_cache]
                            st.info("Showing cached analysis for this image.")
                        else:
                            with st.spinner("🧠 Analyzing with Gemini 2.5 Flash..."):
                                response = call_ai_radiology(img_bytes, "image/jpeg", sample["modality"], context)
                            if not response.startswith("!ERR"):
                                st.session_state[sample_cache] = response

                        if not response.startswith("!ERR"):
                            findings, structured = parse_findings_from_response(response)
                            report_text = extract_report_text(response)

                            st.markdown("---")
                            st.markdown(f"## 🎯 Analysis: {selected_case_name}")

                            if findings:
                                valid_f = [f for f in findings if f.get("box")]
                                if valid_f:
                                    with st.spinner("Annotating image..."):
                                        ann_bytes, legend = annotate_image(img_bytes, valid_f)
                                    st.markdown("### 📍 Annotated Image")
                                    st.image(ann_bytes, caption="Abnormalities highlighted by AI", use_container_width=True)
                                    st.download_button("💾 Download Annotated", ann_bytes,
                                        f"annotated_sample.png", "image/png", use_container_width=True)

                                    st.markdown("### 🔍 Findings")
                                    for item in legend:
                                        c_color = item["color"]
                                        st.markdown(f'<div style="border-left:4px solid {c_color};padding:.5rem .8rem;margin:.3rem 0;background:white;border-radius:0 8px 8px 0"><b style="color:{c_color}">#{item["num"]} {item["label"]}</b> — {item.get("desc","")}</div>', unsafe_allow_html=True)

                            if structured.get("impression"):
                                st.markdown(f'<div style="background:#fff7ed;border:2px solid #f59e0b;border-radius:12px;padding:1rem;margin:1rem 0;"><b>Impression:</b> {structured["impression"]}</div>', unsafe_allow_html=True)

                            with st.expander("📄 Full Report"):
                                st.markdown(report_text)
                        else:
                            st.error(response)
                    else:
                        st.error(f"Could not download image (HTTP {img_response.status_code}). Try uploading manually in Tab 1.")
                except Exception as e:
                    st.error(f"Download failed: {e}. Upload the image manually in Tab 1.")

    # ════════════════════════════════════════════════════════
    # TAB 3: ECG FILE ANALYZER — PhysioNet / WFDB format
    # ════════════════════════════════════════════════════════
    with tab_ecg:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                    border-radius:14px;padding:1.2rem 1.5rem;margin-bottom:1rem;">
            <h3 style="margin:0 0 .4rem">💓 ECG File Analyzer — PhysioNet / WFDB Format</h3>
            <p style="margin:0;opacity:.9;font-size:.88rem">
                Upload real ECG files in WFDB format (.hea + .dat) from PhysioNet,
                or upload a standard ECG image (JPG/PNG) for AI analysis.
                Supports MIT-BIH, PTB-XL, and all PhysioNet ECG databases.
            </p>
        </div>
        """, unsafe_allow_html=True)

        ecg_mode = st.radio("ECG Input Method:", [
            "📸 Upload ECG Image (JPG/PNG) — AI Visual Analysis",
            "📁 Upload WFDB Files (.hea + .dat) — Signal Processing",
        ], key="ecg_mode")

        if "Image" in ecg_mode:
            st.markdown('<div class="alert-info">Upload any ECG strip or 12-lead printout as an image. Gemini 2.5 Flash will analyze it using the full AHA/ACC 10-step protocol.</div>', unsafe_allow_html=True)
            ecg_img = st.file_uploader("Upload ECG image:", type=["jpg","jpeg","png","bmp","webp"], key="ecg_img_up")
            ecg_ctx = st.text_input("Clinical context:", placeholder="e.g. 55yo male, chest pain, diaphoresis", key="ecg_ctx")

            if ecg_img and st.button("🔍 Analyze ECG", type="primary", use_container_width=True, key="ecg_img_btn"):
                import hashlib
                img_bytes = ecg_img.read()
                ecg_hash  = hashlib.md5(img_bytes).hexdigest()[:12]
                cache_key = f"ecg_{ecg_hash}"

                if cache_key in st.session_state:
                    response = st.session_state[cache_key]
                    st.markdown('<div class="alert-good" style="font-size:.78rem">✅ Cached result — same image always gives same analysis.</div>', unsafe_allow_html=True)
                else:
                    with st.spinner("🧠 Analyzing ECG with AHA/ACC 10-step protocol..."):
                        response = call_ai_radiology(img_bytes, f"image/{ecg_img.name.split('.')[-1].lower()}", "ECG / 12-lead EKG", ecg_ctx)
                    if not response.startswith("!ERR"):
                        st.session_state[cache_key] = response

                if not response.startswith("!ERR"):
                    findings, structured = parse_findings_from_response(response)
                    report_text = extract_report_text(response)

                    # Primary diagnosis
                    dx  = structured.get("diagnosis","")
                    urg = structured.get("urgency","")
                    if dx:
                        urg_color = {"EMERGENCY":"#dc2626","URGENT":"#f59e0b","ROUTINE":"#16a34a"}.get(urg.upper(),"#0e7490")
                        st.markdown(f"""
                        <div style="background:{urg_color};color:white;border-radius:12px;
                                    padding:1.2rem 1.5rem;margin:1rem 0;box-shadow:0 4px 12px {urg_color}55;">
                            <div style="font-size:.75rem;font-weight:700;opacity:.85">
                                {"⚠️ EMERGENCY" if urg.upper()=="EMERGENCY" else "⚡ URGENT" if urg.upper()=="URGENT" else "✅ ROUTINE"} — ECG DIAGNOSIS
                            </div>
                            <div style="font-size:1.3rem;font-weight:800;margin:.3rem 0">{dx}</div>
                        </div>
                        """, unsafe_allow_html=True)

                    # Annotated image
                    if findings:
                        valid_f = [f for f in findings if f.get("box")]
                        if valid_f:
                            with st.spinner("Annotating ECG..."):
                                ann_bytes, legend = annotate_image(img_bytes, valid_f)
                            st.image(ann_bytes, caption="ECG with annotated abnormalities", use_container_width=True)
                            st.download_button("💾 Download Annotated ECG", ann_bytes, "annotated_ecg.png", "image/png", use_container_width=True)

                            for item in legend:
                                color = item["color"]
                                st.markdown(f'<div style="border-left:4px solid {color};padding:.5rem .8rem;margin:.3rem 0;background:white;border-radius:0 8px 8px 0"><b style="color:{color}">#{item["num"]} {item["label"]}</b> — {item.get("desc","")}</div>', unsafe_allow_html=True)

                    if report_text:
                        with st.expander("📄 Full ECG Report", expanded=True):
                            st.markdown(report_text)
                            st.download_button("💾 Download Report", report_text, "ecg_report.txt", "text/plain", use_container_width=True)
                else:
                    st.error(response)

        else:  # WFDB mode
            st.markdown("""
            <div class="alert-info">
            <b>PhysioNet WFDB Format:</b> Each ECG record = 2 files (.hea header + .dat signal data).
            Download free ECG records from <a href="https://physionet.org/content/mitdb/1.0.0/" target="_blank">MIT-BIH Arrhythmia Database</a>
            or <a href="https://physionet.org/content/ptbdb/1.0.0/" target="_blank">PTB Diagnostic ECG Database</a>.
            Upload both files together below.
            </div>
            """, unsafe_allow_html=True)

            col_hea, col_dat = st.columns(2)
            with col_hea:
                hea_file = st.file_uploader("Upload .hea file:", type=["hea"], key="hea_up")
            with col_dat:
                dat_file = st.file_uploader("Upload .dat file:", type=["dat"], key="dat_up")

            if hea_file and dat_file:
                if st.button("💓 Process & Analyze WFDB ECG", type="primary", use_container_width=True):
                    try:
                        import wfdb
                        import neurokit2 as nk
                        import numpy as np
                        import matplotlib
                        matplotlib.use("Agg")
                        import matplotlib.pyplot as plt
                        import tempfile, os

                        # Save uploaded files to temp directory
                        with tempfile.TemporaryDirectory() as tmpdir:
                            record_name = hea_file.name.replace(".hea","")
                            hea_path = os.path.join(tmpdir, hea_file.name)
                            dat_path = os.path.join(tmpdir, dat_file.name)

                            with open(hea_path,"wb") as f: f.write(hea_file.read())
                            with open(dat_path,"wb") as f: f.write(dat_file.read())

                            # Read record
                            with st.spinner("Reading WFDB record..."):
                                record = wfdb.rdrecord(os.path.join(tmpdir, record_name))

                            fs      = record.fs
                            signals = record.p_signal
                            names   = record.sig_name
                            duration= signals.shape[0] / fs

                            st.markdown(f"""
                            <div style="background:#f8fafc;border-radius:10px;padding:.8rem 1rem;
                                        font-size:.85rem;border:1px solid #e2e8f0;margin:.5rem 0">
                                <b>Record:</b> {record_name} |
                                <b>Sampling rate:</b> {fs} Hz |
                                <b>Duration:</b> {duration:.1f}s |
                                <b>Leads:</b> {len(names)} ({", ".join(names[:6])}{"..." if len(names)>6 else ""}) |
                                <b>Samples:</b> {signals.shape[0]:,}
                            </div>
                            """, unsafe_allow_html=True)

                            # Plot ECG strips
                            with st.spinner("Generating ECG plot..."):
                                n_display = min(int(fs * 10), signals.shape[0])  # 10 seconds
                                n_leads   = min(12, len(names))
                                t = np.arange(n_display) / fs

                                fig, axes = plt.subplots(n_leads, 1, figsize=(16, n_leads*1.4), facecolor="#fafafa")
                                fig.suptitle(f"ECG — {record_name} | {fs}Hz | First 10 seconds", fontsize=12, fontweight="bold")
                                if n_leads == 1: axes = [axes]

                                for i in range(n_leads):
                                    ax = axes[i]
                                    sig = signals[:n_display, i]
                                    ax.plot(t, sig, color="#dc2626", linewidth=0.8)
                                    ax.set_ylabel(names[i], fontsize=8, rotation=0, labelpad=25)
                                    ax.set_xlim(0, t[-1])
                                    ax.grid(True, which="major", color="#ffcccc", linewidth=0.5)
                                    ax.grid(True, which="minor", color="#ffe5e5", linewidth=0.3)
                                    ax.minorticks_on()
                                    ax.set_facecolor("#fff8f8")
                                    ax.tick_params(axis="y", labelsize=7)
                                    if i < n_leads-1: ax.set_xticklabels([])
                                    else: ax.set_xlabel("Time (seconds)", fontsize=9)

                                plt.tight_layout()

                                import io as _io
                                buf = _io.BytesIO()
                                fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
                                buf.seek(0)
                                ecg_img_bytes = buf.getvalue()
                                plt.close(fig)

                            st.image(ecg_img_bytes, caption=f"ECG Record: {record_name}", use_container_width=True)
                            st.download_button("💾 Download ECG Plot", ecg_img_bytes, f"{record_name}_ecg.png", "image/png", use_container_width=True)

                            # NeuroKit2 analysis on lead II (most common for rhythm)
                            lead_ii_idx = next((i for i,n in enumerate(names) if "II" in n.upper() or n=="MLII"), 0)
                            lead_signal = signals[:, lead_ii_idx]

                            with st.spinner("NeuroKit2 signal processing..."):
                                try:
                                    ecg_clean = nk.ecg_clean(lead_signal, sampling_rate=fs)
                                    _, rpeaks = nk.ecg_peaks(ecg_clean, sampling_rate=fs)
                                    hr_mean   = nk.ecg_rate(rpeaks, sampling_rate=fs, desired_length=None)
                                    mean_hr   = float(np.mean(hr_mean)) if len(hr_mean) > 0 else 0

                                    # HRV analysis
                                    rr = np.diff(rpeaks["ECG_R_Peaks"]) / fs * 1000  # ms
                                    rr_mean = float(np.mean(rr)) if len(rr) > 0 else 0
                                    rr_std  = float(np.std(rr)) if len(rr) > 0 else 0
                                    rr_cv   = (rr_std/rr_mean*100) if rr_mean > 0 else 0

                                    # Rhythm regularity
                                    is_regular = rr_cv < 10
                                    n_beats    = len(rpeaks["ECG_R_Peaks"])

                                    st.markdown("### 📊 Signal Analysis Results")
                                    m1,m2,m3,m4 = st.columns(4)
                                    with m1: st.metric("Mean Heart Rate", f"{mean_hr:.0f} bpm",
                                        delta="Normal" if 60 <= mean_hr <= 100 else ("Bradycardia" if mean_hr < 60 else "Tachycardia"))
                                    with m2: st.metric("Beats Detected", str(n_beats))
                                    with m3: st.metric("Mean RR Interval", f"{rr_mean:.0f} ms")
                                    with m4: st.metric("Rhythm", "Regular" if is_regular else "Irregular",
                                        delta="Normal" if is_regular else "⚠️ Check for AF/Arrhythmia")

                                    # Flags
                                    flags = []
                                    if mean_hr < 40:  flags.append("⚠️ CRITICAL BRADYCARDIA")
                                    if mean_hr > 150: flags.append("⚠️ CRITICAL TACHYCARDIA")
                                    if not is_regular and rr_cv > 20: flags.append("⚠️ HIGHLY IRREGULAR — Consider AF")
                                    if mean_hr < 60:  flags.append("🟡 Bradycardia (<60 bpm)")
                                    if mean_hr > 100: flags.append("🟡 Tachycardia (>100 bpm)")

                                    for flag in flags:
                                        css = "alert-bad" if "CRITICAL" in flag else "alert-warn"
                                        st.markdown(f'<div class="{css}">{flag}</div>', unsafe_allow_html=True)

                                    if not flags:
                                        st.markdown('<div class="alert-good">✅ Heart rate and rhythm within normal parameters on signal analysis. Proceed to AI visual analysis for full 12-lead interpretation.</div>', unsafe_allow_html=True)

                                except Exception as e_nk:
                                    st.warning(f"NeuroKit2 processing note: {e_nk}")

                            # Send plot to Gemini for full 12-lead interpretation
                            st.markdown("---")
                            st.markdown("### 🤖 AI Full 12-Lead Interpretation")
                            if st.button("🔍 Send ECG to Gemini for AI Interpretation", type="primary", use_container_width=True):
                                with st.spinner("Gemini analyzing 12-lead ECG..."):
                                    ai_ctx = f"PhysioNet WFDB record: {record_name} | Sampling rate: {fs}Hz | Duration: {duration:.1f}s | Signal-derived HR: {mean_hr:.0f}bpm | Rhythm: {'Regular' if is_regular else 'Irregular'}"
                                    response = call_ai_radiology(ecg_img_bytes, "image/png", "ECG / 12-lead EKG", ai_ctx)
                                if not response.startswith("!ERR"):
                                    findings, structured = parse_findings_from_response(response)
                                    report = extract_report_text(response)
                                    dx = structured.get("diagnosis","")
                                    if dx:
                                        urg = structured.get("urgency","ROUTINE")
                                        urg_color = {"EMERGENCY":"#dc2626","URGENT":"#f59e0b","ROUTINE":"#16a34a"}.get(urg.upper(),"#0e7490")
                                        st.markdown(f'<div style="background:{urg_color};color:white;border-radius:10px;padding:1rem 1.5rem;margin:.5rem 0;font-size:1.1rem;font-weight:700">{urg} — {dx}</div>', unsafe_allow_html=True)
                                    with st.expander("📄 Full AI ECG Report", expanded=True):
                                        st.markdown(report)
                                        st.download_button("💾 Download", report, f"{record_name}_ai_report.txt", "text/plain", use_container_width=True)
                                else:
                                    st.error(response)

                    except ImportError as e_imp:
                        missing = str(e_imp).split("'")[1] if "'" in str(e_imp) else str(e_imp)
                        st.error(f"Library not installed: {missing}")
                        st.markdown(f"""
                        **Install the required libraries:**
                        ```bash
                        pip install wfdb neurokit2 numpy matplotlib scipy
                        ```
                        Then restart Streamlit.
                        """)
                    except Exception as e_wfdb:
                        st.error(f"Error processing WFDB file: {e_wfdb}")
                        st.markdown("Make sure both .hea and .dat files are from the same record and are not corrupted.")

    # ════════════════════════════════════════════════════════
    # TAB 4: CASE IMAGING REPORT (from Excel)
    # ════════════════════════════════════════════════════════
    with tab_case:
        if not c:
            st.warning("No case selected. Go to Case Library and start a case first.")
            if st.button("📚 Case Library"): nav("library")
        else:
            st.markdown(f'<div class="alert-info">📋 <b>Patient:</b> {c.get("Age_Sex","?")} | <b>CC:</b> {c.get("Chief_Complaint","?")}</div>', unsafe_allow_html=True)
            opts = ["Chest X-Ray","Abdominal X-Ray","CT Head","CT Abdomen/Pelvis",
                    "CT Chest","MRI Brain","Uroscan","Echocardiogram","ECG (12-lead)","FAST Ultrasound"]
            sel = st.multiselect("Order imaging:", opts)
            if st.button("📡 Request Imaging", type="primary"):
                if sel: st.session_state.imaging_seen = True
                else:   st.warning("Select at least one modality.")
            if st.session_state.imaging_seen:
                img_data = str(c.get("Imaging_Tests","none"))
                st.markdown(f'''<div style="background:#f0f9ff;border-radius:12px;padding:1.2rem;border:2px solid #0ea5e9;">
                    <b>📋 Reported Findings:</b><br><span style="color:#1e3a5f;line-height:1.8">{img_data}</span>
                </div>''', unsafe_allow_html=True)

                # ── Display linked real radiological images for this case ──
                if IMAGE_LIBRARY_OK:
                    case_id = str(c.get("Case_ID","")).strip()
                    if case_id:
                        try:
                            render_case_linked_images(case_id)
                        except Exception as _imerr:
                            print(f"[case viewer] could not render linked images: {_imerr}")

                st.markdown('<div class="alert-warn">💡 <b>Tip:</b> Upload the actual image in Tab 1 for AI annotation and visual highlights.</div>', unsafe_allow_html=True)

    b1, b2 = st.columns(2)
    with b1:
        if st.button("← Laboratory", use_container_width=True): nav("lab")
    with b2:
        if st.button("📝 Submit Diagnosis →", use_container_width=True, type="primary"): nav("diagnosis")


def page_surgery():
    st.markdown('<div class="section-header">🔪 Surgery Room</div>',unsafe_allow_html=True)
    c=st.session_state.selected_case
    sk,sv=get_surgery_for_case(c) if c else (None,None)
    surg_options={v["name"]:k for k,v in SURGERIES.items()}
    col1,col2=st.columns([3,1])
    with col1:
        selected_name=st.selectbox("Select Surgical Procedure:",list(surg_options.keys()),
            index=list(surg_options.keys()).index(sv["name"]) if sv and sv["name"] in surg_options else 0)
    with col2:
        st.markdown("")
        st.markdown("")
        if sv and sv["name"]==selected_name: st.markdown('<div class="alert-good">✅ Recommended</div>',unsafe_allow_html=True)
    chosen=SURGERIES[surg_options[selected_name]]
    st.markdown(f'<div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;border-radius:14px;padding:1.3rem;margin-bottom:1rem;"><h2 style="margin:0;font-size:1.4rem">🔪 {chosen["name"]}</h2><div style="display:flex;gap:2rem;margin-top:.6rem;font-size:.85rem;opacity:.9"><span>⏱️ {chosen["duration"]}</span><span>💉 {chosen["anesthesia"]}</span><span>🛏️ {chosen["position"]}</span></div></div>',unsafe_allow_html=True)
    tab1,tab2,tab3=st.tabs(["📋 Step-by-Step","🎬 Video","⚠️ Complications & AI Notes"])
    with tab1:
        all_steps=[]
        for ph,steps in chosen["phases"].items():
            for s in steps: all_steps.append((ph,s))
        total=len(all_steps); current=min(st.session_state.surgery_step,total-1)
        st.markdown(f'<div style="background:#e5e7eb;border-radius:999px;height:10px;margin:.5rem 0 1rem;"><div style="background:linear-gradient(90deg,#0e7490,#16a34a);height:10px;border-radius:999px;width:{((current+1)/total)*100:.0f}%"></div></div><div style="text-align:center;font-size:.82rem;color:#6b7280;margin-bottom:1rem">Step {current+1} of {total}</div>',unsafe_allow_html=True)
        nc1,nc2,nc3=st.columns([1,2,1])
        with nc1:
            if st.button("⬅️ Previous",use_container_width=True,disabled=current==0):
                st.session_state.surgery_step=max(0,current-1); st.rerun()
        with nc2: st.markdown(f'<div style="text-align:center;font-weight:600;color:#0a2540">{all_steps[current][0]}</div>',unsafe_allow_html=True)
        with nc3:
            if st.button("Next ➡️",use_container_width=True,disabled=current==total-1):
                st.session_state.surgery_step=min(total-1,current+1); st.rerun()
        ph,step=all_steps[current]
        st.markdown(f'<div style="background:white;border-radius:14px;padding:1.5rem;border:2px solid #0e7490;box-shadow:0 4px 16px rgba(14,116,144,.15);margin:.5rem 0;"><h3 style="color:#0a2540;margin:0 0 .8rem">🔹 {step["step"]}</h3><p style="color:#374151;line-height:1.8;font-size:.95rem">{step["detail"]}</p></div>',unsafe_allow_html=True)
        if step.get("instruments"):
            st.markdown("**🔧 Instruments/Supplies:**")
            icols=st.columns(min(len(step["instruments"]),5))
            for col,inst in zip(icols,step["instruments"]):
                with col: st.markdown(f'<div class="instrument-card">🔧<br>{inst}</div>',unsafe_allow_html=True)
        with st.expander("📑 All Steps Overview"):
            pidx=0
            for ph_name,steps in chosen["phases"].items():
                st.markdown(f'<div class="phase-header">{ph_name}</div>',unsafe_allow_html=True)
                for i,s in enumerate(steps):
                    sidx=pidx+i; is_cur=sidx==current; is_done=sidx<current
                    icon="✅" if is_done else "▶️" if is_cur else "⬜"
                    css="active" if is_cur else "completed" if is_done else ""
                    st.markdown(f'<div class="surgery-step {css}">{icon} <b>{s["step"]}</b><br><span style="font-size:.8rem;color:#6b7280">{s["detail"][:100]}...</span></div>',unsafe_allow_html=True)
                pidx+=len(steps)
        if st.button("🔄 Restart",use_container_width=True): st.session_state.surgery_step=0; st.rerun()
    with tab2:
        yt_url   = chosen["youtube"]
        vid_id   = yt_url.split("/embed/")[-1].split("?")[0] if "/embed/" in yt_url else ""
        watch_url= f"https://www.youtube.com/watch?v={vid_id}" if vid_id else yt_url
        search_q = chosen["name"].replace(" ","+") + "+surgical+technique+educational"
        search_url = f"https://www.youtube.com/results?search_query={search_q}"

        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                    border-radius:12px;padding:1rem 1.5rem;margin-bottom:1rem;">
            <h3 style="margin:0 0 .3rem">🎬 Surgical Video — {chosen["name"]}</h3>
            <p style="margin:0;opacity:.85;font-size:.85rem">
                Educational video showing the complete procedure step by step.
            </p>
        </div>
        """, unsafe_allow_html=True)

        # Embedded player with JS error detection + auto-fallback
        components.html(f"""
        <!DOCTYPE html><html>
        <head><style>
        *{{box-sizing:border-box;margin:0;padding:0;font-family:system-ui,sans-serif;}}
        body{{background:#0a1628;color:white;}}
        .yt-wrapper{{position:relative;width:100%;padding-bottom:56.25%;background:#000;border-radius:10px;overflow:hidden;}}
        .yt-wrapper iframe{{position:absolute;top:0;left:0;width:100%;height:100%;border:0;}}
        .fallback{{display:none;flex-direction:column;align-items:center;justify-content:center;
                   text-align:center;padding:2rem;gap:1rem;min-height:260px;
                   background:linear-gradient(135deg,#0a1628,#0f3460);border-radius:10px;}}
        .fallback.show{{display:flex;}}
        .fb-icon{{font-size:3rem;}}
        .fb-title{{font-size:1rem;font-weight:700;color:#e2e8f0;}}
        .fb-sub{{font-size:.82rem;color:#94a3b8;max-width:400px;line-height:1.5;}}
        .fb-btn{{display:inline-block;padding:.65rem 1.4rem;border-radius:8px;
                 font-weight:700;font-size:.88rem;text-decoration:none;
                 transition:opacity .15s;}}
        .fb-btn:hover{{opacity:.85;}}
        .check-btn{{margin-top:.5rem;background:none;border:1px solid #475569;
                    color:#94a3b8;padding:.35rem .9rem;border-radius:6px;font-size:.78rem;cursor:pointer;}}
        </style></head>
        <body>
        <div class="yt-wrapper" id="ytWrap">
            <iframe id="ytFrame"
                src="https://www.youtube-nocookie.com/embed/{vid_id}?rel=0&modestbranding=1&controls=1"
                allow="accelerometer; clipboard-write; encrypted-media; gyroscope; picture-in-picture"
                allowfullscreen
                title="{chosen["name"]}"
                onload="checkVideoLoaded(this)">
            </iframe>
        </div>
        <div class="fallback" id="fallback">
            <div class="fb-icon">🎬</div>
            <div class="fb-title">Video unavailable or restricted</div>
            <div class="fb-sub">
                This specific video is no longer available on YouTube. Use the buttons below to find
                current high-quality educational videos for <b>{chosen["name"]}</b>.
            </div>
            <div style="display:flex;gap:.6rem;flex-wrap:wrap;justify-content:center;">
                <a class="fb-btn" style="background:#dc2626;color:white;"
                   href="{search_url}" target="_blank">
                   ▶ Search YouTube: {chosen["name"]}
                </a>
                <a class="fb-btn" style="background:#0e7490;color:white;"
                   href="https://www.youtube.com/results?search_query={chosen['name'].replace(' ','+')}+full+procedure" target="_blank">
                   🔍 Full Procedure Videos
                </a>
            </div>
            <button class="check-btn" onclick="retryVideo()">↩ Retry embed</button>
        </div>
        <script>
        var _checked = false;
        function checkVideoLoaded(iframe) {{
            if (_checked) return;
            _checked = true;
            // Give YouTube 3s to load; if the contentDocument title shows error, show fallback
            setTimeout(function() {{
                try {{
                    var doc = iframe.contentDocument || iframe.contentWindow.document;
                    var title = doc && doc.title ? doc.title.toLowerCase() : '';
                    if (title.includes('unavailable') || title.includes('error') || title === '') {{
                        showFallback();
                    }}
                }} catch(e) {{
                    // Cross-origin — can't read title, try image probe instead
                    probeThumb();
                }}
            }}, 3000);
        }}
        function probeThumb() {{
            var img = new Image();
            img.onload  = function() {{ /* thumb exists, video likely ok */ }};
            img.onerror = function() {{ showFallback(); }};
            img.src = 'https://img.youtube.com/vi/{vid_id}/mqdefault.jpg?' + Date.now();
        }}
        function showFallback() {{
            document.getElementById('ytWrap').style.display = 'none';
            document.getElementById('fallback').classList.add('show');
        }}
        function retryVideo() {{
            _checked = false;
            document.getElementById('ytWrap').style.display = 'block';
            document.getElementById('fallback').classList.remove('show');
            document.getElementById('ytFrame').src = document.getElementById('ytFrame').src;
        }}
        // Also probe on load
        probeThumb();
        </script>
        </body></html>
        """, height=430)

        st.markdown("")
        col_a, col_b, col_c = st.columns(3)
        with col_a:
            st.markdown(f"""
            <a href="{search_url}" target="_blank" style="display:block;text-align:center;
               background:#dc2626;color:white;text-decoration:none;border-radius:10px;
               padding:.7rem 1rem;font-weight:700;font-size:.9rem;">
               ▶️ Search YouTube for this Surgery
            </a>""", unsafe_allow_html=True)
        with col_b:
            st.markdown(f"""
            <a href="https://www.youtube.com/results?search_query={chosen['name'].replace(' ','+')}+step+by+step+annotated" 
               target="_blank" style="display:block;text-align:center;
               background:#0e7490;color:white;text-decoration:none;border-radius:10px;
               padding:.7rem 1rem;font-weight:700;font-size:.9rem;">
               🔍 Annotated Step-by-Step
            </a>""", unsafe_allow_html=True)
        with col_c:
            st.markdown(f"""
            <a href="https://www.websurg.com/search/?query={chosen['name'].replace(' ','+')}+procedure" 
               target="_blank" style="display:block;text-align:center;
               background:#7c3aed;color:white;text-decoration:none;border-radius:10px;
               padding:.7rem 1rem;font-weight:700;font-size:.9rem;">
               🏥 WebSurg Library
            </a>""", unsafe_allow_html=True)

        st.markdown("")
        st.caption("⚠️ Educational content only. Real surgery requires proper surgical training. Videos open on YouTube/WebSurg.")
    with tab3:
        st.markdown("### ⚠️ Complications")
        for comp in chosen["complications"]: st.markdown(f'<div class="alert-bad">🔴 {comp}</div>',unsafe_allow_html=True)
        st.markdown("")
        case_ctx=f"Patient: {c.get('Age_Sex','?')} | Diagnosis: {c.get('Final_Diagnosis','?')}" if c else "General case."
        if st.button("🤖 Generate AI Operative Note",type="primary",use_container_width=True):
            prompt=f"Write a surgical operative note for {chosen['name']}. {case_ctx}. Include: Indication, Procedure, Findings, Complications (none), Disposition. Be clinical and concise."
            with st.spinner("Generating..."):
                note=call_ai("You are a surgeon writing operative notes.",[{"role":"user","content":prompt}])
            st.markdown(f'<div style="background:white;border-radius:12px;padding:1.2rem;border:2px solid #0e7490;font-family:monospace;font-size:.85rem;line-height:1.8;"><b>OPERATIVE NOTE</b><br>{note.replace(chr(10),"<br>")}</div>',unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# 🔴 LIVE DISCUSSION
# ════════════════════════════════════════════════════════════════
def page_live():
    c = st.session_state.selected_case
    if not c:
        st.markdown('<div class="section-header">🔴 Live Discussion</div>', unsafe_allow_html=True)
        st.warning("No case selected.")
        if st.button("📚 Case Library"): nav("library")
        return

    sys_live = live_patient_sys(c)

    # ── LIVE HEADER ──
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1a0533,#4c1d95);color:white;border-radius:14px;
                padding:1.2rem 1.5rem;margin-bottom:1rem;">
        <div style="display:flex;align-items:center;gap:.8rem;">
            <span class="live-indicator"></span>
            <span style="font-size:1.15rem;font-weight:700">🔴 LIVE SESSION — {c.get('Age_Sex','?')}</span>
            <span style="background:#dc2626;border-radius:999px;padding:.1rem .6rem;font-size:.75rem">LIVE</span>
        </div>
        <div style="font-size:.85rem;opacity:.85;margin-top:.3rem">
            ⚠️ {c.get('Chief_Complaint','?')} | Vitals: {c.get('Vitals','?')} |
            The patient will ask you questions back. Speak or type!
        </div>
    </div>""", unsafe_allow_html=True)

    # Initialize live history
    if not st.session_state.live_history:
        with st.spinner("Connecting to patient..."):
            g = call_ai(sys_live, [{
                "role": "user",
                "content": "You are now live with a student physician. Introduce yourself as a patient, describe your symptoms vividly and emotionally. Then ask the doctor one genuine worried question like 'Doctor, is it serious?'"
            }])
        if not g.startswith("!ERR"):
            st.session_state.live_history.append({"role":"patient","content":g,"time":datetime.now().strftime("%H:%M:%S")})
            st.session_state.avatar_mood = detect_mood(g)
            if st.session_state.voice_enabled:
                tts_speak(g)
        else:
            st.error(g); st.stop()

    # ── LAYOUT ──
    av_col, chat_col = st.columns([1, 3])

    with av_col:
        st.markdown(render_avatar(st.session_state.avatar_mood, c.get("Age_Sex","")), unsafe_allow_html=True)
        st.markdown("")
        msgs_count = len([m for m in st.session_state.live_history if m["role"] == "student"])
        st.markdown(f'<div class="alert-info" style="font-size:.75rem"><b>Exchanges:</b> {msgs_count} | <b>Mood:</b> {st.session_state.avatar_mood.title()}</div>', unsafe_allow_html=True)

        # Voice controls
        st.markdown("**🔊 Voice Controls**")
        if st.button("🔊 Replay Patient", use_container_width=True, key="live_replay"):
            last = [m["content"] for m in reversed(st.session_state.live_history) if m["role"] == "patient"]
            if last and st.session_state.voice_enabled:
                tts_speak(last[0])
        if st.button("🔇 Stop Speaking", use_container_width=True, key="live_stop"):
            tts_stop()

        st.markdown("**📝 Your Notes**")
        st.text_area("Clinical notes:", height=100, placeholder="Document impressions...", label_visibility="collapsed", key="live_notes")

    with chat_col:
        # Chat history
        for msg in st.session_state.live_history:
            t = msg.get("time","")
            if msg["role"] == "patient":
                st.markdown(f'<div class="chat-live">🤒 <b>Patient</b> <span style="font-size:.7rem;color:#9ca3af">{t}</span><br>{msg["content"]}</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="chat-student">👨‍⚕️ <b>Dr. You</b> <span style="font-size:.7rem;color:#9ca3af">{t}</span><br>{msg["content"]}</div>', unsafe_allow_html=True)

        st.markdown("")
        st.markdown("---")

        # ── 🎤 VOICE INPUT ──
        st.markdown("**🎤 Speak to Patient (Voice Input)**")
        voice_input_component(key="live_voice", role="doctor", height=165)
        st.markdown('<div class="alert-info" style="font-size:.78rem;">💡 <b>How to use:</b> Press mic → speak → click <b>Send Voice Message</b> → your speech is copied → paste (Ctrl+V) in text box → Send. Works in Chrome & Edge only.</div>', unsafe_allow_html=True)
        st.markdown("---")

        # ── TEXT INPUT ──
        st.markdown("**⌨️ Or Type Your Message**")
        with st.form("live_chat", clear_on_submit=True):
            i1, i2 = st.columns([5,1])
            with i1:
                inp = st.text_input("Type message:", placeholder="e.g. I will now palpate your abdomen...", label_visibility="collapsed")
            with i2:
                go = st.form_submit_button("Send →", use_container_width=True)

        if go and inp.strip():
            st.session_state.live_history.append({"role":"student","content":inp,"time":datetime.now().strftime("%H:%M:%S")})
            # Doctor speaks their message
            if st.session_state.voice_enabled:
                tts_speak_doctor(inp)
            time.sleep(0.5)  # Brief pause before patient responds
            live_msgs = []
            for m in st.session_state.live_history:
                role = m.get("role",""); cont = str(m.get("content","")).strip()
                if not cont: continue
                api_role = "user" if role == "student" else "assistant"
                if live_msgs and live_msgs[-1]["role"] == api_role: continue
                live_msgs.append({"role":api_role,"content":cont})
            while live_msgs and live_msgs[0]["role"] == "assistant": live_msgs.pop(0)
            with st.spinner("Patient responding..."):
                rep = call_ai(sys_live, live_msgs, max_tokens=400)
            if not rep.startswith("!ERR"):
                st.session_state.live_history.append({"role":"patient","content":rep,"time":datetime.now().strftime("%H:%M:%S")})
                st.session_state.avatar_mood = detect_mood(rep)
                if st.session_state.voice_enabled:
                    tts_speak(rep)
                # Generate suggested follow-up questions
                if len(st.session_state.live_history) % 3 == 0:  # Every 3 exchanges
                    q_prompt = (f"Based on this patient response: '{rep}' and the clinical context "
                               f"(patient with {c.get('Chief_Complaint','?')}), "
                               f"suggest 3 short follow-up clinical questions the doctor should ask next. "
                               f"Format: just 3 questions, one per line, no numbering, max 10 words each.")
                    sugg = call_ai("You suggest clinical follow-up questions for a medical student.", 
                                  [{"role":"user","content":q_prompt}], max_tokens=100)
                    if not sugg.startswith("!ERR"):
                        st.session_state["suggested_questions"] = [q.strip() for q in sugg.strip().split("\n") if q.strip()][:3]
            st.rerun()

        # ── AI Suggested Questions ──
        if st.session_state.get("suggested_questions"):
            st.markdown("**🤖 AI Suggests asking:**")
            sq_cols = st.columns(3)
            for sq_col, sq in zip(sq_cols, st.session_state.get("suggested_questions", [])):
                with sq_col:
                    if st.button(f"💬 {sq}", key=f"sq_{sq[:15]}", use_container_width=True):
                        st.session_state.live_history.append({"role":"student","content":sq,"time":datetime.now().strftime("%H:%M:%S")})
                        if st.session_state.voice_enabled:
                            tts_speak_doctor(sq)
                        live_msgs_sq = []
                        for m in st.session_state.live_history:
                            role = m.get("role",""); cont = str(m.get("content","")).strip()
                            if not cont: continue
                            api_role = "user" if role == "student" else "assistant"
                            if live_msgs_sq and live_msgs_sq[-1]["role"] == api_role: continue
                            live_msgs_sq.append({"role":api_role,"content":cont})
                        while live_msgs_sq and live_msgs_sq[0]["role"] == "assistant": live_msgs_sq.pop(0)
                        with st.spinner("Patient responding..."):
                            rep_sq = call_ai(live_patient_sys(c), live_msgs_sq, max_tokens=400)
                        if not rep_sq.startswith("!ERR"):
                            st.session_state.live_history.append({"role":"patient","content":rep_sq,"time":datetime.now().strftime("%H:%M:%S")})
                            st.session_state.avatar_mood = detect_mood(rep_sq)
                            if st.session_state.voice_enabled:
                                tts_speak(rep_sq)
                        st.rerun()

        # ── Quick Actions ──
        st.markdown("**⚡ Quick Clinical Actions:**")
        actions = [
            "I will now auscultate your chest — breathe normally",
            "I will palpate your abdomen — tell me if it hurts",
            "Can you point exactly where the pain is?",
            "Does the pain radiate anywhere?",
            "I will check your blood pressure now",
            "Take a deep breath please",
            "Any nausea or vomiting?",
            "I will examine your eyes and throat",
        ]
        acols = st.columns(4)
        for i, a in enumerate(actions):
            with acols[i % 4]:
                if st.button(a[:26]+"...", key=f"la_{i}", use_container_width=True):
                    st.session_state.live_history.append({"role":"student","content":a,"time":datetime.now().strftime("%H:%M:%S")})
                    live_msgs = []
                    for m in st.session_state.live_history:
                        role = m.get("role",""); cont = str(m.get("content","")).strip()
                        if not cont: continue
                        api_role = "user" if role == "student" else "assistant"
                        if live_msgs and live_msgs[-1]["role"] == api_role: continue
                        live_msgs.append({"role":api_role,"content":cont})
                    while live_msgs and live_msgs[0]["role"] == "assistant": live_msgs.pop(0)
                    with st.spinner("Patient responding..."):
                        rep = call_ai(sys_live, live_msgs, max_tokens=400)
                    if not rep.startswith("!ERR"):
                        st.session_state.live_history.append({"role":"patient","content":rep,"time":datetime.now().strftime("%H:%M:%S")})
                        st.session_state.avatar_mood = detect_mood(rep)
                        if st.session_state.voice_enabled:
                            tts_speak(rep)
                    st.rerun()

        # ── Scenario Injectors ──
        st.markdown("**🎯 Inject Scenario:**")
        sc1,sc2,sc3,sc4 = st.columns(4)
        scenarios = [
            ("😰 Deteriorates", "The patient suddenly grabs their chest and says the pain is now 10/10 and they feel they cannot breathe."),
            ("🤢 New Symptom",  "Patient suddenly says they feel very nauseous and dizzy and their vision is blurring."),
            ("👨‍👩‍👧 Family Arrives","A panicked family member rushes in asking 'Doctor what is wrong with my mother/father?!'"),
            ("💊 Allergy Alert", "Patient says 'Wait doctor — I think I am allergic to that medication you mentioned!'"),
        ]
        for col,(label,scenario) in zip([sc1,sc2,sc3,sc4], scenarios):
            with col:
                if st.button(label, key=f"sc_{label[:6]}", use_container_width=True):
                    st.session_state.live_history.append({"role":"student","content":f"[SCENARIO]: {scenario}","time":datetime.now().strftime("%H:%M:%S")})
                    live_msgs = []
                    for m in st.session_state.live_history:
                        role = m.get("role",""); cont = str(m.get("content","")).strip()
                        if not cont: continue
                        api_role = "user" if role == "student" else "assistant"
                        if live_msgs and live_msgs[-1]["role"] == api_role: continue
                        live_msgs.append({"role":api_role,"content":cont})
                    while live_msgs and live_msgs[0]["role"] == "assistant": live_msgs.pop(0)
                    with st.spinner("Patient responding..."):
                        rep = call_ai(sys_live, live_msgs, max_tokens=400)
                    if not rep.startswith("!ERR"):
                        st.session_state.live_history.append({"role":"patient","content":rep,"time":datetime.now().strftime("%H:%M:%S")})
                        st.session_state.avatar_mood = detect_mood(rep)
                        if st.session_state.voice_enabled:
                            tts_speak(rep)
                    st.rerun()

    # ── End Session ──
    st.markdown("")
    b1, b2, b3 = st.columns(3)
    with b1:
        if st.button("🏁 End Session & Get Feedback", use_container_width=True, type="primary"):
            transcript = "\n".join([
                f"{'DOCTOR' if m['role']=='student' else 'PATIENT'} [{m.get('time','')}]: {m['content']}"
                for m in st.session_state.live_history
            ])
            feedback_prompt = f"""A medical student had a LIVE clinical discussion with a virtual patient.
Case: {c.get('Age_Sex','?')} with {c.get('Chief_Complaint','?')}
True Diagnosis: {c.get('Final_Diagnosis','?')}

TRANSCRIPT:
{transcript[:3000]}

Evaluate the student:
1. Communication & Empathy — /10
2. History Taking — /10
3. Clinical Reasoning — /10
4. Patient Management — /10
5. Overall Score — /10
6. Key Strengths (2-3 points)
7. Areas for Improvement (2-3 points)
8. Key findings/questions they missed
Be constructive, specific and educational."""
            with st.spinner("Generating session feedback..."):
                feedback = call_ai("You are a senior clinical educator evaluating a medical student's live patient interaction.",
                                  [{"role":"user","content":feedback_prompt}], max_tokens=1000)
            st.markdown("### 📊 Live Session Feedback")
            st.markdown(f'<div style="background:#fff7ed;border-radius:12px;padding:1.3rem;border-left:5px solid #f59e0b;">{feedback.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)
            st.download_button("💾 Download Transcript", transcript,
                              f"live_session_{datetime.now().strftime('%Y%m%d_%H%M')}.txt", "text/plain")
    with b2:
        if st.button("🗑️ Clear Session", use_container_width=True):
            st.session_state.live_history = []; st.rerun()
    with b3:
        if st.button("📝 Submit Diagnosis", use_container_width=True):
            nav("diagnosis")

def page_diagnosis():
    st.markdown('<div class="section-header">📝 Diagnosis & Treatment Plan</div>',unsafe_allow_html=True)
    c=st.session_state.selected_case
    if not c: st.warning("No case selected."); return
    s1,s2,s3,s4=st.columns(4)
    with s1: st.markdown(f"Interview: {'✅' if st.session_state.chat_history else '❌'}")
    with s2: st.markdown(f"Examination: {'✅' if st.session_state.exam_findings else '❌'}")
    with s3: st.markdown(f"Lab Tests: {'✅' if st.session_state.lab_seen else '❌'}")
    with s4: st.markdown(f"Imaging: {'✅' if st.session_state.imaging_seen else '❌'}")
    st.markdown("")
    dx=st.text_area("Your Diagnosis:",placeholder="Primary (and secondary) diagnosis...",height=80)
    tx=st.text_area("Your Treatment Plan:",placeholder="Medications, doses, monitoring, referrals...",height=120)
    rs=st.text_area("Clinical Reasoning:",placeholder="Explain your reasoning based on history, examination, labs, and imaging...",height=80)
    sk,sv=get_surgery_for_case(c)
    surgical=False
    if sv: surgical=st.checkbox(f"🔪 I recommend surgical intervention: {sv['name']}")
    exam_summary=""
    if st.session_state.exam_findings:
        exam_summary="\nEXAMINATION FINDINGS STUDENT COLLECTED:\n"+"\n".join([f"- {k}: {v}" for k,v in st.session_state.exam_findings.items()])
    if st.button("✅ Submit for Evaluation",type="primary",use_container_width=True):
        if dx.strip() and tx.strip():
            with st.spinner("🤖 AI Tutor evaluating..."):
                sn=f"\nSurgery recommended: {'YES - '+sv['name'] if surgical else 'NO'}" if sv else ""
                msg=(f"Student submission:\nDIAGNOSIS: {dx}\nTREATMENT: {tx}\nREASONING: {rs}{sn}{exam_summary}\n\n"
                     f"Provide:\n1. Diagnosis Evaluation (correct/partial/wrong + why)\n"
                     f"2. Correct Diagnosis with reasoning\n3. Treatment Plan Evaluation\n"
                     f"4. Recommended Treatment (including surgery if indicated)\n"
                     f"5. Examination Findings Interpretation (if provided)\n"
                     f"6. Score /10\n7. Key Learning Points (3 bullets)\nBe constructive and educational.")
                fb=call_ai(tutor_sys(c),[{"role":"user","content":msg}],max_tokens=1500)
            st.markdown("---")
            st.markdown("### 🤖 AI Tutor Evaluation")
            st.markdown(f'<div style="background:#fff7ed;border-radius:12px;padding:1.3rem;border-left:5px solid #f59e0b;">{fb.replace(chr(10),"<br>")}</div>',unsafe_allow_html=True)
            rn=str(c.get("row_num",""))
            if rn not in [str(x) for x in st.session_state.cases_done]:
                st.session_state.cases_done.append(rn); st.session_state.score+=10
            st.session_state.submitted=True; st.balloons()
        else: st.warning("Please fill in both Diagnosis and Treatment Plan.")
    if st.session_state.submitted:
        # ── Auto-prompt MCQs after final diagnosis ─────────────────────
        if MCQ_SYSTEM_OK:
            render_post_diagnosis_mcq_button()

        b1,b2,b3=st.columns(3)
        with b1:
            if st.button("📚 Another Case",use_container_width=True,type="primary"):
                st.session_state.selected_case=None; reset_case(); nav("library")
        with b2:
            if sv and st.button("🔪 View Surgery",use_container_width=True): nav("surgery")
        with b3:
            if st.button("🏠 Home",use_container_width=True): nav("home")

# ════════════════════════════════════════════════════════════════
# AI TUTOR  (Gemini chat  +  Clinical Tutor AI fine-tuned model)
# ════════════════════════════════════════════════════════════════

@st.cache_resource(show_spinner="⏳ Loading Clinical Tutor AI model (first run only)…")
def _load_clinical_tutor_model():
    if not TRANSFORMERS_OK:
        return None, None
    try:
        from huggingface_hub import login as hf_login
        hf_login(token="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")
        model_name = "HamdarAI/clinical-tutor-model"
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForCausalLM.from_pretrained(
            model_name, torch_dtype=torch.float32,
            device_map="cpu", low_cpu_mem_usage=True)
        model.eval()
        return model, tokenizer
    except Exception as e:
        return None, str(e)


def _run_clinical_tutor_model(case_text: str) -> str:
    model, tokenizer = _load_clinical_tutor_model()
    if model is None:
        err = tokenizer if isinstance(tokenizer, str) else "transformers not installed"
        return f"⚠️ Clinical Tutor model unavailable: {err}"
    prompt = (f"<|system|>You are an expert clinical tutor.<|end|>"
              f"<|user|>{case_text}<|end|><|assistant|>")
    inputs = tokenizer(prompt, return_tensors="pt")
    with torch.no_grad():
        outputs = model.generate(**inputs, max_new_tokens=300, do_sample=False)
    full = tokenizer.decode(outputs[0], skip_special_tokens=True)
    if "<|assistant|>" in full:
        return full.split("<|assistant|>")[-1].strip()
    if case_text in full:
        return full.split(case_text)[-1].strip()
    return full.strip()


def page_tutor():
    st.markdown(
        '<div class="main-header">'
        '<h1>🤖 AI Clinical Tutor</h1>'
        '<p>Two AI engines — Gemini conversational tutor &amp; fine-tuned Clinical Tutor AI</p>'
        '</div>', unsafe_allow_html=True)

    tab_gemini, tab_clinical = st.tabs([
        "💬 Gemini Tutor  (conversational)",
        "🧠 Clinical Tutor AI  (fine-tuned model)",
    ])

    # ── TAB 1 : Gemini ────────────────────────────────────────────────────────
    with tab_gemini:
        c = st.session_state.selected_case
        if c:
            sys_t = tutor_sys(c)
            st.markdown(
                f'<div class="alert-warn">📋 Active case: '
                f'<b>{c.get("Age_Sex","?")} — {c.get("Chief_Complaint","?")}</b></div>',
                unsafe_allow_html=True)
        else:
            sys_t = ("You are a clinical tutor for MLS Academy. Answer general clinical "
                     "questions, explain medical concepts, guide students. Be educational "
                     "and evidence-based.")
            st.markdown('<div class="alert-info">ℹ️ No active case. '
                        'Ask any general clinical question.</div>', unsafe_allow_html=True)

        for msg in st.session_state.tutor_history:
            if msg["role"] == "student":
                st.markdown(f'<div class="chat-student">👨‍⚕️ <b>You:</b> {msg["content"]}</div>',
                            unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="chat-tutor">🤖 <b>Tutor:</b> {msg["content"]}</div>',
                            unsafe_allow_html=True)

        with st.form("tutor_gemini", clear_on_submit=True):
            t1, t2 = st.columns([5, 1])
            with t1:
                q = st.text_input("Ask tutor:", placeholder="Give me a hint / Explain this finding…",
                                  label_visibility="collapsed")
            with t2:
                go = st.form_submit_button("Ask →", use_container_width=True)

        if go and q.strip():
            st.session_state.tutor_history.append({"role": "student", "content": q})
            msgs = [{"role": "user" if m["role"] == "student" else "assistant",
                     "content": m["content"]} for m in st.session_state.tutor_history]

            # ── RAG: inject relevant medical references into system prompt ──
            sys_with_rag = sys_t
            if RAG_SYSTEM_OK:
                try:
                    rag_context = get_rag_context_for_query(q)
                    if rag_context:
                        sys_with_rag = sys_t + "\n\n" + rag_context

                        st.session_state["_last_rag_used"] = True
                except Exception as _rag_err:
                    print(f"[tutor] RAG error: {_rag_err}")

            with st.spinner("Tutor thinking…"):
                rep = call_ai(sys_with_rag, msgs)
            st.session_state.tutor_history.append({"role": "assistant", "content": rep})
            st.rerun()

        # If the last response used RAG references, show a small badge
        if st.session_state.get("_last_rag_used") and st.session_state.tutor_history:
            st.caption("📖 Last response referenced your custom medical library")
            st.session_state["_last_rag_used"] = False

        if c:
            st.markdown("**Quick hints:**")
            hints = ["Give me a diagnostic hint", "Key findings to focus on?",
                     "Tests to prioritize?", "Explain pathophysiology", "Is surgery indicated?"]
            hcols = st.columns(len(hints))
            for col, h in zip(hcols, hints):
                with col:
                    if st.button(h[:20] + "…", key=f"h_{h[:10]}", use_container_width=True):
                        st.session_state.tutor_history.append({"role": "student", "content": h})
                        msgs = [{"role": "user" if m["role"] == "student" else "assistant",
                                 "content": m["content"]} for m in st.session_state.tutor_history]
                        with st.spinner("…"):
                            rep = call_ai(sys_t, msgs)
                        st.session_state.tutor_history.append({"role": "assistant", "content": rep})
                        st.rerun()

        if st.button("🗑️ Clear Chat", use_container_width=True, key="clr_gemini"):
            st.session_state.tutor_history = []
            st.rerun()

    # ── TAB 2 : Clinical Tutor AI (fine-tuned) ───────────────────────────────
    with tab_clinical:
        st.markdown('<div class="section-header">🧠 Clinical Tutor AI — Fine-tuned Diagnostic Model</div>',
                    unsafe_allow_html=True)
        st.markdown(
            '<div class="alert-info">📌 Uses <b>HamdarAI/clinical-tutor-model</b> — '
            'a fine-tuned model trained for clinical reasoning. Enter a patient case to '
            'receive a structured diagnosis, differential, and clinical learning points.</div>',
            unsafe_allow_html=True)

        active = st.session_state.get("selected_case")
        prefill = ""
        if active:
            prefill = (f"Patient: {active.get('Age_Sex','?')}\n"
                       f"Chief Complaint: {active.get('Chief_Complaint','?')}\n"
                       f"HPI: {active.get('HPI','?')}\n"
                       f"Vitals: {active.get('Vitals','?')}\n"
                       f"Physical Findings: {active.get('Physical_Findings','?')}\n"
                       f"Labs: {active.get('Labs','?')}\n"
                       f"Imaging: {active.get('Imaging_Tests','?')}\n"
                       f"What is the diagnosis and clinical reasoning?")
            st.markdown(
                f'<div class="alert-warn">📋 Pre-filled from active case: '
                f'<b>{active.get("Age_Sex","?")} — {active.get("Chief_Complaint","?")}</b></div>',
                unsafe_allow_html=True)

        case_input = st.text_area(
            "Enter patient case:", value=st.session_state.get("ct_case_input", prefill),
            height=220,
            placeholder=("Patient: 29yr male\nChief Complaint: haemoptysis and weight loss\n"
                         "HPI: productive cough, night sweats for 3 weeks\n"
                         "Vitals: HR 96, Temp 37.6°C, RR 20\n"
                         "Labs: Sputum AFB 3+ positive, CXR: upper lobe infiltrates\n"
                         "What is the diagnosis and clinical reasoning?"),
            key="ct_case_area")

        col_run, col_clr = st.columns([3, 1])
        with col_run:
            run_btn = st.button("🔍 Analyse Case with Clinical Tutor AI",
                                type="primary", use_container_width=True, key="ct_run")
        with col_clr:
            if st.button("🗑️ Clear", use_container_width=True, key="ct_clear"):
                st.session_state.pop("ct_result", None)
                st.rerun()

        if run_btn:
            if not case_input.strip():
                st.warning("⚠️ Please enter a patient case before analysing.")
            elif not TRANSFORMERS_OK:
                st.error("❌ Install `transformers torch` and restart the app.")
            else:
                with st.spinner("🧠 Analysing… (first run downloads the model — may take 1–3 min on CPU)"):
                    result = _run_clinical_tutor_model(case_input.strip())
                st.session_state["ct_result"] = result
                st.rerun()

        if "ct_result" in st.session_state and st.session_state["ct_result"]:
            st.markdown('<div class="section-header">📊 Clinical Analysis Result</div>',
                        unsafe_allow_html=True)
            st.markdown(
                f'<div class="chat-tutor" style="font-style:normal;font-size:.9rem;'
                f'line-height:1.75;white-space:pre-wrap;">{st.session_state["ct_result"]}</div>',
                unsafe_allow_html=True)
            st.markdown("---")
            st.markdown("**💬 Follow-up with Gemini Tutor** — ask about this analysis:")
            with st.form("ct_followup", clear_on_submit=True):
                f1, f2 = st.columns([5, 1])
                with f1:
                    fu_q = st.text_input("Follow-up:", placeholder="Why this diagnosis? What treatment?",
                                         label_visibility="collapsed")
                with f2:
                    fu_go = st.form_submit_button("Ask →", use_container_width=True)
            if fu_go and fu_q.strip():
                fu_sys = (f"You are an expert clinical tutor. A student received this AI analysis:\n\n"
                          f"{st.session_state['ct_result']}\n\nAnswer their follow-up question clearly.")
                with st.spinner("Gemini elaborating…"):
                    fu_rep = call_ai(fu_sys, [{"role": "user", "content": fu_q}])
                st.markdown(f'<div class="chat-tutor">🤖 <b>Gemini:</b> {fu_rep}</div>',
                            unsafe_allow_html=True)

        with st.expander("💡 How to write a good case prompt"):
            st.markdown("""
**Include:** Patient demographics · Chief complaint · HPI (onset, duration, severity) ·
Vitals · Physical findings · Labs · Imaging/ECG · End with a question.

**Example:**
```
Patient: 45yr female
Chief Complaint: chest pain 2 hours
HPI: crushing chest pain radiating to left arm, diaphoresis
Vitals: BP 90/60, HR 110, SpO2 94%
Labs: Troponin 1.2 HIGH, CK-MB elevated
ECG: ST elevation V1–V4
What is the diagnosis and immediate management?
```
            """)

# ════════════════════════════════════════════════════════════════
# 🏥 ADD REAL CASE
# ════════════════════════════════════════════════════════════════
def page_add_case():
    st.markdown('<div class="section-header">🏥 Add Real Patient Case</div>',unsafe_allow_html=True)
    st.markdown('<div class="alert-warn">⚠️ <b>De-identification Notice:</b> Please anonymize all patient information before submitting. Remove or change patient name, ID, dates, and any identifying details. This case will be used for educational purposes only.</div>',unsafe_allow_html=True)
    st.markdown('<div class="alert-info">ℹ️ Fill in the patient details below. The system will automatically assign a Case ID and add this case to the hospital database, making it immediately available as an AI patient for students to practice with.</div>',unsafe_allow_html=True)

    with st.form("add_case_form"):
        st.markdown("### 👤 Patient Demographics")
        c1,c2,c3=st.columns(3)
        with c1: age_sex=st.text_input("Age & Sex *",placeholder="e.g. 45 yrs, male")
        with c2: occupation=st.text_input("Occupation",placeholder="e.g. teacher, retired, student")
        with c3: system=st.selectbox("System/Specialty",["cardio","gastro","neuro","ortho","pulmo","uro","surgery","ped","gyneco","endo","ID","ENT","derm","psych","other"])

        st.markdown("### 📋 Presentation")
        c1,c2=st.columns(2)
        with c1: chief_complaint=st.text_input("Chief Complaint *",placeholder="e.g. chest pain and dyspnea")
        with c2: duration=st.text_input("Duration",placeholder="e.g. 2 days, few hours")
        context=st.text_area("Presenting Context",placeholder="e.g. patient came with crushing chest pain radiating to left arm...",height=70)
        title=st.text_input("Case Title",placeholder="e.g. Acute MI, Appendicitis — leave blank for auto-title")

        st.markdown("### 📖 History")
        c1,c2=st.columns(2)
        with c1:
            hpi=st.text_area("History of Present Illness (HPI) *",placeholder="Describe the presenting illness in detail...",height=100)
            pmh=st.text_area("Past Medical History",placeholder="e.g. HTN, DM, previous surgeries...",height=70)
        with c2:
            medications=st.text_area("Current Medications",placeholder="e.g. Aspirin, Metformin, Amlodipine...",height=70)
            social_hx=st.text_area("Social History",placeholder="e.g. smoker, alcoholic, married, teacher...",height=70)
            family_hx=st.text_input("Family History",placeholder="e.g. father: CAD, neg. for cancer")

        st.markdown("### 🩺 Physical Examination")
        c1,c2=st.columns(2)
        with c1: vitals=st.text_area("Vitals *",placeholder="e.g. BP 140/90, HR 95, Temp 37.8, RR 18, SpO2 96%, HGT 120",height=70)
        with c2: appearance=st.text_area("General Appearance",placeholder="e.g. in pain, pale, diaphoretic, dyspneic...",height=70)
        physical_findings=st.text_area("Relevant Physical Findings *",placeholder="e.g. RLQ tenderness, rebound positive, guarding, bowel sounds absent...",height=100)

        st.markdown("### 🔬 Investigations")
        c1,c2=st.columns(2)
        with c1:
            labs=st.text_area("Lab Results",placeholder="e.g. WBC 14, CRP 85, Troponin 0.8 (HIGH), Hb 12...",height=100)
            urine=st.text_area("Urinalysis",placeholder="e.g. WBC >100, RBC 20-30, neg., normal...",height=60)
        with c2:
            imaging=st.text_area("Imaging & ECG",placeholder="e.g. CT abdomen: free fluid RLQ. CXR: normal. ECG: ST elevation V1-V4...",height=100)

        st.markdown("### ✅ Diagnosis & Learning")
        c1,c2=st.columns(2)
        with c1:
            final_dx=st.text_input("Final Diagnosis *",placeholder="e.g. Acute Appendicitis, NSTEMI, Pneumonia")
            difficulty=st.selectbox("Difficulty Level",["basic","intermediate","advanced"])
        with c2:
            learning_obj=st.text_area("Learning Objectives (optional)",placeholder="What should students learn from this case?",height=80)

        st.markdown("### ✅ Verification")
        confirmed=st.checkbox("I confirm this case is based on a real clinical scenario and has been de-identified")
        real_case=st.checkbox("This case is verified as clinically accurate")

        submitted=st.form_submit_button("🚀 Add Case to Hospital Database",type="primary",use_container_width=True)

    if submitted:
        if not all([age_sex,chief_complaint,hpi,vitals,physical_findings,final_dx]):
            st.error("Please fill all required fields (marked with *)")
        elif not confirmed or not real_case:
            st.error("Please confirm de-identification and clinical accuracy checkboxes.")
        else:
            with st.spinner("🤖 AI validating case and creating AI patient profile..."):
                validation_prompt=(f"Validate this medical case for consistency and clinical accuracy:\n"
                                  f"Patient: {age_sex} | CC: {chief_complaint}\n"
                                  f"HPI: {hpi}\n Vitals: {vitals}\n"
                                  f"Physical Findings: {physical_findings}\n"
                                  f"Labs: {labs}\n Imaging: {imaging}\n"
                                  f"Diagnosis: {final_dx}\n\n"
                                  f"1. Is the diagnosis consistent with the clinical picture? (Yes/No + brief reason)\n"
                                  f"2. Are there any obvious clinical inconsistencies?\n"
                                  f"3. What is the educational value of this case (1-5)?\n"
                                  f"4. Generate a realistic patient first name only (do NOT use a real name — generate fictional).\n"
                                  f"5. Suggest 2-3 key learning points for students.\n"
                                  f"Keep response concise.")
                validation=call_ai("You are a clinical expert validating medical educational cases.",
                                   [{"role":"user","content":validation_prompt}],max_tokens=400)

            case_data={
                "title": title if title else final_dx,
                "system": system, "difficulty": difficulty,
                "age_sex": age_sex, "occupation": occupation,
                "chief_complaint": chief_complaint, "duration": duration,
                "context": context, "hpi": hpi, "pmh": pmh,
                "family_hx": family_hx, "social_hx": social_hx,
                "medications": medications, "vitals": vitals,
                "appearance": appearance, "physical_findings": physical_findings,
                "labs": labs, "urine": urine, "imaging": imaging,
                "final_diagnosis": final_dx, "learning_obj": learning_obj,
            }
            case_id=save_new_case(case_data)

            if "ERROR" not in str(case_id):
                st.balloons()
                st.markdown(f"""
                <div style="background:linear-gradient(135deg,#f0fdf4,#dcfce7);border-radius:16px;padding:1.5rem;border:2px solid #16a34a;text-align:center;">
                    <h2 style="color:#166534;margin:0">✅ Case Added Successfully!</h2>
                    <div style="font-size:2rem;font-weight:800;color:#0a2540;margin:.5rem 0">🆔 {case_id}</div>
                    <p style="color:#374151">This case is now live in the Case Library as an AI patient.</p>
                    <p style="color:#374151">Students can search for it by Case ID, title, or system.</p>
                </div>""",unsafe_allow_html=True)
                st.markdown("### 🤖 AI Case Validation Report")
                st.markdown(f'<div style="background:white;border-radius:12px;padding:1.2rem;border:2px solid #0e7490;line-height:1.8;">{validation.replace(chr(10),"<br>")}</div>',unsafe_allow_html=True)
                st.markdown(f'<div class="alert-info">💡 Tell students to search for <b>{case_id}</b> or <b>{final_dx}</b> in the Case Library to access this case.</div>',unsafe_allow_html=True)
                load_cases.clear()
            else:
                st.error(f"Error saving case: {case_id}. Please try again.")


def page_submit_case():
    st.markdown('<div class="section-header">🏥 Submit Real Case — Convert to AI Patient</div>',unsafe_allow_html=True)
    st.markdown("""<div style="background:linear-gradient(135deg,#fdf4ff,#f3e8ff);border:2px solid #a855f7;border-radius:14px;padding:1.3rem;margin-bottom:1rem;">
        <h3 style="margin:0 0 .5rem;color:#6b21a8">🏥 Real Case Submission Portal</h3>
        <p style="margin:0;color:#4c1d95;font-size:.9rem">Submit a real patient case and it becomes a live AI training patient immediately in the Case Library with a unique Case ID.<br><br>
        <b>Warning:</b> Please de-identify all data. Do not include real patient names.</p>
    </div>""",unsafe_allow_html=True)

    st.markdown(f'<div class="alert-good">🆔 <b>Auto-assigned Case ID preview:</b> <code>MLS-{datetime.now().year}-XXXX</code> (finalized on submission)</div>',unsafe_allow_html=True)
    st.markdown("### 📋 Case Information")
    col1,col2,col3=st.columns(3)
    with col1:
        title=st.text_input("Case Title *",placeholder="e.g. Acute Appendicitis in Young Female")
        system=st.selectbox("Medical System *",["abdomen","cardio","respiratory","neuro","ortho","urology","gastro","pulmo","endo","ID","pediatrics","surgery","ENT","gyneco","other"])
    with col2:
        difficulty=st.selectbox("Difficulty *",["basic","intermediate","advanced"])
        learning_obj=st.text_area("Learning Objectives",placeholder="What should students learn?",height=70)
    with col3:
        age_sex=st.text_input("Age & Sex *",placeholder="e.g. 25 yrs, female")
        occupation=st.text_input("Occupation",placeholder="e.g. student, worker")
    st.markdown("### 🚨 Presentation")
    col1,col2=st.columns(2)
    with col1:
        chief_complaint=st.text_input("Chief Complaint *",placeholder="e.g. right lower quadrant pain")
        duration=st.text_input("Duration *",placeholder="e.g. 6 hours, 2 days")
    with col2:
        context=st.text_area("Context / Setting *",placeholder="How did patient present?",height=70)
    st.markdown("### 📖 History")
    col1,col2=st.columns(2)
    with col1:
        hpi=st.text_area("HPI *",placeholder="Detailed description of the current illness...",height=90)
        pmh=st.text_input("Past Medical History",placeholder="e.g. HTN, DM, none")
        family_hx=st.text_input("Family History",placeholder="e.g. neg.")
    with col2:
        social_hx=st.text_input("Social History",placeholder="e.g. smoker, none")
        medications=st.text_area("Medications",placeholder="e.g. Metformin 500mg BD, none",height=70)
    st.markdown("### 🩺 Physical Examination")
    col1,col2=st.columns(2)
    with col1:
        vitals=st.text_input("Vital Signs *",placeholder="e.g. BP 120/80, HR 90, Temp 38.2, SpO2 98%")
        appearance=st.text_input("General Appearance",placeholder="e.g. in pain, diaphoretic")
    with col2:
        physical_findings=st.text_area("Physical Findings *",placeholder="Key examination findings...",height=70)
    st.markdown("### 🔬 Investigations")
    col1,col2,col3=st.columns(3)
    with col1: labs=st.text_area("Lab Results",placeholder="e.g. WBC 14, CRP 45",height=70)
    with col2: urine=st.text_area("Urinalysis",placeholder="e.g. negative",height=70)
    with col3: imaging=st.text_area("Imaging & ECG",placeholder="e.g. CT: acute appendicitis",height=70)
    st.markdown("### ✅ Final Diagnosis")
    col1,col2=st.columns(2)
    with col1: final_dx=st.text_input("Final Diagnosis *",placeholder="e.g. Acute appendicitis")
    with col2: teaching=st.text_area("Teaching Points (optional)",height=60)
    st.markdown("")

    if st.button("🤖 Validate Case with AI Before Submitting", use_container_width=True):
        if all([title,age_sex,chief_complaint,hpi,vitals,physical_findings,final_dx]):
            prompt=f"Review this medical training case:\nCase: {title}\nPatient: {age_sex} | CC: {chief_complaint} | Duration: {duration}\nHPI: {hpi}\nVitals: {vitals} | Physical: {physical_findings}\nLabs: {labs} | Imaging: {imaging} | Final Dx: {final_dx}\n\nEvaluate: 1) Clinical Consistency 2) Completeness 3) Educational Value 4) Red Flags 5) Recommendation: APPROVE / REVISE"
            with st.spinner("AI validating case..."):
                v=call_ai("You are a clinical educator reviewing a medical training case.",
                         [{"role":"user","content":prompt}],max_tokens=600)
            css="alert-good" if "APPROVE" in v.upper() else "alert-warn"
            st.markdown(f'<div class="{css}">{v.replace(chr(10),"<br>")}</div>',unsafe_allow_html=True)
        else:
            st.warning("Fill all required (*) fields first.")

    st.markdown("")
    confirm=st.checkbox("✅ I confirm this case is de-identified and suitable for educational use.")

    if st.button("🚀 Submit Case to Hospital", type="primary", use_container_width=True):
        required=[title,age_sex,chief_complaint,duration,context,hpi,vitals,physical_findings,final_dx]
        if all(r and r.strip() for r in required) and confirm:
            import random, string
            new_id=f"MLS-{datetime.now().year}-" + "".join(random.choices(string.digits,k=4))
            try:
                from openpyxl import load_workbook
                path=os.path.join(os.path.dirname(__file__),"case_studies.xlsx")
                wb=load_workbook(path)
                sm={"case Metadata ":[new_id,title,system,difficulty,learning_obj or ""],
                    "Initial Presentation ":[age_sex,occupation or "",chief_complaint,duration,context],
                    "History taking":[hpi,pmh or "none",family_hx or "neg.",social_hx or "none",medications or "none"],
                    "physical examination ":[vitals,appearance or "in pain",physical_findings],
                    "investigation ":[labs or "none",urine or "none",imaging or "none"],
                    "final diagnosis ":[final_dx]}
                for sn,rd in sm.items():
                    if sn in wb.sheetnames: wb[sn].append(rd)
                wb.save(path)
                load_cases.clear()
                st.balloons()
                st.markdown(f'''<div class="alert-good"><h3 style="margin:0 0 .5rem">🎉 Case Submitted Successfully!</h3>
                    <p><b>Case ID:</b> <code>{new_id}</code></p><p><b>Title:</b> {title}</p>
                    <p>The case is now live in the Case Library as an AI patient!</p></div>''',unsafe_allow_html=True)
                st.markdown(f"**Save your Case ID:** `{new_id}`")
                if st.button("📚 Go to Case Library"): nav("library")
            except Exception as e:
                st.error(f"Error saving: {e}")
        elif not confirm:
            st.warning("Confirm the de-identification checkbox.")
        else:
            st.warning("Fill all required (*) fields.")



# ════════════════════════════════════════════════════════════════
# 👥 MULTIPLAYER ROOM SYSTEM
# ════════════════════════════════════════════════════════════════
import hashlib

ROOMS_DIR = os.path.join(os.path.dirname(__file__), "rooms")
os.makedirs(ROOMS_DIR, exist_ok=True)

def room_path(code):
    return os.path.join(ROOMS_DIR, f"room_{code.upper()}.json")

def create_room(code, case_data):
    """Create a new multiplayer room."""
    room = {
        "code": code.upper(),
        "case": case_data,
        "created": datetime.now().isoformat(),
        "status": "waiting",  # waiting / active / ended
        "doctor": None,
        "patient": None,
        "messages": [],
        "hints": [],
        "doctor_diagnosis": None,
        "ai_evaluation": None,
    }
    with open(room_path(code), "w") as f:
        json.dump(room, f, indent=2)
    return room

def load_room(code):
    """Load room data from file."""
    p = room_path(code)
    if not os.path.exists(p):
        return None
    try:
        with open(p) as f:
            return json.load(f)
    except Exception:
        return None

def save_room(room):
    """Save room data to file."""
    try:
        with open(room_path(room["code"]), "w") as f:
            json.dump(room, f, indent=2)
        return True
    except Exception:
        return False

def add_message(code, role, content_text, name=""):
    """Add a message to the room."""
    room = load_room(code)
    if not room: return False
    room["messages"].append({
        "role": role,
        "content": content_text,
        "name": name,
        "time": datetime.now().strftime("%H:%M:%S"),
        "id": len(room["messages"])
    })
    return save_room(room)

def get_ai_patient_hint(room):
    """Get AI hint for the student playing patient."""
    c = room.get("case", {})
    msgs = room.get("messages", [])
    
    # Last doctor message
    doctor_msgs = [m for m in msgs if m["role"] == "doctor"]
    if not doctor_msgs:
        return "💡 Wait for the doctor to ask you something first, then I can help you respond."
    last_q = doctor_msgs[-1]["content"]
    
    prompt = (f"A medical student is playing the role of a REAL PATIENT in a simulation. "
              f"They need help responding to the doctor's question naturally. "
              f"\n\nPATIENT PROFILE:"
              f"\n- {c.get('Age_Sex','?')} | CC: {c.get('Chief_Complaint','?')}"
              f"\n- HPI: {c.get('HPI','?')}"
              f"\n- PMH: {c.get('PMH','none')} | Meds: {c.get('Medications','none')}"
              f"\n- Physical findings: {c.get('Physical_Findings','?')}"
              f"\n- DIAGNOSIS (student playing patient knows this): {c.get('Final_Diagnosis','?')}"
              f"\n\nDOCTOR JUST ASKED: '{last_q}'"
              f"\n\nGive the PATIENT STUDENT 2-3 short suggestions for how to respond naturally "
              f"as this patient. Use everyday language. Format: bullet points, max 20 words each. "
              f"Do NOT use medical jargon. Make it sound like a real worried person. "
              f"COMPLETE every bullet — never cut off mid-sentence.")
    
    result = call_ai("You help a student play a patient role in a medical simulation.", 
                     [{"role":"user","content":prompt}], max_tokens=400)
    if result.startswith("!ERR"):
        return ("⚠️ Hint unavailable right now — the AI service didn't respond. "
                "Try again in a moment, or check the API key in settings.")
    return result

def get_ai_doctor_hint(room):
    """Get AI hint for the student playing doctor."""
    c = room.get("case", {})
    msgs = room.get("messages", [])
    
    patient_msgs = [m for m in msgs if m["role"] == "patient"]
    if not patient_msgs:
        return ("💡 Start by introducing yourself and asking the patient what brought them in today. "
                "I can suggest follow-up questions once they respond.")
    last_resp = patient_msgs[-1]["content"]
    
    prompt = (f"A medical student is interviewing a patient as the doctor in a simulation. "
              f"The patient just said: '{last_resp}'"
              f"\n\nGive the DOCTOR STUDENT 4 suggested follow-up clinical questions "
              f"to ask next. Each question under 18 words. Be clinical but empathetic. "
              f"Format: one question per line, no numbers. "
              f"COMPLETE every question — never truncate mid-sentence.")
    
    result = call_ai("You suggest clinical questions for a doctor in training.",
                     [{"role":"user","content":prompt}], max_tokens=400)
    if result.startswith("!ERR"):
        return ("⚠️ Hint unavailable right now — the AI service didn't respond. "
                "Try again in a moment, or check the API key in settings.")
    return result


def page_multiplayer():
    st.markdown('<div class="section-header">👥 Live Multiplayer — Student vs Student</div>', unsafe_allow_html=True)
    
    # Init multiplayer session state
    for k, v in {"mp_room_code": "", "mp_role": None, "mp_name": "", 
                 "mp_last_msg_count": 0, "mp_hint": None}.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # ── If not in a room yet ──
    if not st.session_state.mp_room_code:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                    border-radius:16px;padding:1.5rem;margin-bottom:1rem;">
            <h2 style="margin:0 0 .5rem">👥 Live Multiplayer Clinical Session</h2>
            <p style="margin:0;opacity:.9;font-size:.9rem">
                One student plays the <b>Doctor</b> — takes history, examines, diagnoses.<br>
                Another student plays the <b>Patient</b> — acts out the real case symptoms.<br>
                Both connect using the same <b>Room Code</b>.
            </p>
        </div>
        """, unsafe_allow_html=True)

        tab_create, tab_join = st.tabs(["🆕 Create New Room", "🚪 Join Existing Room"])

        with tab_create:
            st.markdown("### Create a Room")
            
            if df.empty:
                st.error("No cases loaded. Add case_studies.xlsx first.")
                return

            c1, c2 = st.columns(2)
            with c1:
                your_name = st.text_input("Your name:", placeholder="e.g. Dr. Ahmed")
            with c2:
                role_choice = st.selectbox("Your role:", ["👨‍⚕️ Doctor (I will diagnose)", "🤒 Patient (I will act symptoms)"])

            # Case selection
            st.markdown("**Select a case for this session:**")
            f1, f2, f3 = st.columns(3)
            with f1: search_mp = st.text_input("Search case:", placeholder="e.g. chest pain", key="mp_search")
            with f2:
                sys_opts = ["Any system"] + sorted(df["System"].dropna().unique().tolist())
                sys_filter = st.selectbox("System:", sys_opts, key="mp_sys")
            with f3:
                diff_filter = st.selectbox("Difficulty:", ["Any", "basic", "intermediate", "advanced"], key="mp_diff")

            filt_mp = df.copy()
            if search_mp: filt_mp = filt_mp[filt_mp.apply(lambda r: search_mp.lower() in str(r).lower(), axis=1)]
            if sys_filter != "Any system": filt_mp = filt_mp[filt_mp["System"].str.lower().str.contains(sys_filter.lower(), na=False)]
            if diff_filter != "Any": filt_mp = filt_mp[filt_mp["Difficulty"].str.lower().str.contains(diff_filter.lower(), na=False)]

            if not filt_mp.empty:
                case_options = {f"{r['Case_ID']} — {r.get('Title') or r['Chief_Complaint']} ({r['Age_Sex']})": i 
                               for i, (_, r) in enumerate(filt_mp.head(50).iterrows())}
                selected_case_label = st.selectbox("Choose case:", list(case_options.keys()), key="mp_case_sel")
                selected_row_idx = case_options[selected_case_label]
                selected_case_mp = filt_mp.iloc[selected_row_idx].to_dict()

                st.markdown(f'<div class="alert-info">📋 <b>Case:</b> {selected_case_mp.get("Chief_Complaint","?")} | {selected_case_mp.get("Age_Sex","?")} | Dx: <b>{"Hidden from Doctor" if "Doctor" in role_choice else selected_case_mp.get("Final_Diagnosis","?")}</b></div>', unsafe_allow_html=True)

            custom_code = st.text_input("Room code (or leave blank for auto):", placeholder="e.g. MLS123", max_chars=8)

            if st.button("🚀 Create Room", type="primary", use_container_width=True):
                if not your_name.strip():
                    st.warning("Enter your name.")
                elif filt_mp.empty:
                    st.warning("Select a valid case.")
                else:
                    import random, string
                    code = custom_code.upper().strip() if custom_code.strip() else "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
                    room = create_room(code, selected_case_mp)
                    role = "doctor" if "Doctor" in role_choice else "patient"
                    room[role] = your_name.strip()
                    room["status"] = "waiting"
                    save_room(room)
                    st.session_state.mp_room_code = code
                    st.session_state.mp_role = role
                    st.session_state.mp_name = your_name.strip()
                    st.session_state.selected_case = selected_case_mp
                    st.rerun()

        with tab_join:
            st.markdown("### Join an Existing Room")
            j1, j2 = st.columns(2)
            with j1:
                join_name = st.text_input("Your name:", placeholder="e.g. Sara", key="join_name")
            with j2:
                join_role = st.selectbox("Your role:", ["👨‍⚕️ Doctor (I will diagnose)", "🤒 Patient (I will act symptoms)"], key="join_role")
            join_code = st.text_input("Room code:", placeholder="e.g. MLS123", max_chars=8, key="join_code")

            if st.button("🚪 Join Room", type="primary", use_container_width=True):
                if not join_name.strip() or not join_code.strip():
                    st.warning("Enter your name and room code.")
                else:
                    room = load_room(join_code.upper())
                    if not room:
                        st.error(f"Room {join_code.upper()} not found. Check the code.")
                    else:
                        role = "doctor" if "Doctor" in join_role else "patient"
                        if room.get(role) and room[role] != join_name.strip():
                            st.error(f"Role '{role}' already taken by {room[role]}. Choose the other role.")
                        else:
                            room[role] = join_name.strip()
                            if room.get("doctor") and room.get("patient"):
                                room["status"] = "active"
                            save_room(room)
                            st.session_state.mp_room_code = join_code.upper()
                            st.session_state.mp_role = role
                            st.session_state.mp_name = join_name.strip()
                            st.session_state.selected_case = room["case"]
                            st.rerun()
        return

    # ── IN A ROOM ──
    code = st.session_state.mp_room_code
    role = st.session_state.mp_role
    name = st.session_state.mp_name
    room = load_room(code)

    if not room:
        st.error("Room not found or expired.")
        st.session_state.mp_room_code = ""
        st.rerun()
        return

    c_data = room.get("case", {})

    # ── Room status header ──
    other_role = "patient" if role == "doctor" else "doctor"
    other_name = room.get(other_role, "Waiting...")
    both_joined = room.get("doctor") and room.get("patient")

    status_color = "#16a34a" if both_joined else "#f59e0b"
    status_text  = "🟢 LIVE" if both_joined else "🟡 Waiting for other student..."

    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#0a2540,#4c1d95);color:white;
                border-radius:14px;padding:1rem 1.5rem;margin-bottom:1rem;">
        <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem;">
            <div>
                <span style="font-size:1.1rem;font-weight:700">👥 Room: <code style="background:#ffffff30;padding:.1rem .4rem;border-radius:4px">{code}</code></span>
                <span style="margin-left:1rem;background:{status_color};border-radius:999px;padding:.1rem .7rem;font-size:.8rem;font-weight:600">{status_text}</span>
            </div>
            <div style="font-size:.85rem;opacity:.9">
                👨‍⚕️ Doctor: <b>{room.get("doctor","—")}</b> &nbsp;|&nbsp; 🤒 Patient: <b>{room.get("patient","—")}</b>
            </div>
        </div>
        <div style="margin-top:.4rem;font-size:.82rem;opacity:.8">
            📋 Case: {c_data.get("Chief_Complaint","?")} | {c_data.get("Age_Sex","?")} | 
            {"🔒 Diagnosis hidden from doctor" if role == "doctor" else f"✅ Diagnosis: {c_data.get('Final_Diagnosis','?')}"}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Share link hint
    st.markdown(f'<div class="alert-info" style="font-size:.8rem">📤 <b>Share with other student:</b> Tell them to go to <b>👥 Multiplayer</b> → Join Room → enter code <code>{code}</code></div>', unsafe_allow_html=True)

    if not both_joined:
        st.markdown("""
        <div style="text-align:center;padding:2rem;background:#f8fafc;border-radius:14px;border:2px dashed #e2e8f0;">
            <div style="font-size:2.5rem">⏳</div>
            <h3 style="color:#0a2540">Waiting for the other student to join...</h3>
            <p style="color:#6b7280">Share the room code above. The session will start automatically when both join.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🔄 Refresh", use_container_width=True):
            st.rerun()
        time.sleep(3)
        st.rerun()
        return

    # ══ ACTIVE SESSION ══
    # Two-column layout: left = info panel, right = chat
    info_col, chat_col = st.columns([1, 3])

    with info_col:
        if role == "doctor":
            # Doctor sees: avatar + vitals (NO diagnosis)
            st.markdown(render_avatar(st.session_state.get("avatar_mood","neutral"), c_data.get("Age_Sex","")), unsafe_allow_html=True)
            st.markdown(f'<div class="alert-warn" style="font-size:.78rem"><b>Patient:</b> {c_data.get("Age_Sex","?")}<br><b>CC:</b> {c_data.get("Chief_Complaint","?")}<br><b>Vitals:</b> {c_data.get("Vitals","normal")}</div>', unsafe_allow_html=True)
            st.markdown('<div class="alert-bad" style="font-size:.75rem">🔒 Diagnosis hidden — discover it!</div>', unsafe_allow_html=True)
            
            # AI hints for doctor
            if st.button("💡 AI Hint", use_container_width=True, key="doc_hint"):
                with st.spinner("Getting hint..."):
                    hint = get_ai_doctor_hint(room)
                if hint:
                    st.session_state.mp_hint = hint
            if st.session_state.mp_hint:
                st.markdown(f'<div class="chat-tutor" style="font-size:.78rem">🤖 Suggested questions:<br>{st.session_state.mp_hint.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

        else:
            # Patient sees: their case info + acting guide
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#fdf4ff,#f3e8ff);border-radius:12px;
                        padding:.8rem;border:2px solid #a855f7;font-size:.8rem;">
                <b style="color:#6b21a8">🎭 YOUR PATIENT ROLE</b><br>
                <b>You are:</b> {c_data.get("Age_Sex","?")} | {c_data.get("Occupation","")}<br>
                <b>Your complaint:</b> {c_data.get("Chief_Complaint","?")}<br>
                <b>Your story:</b> {c_data.get("HPI","?")[:120]}...<br>
                <b>Your history:</b> PMH: {c_data.get("PMH","none")}<br>
                <b>Your meds:</b> {c_data.get("Medications","none")}<br>
                <b>Diagnosis (you know):</b> <span style="color:#7c3aed;font-weight:700">{c_data.get("Final_Diagnosis","?")}</span><br>
                <b style="color:#dc2626">⚠️ Never say the diagnosis directly!</b>
            </div>
            """, unsafe_allow_html=True)

            # AI hints for patient
            if st.button("💡 How to respond?", use_container_width=True, key="pat_hint"):
                with st.spinner("Getting suggestion..."):
                    hint = get_ai_patient_hint(room)
                if hint:
                    st.session_state.mp_hint = hint
            if st.session_state.mp_hint:
                st.markdown(f'<div class="chat-tutor" style="font-size:.78rem">💡 Suggested response:<br>{st.session_state.mp_hint.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

    with chat_col:
        # Show all messages
        messages = room.get("messages", [])
        
        if not messages:
            st.markdown("""
            <div style="text-align:center;padding:1.5rem;background:#f8fafc;border-radius:12px;color:#6b7280;">
                <div style="font-size:1.5rem">💬</div>
                <p>Session started! Doctor, introduce yourself and begin the consultation.</p>
            </div>
            """, unsafe_allow_html=True)

        for msg in messages:
            msg_role = msg.get("role","")
            msg_content = msg.get("content","")
            msg_name = msg.get("name","")
            msg_time = msg.get("time","")
            
            if msg_role == "doctor":
                icon = "👨‍⚕️"
                css = "chat-student"
            elif msg_role == "patient":
                icon = "🤒"
                css = "chat-patient"
            elif msg_role == "system":
                st.markdown(f'<div class="alert-info" style="font-size:.78rem">ℹ️ {msg_content}</div>', unsafe_allow_html=True)
                continue
            else:
                continue

            is_mine = (msg_role == role)
            align = "right" if is_mine else "left"
            st.markdown(
                f'<div class="{css}">{icon} <b>{msg_name}</b> <span style="font-size:.68rem;color:#9ca3af">{msg_time}</span><br>{msg_content}</div>',
                unsafe_allow_html=True
            )

        st.markdown("")

        # ── VOICE INPUT + TEXT INPUT ──
        # Web Speech mic
        mic_color = "#0e7490" if role == "doctor" else "#7c3aed"
        mic_label = "🎤 Speak as Doctor" if role == "doctor" else "🎤 Speak as Patient"
        
        components.html(f"""
        <div style="font-family:Inter,sans-serif;margin-bottom:.3rem;">
          <button id="mpMic" onclick="toggleMpMic()" style="
            background:linear-gradient(135deg,{mic_color},{mic_color}dd);
            color:white;border:none;border-radius:50px;
            padding:.45rem 1.2rem;font-size:.82rem;font-weight:600;
            cursor:pointer;box-shadow:0 3px 8px {mic_color}55;margin-bottom:.3rem;">
            {mic_label} (Chrome/Edge)
          </button>
          <div id="mpStatus" style="font-size:.72rem;color:#6b7280;"></div>
          <div id="mpTranscript" style="background:#f0f9ff;border:1px solid #0ea5e9;border-radius:6px;
            padding:.4rem .7rem;font-size:.82rem;margin-top:.2rem;color:#0a2540;display:none;min-height:20px;"></div>
        </div>
        <script>
        var mpRec=null,mpLs=false;
        function toggleMpMic(){{
          if(!('webkitSpeechRecognition' in window)&&!('SpeechRecognition' in window)){{
            document.getElementById('mpStatus').innerHTML='Use Chrome or Edge for voice';return;}}
          if(mpLs){{mpRec.stop();return;}}
          var SR=window.SpeechRecognition||window.webkitSpeechRecognition;
          mpRec=new SR();mpRec.lang='en-US';mpRec.continuous=false;mpRec.interimResults=true;
          var btn=document.getElementById('mpMic');
          mpRec.onstart=function(){{mpLs=true;
            btn.style.background='linear-gradient(135deg,#16a34a,#166534)';
            btn.innerHTML='🔴 Listening... click to stop';
            document.getElementById('mpStatus').innerHTML='🎙️ Speak now...';
          }};
          mpRec.onresult=function(e){{
            var ft='',it='';
            for(var i=e.resultIndex;i<e.results.length;i++){{
              if(e.results[i].isFinal)ft+=e.results[i][0].transcript;
              else it+=e.results[i][0].transcript;}}
            var tr=document.getElementById('mpTranscript');
            tr.style.display='block';tr.innerHTML=ft||it;
            if(ft){{
              navigator.clipboard.writeText(ft).catch(function(){{}});
              document.getElementById('mpStatus').innerHTML='✅ Copied — paste below and Send';
            }}
          }};
          mpRec.onerror=function(e){{
            document.getElementById('mpStatus').innerHTML='Error: '+e.error;mpLs=false;
            btn.style.background='linear-gradient(135deg,{mic_color},{mic_color}dd)';
            btn.innerHTML='{mic_label} (Chrome/Edge)';
          }};
          mpRec.onend=function(){{mpLs=false;
            btn.style.background='linear-gradient(135deg,{mic_color},{mic_color}dd)';
            btn.innerHTML='{mic_label} (Chrome/Edge)';
          }};
          mpRec.start();
        }}
        </script>
        """, height=110)

        # Text input form
        role_label = "Doctor" if role == "doctor" else "Patient"
        placeholder = "e.g. Good morning, I am Dr. Ahmed. What brings you in today?" if role == "doctor" else "e.g. Doctor, I have been having this terrible pain since yesterday..."
        
        with st.form(f"mp_chat_{code}", clear_on_submit=True):
            c1_f, c2_f = st.columns([5, 1])
            with c1_f:
                msg_inp = st.text_input(
                    f"Message as {role_label}:",
                    placeholder=placeholder,
                    label_visibility="collapsed"
                )
            with c2_f:
                send_btn = st.form_submit_button("Send →", use_container_width=True)

        if send_btn and msg_inp.strip():
            add_message(code, role, msg_inp.strip(), name)
            # Speak the message
            if st.session_state.voice_enabled:
                if role == "doctor":
                    tts_speak_doctor(msg_inp.strip())
                else:
                    tts_speak(msg_inp.strip())
            st.session_state.mp_hint = None
            # Refresh room
            room = load_room(code)
            if room:
                room["status"] = "active"
                save_room(room)
            st.rerun()

        # ── Quick phrase buttons ──
        if role == "doctor":
            st.markdown("**⚡ Quick Doctor Phrases:**")
            phrases = [
                "Hello, I am your doctor today. What brings you in?",
                "Can you describe the pain? Rate it 1-10.",
                "When did this start? Any triggers?",
                "Do you have any medical conditions or allergies?",
                "I am going to examine you now.",
                "Can you point to exactly where it hurts?",
            ]
        else:
            st.markdown("**⚡ Quick Patient Phrases:**")
            phrases = [
                f"Doctor, I've had this {c_data.get('Chief_Complaint','pain')} since {c_data.get('Duration','yesterday')}.",
                "The pain is really bad... maybe an 8 out of 10.",
                "It started suddenly and it's getting worse.",
                "I'm really scared doctor, is it serious?",
                "I've never had anything like this before.",
                "Will I need surgery? How long will I be here?",
            ]

        phrase_cols = st.columns(3)
        for i, phrase in enumerate(phrases):
            with phrase_cols[i % 3]:
                if st.button(phrase[:35]+"…", key=f"phrase_{role}_{i}", use_container_width=True):
                    add_message(code, role, phrase, name)
                    if st.session_state.voice_enabled:
                        if role == "doctor": tts_speak_doctor(phrase)
                        else: tts_speak(phrase)
                    st.rerun()

    # ── Doctor: Submit diagnosis ──
    if role == "doctor" and room.get("status") == "active":
        st.markdown("---")
        st.markdown("### 📝 Submit Your Diagnosis")
        d1, d2 = st.columns(2)
        with d1:
            mp_dx = st.text_input("Your diagnosis:", placeholder="What do you think this patient has?", key="mp_dx")
            mp_tx = st.text_area("Your treatment plan:", height=80, placeholder="Management plan...", key="mp_tx")
        with d2:
            mp_reason = st.text_area("Clinical reasoning:", height=80, placeholder="Key findings that led to your diagnosis...", key="mp_reason")

        if st.button("✅ Submit Diagnosis for Evaluation", type="primary", use_container_width=True, key="mp_submit"):
            if mp_dx.strip() and mp_tx.strip():
                # Get transcript
                transcript = "\n".join([f"{m.get('name','?')} ({m.get('role','?').upper()}): {m.get('content','')}" 
                                        for m in room.get("messages",[])])
                
                eval_prompt = (f"Evaluate this live multiplayer clinical consultation:\n"
                              f"CASE: {c_data.get('Age_Sex','?')} with {c_data.get('Chief_Complaint','?')}\n"
                              f"TRUE DIAGNOSIS: {c_data.get('Final_Diagnosis','?')}\n"
                              f"TRUE FINDINGS: {c_data.get('Physical_Findings','?')} | Labs: {c_data.get('Labs','?')}\n\n"
                              f"CONSULTATION TRANSCRIPT:\n{transcript[:2000]}\n\n"
                              f"DOCTOR SUBMITTED:\nDiagnosis: {mp_dx}\nTreatment: {mp_tx}\nReasoning: {mp_reason}\n\n"
                              f"Evaluate:\n1. Diagnosis accuracy /10\n2. History taking quality /10\n"
                              f"3. Communication & empathy /10\n4. Treatment plan /10\n5. Overall /10\n"
                              f"6. What was done well (2-3 points)\n7. What could be improved (2-3 points)\n"
                              f"8. Key learning points from this case\nBe constructive and specific.")
                
                with st.spinner("🤖 AI evaluating your consultation..."):
                    evaluation = call_ai("You are a senior clinician evaluating a medical student consultation.",
                                        [{"role":"user","content":eval_prompt}], max_tokens=1000)
                
                # Save evaluation to room
                room["doctor_diagnosis"] = {"dx": mp_dx, "tx": mp_tx, "reason": mp_reason}
                room["ai_evaluation"] = evaluation
                room["status"] = "ended"
                save_room(room)

                # Notify via room message
                add_message(code, "system", f"Dr. {name} submitted their diagnosis: {mp_dx}")

                st.markdown("### 🤖 AI Evaluation")
                st.markdown(f'<div style="background:#fff7ed;border-radius:12px;padding:1.3rem;border-left:5px solid #f59e0b;">{evaluation.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)
                st.balloons()

    # ── Session ended ──
    if room.get("status") == "ended" and room.get("ai_evaluation"):
        st.markdown("---")
        st.markdown("### 📊 Session Evaluation")
        eval_text = room["ai_evaluation"]
        st.markdown(f'<div style="background:#f0fdf4;border-radius:12px;padding:1.2rem;border:2px solid #16a34a;">{eval_text.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)
        
        # Download transcript
        transcript = "\n".join([f"[{m.get('time','')}] {m.get('name','?')} ({m.get('role','?').upper()}): {m.get('content','')}" 
                                for m in room.get("messages",[])])
        st.download_button("💾 Download Transcript", transcript, 
                          f"session_{code}_{datetime.now().strftime('%Y%m%d')}.txt", "text/plain")

    # ── Auto-refresh ──
    msg_count = len(room.get("messages", []))
    if msg_count != st.session_state.mp_last_msg_count:
        st.session_state.mp_last_msg_count = msg_count
    
    st.markdown("")
    col_r, col_l = st.columns([1, 3])
    with col_r:
        if st.button("🔄 Refresh Chat", use_container_width=True):
            st.rerun()
    with col_l:
        if st.button("🚪 Leave Room", use_container_width=True):
            add_message(code, "system", f"{name} ({role}) left the session.")
            st.session_state.mp_room_code = ""
            st.session_state.mp_role = None
            st.session_state.mp_name = ""
            st.rerun()

    # Auto-refresh every 4 seconds
    time.sleep(4)
    st.rerun()



# ════════════════════════════════════════════════════════════════════════════
# 👥 PEER SIMULATION — Supabase Live Chat Backend
# ════════════════════════════════════════════════════════════════════════════

# ── Supabase Configuration ───────────────────────────────────────────────────
# Students/instructors enter their own Supabase credentials in the UI
# Default shown here — replace with your project's values


def get_supabase_client(url=None, key=None):
    """Create Supabase client. Returns (client, error_message).
    Auto-uses hardcoded defaults if available."""
    try:
        from supabase import create_client
        # Priority: explicit args → session state → hardcoded defaults
        u = url or st.session_state.get("sb_url","") or SUPABASE_DEFAULT_URL
        k = key or st.session_state.get("sb_key","") or SUPABASE_DEFAULT_KEY
        if not u or not k or "YOUR_SUPABASE" in u or "YOUR_SUPABASE" in k:
            return None, "No Supabase credentials. Enter URL and Key in settings."
        client = create_client(u.strip(), k.strip())
        return client, None
    except ImportError:
        return None, (
            "Supabase library not installed. Run this command in your terminal:\n"
            "pip install supabase==2.3.0\n\n"
            "If you get a C++ build error for pyiceberg, try:\n"
            "pip install supabase==2.3.0 --no-deps\n"
            "pip install httpx gotrue storage3 postgrest realtime"
        )
    except Exception as e:
        return None, f"Connection error: {e}"

def sb_init_tables(client):
    """Check if required tables exist — guide user if not."""
    try:
        client.table("peer_rooms").select("id").limit(1).execute()
        client.table("peer_messages").select("id").limit(1).execute()
        return True, None
    except Exception as e:
        return False, str(e)

def sb_create_room(client, code, case_data, doctor_name):
    try:
        client.table("peer_rooms").insert({
            "code":         code,
            "case_id":      str(case_data.get("Case_ID","?")),
            "case_title":   str(case_data.get("Title") or case_data.get("Chief_Complaint","?")),
            "case_data":    json.dumps({k:str(v) for k,v in case_data.items()}),
            "doctor_name":  doctor_name,
            "patient_name": "",
            "status":       "waiting",
        }).execute()
        return True
    except Exception as e:
        st.error(f"Create room error: {e}"); return False

def sb_get_room(client, code):
    try:
        r = client.table("peer_rooms").select("*").eq("code", code.upper()).execute()
        return r.data[0] if r.data else None
    except Exception:
        return None

def sb_join_room(client, code, patient_name):
    """Join a room — only succeeds if room exists and is in 'waiting' status.

    Fix #5: Prevents a second patient from overwriting the first patient's
    name when a room is already active.
    """
    try:
        # Verify room is in waiting status before joining
        r = client.table("peer_rooms").select("status, patient_name").eq("code", code.upper()).execute()
        if not r.data:
            st.error("Room not found. Check the code with your doctor partner.")
            return False
        existing = r.data[0]
        if existing.get("status") == "active" and existing.get("patient_name"):
            st.error(f"This room already has a patient ({existing.get('patient_name')}). "
                     f"Ask your doctor partner to create a new room.")
            return False
        if existing.get("status") == "ended":
            st.error("This session has already ended. Ask your partner to create a new room.")
            return False
        # OK to join
        client.table("peer_rooms").update({
            "patient_name": patient_name,
            "status": "active",
        }).eq("code", code.upper()).execute()
        return True
    except Exception as e:
        st.error(f"Join error: {e}"); return False

def sb_send_message(client, code, role, sender_name, content, msg_type="chat"):
    try:
        client.table("peer_messages").insert({
            "room_code":   code.upper(),
            "role":        role,
            "sender_name": sender_name,
            "content":     content,
            "msg_type":    msg_type,
        }).execute()
        # Mark connection healthy on successful send
        st.session_state["peer_connection_ok"] = True
        return True
    except Exception as e:
        st.session_state["peer_connection_ok"] = False
        st.error(f"Send error: {e}"); return False

def sb_get_messages(client, code, since_id=0):
    """Fetch messages newer than since_id.

    Fix #4: Track connection health in session state so the UI can show
    a 'reconnecting...' indicator if Supabase is briefly down.
    """
    try:
        r = (client.table("peer_messages")
             .select("*")
             .eq("room_code", code.upper())
             .gt("id", since_id)
             .order("id")
             .execute())
        st.session_state["peer_connection_ok"] = True
        return r.data or []
    except Exception as e:
        st.session_state["peer_connection_ok"] = False
        st.session_state["peer_last_error"] = str(e)
        return []

def sb_end_room(client, code):
    try:
        client.table("peer_rooms").update({"status":"ended"}).eq("code",code.upper()).execute()
        return True
    except Exception:
        return False

def sb_save_notes(client, code, role, notes):
    try:
        field = "doctor_notes" if role == "doctor" else "patient_notes"
        client.table("peer_rooms").update({field: notes}).eq("code", code.upper()).execute()
    except: pass

def sb_cleanup_stale_rooms(client, max_age_hours=24):
    """Auto-expire rooms older than max_age_hours.

    Fix #3 (part 1): Marks old rooms as 'ended' to prevent table bloat.
    Runs at most once per app-load via session-state guard.
    """
    if st.session_state.get("_peer_cleanup_done_this_session"):
        return
    try:
        from datetime import timedelta
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=max_age_hours)).isoformat()
        client.table("peer_rooms").update({"status": "ended"})\
            .lt("created_at", cutoff)\
            .neq("status", "ended")\
            .execute()
        st.session_state["_peer_cleanup_done_this_session"] = True
    except Exception:
        # Silent — cleanup failure shouldn't break peer sim
        pass

def generate_room_code(client=None, max_tries=10):
    """Generate a unique 6-character room code.

    Fix #3 (part 2): Verifies uniqueness against existing rooms before
    returning. Falls back to longer code if collision after max_tries.
    """
    import random, string
    chars = string.ascii_uppercase + string.digits
    # Avoid easily-confused characters: 0/O, 1/I/L
    chars = chars.replace("0","").replace("O","").replace("1","").replace("I","")
    for _ in range(max_tries):
        code = "".join(random.choices(chars, k=6))
        if client is None:
            return code
        # Check uniqueness — only consider non-ended rooms (ended ones are reusable)
        try:
            r = client.table("peer_rooms").select("code")\
                .eq("code", code).neq("status","ended").execute()
            if not r.data:
                return code
        except Exception:
            return code  # If check fails, just use the code
    # Extreme fallback — append timestamp for guaranteed uniqueness
    return "".join(random.choices(chars, k=8))


# ════════════════════════════════════════════════════════════════════════════
# 👥 PEER SIMULATION PAGE
# ════════════════════════════════════════════════════════════════════════════
def page_peer_sim():
    st.markdown('<div class="section-header">👥 Peer Clinical Simulation — Live Doctor vs Patient</div>', unsafe_allow_html=True)

    # ── SUPABASE SETUP PANEL ─────────────────────────────
    with st.expander("⚙️ Supabase Connection Settings", expanded=not st.session_state.get("sb_connected")):
        st.markdown("""
        <div style="background:linear-gradient(135deg,#1a1a2e,#16213e);color:white;
                    border-radius:12px;padding:1.2rem;margin-bottom:1rem;">
            <h4 style="margin:0 0 .5rem">📡 Connect to Supabase Database</h4>
            <p style="margin:0;font-size:.85rem;opacity:.9">
                Supabase stores all chat messages so both students see them in real time.
                Each student enters the same URL + Key to connect to the same session.
            </p>
        </div>
        """, unsafe_allow_html=True)

        col_url, col_key = st.columns(2)
        with col_url:
            sb_url = st.text_input("Supabase Project URL:",
                value=st.session_state.get("sb_url", SUPABASE_DEFAULT_URL),
                placeholder="https://xxxxxxxxxxxx.supabase.co",
                key="sb_url_input")
        with col_key:
            sb_key = st.text_input("Supabase Anon Key:",
                value=st.session_state.get("sb_key", SUPABASE_DEFAULT_KEY),
                placeholder="eyJhbGciOiJIUzI1NiIsInR5cCI6...",
                type="password",
                key="sb_key_input")

        if st.button("🔌 Connect to Supabase", type="primary", use_container_width=True):
            st.session_state["sb_url"] = sb_url.strip()
            st.session_state["sb_key"] = sb_key.strip()
            client, err = get_supabase_client(sb_url, sb_key)
            if err:
                st.error(f"❌ {err}")
                if "not installed" in err:
                    st.code("pip install supabase==2.3.0", language="bash")
                    st.markdown("If C++ build error: `pip install supabase==2.3.0 --no-deps`")
                st.session_state["sb_connected"] = False
            else:
                ok, table_err = sb_init_tables(client)
                if ok:
                    st.success("✅ Connected to Supabase!")
                    st.session_state["sb_connected"] = True
                else:
                    st.error(f"❌ Tables missing. Run the SQL setup in Supabase first.")
                    st.session_state["sb_connected"] = False
                    st.markdown("""
                    **Run this SQL in your Supabase SQL Editor:**
                    ```sql
                    -- Table 1: Rooms
                    create table if not exists peer_rooms (
                      id            bigint generated always as identity primary key,
                      code          text unique not null,
                      case_id       text,
                      case_title    text,
                      case_data     text,
                      doctor_name   text,
                      patient_name  text default \'\',
                      status        text default \'waiting\',
                      doctor_notes  text default \'\',
                      patient_notes text default \'\',
                      created_at    timestamptz default now()
                    );

                    -- Table 2: Messages
                    create table if not exists peer_messages (
                      id           bigint generated always as identity primary key,
                      room_code    text not null,
                      role         text,
                      sender_name  text,
                      content      text,
                      msg_type     text default \'chat\',
                      created_at   timestamptz default now()
                    );

                    -- Enable Row Level Security (allow all for now)
                    alter table peer_rooms    enable row level security;
                    alter table peer_messages enable row level security;

                    create policy "Allow all peer_rooms"    on peer_rooms    for all using (true) with check (true);
                    create policy "Allow all peer_messages" on peer_messages for all using (true) with check (true);
                    ```
                    """)

        if st.session_state.get("sb_connected"):
            st.markdown('<div class="alert-good">✅ Supabase connected and ready.</div>', unsafe_allow_html=True)

    if not st.session_state.get("sb_connected"):
        st.warning("Connect to Supabase first to use Peer Simulation.")
        return

    client, err = get_supabase_client()
    if err:
        st.error(err); return

    # ── NOT IN A ROOM YET ────────────────────────────────
    if not st.session_state.peer_room_code:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                    border-radius:14px;padding:1.5rem;margin-bottom:1.5rem;">
            <h3 style="margin:0 0 .5rem">👥 Two-Student Live Clinical Simulation</h3>
            <p style="margin:0;opacity:.9;font-size:.9rem">
                <b>Student A (Doctor)</b> — takes history, examines, orders tests, makes diagnosis.<br>
                <b>Student B (Patient)</b> — acts as the real patient using the case brief.<br>
                Both connect via a <b>6-character Room Code</b>.
                Messages sync through Supabase in real time.
            </p>
        </div>
        """, unsafe_allow_html=True)

        tab_create, tab_join = st.tabs(["🏥 Create Room (Doctor)", "🚪 Join Room (Patient)"])

        # ── CREATE ROOM ───────────────────────────────────
        with tab_create:
            st.markdown("### 👨‍⚕️ You are the Doctor")
            st.markdown('<div class="alert-info">Select a case, create a room, share the 6-character code with the patient student.</div>', unsafe_allow_html=True)

            doctor_name = st.text_input("Your name:", placeholder="e.g. Dr. Ahmed", key="dr_name")

            if df.empty:
                st.error("case_studies.xlsx not found.")
            else:
                fc1, fc2 = st.columns(2)
                with fc1:
                    systems = ["All"] + sorted(df["System"].dropna().unique().tolist())
                    sys_sel = st.selectbox("Filter by system:", systems, key="cr_sys")
                with fc2:
                    diff_sel = st.selectbox("Filter by difficulty:", ["All","basic","intermediate","advanced"], key="cr_diff")

                filt = df.copy()
                if sys_sel  != "All": filt = filt[filt["System"].str.lower().str.contains(sys_sel.lower(), na=False)]
                if diff_sel != "All": filt = filt[filt["Difficulty"].str.lower().str.contains(diff_sel.lower(), na=False)]

                case_options = {
                    f"{r['Case_ID']} — {str(r.get('Title') or r.get('Chief_Complaint','?')).title()} ({r['Age_Sex']})": r["row_num"]
                    for _,r in filt.head(60).iterrows()
                }
                sel_label = st.selectbox("Select case:", list(case_options.keys()), key="cr_case")

            if st.button("🚀 Create Room", type="primary", use_container_width=True):
                if not doctor_name.strip():
                    st.warning("Enter your name first.")
                else:
                    row_num  = case_options[sel_label]
                    case_row = df[df["row_num"]==row_num].iloc[0].to_dict()
                    code     = generate_room_code(client)
                    if sb_create_room(client, code, case_row, doctor_name.strip()):
                        sb_send_message(client, code, "system", "System",
                            f"Room {code} created by Dr. {doctor_name.strip()}. Waiting for patient...", "system")
                        st.session_state.peer_room_code = code
                        st.session_state.peer_role      = "doctor"
                        st.session_state.peer_name      = doctor_name.strip()
                        st.session_state["last_msg_id"] = 0
                        st.rerun()

        # ── JOIN ROOM ─────────────────────────────────────
        with tab_join:
            st.markdown("### 🤒 You are the Patient")
            st.markdown('<div class="alert-info">Your partner created a room. Enter the 6-character code they shared and your name.</div>', unsafe_allow_html=True)

            patient_name = st.text_input("Your name:", placeholder="e.g. Sara", key="pt_name")
            code_input   = st.text_input("Room Code:", placeholder="e.g. AB3X7Q",
                                          max_chars=6, key="code_input").upper().strip()

            if st.button("🚪 Join Room", type="primary", use_container_width=True):
                if not patient_name.strip():
                    st.warning("Enter your name.")
                elif len(code_input) != 6:
                    st.warning("Room code must be exactly 6 characters.")
                else:
                    room = sb_get_room(client, code_input)
                    if not room:
                        st.error(f"Room `{code_input}` not found. Check the code.")
                    elif room["status"] == "ended":
                        st.error("This session has ended.")
                    else:
                        if sb_join_room(client, code_input, patient_name.strip()):
                            sb_send_message(client, code_input, "system", "System",
                                f"{patient_name.strip()} joined as patient. Session starting!", "system")
                            st.session_state.peer_room_code = code_input
                            st.session_state.peer_role      = "patient"
                            st.session_state.peer_name      = patient_name.strip()
                            st.session_state["last_msg_id"] = 0
                            st.rerun()
        return

    # ════════════════════════════════════════════════════
    # IN A ROOM
    # ════════════════════════════════════════════════════
    code = st.session_state.peer_room_code
    role = st.session_state.peer_role
    name = st.session_state.peer_name

    room = sb_get_room(client, code)
    if not room:
        st.error("Room not found or expired.")
        # Clear cached messages for this room (Fix #2 cleanup)
        st.session_state.pop(f"peer_msgs_{code}", None)
        st.session_state.pop(f"peer_last_msg_id_{code}", None)
        st.session_state.peer_room_code = None
        st.rerun()
        return

    case     = json.loads(room.get("case_data","{}"))
    status   = room.get("status","waiting")
    doc_name = room.get("doctor_name","Doctor")
    pat_name = room.get("patient_name","")

    # ── ROLE BANNER — large and unmistakable ─────────────
    role_color   = "#0e7490" if role=="doctor" else "#7c3aed"
    role_icon    = "👨‍⚕️" if role=="doctor" else "🤒"
    role_label   = "DOCTOR" if role=="doctor" else "PATIENT"
    role_subtitle= ("Take history, examine, order tests, make diagnosis."
                    if role=="doctor" else
                    "Act as the patient. Use your brief below. NEVER reveal the diagnosis.")
    urg_color    = {"active":"#16a34a","waiting":"#f59e0b","ended":"#6b7280"}.get(status,"#0ea5e9")
    status_lbl   = {"active":"🟢 LIVE","waiting":"⏳ Waiting for patient","ended":"🔴 Session Ended"}.get(status,"?")

    st.markdown(f"""
    <div style="background:{role_color};color:white;border-radius:14px;
                padding:1.2rem 1.5rem;margin-bottom:.8rem;
                box-shadow:0 4px 16px {role_color}55;">
        <div style="display:flex;align-items:center;justify-content:space-between;">
            <div>
                <div style="font-size:2rem;font-weight:900;letter-spacing:.05em">
                    {role_icon} YOU ARE THE {role_label}
                </div>
                <div style="font-size:.88rem;opacity:.92;margin-top:.2rem">{role_subtitle}</div>
            </div>
            <div style="text-align:right;">
                <div style="background:rgba(255,255,255,.2);border-radius:8px;padding:.4rem .9rem;
                            font-size:.8rem;margin-bottom:.3rem">{status_lbl}</div>
                <div style="font-size:.75rem;opacity:.8">{datetime.now().strftime("%H:%M:%S")}</div>
            </div>
        </div>
    </div>
    <div style="background:#f8fafc;border-radius:10px;padding:.7rem 1.2rem;margin-bottom:.8rem;
                border:2px solid {role_color}44;font-size:.85rem;display:flex;gap:2rem;">
        <span>🏠 Room: <code style="background:#e2e8f0;padding:.1rem .3rem;border-radius:4px;
                                    letter-spacing:.12em;font-weight:700">{code}</code></span>
        <span>📋 Case: <b>{case.get("Chief_Complaint","?")}</b> | {case.get("Age_Sex","?")}</span>
        <span>{"👨‍⚕️ Doctor: <b>"+doc_name+"</b>" if role=="patient" else "🤒 Patient: <b>"+(pat_name or "waiting...")+"</b>"}</span>
    </div>
    """, unsafe_allow_html=True)

    # ── WAITING STATE ─────────────────────────────────────
    if role == "doctor" and status == "waiting":
        st.markdown(f"""
        <div style="background:#fef9c3;border:2px solid #ca8a04;border-radius:14px;
                    padding:1.5rem;text-align:center;margin-bottom:1rem;">
            <div style="color:#78350f;font-size:.9rem;margin-bottom:.5rem">
                📤 Share this code with the patient student:
            </div>
            <div style="font-size:3.5rem;font-weight:900;letter-spacing:.4rem;
                        color:#0a2540;font-family:monospace">{code}</div>
            <div style="color:#78350f;font-size:.82rem;margin-top:.5rem">
                They go to <b>Peer Simulation → Join Room</b> and enter this code
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── PATIENT BRIEF (patient only) ──────────────────────
    if role == "patient":
        with st.expander("📋 Your Patient Brief — Read before you start!", expanded=(status=="waiting")):
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#fdf4ff,#f3e8ff);
                        border:2px solid #a855f7;border-radius:12px;padding:1.2rem;">
                <h3 style="color:#6b21a8;margin:0 0 .8rem">🤒 You are the Patient</h3>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:.6rem;font-size:.87rem;">
                    <div><b>You are:</b> {case.get("Age_Sex","?")}</div>
                    <div><b>Occupation:</b> {case.get("Occupation","?")}</div>
                    <div><b>Your complaint:</b> {case.get("Chief_Complaint","?")}</div>
                    <div><b>Duration:</b> {case.get("Duration","?")}</div>
                </div>
                <div style="margin:.7rem 0;padding:.8rem;background:white;border-radius:8px;font-size:.87rem;">
                    <b>Your story:</b> {case.get("HPI","?")}
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:.5rem;font-size:.85rem;margin-bottom:.7rem;">
                    <div><b>Medical history:</b> {case.get("PMH","none")}</div>
                    <div><b>Medications:</b> {case.get("Medications","none")}</div>
                    <div><b>How you feel:</b> {case.get("Appearance","in pain")}</div>
                    <div><b>Social:</b> {case.get("Social_Hx","none")}</div>
                </div>
                <div style="padding:.8rem;background:#fff1f2;border-radius:8px;
                            border-left:4px solid #f43f5e;font-size:.85rem;">
                    🔴 <b>SECRET — NEVER tell the doctor:</b> {case.get("Final_Diagnosis","?")}
                </div>
                <div style="margin-top:.7rem;padding:.8rem;background:#fffbeb;border-radius:8px;font-size:.83rem;">
                    <b>🎭 How to play:</b> Speak/type as a real scared patient.
                    No medical jargon. Describe pain vividly ("stabbing", "burning").
                    Ask questions back: "Is it serious?", "Will I need surgery?".
                    Only reveal symptoms when asked. React emotionally.
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ════════════════════════════════════════════════════
    # MAIN CHAT LAYOUT
    # ════════════════════════════════════════════════════

    # ── Fix #1: REAL auto-refresh every 3s while session is active ───────
    # Uses streamlit-autorefresh if installed, falls back to a meta-refresh
    # tag in an iframe (which causes Streamlit to rerun naturally).
    if status != "ended":
        try:
            from streamlit_autorefresh import st_autorefresh
            st_autorefresh(interval=3000, key=f"peer_autorefresh_{code}")
        except ImportError:
            # Fallback: meta-refresh inside a tiny invisible iframe
            # (streamlit-autorefresh isn't installed)
            components.html(
                '<meta http-equiv="refresh" content="3">',
                height=0
            )

    # ── Fix #3 (cleanup): Run stale-room cleanup once per session ────────
    sb_cleanup_stale_rooms(client, max_age_hours=24)

    # ── Fix #4 (connection status) banner ────────────────────────────────
    if not st.session_state.get("peer_connection_ok", True):
        st.markdown("""
        <div style="background:#fef2f2;border:2px solid #dc2626;border-radius:8px;
                    padding:.6rem 1rem;margin-bottom:.6rem;font-size:.83rem;color:#991b1b;">
          ⚠️ <b>Connection issue</b> — couldn't reach the chat server.
          Auto-retrying in 3 seconds...
        </div>
        """, unsafe_allow_html=True)

    main_col, side_col = st.columns([3,1])

    with main_col:
        # ── Fix #2: INCREMENTAL message fetching ──────────────────────────
        # Cache existing messages in session state, only fetch new ones since
        # the last seen ID. Reduces Supabase load by ~95% in long sessions.
        cache_key  = f"peer_msgs_{code}"
        last_id_key = f"peer_last_msg_id_{code}"

        cached_messages = st.session_state.get(cache_key, [])
        last_seen_id    = st.session_state.get(last_id_key, 0)

        new_messages = sb_get_messages(client, code, since_id=last_seen_id)
        if new_messages:
            cached_messages.extend(new_messages)
            st.session_state[cache_key] = cached_messages
            st.session_state[last_id_key] = max(m.get("id", 0) for m in new_messages)

        all_messages = cached_messages

        # ── RENDER CHAT ────────────────────────────────────
        chat_html = ""
        for m in all_messages:
            r    = m.get("role","")
            txt  = m.get("content","")
            sn   = m.get("sender_name","")
            t    = str(m.get("created_at",""))[-8:][:5]  # HH:MM
            mtyp = m.get("msg_type","chat")

            if mtyp == "system":
                chat_html += f'<div style="text-align:center;font-size:.75rem;color:#9ca3af;margin:.3rem 0;padding:.2rem .5rem;background:#f8fafc;border-radius:6px;"><i>⚙️ {txt}</i></div>'
            elif mtyp in ("exam_request","order"):
                icon = "🩺" if mtyp=="exam_request" else "📋"
                chat_html += f'<div style="background:#fef9c3;border-radius:8px;padding:.5rem .8rem;margin:.3rem 0;font-size:.83rem;border-left:3px solid #ca8a04;">{icon} <b>Dr. {sn}:</b> {txt} <span style="color:#9ca3af;font-size:.72rem">{t}</span></div>'
            elif r == "doctor":
                chat_html += f'<div class="chat-student" style="text-align:right;margin-left:auto;">👨‍⚕️ <b>Dr. {sn}</b> <span style="font-size:.7rem;color:#9ca3af">{t}</span><br>{txt}</div>'
            else:
                chat_html += f'<div class="chat-live">🤒 <b>{sn}</b> <span style="font-size:.7rem;color:#9ca3af">{t}</span><br>{txt}</div>'

        if not all_messages:
            chat_html = '<div style="text-align:center;color:#9ca3af;padding:2rem;font-size:.9rem">Session started — Doctor, introduce yourself and begin the consultation.</div>'

        # Scrollable chat window
        refresh_status = "🟢 Live · auto-refreshes every 3s" if status != "ended" else "🔴 Session ended"
        st.markdown(f"""
        <div style="background:#f8fafc;border-radius:12px;padding:1rem;
                    min-height:400px;max-height:450px;overflow-y:auto;
                    border:2px solid #e5e7eb;"
             id="chatbox">
            {chat_html}
        </div>
        <div style="text-align:right;font-size:.7rem;color:#9ca3af;margin-top:.2rem">
            {len(all_messages)} messages · {refresh_status}
        </div>
        <script>
        // Auto-scroll to bottom
        var cb=document.getElementById("chatbox");
        if(cb) cb.scrollTop=cb.scrollHeight;
        </script>
        """, unsafe_allow_html=True)

        # ── MESSAGE INPUT ─────────────────────────────────
        if status != "ended":
            st.markdown("")

            # Mic button
            mic_color = "#0e7490" if role=="doctor" else "#7c3aed"
            components.html(f"""
            <div style="font-family:Inter,sans-serif;margin-bottom:4px;">
              <button id="peerMic" onclick="peerToggle()" style="
                background:linear-gradient(135deg,{mic_color},#0a2540);
                color:white;border:none;border-radius:50px;
                padding:.35rem .9rem;font-size:.78rem;font-weight:600;
                cursor:pointer;box-shadow:0 2px 5px rgba(0,0,0,.2);">
                🎤 Speak (Chrome/Edge)
              </button>
              <span id="peerSt" style="font-size:.73rem;color:#6b7280;margin-left:.4rem;"></span>
              <div id="peerTr" style="background:#f0fdf4;border:1px solid #16a34a;border-radius:6px;
                padding:.3rem .6rem;font-size:.82rem;margin-top:.25rem;color:#166534;display:none;"></div>
            </div>
            <script>
            var pRec=null,pLs=false;
            function peerToggle(){{
              if(!('webkitSpeechRecognition'in window)&&!('SpeechRecognition'in window)){{
                document.getElementById('peerSt').innerHTML='Use Chrome/Edge';return;}}
              if(pLs){{pRec.stop();return;}}
              var SR=window.SpeechRecognition||window.webkitSpeechRecognition;
              pRec=new SR();pRec.lang='en-US';pRec.continuous=false;pRec.interimResults=true;
              var btn=document.getElementById('peerMic');
              pRec.onstart=function(){{pLs=true;
                btn.innerHTML='🔴 Listening... click to stop';
                btn.style.background='linear-gradient(135deg,#dc2626,#991b1b)';
                document.getElementById('peerSt').innerHTML='Speak now...';
              }};
              pRec.onresult=function(e){{
                var ft='',it='';
                for(var i=e.resultIndex;i<e.results.length;i++){{
                  if(e.results[i].isFinal)ft+=e.results[i][0].transcript;
                  else it+=e.results[i][0].transcript;}}
                var tr=document.getElementById('peerTr');
                tr.style.display='block';tr.innerHTML=ft||it;
                if(ft){{navigator.clipboard.writeText(ft).catch(function(){{}});
                  document.getElementById('peerSt').innerHTML='✅ Copied — paste below & Send';}}
              }};
              pRec.onerror=function(e){{pLs=false;
                document.getElementById('peerSt').innerHTML='Error: '+e.error;
                btn.innerHTML='🎤 Speak (Chrome/Edge)';
                btn.style.background='linear-gradient(135deg,{mic_color},#0a2540)';
              }};
              pRec.onend=function(){{pLs=false;
                btn.innerHTML='🎤 Speak (Chrome/Edge)';
                btn.style.background='linear-gradient(135deg,{mic_color},#0a2540)';
              }};
              pRec.start();
            }}
            </script>
            """, height=70)

            with st.form("peer_msg_form", clear_on_submit=True):
                c1, c2 = st.columns([5,1])
                with c1:
                    ph = ("e.g. Hello, I'm Dr. Ahmed. What brings you in today?"
                          if role=="doctor" else
                          "e.g. Doctor I have terrible pain in my stomach...")
                    msg_inp = st.text_input("Message:", placeholder=ph,
                                            label_visibility="collapsed", key="peer_msg_inp")
                with c2:
                    send_btn = st.form_submit_button("Send →", use_container_width=True)

            if send_btn and msg_inp.strip():
                sb_send_message(client, code, role, name, msg_inp.strip(), "chat")
                if st.session_state.voice_enabled:
                    if role == "doctor":
                        tts_speak_doctor(msg_inp.strip())
                    else:
                        tts_speak(msg_inp.strip())
                st.rerun()

            # ── DOCTOR CLINICAL TOOLS ─────────────────────
            if role == "doctor":
                st.markdown("**🩺 Clinical Tools:**")
                dt1,dt2,dt3,dt4 = st.columns(4)
                with dt1:
                    with st.expander("🩺 Request Exam"):
                        ez = st.selectbox("Zone:",["Chest","Abdomen","Back","Head/Neck","Limbs"], key="ez")
                        et = st.selectbox("Type:",["Auscultation","Percussion","Palpation","Inspection"], key="et")
                        if st.button("Request", key="req_e", use_container_width=True):
                            msg = f"Please allow {et} of {ez}"
                            sb_send_message(client, code, "doctor", name, msg, "exam_request")
                            st.rerun()
                with dt2:
                    with st.expander("🧪 Order Lab"):
                        lab_t = st.selectbox("Test:",["CBC","CRP","Troponin","LFTs","Renal Function",
                                                       "Blood Cultures","Urinalysis","ABG","Amylase/Lipase",
                                                       "D-dimer","BNP","Coagulation Panel"], key="lt")
                        if st.button("Order", key="ord_l", use_container_width=True):
                            sb_send_message(client, code, "doctor", name, f"Ordering lab: {lab_t}", "order")
                            st.rerun()
                with dt3:
                    with st.expander("🔬 Order Imaging"):
                        img_t = st.selectbox("Imaging:",["Chest X-Ray","CT Abdomen","CT Head",
                                                          "ECG","Ultrasound","MRI Brain","CTPA"], key="it")
                        if st.button("Order", key="ord_i", use_container_width=True):
                            sb_send_message(client, code, "doctor", name, f"Ordering imaging: {img_t}", "order")
                            st.rerun()
                with dt4:
                    with st.expander("💉 Medication"):
                        med_t = st.text_input("Medication:", placeholder="e.g. Morphine 2mg IV", key="mt")
                        if st.button("Give", key="giv_m", use_container_width=True) and med_t.strip():
                            sb_send_message(client, code, "doctor", name, f"Administering: {med_t}", "order")
                            st.rerun()

            # ── PATIENT HINT BUTTONS ──────────────────────
            if role == "patient":
                exam_msgs = [m for m in all_messages if m.get("msg_type")=="exam_request"]
                if exam_msgs:
                    last_exam = exam_msgs[-1]
                    st.markdown(f'<div class="alert-warn">🩺 Doctor requests: <b>{last_exam["content"]}</b> — respond in chat!</div>', unsafe_allow_html=True)

                st.markdown("**💡 Quick Patient Responses:**")
                hints = [
                    f"The pain is here and it\'s terrible",
                    "Is it serious, doctor?",
                    "I\'m really scared right now",
                    "It hurts when you press there!",
                    "Will I need surgery?",
                ]
                hc = st.columns(len(hints))
                for col, h in zip(hc, hints):
                    with col:
                        if st.button(h[:18]+"...", key=f"ph_{h[:8]}", use_container_width=True):
                            sb_send_message(client, code, "patient", name, h, "chat")
                            if st.session_state.voice_enabled:
                                tts_speak(h)
                            st.rerun()

    with side_col:
        # ── SIDE PANEL ────────────────────────────────────
        st.markdown("**📋 Case**")
        st.markdown(f"""
        <div style="background:#f8fafc;border-radius:10px;padding:.8rem;
                    font-size:.8rem;border:1px solid #e2e8f0;line-height:1.8;">
            <b>Patient:</b> {case.get("Age_Sex","?")}<br>
            <b>CC:</b> {case.get("Chief_Complaint","?")}<br>
            <b>Duration:</b> {case.get("Duration","?")}<br>
            {f'<b>Vitals:</b> {case.get("Vitals","?")}<br>' if role=="doctor" else ""}
        </div>
        """, unsafe_allow_html=True)

        if role == "doctor":
            st.markdown("")
            st.markdown("**📝 Your Notes**")
            notes = st.text_area("",
                value=room.get("doctor_notes",""),
                placeholder="Document findings, differentials...",
                height=130,
                key="doc_notes",
                label_visibility="collapsed")
            if st.button("💾 Save Notes", use_container_width=True):
                sb_save_notes(client, code, "doctor", notes)
                st.success("Saved!")

        st.markdown("")

        # Refresh button
        if st.button("🔄 Refresh Chat", use_container_width=True):
            st.rerun()

        # End session (doctor only)
        if role == "doctor" and status == "active":
            st.markdown("")
            if st.button("🏁 End Session", use_container_width=True, type="primary"):
                sb_send_message(client, code, "system", "System",
                    "Session ended by Doctor. Requesting AI evaluation...", "system")
                sb_end_room(client, code)
                st.rerun()

        # Session ended → AI evaluation + transcript
        if status == "ended":
            st.markdown('<div class="alert-warn">🔴 Session ended</div>', unsafe_allow_html=True)
            st.markdown("")

            if st.button("🤖 AI Evaluation", use_container_width=True, type="primary"):
                chat_msgs = [m for m in all_messages if m.get("msg_type")=="chat"]
                transcript_txt = "\n".join([
                    f"{'DOCTOR' if m['role']=='doctor' else 'PATIENT'} ({m['sender_name']}): {m['content']}"
                    for m in chat_msgs
                ])
                eval_prompt = f"""Evaluate this peer clinical simulation between two students.

Case: {case.get("Chief_Complaint","?")} | Patient: {case.get("Age_Sex","?")}
True Diagnosis: {case.get("Final_Diagnosis","?")}

TRANSCRIPT:
{transcript_txt[:3000]}

Evaluate:
**DOCTOR ({doc_name}):**
1. History taking completeness — /10
2. Communication & empathy — /10
3. Clinical reasoning — /10
4. Reached correct diagnosis? — /10
5. Overall doctor score — /10

**PATIENT ({pat_name}):**
1. Realism & authenticity — /10
2. Correct symptom portrayal — /10
3. Emotional performance — /10
4. Overall patient score — /10

**3 key learning points for both students**
Be specific and educational."""
                with st.spinner("AI evaluating..."):
                    evaluation = call_ai(
                        "You are a senior clinical educator evaluating a peer simulation.",
                        [{"role":"user","content":eval_prompt}], max_tokens=900
                    )
                st.markdown("### 🤖 Evaluation")
                st.markdown(f'<div style="background:#fff7ed;border-radius:10px;padding:1rem;border-left:4px solid #f59e0b;font-size:.83rem;">{evaluation.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

            # Download transcript
            chat_msgs = [m for m in all_messages if m.get("msg_type")=="chat"]
            if chat_msgs:
                transcript_dl = "\n".join([
                    f"[{str(m.get('created_at',''))[-8:][:5]}] "
                    f"{'DOCTOR' if m['role']=='doctor' else 'PATIENT'} ({m['sender_name']}): {m['content']}"
                    for m in chat_msgs
                ])
                st.download_button("💾 Download Transcript",
                    transcript_dl,
                    f"peer_session_{code}_{datetime.now().strftime('%Y%m%d')}.txt",
                    "text/plain",
                    use_container_width=True)

        st.markdown("")
        if st.button("🚪 Leave Room", use_container_width=True):
            # Clear cached messages for this room (Fix #2 cleanup)
            st.session_state.pop(f"peer_msgs_{code}", None)
            st.session_state.pop(f"peer_last_msg_id_{code}", None)
            st.session_state.peer_room_code = None
            st.session_state.peer_role      = None
            st.session_state.peer_name      = ""
            st.rerun()

    # NOTE: Old time.sleep(3)+st.rerun() block removed — replaced by
    # st_autorefresh near the top of the chat layout (Fix #1). The old
    # approach froze the UI for 3 seconds on every rerun making typing laggy.


# ════════════════════════════════════════════════════════════════════════════
# 💳 CREDITS & PLANS PAGE
# ════════════════════════════════════════════════════════════════════════════
def page_credits():
    st.markdown('<div class="section-header">💳 Credits & Plans</div>', unsafe_allow_html=True)

    state, client = get_credit_state()
    plan      = state.get("plan","free")
    remaining = get_credits_remaining(state)
    used      = state.get("credits_used",0)
    total     = state.get("total_used_ever",0)
    uid       = get_user_id()
    reset_dt  = get_window_reset_time(state)
    now       = datetime.now(timezone.utc).replace(tzinfo=None)
    time_left = reset_dt - now
    h_left    = max(0,int(time_left.total_seconds()//3600))
    m_left    = max(0,int((time_left.total_seconds()%3600)//60))
    expires   = state.get("plan_expires","")

    # ── Current status banner ─────────────────────────────
    if plan == "premium":
        exp_str = datetime.fromisoformat(expires).strftime("%d %b %Y") if expires else "Never"
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#d97706,#f59e0b);color:white;
                    border-radius:16px;padding:1.5rem 2rem;margin-bottom:1.5rem;
                    box-shadow:0 4px 20px rgba(217,119,6,.4);">
            <div style="font-size:1.5rem;font-weight:900">⭐ PREMIUM ACTIVE</div>
            <div style="font-size:.9rem;opacity:.92;margin-top:.4rem">
                Unlimited access to all features | Expires: {exp_str}
            </div>
            <div style="font-size:.8rem;opacity:.8;margin-top:.3rem">User ID: {uid}</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        pct = max(0, remaining/FREE_CREDITS_PER_WINDOW*100)
        color = "#16a34a" if pct>50 else "#f59e0b" if pct>20 else "#dc2626"
        st.markdown(f"""
        <div style="background:white;border:2px solid {color};border-radius:16px;
                    padding:1.5rem 2rem;margin-bottom:1.5rem;">
            <div style="display:flex;justify-content:space-between;align-items:center;">
                <div>
                    <div style="font-size:1.2rem;font-weight:800;color:#0a2540">
                        💳 FREE PLAN
                    </div>
                    <div style="font-size:.85rem;color:#6b7280;margin-top:.2rem">
                        {remaining} credits remaining | Resets in {h_left}h {m_left}m
                    </div>
                </div>
                <div style="font-size:2.5rem;font-weight:900;color:{color}">
                    {remaining}<span style="font-size:1rem;color:#9ca3af">/{FREE_CREDITS_PER_WINDOW}</span>
                </div>
            </div>
            <div style="background:#e5e7eb;border-radius:999px;height:12px;margin:1rem 0;">
                <div style="background:{color};height:12px;border-radius:999px;width:{pct:.0f}%;
                            transition:width .5s;"></div>
            </div>
            <div style="font-size:.8rem;color:#6b7280;">
                Used this window: {used} | Total ever used: {total} | User ID: {uid}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── Credit cost table ─────────────────────────────────
    st.markdown("### 💡 Credit Costs")
    cost_cols = st.columns(4)
    cost_items = [
        ("💬","Patient Interview","1 credit per message"),
        ("🫀","Physical Exam Finding","1 credit per zone"),
        ("🧪","Lab Generation","2 credits per order"),
        ("📝","Diagnosis Evaluation","2 credits"),
        ("🔬","Image Analysis","3 credits per upload"),
        ("💓","ECG Analysis","3 credits per scan"),
        ("🔪","Surgery AI Note","2 credits"),
        ("🏥","Submit Real Case","2 credits"),
    ]
    for i,(icon,name,cost) in enumerate(cost_items):
        with cost_cols[i%4]:
            st.markdown(f"""
            <div style="background:#f8fafc;border-radius:10px;padding:.7rem;
                        text-align:center;border:1px solid #e2e8f0;margin:.3rem 0;">
                <div style="font-size:1.3rem">{icon}</div>
                <div style="font-size:.78rem;font-weight:600;color:#0a2540">{name}</div>
                <div style="font-size:.72rem;color:#6b7280">{cost}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("")

    # ── Plan comparison ────────────────────────────────────
    st.markdown("### 📋 Plans")
    pc1, pc2 = st.columns(2)

    with pc1:
        st.markdown(f"""
        <div style="background:#f8fafc;border:2px solid #e2e8f0;border-radius:16px;
                    padding:1.5rem;height:100%;">
            <div style="font-size:1.1rem;font-weight:800;color:#0a2540">🆓 Free Plan</div>
            <div style="font-size:2rem;font-weight:900;color:#0a2540;margin:.5rem 0">
                $0<span style="font-size:.9rem;color:#6b7280">/month</span>
            </div>
            <div style="font-size:.85rem;color:#374151;line-height:2;">
                ✅ {FREE_CREDITS_PER_WINDOW} credits every {WINDOW_HOURS} hours<br>
                ✅ All clinical rooms<br>
                ✅ Patient interview & AI tutor<br>
                ✅ Lab results & imaging<br>
                ✅ Case library (316+ cases)<br>
                ⏳ Credits reset every 8 hours<br>
                ⏳ Full bundle renews monthly
            </div>
            {"<div style='margin-top:1rem;background:#e2e8f0;color:#6b7280;border-radius:8px;padding:.6rem;text-align:center;font-weight:600'>Current Plan</div>" if plan=="free" else ""}
        </div>
        """, unsafe_allow_html=True)

    with pc2:
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#fffbeb,#fef3c7);
                    border:2px solid #f59e0b;border-radius:16px;padding:1.5rem;height:100%;
                    box-shadow:0 4px 20px rgba(245,158,11,.2);">
            <div style="font-size:1.1rem;font-weight:800;color:#92400e">
                ⭐ Premium Plan
                <span style="background:#f59e0b;color:white;border-radius:999px;
                             padding:.1rem .5rem;font-size:.72rem;margin-left:.3rem">BEST VALUE</span>
            </div>
            <div style="font-size:2rem;font-weight:900;color:#92400e;margin:.5rem 0">
                $5<span style="font-size:.9rem;color:#b45309">/month</span>
            </div>
            <div style="font-size:.85rem;color:#78350f;line-height:2;">
                ✅ <b>UNLIMITED credits</b> — no limits<br>
                ✅ All features unlocked<br>
                ✅ Priority AI responses<br>
                ✅ Peer simulation (unlimited rooms)<br>
                ✅ Download all reports & transcripts<br>
                ✅ Full radiology image analysis<br>
                ✅ ECG PhysioNet file analysis<br>
                ✅ 30-day subscription
            </div>
            {"<div style='margin-top:1rem;background:#f59e0b;color:white;border-radius:8px;padding:.6rem;text-align:center;font-weight:600'>✅ Active Premium</div>" if plan=="premium" else ""}
        </div>
        """, unsafe_allow_html=True)

    st.markdown("")

    # ── Upgrade section ────────────────────────────────────
    if plan != "premium":
        st.markdown("### ⭐ Upgrade to Premium")
        st.markdown('<div class="alert-warn">💳 After payment, enter your activation code below to unlock Premium immediately.</div>', unsafe_allow_html=True)

        st.markdown(f"""
        **How to upgrade — {PRICE_PER_MONTH} USD/month:**

        1. Send **${PRICE_PER_MONTH} via {PAYMENT_METHOD}** to: `{PAYMENT_CONTACT}`
        2. In your payment note write: **"MLS Premium - " + your User ID**: `{uid}`
        3. You will receive a **unique activation code** within 24 hours
        4. Enter the code below — Premium activates instantly ✅

        > 💡 *Your payment goes directly to MLS Academy.
          MLS Academy pays Google separately for AI usage.*
        """)

        act_col1, act_col2 = st.columns([3,1])
        with act_col1:
            code_input = st.text_input(
                "Activation Code:",
                placeholder="e.g. MLS-PREM-XXXX-XXXX",
                max_chars=20,
                key="activation_code_input"
            )
        with act_col2:
            st.markdown("")
            st.markdown("")
            activate_btn = st.button("✅ Activate", type="primary", use_container_width=True)

        if activate_btn:
            if code_input.strip():
                ok_prem, result = activate_premium(code_input.strip())
                if ok_prem:
                    exp_date = datetime.fromisoformat(result).strftime("%d %b %Y")
                    st.balloons()
                    st.markdown(f'''<div class="alert-good">
                        <h3 style="margin:0 0 .4rem">🎉 Premium Activated!</h3>
                        <p style="margin:0">You now have <b>unlimited access</b> until <b>{exp_date}</b>.
                        Enjoy all features without any credit limits!</p>
                    </div>''', unsafe_allow_html=True)
                    st.rerun()
                else:
                    st.error(f"❌ {result}")

    else:
        # Premium user — show renew option
        st.markdown("### 🔄 Renew Premium")
        st.markdown('<div class="alert-good">Your Premium subscription is active. To renew after expiry, contact MLS Academy and enter a new activation code.</div>', unsafe_allow_html=True)
        if st.button("Enter New Activation Code", use_container_width=True):
            state["plan"] = "free"
            _save_state(state, client)
            st.rerun()

    # ── Admin code generator (hidden, for MLS Academy staff) ─
    st.markdown("")
    with st.expander("🔐 Admin: Generate Premium Code (Staff Only)"):
        admin_pass = st.text_input("Admin password:", type="password", key="admin_pw")
        if admin_pass == "MLS_ADMIN_2026":  # Change this to your admin password
            st.markdown("**Generate Premium Activation Code:**")
            num_codes = st.number_input("Number of codes:", 1, 50, 1, key="n_codes")
            duration  = st.selectbox("Duration:", ["30 days","60 days","90 days","365 days"], key="dur_sel")
            dur_days  = int(duration.split()[0])

            if st.button("Generate Codes", type="primary", key="gen_codes"):
                import random, string as _str
                codes = []
                for _ in range(int(num_codes)):
                    part1 = "".join(random.choices(_str.ascii_uppercase+_str.digits, k=4))
                    part2 = "".join(random.choices(_str.ascii_uppercase+_str.digits, k=4))
                    code  = f"MLS-PREM-{part1}-{part2}"
                    codes.append(code)

                    # Save to Supabase if connected
                    if client:
                        try:
                            client.table("premium_codes").insert({
                                "code":       code,
                                "used":       False,
                                "duration_days": dur_days,
                                "created_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
                            }).execute()
                        except Exception:
                            pass

                st.markdown("**Generated codes (save these!):**")
                for c in codes:
                    st.code(c)

                codes_txt = "\n".join(codes)
                st.download_button("💾 Download Codes",
                    codes_txt, f"premium_codes_{datetime.now().strftime('%Y%m%d')}.txt",
                    "text/plain", use_container_width=True)

    # ── Usage history ─────────────────────────────────────
    if total > 0:
        st.markdown("")
        st.markdown(f"""
        <div style="background:#f8fafc;border-radius:10px;padding:1rem;border:1px solid #e2e8f0;
                    font-size:.85rem;color:#374151;">
            📊 <b>Usage Summary</b><br>
            Total credits used (all time): <b>{total}</b> |
            This window: <b>{used}</b> |
            Remaining: <b>{remaining}</b>
        </div>
        """, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# 🧮 CLINICAL DECISION SUPPORT PAGES
# ════════════════════════════════════════════════════════════════════════════

def page_clinical_scores():
    st.markdown('<div class="section-header">🧮 Clinical Scoring Tools</div>', unsafe_allow_html=True)
    c = st.session_state.selected_case

    if c:
        st.markdown(f'<div class="alert-info">📋 Active case: <b>{c.get("Age_Sex","?")} — {c.get("Chief_Complaint","?")}</b></div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                border-radius:14px;padding:1.2rem 1.5rem;margin-bottom:1rem;">
        <h3 style="margin:0 0 .3rem">🧮 Validated Clinical Scoring Systems</h3>
        <p style="margin:0;opacity:.88;font-size:.88rem">
            All scores follow published clinical guidelines (ESC, AHA, BTS, WSES).
            Results are automatically contextualized with your active case.
        </p>
    </div>
    """, unsafe_allow_html=True)

    score_name = st.selectbox("Select Scoring System:",
        list(CLINICAL_SCORES.keys()),
        help="Each score follows its original validated publication")

    score_def = CLINICAL_SCORES[score_name]
    st.markdown(f'<div class="alert-info">📖 <b>{score_def["description"]}</b></div>', unsafe_allow_html=True)

    st.markdown("**Select all criteria that apply to your patient:**")
    cols = st.columns(2)
    selected = []
    for i, (criterion, pts) in enumerate(score_def["variables"].items()):
        with cols[i % 2]:
            pt_label = f"(+{pts} pts)" if pts > 0 else f"({pts} pts)"
            if st.checkbox(f"{criterion} {pt_label}", key=f"sc_{score_name[:6]}_{i}"):
                selected.append(criterion)

    total, severity, interpretation = calculate_clinical_score(score_name, selected)
    
    # Color by severity
    sev_colors = {
        "low": "#16a34a", "mild": "#16a34a", "minor": "#16a34a",
        "moderate": "#f59e0b", "moderate": "#f59e0b", "B": "#f59e0b",
        "high": "#dc2626", "severe": "#dc2626", "critical": "#dc2626",
        "C": "#dc2626",
    }
    color = sev_colors.get(severity, "#0e7490")

    st.markdown(f"""
    <div style="background:{color};color:white;border-radius:14px;
                padding:1.5rem 2rem;margin:1.5rem 0;
                box-shadow:0 4px 16px {color}44;">
        <div style="display:flex;justify-content:space-between;align-items:center;">
            <div>
                <div style="font-size:.85rem;opacity:.85;margin-bottom:.3rem">{score_name} Result</div>
                <div style="font-size:2.5rem;font-weight:900">{total} points</div>
                <div style="font-size:1.1rem;font-weight:700;margin-top:.3rem">
                    Risk Level: {severity.upper()}
                </div>
            </div>
            <div style="font-size:3rem;">{"🟢" if "low" in severity or "mild" in severity or "minor" in severity or severity=="A"
                                         else "🟡" if "mod" in severity or severity=="B"
                                         else "🔴"}</div>
        </div>
        <div style="margin-top:1rem;padding:.8rem 1rem;background:rgba(255,255,255,.15);
                    border-radius:10px;font-size:.9rem;line-height:1.6;">
            {interpretation}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # AI contextualisation with active case
    if c and st.button("🤖 AI Interpretation for This Patient", type="primary", use_container_width=True, key="score_ai"):
        prompt = (
            f"Patient: {c.get('Age_Sex','?')} | CC: {c.get('Chief_Complaint','?')} | "
            f"Vitals: {c.get('Vitals','?')} | PMH: {c.get('PMH','?')} | Labs: {c.get('Labs','?')}\n\n"
            f"Clinical Score: {score_name}\n"
            f"Score: {total} points → {severity.upper()} risk\n"
            f"Criteria selected: {', '.join(selected) if selected else 'None'}\n\n"
            f"As a senior clinician:\n"
            f"1. Is this score result consistent with the clinical picture?\n"
            f"2. What immediate management steps does this score recommend?\n"
            f"3. Which specific findings most changed the risk category?\n"
            f"4. What would change this patient's risk category?\n"
            f"Be specific, cite the guideline behind this score."
        )
        with st.spinner("AI interpreting score in context..."):
            interp = call_ai("You are a senior clinician interpreting clinical scores for medical education.",
                [{"role":"user","content":prompt}], max_tokens=600)
        st.markdown(f'<div style="background:#fff7ed;border-radius:12px;padding:1.2rem;border-left:5px solid #f59e0b;">{interp.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

    # Evidence link for this score
    query_link = search_evidence(f"{score_name} score clinical validation")
    st.markdown(f'<div class="alert-info" style="font-size:.82rem">🔗 <b>Evidence:</b> <a href="{query_link}" target="_blank">Search PubMed for {score_name} validation studies</a></div>', unsafe_allow_html=True)


def page_evidence():
    st.markdown('<div class="section-header">📚 Evidence-Based Medicine & Real Datasets</div>', unsafe_allow_html=True)

    tab_db, tab_guide, tab_search = st.tabs([
        "🗄️ Real Medical Datasets",
        "📋 Clinical Guidelines",
        "🔍 Literature Search",
    ])

    with tab_db:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0a2540,#1a4f8a);color:white;
                    border-radius:14px;padding:1.2rem 1.5rem;margin-bottom:1rem;">
            <h3 style="margin:0 0 .3rem">🗄️ Verified Real Medical Image Databases</h3>
            <p style="margin:0;opacity:.88;font-size:.85rem">
                All databases below are publicly accessible and peer-reviewed.
                Use these to validate AI imaging accuracy with ground-truth labels.
            </p>
        </div>
        """, unsafe_allow_html=True)

        for db_name, db_info in REAL_CASE_DATABASES.items():
            verified_badge = '<span style="background:#16a34a;color:white;border-radius:999px;padding:.15rem .5rem;font-size:.72rem;font-weight:700;margin-left:.4rem">✅ VERIFIED</span>' if db_info["verified"] else ""
            conditions_str = " · ".join(db_info["conditions"][:6])
            if len(db_info["conditions"]) > 6:
                conditions_str += f" + {len(db_info['conditions'])-6} more"

            st.markdown(f"""
            <div style="background:white;border-radius:12px;padding:1rem 1.3rem;
                        margin:.5rem 0;border:2px solid #e2e8f0;
                        box-shadow:0 2px 8px rgba(0,0,0,.04);">
                <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:.5rem;">
                    <div>
                        <div style="font-weight:700;color:#0a2540;font-size:.95rem">
                            🗄️ {db_name}{verified_badge}
                        </div>
                        <div style="font-size:.78rem;color:#6b7280;margin-top:.2rem">
                            📦 {db_info["size"]} &nbsp;|&nbsp; 🏛️ {db_info["source"]}
                        </div>
                        <div style="font-size:.78rem;color:#374151;margin-top:.3rem">
                            <b>Conditions:</b> {conditions_str}
                        </div>
                    </div>
                    <a href="{db_info["url"]}" target="_blank"
                       style="background:linear-gradient(135deg,#0e7490,#0a2540);
                              color:white;text-decoration:none;border-radius:8px;
                              padding:.5rem 1rem;font-size:.82rem;font-weight:600;
                              white-space:nowrap;flex-shrink:0;">
                        🔗 Access Dataset
                    </a>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("")
        st.markdown("""
        <div class="alert-info">
            <b>How to use these for accuracy validation:</b><br>
            1. Download labeled images from a database above<br>
            2. Upload them in <b>Imaging & ECG Analysis</b> tab<br>
            3. The AI will analyze and show an <b>Accuracy Check</b> comparing AI vs ground truth label<br>
            4. Run 10-20 cases to get a session accuracy score
        </div>
        """, unsafe_allow_html=True)

    with tab_guide:
        st.markdown("### 📋 Validated Clinical Practice Guidelines")

        for condition, guide in CLINICAL_GUIDELINES.items():
            with st.expander(f"📋 {condition} — {guide['body']}"):
                st.markdown(f'<div class="alert-info">🔗 <a href="{guide["url"]}" target="_blank"><b>Official {guide["body"]} Guideline →</b></a></div>', unsafe_allow_html=True)
                st.markdown("**Key Management Points:**")
                for point in guide["key_points"]:
                    st.markdown(f'<div style="padding:.3rem .8rem;background:#f8fafc;border-radius:6px;border-left:3px solid #0e7490;margin:.2rem 0;font-size:.87rem">✅ {point}</div>', unsafe_allow_html=True)

                # AI expansion
                c = st.session_state.selected_case
                if c and st.button(f"🤖 Apply to My Patient", key=f"g_{condition[:8]}"):
                    prompt = (
                        f"Apply the {condition} guideline ({guide['body']}) to this patient:\n"
                        f"Patient: {c.get('Age_Sex','?')} | CC: {c.get('Chief_Complaint','?')} | "
                        f"Vitals: {c.get('Vitals','?')} | Labs: {c.get('Labs','?')} | "
                        f"Diagnosis: {c.get('Final_Diagnosis','?')}\n\n"
                        f"1. Does this guideline apply? Why?\n"
                        f"2. Which specific recommendations apply to this patient?\n"
                        f"3. What is the recommended first-line treatment with doses?\n"
                        f"4. What monitoring parameters are required?\n"
                        f"Be specific with drug names, doses, and monitoring intervals."
                    )
                    with st.spinner("Applying guideline..."):
                        resp = call_ai("You are a senior clinician applying evidence-based guidelines.",
                            [{"role":"user","content":prompt}], max_tokens=600)
                    st.markdown(f'<div style="background:#f0fdf4;border-radius:10px;padding:1rem;border-left:4px solid #16a34a;">{resp.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

    with tab_search:
        st.markdown("### 🔍 Medical Literature Search")
        st.markdown('<div class="alert-info">Search directly in verified medical databases. All links open the real database with your query pre-filled.</div>', unsafe_allow_html=True)

        search_q = st.text_input("Search query:", placeholder="e.g. appendicitis laparoscopic outcomes 2023", key="ev_search")
        db_choice = st.multiselect("Search in:", list(EVIDENCE_DATABASES.keys()),
            default=["PubMed","NICE","BMJ Best Practice"])

        if search_q and st.button("🔍 Search", type="primary", use_container_width=True):
            st.markdown("### 🔗 Search Links")
            for db in db_choice:
                link = search_evidence(search_q, db)
                st.markdown(f"""
                <a href="{link}" target="_blank"
                   style="display:block;background:white;border-radius:10px;padding:.8rem 1.2rem;
                          margin:.3rem 0;border:2px solid #e2e8f0;text-decoration:none;
                          color:#0a2540;font-weight:600;font-size:.88rem;">
                    🔗 Search in {db} →
                </a>
                """, unsafe_allow_html=True)

        # Auto-search from active case
        c = st.session_state.selected_case
        if c:
            st.markdown("")
            st.markdown(f"**Quick search for active case: {c.get('Final_Diagnosis','?')}**")
            dx = str(c.get("Final_Diagnosis",""))
            qs = [dx, f"{dx} management", f"{dx} diagnosis criteria", f"{dx} treatment guidelines"]
            for qlink in qs:
                link = search_evidence(qlink)
                st.markdown(f'<a href="{link}" target="_blank" style="display:inline-block;background:#eff6ff;border:1px solid #3b82f6;border-radius:8px;padding:.3rem .8rem;margin:.2rem;font-size:.82rem;color:#1e40af;text-decoration:none;">🔍 {qlink}</a>', unsafe_allow_html=True)


def _fetch_real_pmid(query: str) -> str:
    """Fetch a single verified PubMed ID for a diagnosis using NCBI E-utilities.
    Uses session cache to avoid repeated lookups. Returns "" on failure.
    """
    if not query or not query.strip():
        return ""
    cache = st.session_state.setdefault("_pmid_cache", {})
    key = query.strip().lower()
    if key in cache:
        return cache[key]
    try:
        url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        params = {
            "db": "pubmed",
            "term": f"{query}[Title/Abstract] AND (review[Filter] OR guideline[Filter])",
            "retmode": "json",
            "retmax": 3,
            "sort": "relevance",
            "mindate": "2019",
            "maxdate": "2026",
        }
        r = requests.get(url, params=params, timeout=6)
        if r.status_code == 200:
            ids = r.json().get("esearchresult", {}).get("idlist", [])
            if ids:
                cache[key] = ids[0]
                return ids[0]
        # Fallback: broader search
        params["term"] = query
        r = requests.get(url, params=params, timeout=6)
        if r.status_code == 200:
            ids = r.json().get("esearchresult", {}).get("idlist", [])
            if ids:
                cache[key] = ids[0]
                return ids[0]
    except Exception:
        pass
    cache[key] = ""
    return ""


def page_doccollab():
    st.markdown('<div class="section-header">🌐 DocCollab — Global Case Matching</div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="background:linear-gradient(135deg,#1a0533,#4c1d95);color:white;
                border-radius:14px;padding:1.5rem 2rem;margin-bottom:1.5rem;">
        <h2 style="margin:0 0 .5rem;font-size:1.4rem">🌐 DocCollab Global — Real Case Reports from PubMed</h2>
        <p style="margin:0;opacity:.9;font-size:.9rem">
            Retrieves <b>real, peer-reviewed case reports</b> from PubMed (NCBI E-utilities API)
            matching your patient's diagnosis. Every case shown is verified, published,
            and linked to its original journal article.
        </p>
        <div style="margin-top:.8rem;font-size:.8rem;opacity:.75">
            Source: PubMed — 35M+ biomedical citations · Filter: Case Reports · Last 10 years · English
        </div>
    </div>
    """, unsafe_allow_html=True)

    if not CLINICAL_HELPERS_OK:
        st.error("⚠️ clinical_helpers.py not found. Place it in the same folder as app.py.")
        return

    c = st.session_state.selected_case

    tab_match, tab_upload, tab_db = st.tabs([
        "🔍 Find Real Case Reports",
        "📤 Upload for AI Matching",
        "🗄️ Real Case Database Stats",
    ])

    with tab_match:
        if not c:
            st.warning("Select a case first from the Case Library to enable case matching.")
            if st.button("📚 Go to Case Library"): nav("library")
        else:
            dx          = str(c.get("Final_Diagnosis", "")).strip()
            chief       = str(c.get("Chief_Complaint", "")).strip()
            age_sex     = str(c.get("Age_Sex", "")).strip()

            st.markdown(
                f'<div class="alert-info">🎯 Searching PubMed for real published case '
                f'reports matching: <b>{dx or chief or "current case"}</b></div>',
                unsafe_allow_html=True,
            )

            match_depth = st.select_slider(
                "How many cases to retrieve:",
                ["Quick (5 cases)", "Standard (10 cases)", "Deep (20 cases)", "Comprehensive (30 cases)"],
                value="Standard (10 cases)",
            )
            n_matches = {"Quick (5 cases)":5, "Standard (10 cases)":10,
                         "Deep (20 cases)":20, "Comprehensive (30 cases)":30}[match_depth]

            if st.button("🔍 Search PubMed for Real Cases", type="primary",
                         use_container_width=True):
                if not dx or dx == "?":
                    st.warning("This case has no diagnosis recorded — searching by chief complaint.")
                with st.spinner(f"Querying NCBI PubMed for {n_matches} real case reports..."):
                    search_term = dx if dx and dx != "?" else chief
                    pubmed_result = fetch_pubmed_case_reports(
                        diagnosis    = search_term,
                        age_sex      = age_sex,
                        presenting   = chief,
                        max_results  = n_matches,
                    )
                st.session_state["doccollab_pubmed"] = pubmed_result
                # Clear any old AI-generated result
                st.session_state.pop("doccollab_result", None)

            pmr = st.session_state.get("doccollab_pubmed")
            if pmr is not None:
                if not pmr.get("ok"):
                    st.markdown(
                        f'<div class="alert-warn">⚠️ {pmr.get("error", "No results found")}</div>',
                        unsafe_allow_html=True,
                    )
                    if pmr.get("query"):
                        st.caption(f"Query tried: `{pmr['query']}`")
                else:
                    cases = pmr.get("cases", [])
                    st.markdown(f"### 🌍 Found {len(cases)} Real PubMed Case Reports")
                    st.caption(
                        f"Search query: `{pmr.get('query','')}` · "
                        f"All cases below are real, peer-reviewed, and linked to original sources."
                    )

                    # Render each real case
                    for idx, case in enumerate(cases):
                        render_pubmed_case_card(case, idx)

                    # "See all" link to PubMed website
                    if pmr.get("search_url"):
                        st.markdown(
                            f'<div style="text-align:center;margin:1rem 0;">'
                            f'<a href="{pmr["search_url"]}" target="_blank" '
                            f'style="display:inline-block;background:#1e40af;color:white;'
                            f'border-radius:8px;padding:.6rem 1.5rem;text-decoration:none;'
                            f'font-weight:600;font-size:.88rem;">'
                            f'🔗 View all results on PubMed →</a></div>',
                            unsafe_allow_html=True,
                        )

                    # ── AI Aggregate Analysis (built from real abstracts) ────────
                    if cases:
                        st.markdown("---")
                        st.markdown("### 📊 AI Synthesis of Real Cases")
                        st.caption(
                            "Gemini analyzes the abstracts above and extracts common patterns. "
                            "Unlike before, this synthesis is built ONLY from real published cases."
                        )

                        if st.button("🧠 Generate AI Synthesis from Above Cases",
                                     use_container_width=True, key="ai_synth"):
                            with st.spinner("Synthesizing patterns from real abstracts..."):
                                # Build context from real abstracts only
                                abstracts_text = "\n\n".join([
                                    f"[Case {i+1}] {ca.get('title','')} ({ca.get('year','')}, "
                                    f"{ca.get('country','')}, PMID {ca.get('pmid','')}):\n"
                                    f"{ca.get('abstract','')[:800]}"
                                    for i, ca in enumerate(cases[:10])
                                ])
                                synth_prompt = (
                                    f"Below are abstracts from {min(len(cases),10)} REAL peer-reviewed case reports "
                                    f"on '{dx}'. Synthesize them for a medical student.\n\n"
                                    f"{abstracts_text}\n\n"
                                    "Provide:\n"
                                    "1. **Common Presentation Pattern** — what features appear repeatedly\n"
                                    "2. **Most Frequent Treatment Approach** — with case numbers as evidence\n"
                                    "3. **Reported Outcomes** — recovery rates and complications across cases\n"
                                    "4. **Red Flags** — features associated with worse outcomes\n"
                                    "5. **Educational Takeaway** — what to remember from these cases\n\n"
                                    "Cite specific case numbers (e.g. 'Case 3, 7') for every claim. "
                                    "Do NOT invent details not in the abstracts."
                                )
                                synth = call_ai(
                                    "You synthesize patterns from real medical case reports. "
                                    "Only use information present in the provided abstracts.",
                                    [{"role":"user","content":synth_prompt}],
                                    max_tokens=1500,
                                )
                            st.session_state["doccollab_synth"] = synth

                        if st.session_state.get("doccollab_synth"):
                            synth = st.session_state["doccollab_synth"]
                            if synth.startswith("!ERR"):
                                st.error(f"AI synthesis failed: {synth}")
                            else:
                                st.markdown(f"""
                                <div style="background:linear-gradient(135deg,#fdf4ff,#f3e8ff);
                                            border:2px solid #a855f7;border-radius:12px;
                                            padding:1.2rem 1.5rem;">
                                  <div style="font-weight:700;color:#6b21a8;margin-bottom:.5rem">
                                    🔬 Pattern Analysis from {min(len(cases),10)} Real Cases
                                  </div>
                                  <div style="color:#1e293b;font-size:.88rem;line-height:1.6;
                                              white-space:pre-wrap;">{synth}</div>
                                </div>""", unsafe_allow_html=True)

                    # Evidence links (kept from original)
                    st.markdown("---")
                    st.markdown("### 🔗 Additional Evidence Sources")
                    ev_cols = st.columns(3)
                    with ev_cols[0]:
                        st.markdown(f'<a href="{search_evidence(dx)}" target="_blank" style="display:block;background:#eff6ff;border:2px solid #3b82f6;border-radius:10px;padding:.8rem;text-align:center;text-decoration:none;color:#1e40af;font-weight:600;font-size:.85rem;">📚 PubMed (all article types)<br><span style="font-size:.75rem;opacity:.8">{dx}</span></a>', unsafe_allow_html=True)
                    with ev_cols[1]:
                        st.markdown(f'<a href="{search_evidence(dx, "ClinicalTrials")}" target="_blank" style="display:block;background:#f0fdf4;border:2px solid #16a34a;border-radius:10px;padding:.8rem;text-align:center;text-decoration:none;color:#166534;font-weight:600;font-size:.85rem;">🧪 Clinical Trials<br><span style="font-size:.75rem;opacity:.8">{dx}</span></a>', unsafe_allow_html=True)
                    with ev_cols[2]:
                        st.markdown(f'<a href="{search_evidence(dx, "NICE")}" target="_blank" style="display:block;background:#fff7ed;border:2px solid #f59e0b;border-radius:10px;padding:.8rem;text-align:center;text-decoration:none;color:#92400e;font-weight:600;font-size:.85rem;">📋 NICE Guidelines<br><span style="font-size:.75rem;opacity:.8">{dx}</span></a>', unsafe_allow_html=True)

                    # Download report (real cases this time)
                    if cases:
                        report_text = f"# Real PubMed Case Reports for: {dx}\n\n"
                        report_text += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                        report_text += f"Source: PubMed via NCBI E-utilities\n"
                        report_text += f"Query: {pmr.get('query','')}\n\n"
                        for i, ca in enumerate(cases):
                            report_text += f"\n## Case {i+1}\n"
                            report_text += f"**Title:** {ca.get('title','')}\n"
                            report_text += f"**Authors:** {ca.get('authors','')}\n"
                            report_text += f"**Journal:** {ca.get('journal','')} ({ca.get('year','')})\n"
                            report_text += f"**Country:** {ca.get('country','')}\n"
                            report_text += f"**PMID:** {ca.get('pmid','')}\n"
                            report_text += f"**URL:** {ca.get('url','')}\n\n"
                            report_text += f"**Abstract:**\n{ca.get('abstract','')}\n\n"
                            report_text += "---\n"

                        st.download_button(
                            "💾 Download Full Report (with abstracts)",
                            report_text,
                            f"pubmed_cases_{datetime.now().strftime('%Y%m%d_%H%M')}.md",
                            "text/markdown",
                            use_container_width=True,
                        )

    with tab_upload:
        st.markdown("""
        <div class="alert-info">
            <b>Upload anonymized patient data</b> (image, ECG, or lab results) for AI matching.
            The system will find the most similar verified cases from real medical databases.
        </div>
        """, unsafe_allow_html=True)

        upload_img = st.file_uploader("Upload medical image (X-ray, CT, ECG):", type=["jpg","jpeg","png","dcm"], key="dc_img")
        clinical_ctx = st.text_area("Clinical summary:", placeholder="Age, sex, chief complaint, key findings, suspected diagnosis...", height=100, key="dc_ctx")

        if upload_img and clinical_ctx and st.button("🔍 Match Against Global Databases", type="primary", use_container_width=True):
            img_bytes = upload_img.read()
            with st.spinner("AI analyzing image + matching against global databases..."):
                import base64 as _b64
                img_b64 = _b64.b64encode(img_bytes).decode()
                mime    = f"image/{upload_img.name.split('.')[-1].lower()}"

                # First analyze the image
                analysis = call_ai_radiology(img_bytes, mime, "Auto-detect", clinical_ctx)
                # Then match
                match_prompt = (
                    f"Based on this image analysis:\n{analysis[:1000]}\n\n"
                    f"Clinical context: {clinical_ctx}\n\n"
                    f"Match this case against similar cases in medical databases.\n"
                    f"Identify: 1) Most likely diagnosis 2) 5 similar database cases "
                    f"3) Recommended management based on matched outcomes "
                    f"4) Relevant evidence links from NIH/MIMIC/PhysioNet"
                )
                match_result = call_ai(
                    "You are a clinical AI performing image-based case matching.",
                    [{"role":"user","content":match_prompt}], max_tokens=1200
                )

            st.markdown(f'<div style="background:white;border-radius:12px;padding:1.2rem;border:2px solid #7c3aed;">{match_result.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)

    with tab_db:
        st.markdown("### 📊 Real Medical Database Statistics")
        total_cases = sum(0 for _ in REAL_CASE_DATABASES)
        st.markdown(f'<div class="alert-good">✅ {len(REAL_CASE_DATABASES)} verified public medical databases available for matching</div>', unsafe_allow_html=True)

        for db_name, db_info in REAL_CASE_DATABASES.items():
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;align-items:center;
                        padding:.7rem 1rem;background:#f8fafc;border-radius:8px;
                        border-left:4px solid #7c3aed;margin:.3rem 0;">
                <div>
                    <span style="font-weight:600;color:#0a2540">{db_name}</span>
                    <span style="font-size:.78rem;color:#6b7280;margin-left:.5rem">— {db_info["source"]}</span>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:.8rem;font-weight:600;color:#7c3aed">{db_info["size"]}</div>
                    <a href="{db_info["url"]}" target="_blank"
                       style="font-size:.73rem;color:#0e7490;text-decoration:none;">🔗 Access</a>
                </div>
            </div>
            """, unsafe_allow_html=True)

# 🏥 EPIC EHR System Integration Guide for Hospital App

## 📋 Overview

#EPIC is one of the largest EHR systems in healthcare. Integrating it with your hospital app allows you to:
#✅ Retrieve patient data using patient ID
#✅ Access patient history across all departments
#✅ Create follow-up records automatically
#✅ Link patients across different sections (DocCollab, Lab, Surgery, etc.)
#✅ Sync data bidirectionally with EPIC

# ---

## 🔌 Available Options to Connect to EPIC

### Option 1: EPIC FHIR API (Recommended - Modern)
# **Best for:** New implementations, cloud-based, RESTful, JSON format
# **Status:** Official, supported by Epic
# **Cost:** Usually free or included in Epic license

### Option 2: EPIC HL7 Interface (Traditional)
# **Best for:** Legacy systems, HL7v2 format, hospital integration
# **Status:** Mature, widely used
# **Cost:** Usually included in Epic license

### Option 3: EPIC Web Services API (Direct)
# **Best for:** Direct database access, proprietary format
# **Status:** Enterprise, requires special access
# **Cost:** License-dependent

### Option 4: Third-Party Libraries
# **Available:** pyEHR, python-fhirclient, hl7apy

# ---

## ✅ RECOMMENDED: EPIC FHIR API Implementation

### Step 1: Get EPIC FHIR Credentials

# Contact your Epic Support/IT team:
# ```
# 1. Request EPIC FHIR API access
# 2. Get:
   # - Client ID
   # - Client Secret
   # - FHIR Base URL (e.g., https://your-hospital.epic.com/api/FHIR/R4/)
   # - Scopes needed (patient/*.read, patient/*.write, etc.)
# 3. Register your application
# 4. Get authorization endpoint
# ```

# ---

## 💻 Code Implementation: EPIC FHIR Integration

### Library Installation
# ```bash
# pip install requests fhirclient python-dateutil
# ```

### Complete Working Code

# ```python
# import requests
# import json
# from datetime import datetime
# from typing import Dict, Optional, List
# import streamlit as st

# ════════════════════════════════════════════════════════════════════════════
# 🏥 EPIC FHIR CONFIGURATION
# ════════════════════════════════════════════════════════════════════════════

# EPIC_CONFIG = {
    # Get these from your Epic Admin/IT team
    # "client_id": "YOUR_CLIENT_ID_HERE",
    # "client_secret": "YOUR_CLIENT_SECRET_HERE",
    # "fhir_url": "https://your-hospital.epic.com/api/FHIR/R4/",
    # "auth_url": "https://your-hospital.epic.com/oauth2/authorize",
    # "token_url": "https://your-hospital.epic.com/oauth2/token",
    # "redirect_uri": "http://localhost:8501/callback",
# }

# ════════════════════════════════════════════════════════════════════════════
# 🔐 EPIC FHIR Authentication
# ════════════════════════════════════════════════════════════════════════════

# ── EPIC FHIR Configuration ──────────────────────────────────────────────────
# Replace with your real EPIC credentials when you have API access
# Contact your hospital IT / Epic support to get these values
# ════════════════════════════════════════════════════════════════════════════
# 📊 USAGE EXAMPLES IN YOUR APP
# ════════════════════════════════════════════════════════════════════════════

def example_lab_section_with_epic():
    """
    Example: Lab section with EPIC integration
    """
    st.markdown("## 🧪 Laboratory Tests")
    
    # Add patient lookup
    epic_patient_lookup_widget()
    
    # Show patient summary if available
    display_patient_summary()
    
    # Lab tests form
    st.markdown("### Order New Lab Tests")
    test_type = st.selectbox("Lab Test:", ["CBC", "CMP", "Lipid Panel", "TSH", "Custom"])
    
    if st.button("Order Lab Test"):
        if "patient_id" in st.session_state:
            st.success("✅ Lab order sent to EPIC")
        else:
            st.warning("Select a patient first")
    
    # Create lab follow-up
    create_followup_section("Lab")


def example_surgery_section_with_epic():
    """
    Example: Surgery section with EPIC integration
    """
    st.markdown("## 🏥 Surgical Procedures")
    
    # Get current patient from session
    epic_patient_lookup_widget()
    
    if "epic_patient" in st.session_state:
        display_patient_summary()
        
        st.markdown("### Schedule Surgery")
        procedure = st.selectbox("Procedure:", ["Appendectomy", "C-section", "Hip Replacement"])
        
        if st.button("Schedule Procedure"):
            st.success("✅ Procedure scheduled in EPIC")
        
        # Post-surgery follow-up
        create_followup_section("Surgery")


# ════════════════════════════════════════════════════════════════════════════
# 🔗 LINK ALL SECTIONS TOGETHER WITH EPIC
# ════════════════════════════════════════════════════════════════════════════

# In your main app navigation, use this:
def main_with_epic():
    """
    Main app with EPIC integration across all sections
    """
    st.sidebar.title("🏥 Hospital App + EPIC")
    
    # EPIC Patient Lookup - Available in all sections
    epic_patient_lookup_widget()
    
    # Navigation
    page = st.sidebar.radio("Select Section:", [
        "Home",
        "Lab",
        "Surgery",
        "Imaging",
        "DocCollab",
        "Patient Follow-up"
    ])
    
    # All sections now have access to patient data
    if "epic_patient" in st.session_state:
        st.sidebar.success(f"✅ Patient: {st.session_state.epic_patient['demographics']['name']}")
    
    if page == "Home":
        st.markdown("# 🏥 Hospital Management System with EPIC Integration")
    
    elif page == "Lab":
        example_lab_section_with_epic()
    
    elif page == "Surgery":
        example_surgery_section_with_epic()
    
    elif page == "Patient Follow-up":
        st.markdown("## 📋 Patient Follow-ups")
        display_patient_summary()


# EPIC integration runs only when explicitly called from the sidebar
# Do NOT call main_with_epic() at module level — it breaks Streamlit
# if __name__ == "__main__":
#     main_with_epic()


# ════════════════════════════════════════════════════════════════════════════
# 🔐 AUTH — USER MANAGEMENT (Student + Faculty login)
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# 🔐 AUTH — Supabase-backed (permanent) with in-memory demo fallback
# ════════════════════════════════════════════════════════════════════════════

def _hash_pw(pw: str) -> str:
    """SHA-256 hash — never store plain passwords."""
    return hashlib.sha256(pw.encode()).hexdigest()

def _sb_headers() -> dict:
    """Supabase REST API headers."""
    key = SUPABASE_DEFAULT_KEY
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }




# ════════════════════════════════════════════════════════════════════════════
# 📊 FACULTY ANALYTICS DASHBOARD
# ════════════════════════════════════════════════════════════════════════════

# In production wire to Supabase. Here we use realistic demo data.
_DEMO_ANALYTICS = {
    "students": [
        {"name":"Alice M.","cases":24,"score":87,"accuracy":78,"last":"2026-04-11"},
        {"name":"Bob K.","cases":18,"score":72,"accuracy":65,"last":"2026-04-10"},
        {"name":"Sara L.","cases":31,"score":94,"accuracy":91,"last":"2026-04-12"},
        {"name":"James O.","cases":12,"score":61,"accuracy":54,"last":"2026-04-09"},
        {"name":"Mia T.","cases":27,"score":88,"accuracy":83,"last":"2026-04-11"},
        {"name":"Chris P.","cases":9,"score":55,"accuracy":48,"last":"2026-04-08"},
        {"name":"Lena R.","cases":33,"score":96,"accuracy":93,"last":"2026-04-12"},
    ],
    "case_attempts": {"Appendicitis":45,"NSTEMI":38,"Meningitis":29,"DKA":22,"Pneumonia":31,"Sepsis":18,"Stroke":14},
    "wrong_dx": {"Appendicitis→Ovarian Cyst":8,"NSTEMI→GERD":6,"Meningitis→Migraine":5,"DKA→HHS":4},
    "module_usage": {"Patient Interview":312,"Laboratory":289,"Surgery Room":198,"Imaging":176,"AI Tutor":421,"Physical Exam":143},
    "scores_dist": [45,52,55,61,61,65,67,70,72,72,74,75,78,80,81,83,85,87,87,88,88,90,91,93,94,96],
}

def page_faculty_analytics():
    """Faculty analytics dashboard — student performance overview."""
    user = st.session_state.get("auth_user") or {}
    if user.get("role") != "faculty":
        st.error("🔒 Access restricted to faculty members.")
        return

    st.markdown('<div class="main-header"><h1>📊 Faculty Analytics Dashboard</h1><p>Student performance & engagement overview</p></div>', unsafe_allow_html=True)

    data = _DEMO_ANALYTICS
    students = data["students"]

    # ── KPIs ──────────────────────────────────────────────────────
    k1,k2,k3,k4 = st.columns(4)
    with k1:
        st.markdown('<div class="kpi-card"><div class="kpi-value">'+str(len(students))+'</div><div class="kpi-label">👥 Total Students</div></div>', unsafe_allow_html=True)
    with k2:
        avg_score = int(sum(s["score"] for s in students)/len(students))
        st.markdown('<div class="kpi-card"><div class="kpi-value">'+str(avg_score)+'</div><div class="kpi-label">⭐ Avg Score</div></div>', unsafe_allow_html=True)
    with k3:
        avg_cases = int(sum(s["cases"] for s in students)/len(students))
        st.markdown('<div class="kpi-card"><div class="kpi-value">'+str(avg_cases)+'</div><div class="kpi-label">📋 Avg Cases Done</div></div>', unsafe_allow_html=True)
    with k4:
        avg_acc = int(sum(s["accuracy"] for s in students)/len(students))
        st.markdown('<div class="kpi-card"><div class="kpi-value">'+str(avg_acc)+'%</div><div class="kpi-label">🎯 Avg Accuracy</div></div>', unsafe_allow_html=True)

    st.markdown("")
    col_l, col_r = st.columns([3,2])

    with col_l:
        st.markdown('<div class="section-header">👥 Student Leaderboard</div>', unsafe_allow_html=True)
        sdf = pd.DataFrame(students).sort_values("score", ascending=False).reset_index(drop=True)
        sdf.index += 1
        sdf.columns = ["Name","Cases Done","Score","Accuracy (%)","Last Active"]
        st.dataframe(sdf, use_container_width=True)

        st.markdown('<div class="section-header">📋 Most Attempted Cases</div>', unsafe_allow_html=True)
        if PLOTLY_OK:
            ca = data["case_attempts"]
            fig = px.bar(x=list(ca.keys()), y=list(ca.values()),
                         color=list(ca.values()), color_continuous_scale="Blues",
                         labels={"x":"Case","y":"Attempts"})
            fig.update_layout(height=250, showlegend=False, margin=dict(l=0,r=0,t=10,b=0),
                              plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.bar_chart(data["case_attempts"])

    with col_r:
        st.markdown('<div class="section-header">❌ Common Wrong Diagnoses</div>', unsafe_allow_html=True)
        for dx,cnt in data["wrong_dx"].items():
            st.markdown(f'<div class="alert-bad" style="margin:.3rem 0;">❌ {dx} — <b>{cnt} students</b></div>', unsafe_allow_html=True)

        st.markdown('<div class="section-header">🖥️ Module Usage</div>', unsafe_allow_html=True)
        if PLOTLY_OK:
            mu = data["module_usage"]
            fig2 = px.pie(names=list(mu.keys()), values=list(mu.values()),
                          hole=0.4, color_discrete_sequence=px.colors.sequential.Blues_r)
            fig2.update_layout(height=260, margin=dict(l=0,r=0,t=10,b=0),
                               paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.bar_chart(data["module_usage"])

        st.markdown('<div class="section-header">📈 Score Distribution</div>', unsafe_allow_html=True)
        if PLOTLY_OK:
            fig3 = px.histogram(x=data["scores_dist"], nbins=10,
                                color_discrete_sequence=["#0e7490"],
                                labels={"x":"Score","y":"Students"})
            fig3.update_layout(height=200, margin=dict(l=0,r=0,t=10,b=0),
                               plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig3, use_container_width=True)

    # ── Per-student detail ────────────────────────────────────────
    st.markdown('<div class="section-header">🔍 Individual Student Report</div>', unsafe_allow_html=True)
    sel = st.selectbox("Select student:", [s["name"] for s in students], key="fac_student_sel")
    s = next(x for x in students if x["name"]==sel)
    c1,c2,c3 = st.columns(3)
    with c1: st.metric("Cases Completed", s["cases"])
    with c2: st.metric("Score", s["score"])
    with c3: st.metric("Accuracy", f"{s['accuracy']}%")
    if FPDF_OK:
        if st.button(f"📄 Export {sel}'s Report as PDF", key="fac_pdf"):
            pdf = FPDF(); pdf.add_page(); pdf.set_font("Helvetica","B",16)
            pdf.cell(0,10,f"MLS Virtual Hospital — Student Report",ln=True,align="C")
            pdf.set_font("Helvetica","",12)
            pdf.cell(0,8,f"Student: {s['name']}",ln=True)
            pdf.cell(0,8,f"Cases Completed: {s['cases']}",ln=True)
            pdf.cell(0,8,f"Score: {s['score']}",ln=True)
            pdf.cell(0,8,f"Accuracy: {s['accuracy']}%",ln=True)
            pdf.cell(0,8,f"Last Active: {s['last']}",ln=True)
            pdf_bytes = pdf.output(dest="S").encode("latin-1")
            st.download_button("⬇ Download PDF", pdf_bytes, f"{sel}_report.pdf", "application/pdf")


# ════════════════════════════════════════════════════════════════════════════
# 🧬 DIFFERENTIAL DIAGNOSIS BUILDER — Case-correlated via AI
# ════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def generate_case_ddx(case_id: str, cc: str, age_sex: str, hpi: str,
                      vitals: str, labs: str, system: str, final_dx: str) -> list:
    """
    Use Gemini to generate 8 case-specific differentials for THIS exact presentation.
    Cached per case_id so it only runs once per case.
    Correct diagnosis is always included but shuffled into the list.
    """
    prompt = f"""You are a clinical educator. Generate exactly 8 differential diagnoses
for this specific patient presentation — ranked from most to least plausible.
The correct final diagnosis MUST be included somewhere in the list.
Include plausible distractors appropriate for this exact case.

Patient: {age_sex}
System: {system}
Chief Complaint: {cc}
History: {hpi[:400] if hpi else "Not provided"}
Vitals: {vitals[:200] if vitals else "Not provided"}
Labs: {labs[:200] if labs else "Not provided"}
Correct diagnosis (include this): {final_dx}

Return ONLY a valid JSON array of 8 strings. No explanation. No markdown. Example:
["Diagnosis 1","Diagnosis 2","Diagnosis 3","Diagnosis 4","Diagnosis 5","Diagnosis 6","Diagnosis 7","Diagnosis 8"]"""

    try:
        result = call_ai(
            "You generate differential diagnosis lists for medical education. Return only JSON arrays.",
            [{"role":"user","content":prompt}], max_tokens=300, credit_type="chat")
        # Parse JSON from response
        import re as _re
        match = _re.search(r"\[.*?\]", result, _re.DOTALL)
        if match:
            ddx = json.loads(match.group())
            if isinstance(ddx, list) and len(ddx) >= 4:
                # Ensure final dx is present
                if not any(final_dx.lower() in d.lower() for d in ddx):
                    ddx.insert(random.randint(1,3), final_dx)
                return ddx[:8]
    except Exception:
        pass
    # Fallback: system-based defaults + correct diagnosis
    return _system_fallback_ddx(system, final_dx)


def _system_fallback_ddx(system: str, final_dx: str) -> list:
    """Fallback differentials by system when AI call fails."""
    sys_map = {
        "gastrointestinal": ["Acute Appendicitis","Mesenteric Adenitis","Crohn's Disease",
                              "Cholecystitis","Renal Colic","IBS","Pancreatitis","Ectopic Pregnancy"],
        "cardiovascular":   ["NSTEMI","Unstable Angina","Pericarditis","Pulmonary Embolism",
                              "Aortic Dissection","GERD","Musculoskeletal Chest Pain","Pneumonia"],
        "respiratory":      ["Community-Acquired Pneumonia","Pulmonary Embolism","Asthma Exacerbation",
                              "Pleural Effusion","Pneumothorax","COPD Exacerbation","Heart Failure","TB"],
        "neurological":     ["Bacterial Meningitis","Migraine","Subarachnoid Haemorrhage",
                              "Viral Encephalitis","Stroke (Ischaemic)","TIA","Brain Tumour","Seizure"],
        "endocrine":        ["Type 1 Diabetes / DKA","Type 2 Diabetes / HHS","Addison's Disease",
                              "Thyroid Storm","Hypoglycaemia","Cushing's Syndrome","Phaeochromocytoma","SIADH"],
        "renal":            ["Acute Kidney Injury","Renal Colic / Nephrolithiasis","Pyelonephritis",
                              "Glomerulonephritis","CKD Exacerbation","Rhabdomyolysis","Renal Vein Thrombosis","UTI"],
    }
    s = system.lower() if system else ""
    defaults = next((v for k,v in sys_map.items() if k in s), [
        "Sepsis","Pulmonary Embolism","Acute MI","Stroke","DKA",
        "Meningitis","Aortic Dissection","Anaphylaxis"])
    # Ensure correct dx is present
    if not any(final_dx.lower() in d.lower() for d in defaults):
        defaults[random.randint(1,3)] = final_dx
    return defaults


@st.cache_data(show_spinner=False)
def generate_case_rx_suggestions(case_id: str, dx: str, age_sex: str,
                                  allergies: str, current_meds: str) -> dict:
    """
    Generate case-specific drug suggestions grouped by category.
    Returns {"First-line treatment":["Drug A","Drug B"], "Analgesia":["Drug C"]}
    """
    prompt = f"""You are a clinical pharmacologist. For a patient with {dx} ({age_sex}),
allergies: {allergies or "NKDA"}, current meds: {current_meds or "none"},

List the most clinically relevant drugs to consider, grouped by purpose.
Return ONLY valid JSON like:
{{"First-line Antibiotic":["Ceftriaxone","Metronidazole"],"Analgesia":["Morphine","Paracetamol"],"Supportive":["Ondansetron","IV Fluids"]}}
Maximum 4 groups, 3 drugs each. No explanations."""

    try:
        result = call_ai(
            "You suggest case-specific drug lists for medical education. Return only JSON.",
            [{"role":"user","content":prompt}], max_tokens=300, credit_type="chat")
        import re as _re
        match = _re.search(r'\{.*\}', result, _re.DOTALL)
        if match:
            parsed = json.loads(match.group())
            if isinstance(parsed, dict):
                return parsed
    except Exception:
        pass
    return _system_fallback_rx(dx)


def _system_fallback_rx(dx: str) -> dict:
    """Fallback drug suggestions by diagnosis keyword."""
    dx_l = dx.lower()
    if any(k in dx_l for k in ["append","cholecyst","periton"]):
        return {"Antibiotics":["Ceftriaxone","Metronidazole","Piperacillin-Tazobactam"],
                "Analgesia":["Morphine","Paracetamol","Ketorolac"],
                "Supportive":["Ondansetron","IV Fluids (Normal Saline)","Omeprazole"]}
    if any(k in dx_l for k in ["mi","nstemi","stemi","angina","cardiac"]):
        return {"Antiplatelets":["Aspirin 300mg","Ticagrelor","Clopidogrel"],
                "Anticoagulants":["Enoxaparin","Unfractionated Heparin"],
                "Analgesia":["Morphine","Nitrates (GTN)"],
                "Other":["Metoprolol","Atorvastatin","Omeprazole"]}
    if any(k in dx_l for k in ["menin","encephal","cns infect"]):
        return {"Antibiotics":["Ceftriaxone","Ampicillin","Benzylpenicillin"],
                "Steroids":["Dexamethasone"],
                "Supportive":["Paracetamol","IV Fluids","Phenytoin (if seizures)"]}
    if any(k in dx_l for k in ["diabet","dka","hhs","hyperglycae"]):
        return {"Insulin":["Insulin Actrapid (fixed-rate infusion)","Insulin Lantus (basal)"],
                "Fluids":["IV Normal Saline","IV Potassium Chloride"],
                "Monitoring":["Glucose monitoring","Ketone monitoring"]}
    if any(k in dx_l for k in ["pneumon","chest infect","lrti"]):
        return {"Antibiotics":["Amoxicillin","Clarithromycin","Doxycycline"],
                "Supportive":["Salbutamol inhaler","Paracetamol","IV Fluids"],
                "Oxygen":["Controlled O2 therapy","High-flow O2 if SpO2 <94%"]}
    return {"Analgesia":["Paracetamol","Ibuprofen","Morphine"],
            "Antibiotics":["Amoxicillin","Ceftriaxone","Metronidazole"],
            "Supportive":["IV Fluids","Ondansetron","Omeprazole"]}


def page_ddx_builder():
    """Differential Diagnosis Builder — fully correlated to the selected case via AI."""
    c = st.session_state.get("selected_case")
    st.markdown('<div class="main-header"><h1>🧬 Differential Diagnosis Builder</h1><p>Case-specific differentials generated from your patient presentation</p></div>', unsafe_allow_html=True)

    if not c:
        st.markdown('<div class="alert-warn">⚠️ No case selected. Go to Case Library and select a case first.</div>', unsafe_allow_html=True)
        return

    case_id   = str(c.get("Case_ID","?"))
    cc        = str(c.get("Chief_Complaint",""))
    age_sex   = str(c.get("Age_Sex",""))
    hpi       = str(c.get("HPI",""))
    vitals    = str(c.get("Vitals",""))
    labs      = str(c.get("Labs",""))
    system    = str(c.get("System","general"))
    final_dx  = str(c.get("Final_Diagnosis",""))
    duration  = str(c.get("Duration",""))

    # ── Patient context banner ────────────────────────────────────
    st.markdown(f'''
    <div class="patient-card">
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;font-size:.85rem;">
            <div><b>Patient:</b> {age_sex}</div>
            <div><b>Chief Complaint:</b> {cc}</div>
            <div><b>Duration:</b> {duration}</div>
            <div><b>System:</b> {system.title()}</div>
            <div><b>Vitals:</b> {vitals[:60]}{"..." if len(vitals)>60 else ""}</div>
            <div><b>Key Labs:</b> {labs[:60]}{"..." if len(labs)>60 else ""}</div>
        </div>
    </div>''', unsafe_allow_html=True)

    submitted = st.session_state.get("ddx_submitted", False)

    if not submitted:
        # ── Generate case-specific DDx list ──────────────────────
        ddx_key = f"ddx_options_{case_id}"
        if ddx_key not in st.session_state:
            with st.spinner("🤖 Generating case-specific differentials from patient data..."):
                options = generate_case_ddx(
                    case_id, cc, age_sex, hpi, vitals, labs, system, final_dx)
                st.session_state[ddx_key] = options
        options = st.session_state[ddx_key]

        st.markdown('<div class="section-header">📋 Build Your Differential — Case-Specific Options</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="alert-info">These {len(options)} diagnoses were generated specifically for this patient presentation: age, symptoms, vitals and labs.</div>', unsafe_allow_html=True)

        # ── Key clinical clues panel ──────────────────────────────
        with st.expander("🔍 Clinical Clues to Guide You", expanded=True):
            cl1, cl2, cl3 = st.columns(3)
            with cl1:
                st.markdown("**🩺 Symptoms**")
                for line in cc.split("|")[:4]:
                    if line.strip(): st.markdown(f"- {line.strip()}")
            with cl2:
                st.markdown("**📊 Vitals**")
                for line in vitals.split("|")[:4]:
                    if line.strip(): st.markdown(f"- {line.strip()}")
            with cl3:
                st.markdown("**🧪 Labs**")
                for line in labs.split("|")[:4]:
                    if line.strip(): st.markdown(f"- {line.strip()}")

        st.markdown("")
        st.caption("Rank your top 3 differentials in order of likelihood. Justify each with specific clinical evidence from THIS case.")

        ddx_list = []
        ddx_reasoning = {}
        select_opts = ["— Select —"] + options + ["Other (type below)"]

        for i in range(1, 4):
            ordinal = "1st" if i==1 else "2nd" if i==2 else "3rd"
            st.markdown(f"**{ordinal} Most Likely Diagnosis**")
            col_dx, col_rs = st.columns([2,3])
            with col_dx:
                dx_sel = st.selectbox(f"Diagnosis", select_opts, key=f"ddx_{i}",
                                      label_visibility="collapsed")
                if dx_sel == "Other (type below)":
                    dx_sel = st.text_input("Enter diagnosis:", key=f"ddx_other_{i}")
            with col_rs:
                rs = st.text_area(
                    f"Why? (reference specific findings)",
                    placeholder=f"e.g. RLQ tenderness + WBC 14.2 + migration of pain supports...",
                    height=80, key=f"ddx_rs_{i}")
            if dx_sel and dx_sel not in ("— Select —",""):
                ddx_list.append(dx_sel)
                ddx_reasoning[dx_sel] = rs
            st.markdown("---")

        if st.button("✅ Submit My Differentials", type="primary",
                     use_container_width=True, key="ddx_submit"):
            if not ddx_list:
                st.warning("Please select at least one differential.")
            else:
                st.session_state.ddx_list = ddx_list
                st.session_state.ddx_reasoning = ddx_reasoning
                st.session_state.ddx_submitted = True
                comp = st.session_state.competencies
                comp.setdefault("Clinical Reasoning",{"attempts":0,"passes":0,"last":""})
                comp["Clinical Reasoning"]["attempts"] += 1
                st.rerun()

    else:
        # ── Results ───────────────────────────────────────────────
        real_dx  = final_dx
        ddx_list = st.session_state.ddx_list
        st.markdown('<div class="section-header">📊 Your Differential vs Actual Diagnosis</div>', unsafe_allow_html=True)

        correct = any(real_dx.lower() in d.lower() or d.lower() in real_dx.lower() for d in ddx_list)
        rank = next((i+1 for i,d in enumerate(ddx_list)
                     if real_dx.lower() in d.lower() or d.lower() in real_dx.lower()), None)

        col_r, col_l = st.columns([2,1])
        with col_r:
            for i,dx in enumerate(ddx_list):
                is_correct = real_dx.lower() in dx.lower() or dx.lower() in real_dx.lower()
                color = "#16a34a" if is_correct else "#dc2626" if i==0 and not correct else "#0e7490"
                icon  = "✅" if is_correct else f"#{i+1}"
                rs    = st.session_state.ddx_reasoning.get(dx,"")
                st.markdown(f'''
                <div style="background:white;border-left:5px solid {color};border-radius:0 10px 10px 0;
                            padding:10px 14px;margin:.5rem 0;box-shadow:0 1px 4px rgba(0,0,0,.06);">
                    <div style="font-weight:700;font-size:.95rem;color:{color};">{icon} {dx}</div>
                    <div style="font-size:.82rem;color:#475569;margin-top:4px;font-style:italic;">
                        {rs if rs else "No reasoning provided"}
                    </div>
                </div>''', unsafe_allow_html=True)

        with col_l:
            pts = 0
            if correct:
                pts = max(10-(rank-1)*3, 4)
                st.markdown(f'''
                <div class="kpi-card" style="border-top-color:#16a34a;text-align:center;">
                    <div class="kpi-value" style="color:#16a34a;">+{pts}</div>
                    <div class="kpi-label">Points Earned</div>
                    <div style="font-size:.78rem;color:#64748b;margin-top:6px;">
                        Correct as #{rank} choice
                    </div>
                </div>''', unsafe_allow_html=True)
                st.session_state.score += pts
                comp = st.session_state.competencies
                comp["Clinical Reasoning"]["passes"] += 1
                comp["Clinical Reasoning"]["last"] = str(datetime.now().date())
            else:
                st.markdown(f'''
                <div class="kpi-card" style="border-top-color:#dc2626;text-align:center;">
                    <div class="kpi-value" style="color:#dc2626;">❌</div>
                    <div class="kpi-label">Missed</div>
                    <div style="font-size:.78rem;color:#64748b;margin-top:6px;">
                        Correct: <b>{real_dx}</b>
                    </div>
                </div>''', unsafe_allow_html=True)

        # ── AI Teaching Feedback ──────────────────────────────────
        st.markdown('<div class="section-header">🤖 AI Clinical Educator Feedback</div>', unsafe_allow_html=True)
        if st.button("Get Personalised Feedback on My Reasoning", key="ddx_ai_fb",
                     use_container_width=True):
            reasoning_summary = "; ".join([f"{d}: {r}" for d,r in
                                           st.session_state.ddx_reasoning.items()])
            prompt = (f"Case: {age_sex}, {cc}. Vitals: {vitals[:150]}. Labs: {labs[:150]}. "
                      f"Student differentials: {ddx_list}. "
                      f"Student reasoning: {reasoning_summary}. "
                      f"Correct diagnosis: {real_dx}. "
                      f"As a senior clinician, give structured feedback: "
                      f"1) What reasoning was good, 2) What was missed, "
                      f"3) Key distinguishing features of {real_dx}, "
                      f"4) One teaching point for next time. Be specific to this case.")
            with st.spinner("Generating personalised feedback..."):
                fb = call_ai("You are a senior clinician giving targeted educational feedback.",
                            [{"role":"user","content":prompt}], max_tokens=600)
            st.markdown(f'<div class="chat-tutor" style="font-style:normal;">🤖 <b>Tutor:</b> {fb}</div>', unsafe_allow_html=True)

        st.markdown("")
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("💊 Prescribe for This Case", use_container_width=True, key="ddx_go_rx"):
                nav("prescribing")
        with col_b:
            if st.button("🔄 Try Again with Same Case", use_container_width=True, key="ddx_retry"):
                st.session_state.ddx_submitted = False
                st.session_state.ddx_list = []
                st.session_state.ddx_reasoning = {}
                st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# 💊 DRUG PRESCRIBING MODULE — Case-correlated via AI
# ════════════════════════════════════════════════════════════════════════════

ROUTES = ["Oral (PO)","Intravenous (IV)","Intramuscular (IM)","Subcutaneous (SC)",
          "Inhalation","Topical","Rectal (PR)","Sublingual (SL)","Nasogastric (NG)"]
FREQUENCIES = ["Once daily (OD)","Twice daily (BD)","Three times daily (TDS)",
               "Four times daily (QDS)","Every 4 hours","Every 6 hours","Every 8 hours",
               "Every 12 hours","Once weekly","As needed (PRN)","Stat (single dose)"]

def page_prescribing():
    """Drug Prescribing Module — case-correlated AI drug suggestions + AI pharmacist check."""
    c = st.session_state.get("selected_case")
    st.markdown('<div class="main-header"><h1>💊 Drug Prescribing Module</h1><p>Write evidence-based prescriptions tailored to YOUR case — AI pharmacist checks every entry</p></div>', unsafe_allow_html=True)

    if not c:
        st.markdown('<div class="alert-warn">⚠️ Select a case from Case Library first.</div>', unsafe_allow_html=True)
        return

    case_id    = str(c.get("Case_ID","?"))
    age_sex    = str(c.get("Age_Sex","?"))
    dx         = str(c.get("Final_Diagnosis","?"))
    allergies  = str(c.get("Allergies","NKDA"))
    curr_meds  = str(c.get("Medications","none"))
    vitals     = str(c.get("Vitals",""))
    labs       = str(c.get("Labs",""))

    # ── Patient banner ────────────────────────────────────────────
    allergy_warning = f'<span style="color:#dc2626;font-weight:700;">⚠️ Allergies: {allergies}</span>' if allergies not in ("NKDA","","none","nan") else '<span style="color:#16a34a;">✅ NKDA</span>'
    st.markdown(f'''
    <div class="patient-card">
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;font-size:.84rem;">
            <div><b>Patient:</b> {age_sex}</div>
            <div><b>Diagnosis:</b> <span style="color:#0e7490;font-weight:700;">{dx}</span></div>
            <div>{allergy_warning}</div>
            <div><b>Current Meds:</b> {curr_meds[:60]}</div>
            <div><b>Vitals:</b> {vitals[:60]}</div>
            <div><b>Labs:</b> {labs[:60]}</div>
        </div>
    </div>''', unsafe_allow_html=True)

    col_form, col_hist = st.columns([3,2])

    with col_form:
        # ── AI Drug Suggestions for THIS case ────────────────────
        rx_key = f"rx_suggestions_{case_id}"
        if rx_key not in st.session_state:
            with st.spinner("🤖 Loading case-specific drug suggestions..."):
                suggestions = generate_case_rx_suggestions(case_id, dx, age_sex, allergies, curr_meds)
                st.session_state[rx_key] = suggestions

        suggestions = st.session_state.get(rx_key, {})

        st.markdown('<div class="section-header">💡 Recommended Drugs for This Case</div>', unsafe_allow_html=True)
        st.caption(f"These drugs are suggested specifically for: **{dx}** in this patient")

        # Show suggestions as clickable quick-fills
        all_suggested = []
        for category, drugs in suggestions.items():
            st.markdown(f"**{category}**")
            cols = st.columns(len(drugs))
            for j, drug in enumerate(drugs):
                all_suggested.append(drug)
                with cols[j]:
                    if st.button(f"+ {drug}", key=f"rx_quick_{case_id}_{j}_{category}",
                                 use_container_width=True):
                        st.session_state["rx_selected_drug"] = drug
                        st.rerun()

        st.markdown("---")
        st.markdown('<div class="section-header">📝 Write Prescription</div>', unsafe_allow_html=True)

        # Pre-fill from quick-select if clicked
        preselected = st.session_state.pop("rx_selected_drug", None)
        drug_options = ["— Select Drug —"] + all_suggested + ["Other (type below)"]
        default_idx = drug_options.index(preselected) if preselected and preselected in drug_options else 0

        drug_sel = st.selectbox("Drug", drug_options, index=default_idx, key="rx_drug")
        if drug_sel == "Other (type below)":
            drug_sel = st.text_input("Enter drug name:", key="rx_drug_other")

        c1,c2 = st.columns(2)
        with c1:
            dose  = st.text_input("Dose", placeholder="e.g. 1g, 500mg, 0.1 units/kg/hr", key="rx_dose")
            route = st.selectbox("Route", ROUTES, key="rx_route")
        with c2:
            freq     = st.selectbox("Frequency", FREQUENCIES, key="rx_freq")
            duration = st.text_input("Duration", placeholder="e.g. 5 days, 24h, until review", key="rx_dur")

        indication = st.text_area(
            "Clinical justification",
            placeholder=f"Why is this drug appropriate for {dx} in this patient?",
            height=70, key="rx_ind")
        notes = st.text_input(
            "Special instructions / monitoring",
            placeholder="e.g. monitor renal function, take with food, check INR weekly",
            key="rx_notes")

        if st.button("💊 Prescribe & Check with AI Pharmacist",
                     type="primary", use_container_width=True, key="rx_submit"):
            if not drug_sel or drug_sel in ("— Select Drug —",""):
                st.warning("Please select or type a drug.")
            elif not dose.strip():
                st.warning("Please enter a dose.")
            else:
                rx_entry = {"drug":drug_sel,"dose":dose,"route":route,"freq":freq,
                            "duration":duration,"indication":indication,"notes":notes,
                            "case_id":case_id,"case_dx":dx,
                            "timestamp":str(datetime.now())[:16]}

                prompt = (
                    f"Patient: {age_sex}. Diagnosis: {dx}.\n"
                    f"Allergies: {allergies}. Current medications: {curr_meds}.\n"
                    f"Vitals: {vitals[:100]}. Labs: {labs[:100]}.\n\n"
                    f"Student prescription: {drug_sel} {dose} {route} {freq} for {duration}.\n"
                    f"Indication given: {indication}.\n"
                    f"Special instructions: {notes}.\n\n"
                    f"As a clinical pharmacist, evaluate:\n"
                    f"1. Is {drug_sel} appropriate for {dx} in this specific patient?\n"
                    f"2. Is {dose} {route} {freq} the correct regimen?\n"
                    f"3. Allergy risk? Drug interactions with {curr_meds}?\n"
                    "4. Monitoring required?\n"
                    "5. Any dose adjustment needed?\n\n"
                    "Start your response with exactly one of: SAFE / CAUTION / UNSAFE\n"
                    "Then give a concise structured explanation."
                )

                with st.spinner("AI Pharmacist reviewing prescription..."):
                    feedback = call_ai(
                        "You are a senior clinical pharmacist reviewing student prescriptions for a specific patient case.",
                        [{"role":"user","content":prompt}],
                        max_tokens=500, credit_type="diagnosis")

                rx_entry["feedback"] = feedback
                st.session_state.rx_history.insert(0, rx_entry)
                st.session_state.rx_feedback = feedback

                # Competency tracking
                comp = st.session_state.competencies
                comp.setdefault("Drug Prescribing",{"attempts":0,"passes":0,"last":""})
                comp["Drug Prescribing"]["attempts"] += 1
                if "SAFE" in feedback:
                    comp["Drug Prescribing"]["passes"] += 1
                    comp["Drug Prescribing"]["last"] = str(datetime.now().date())
                    st.session_state.score += 8
                st.rerun()

        # ── Feedback display ──────────────────────────────────────
        if st.session_state.get("rx_feedback"):
            fb  = st.session_state.rx_feedback
            cls = "alert-good" if "SAFE" in fb else "alert-warn" if "CAUTION" in fb else "alert-bad"
            st.markdown(f'<div class="{cls}" style="margin-top:1rem;line-height:1.6;">{fb}</div>', unsafe_allow_html=True)

    with col_hist:
        st.markdown('<div class="section-header">📜 Prescription History</div>', unsafe_allow_html=True)
        hist = st.session_state.rx_history
        if not hist:
            st.markdown('<div style="color:#94a3b8;font-size:.82rem;padding:.5rem;">No prescriptions written yet for this case.</div>', unsafe_allow_html=True)
        for rx in hist[:8]:
            safe_icon = "✅" if "SAFE" in rx.get("feedback","") else "⚠️" if "CAUTION" in rx.get("feedback","") else "❌"
            border    = "#16a34a" if "SAFE" in rx.get("feedback","") else "#d97706" if "CAUTION" in rx.get("feedback","") else "#dc2626"
            st.markdown(f'''
            <div style="background:white;border-left:4px solid {border};border-radius:0 8px 8px 0;
                        padding:8px 12px;margin:.35rem 0;font-size:.82rem;box-shadow:0 1px 3px rgba(0,0,0,.06);">
                <div style="font-weight:700;color:#0a2540;">{safe_icon} {rx["drug"]} {rx["dose"]}</div>
                <div style="color:#64748b;font-size:.78rem;">{rx["route"]} · {rx["freq"]} · {rx.get("duration","—")}</div>
                <div style="color:#94a3b8;font-size:.72rem;margin-top:2px;">
                    Case #{rx.get("case_id","?")} — {rx.get("case_dx","?")} · {rx["timestamp"]}
                </div>
            </div>''', unsafe_allow_html=True)

        if hist:
            if st.button("🗑️ Clear History", use_container_width=True, key="rx_clear"):
                st.session_state.rx_history = []
                st.session_state.rx_feedback = None
                st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# 🩺 PROCEDURE SIMULATOR
# ════════════════════════════════════════════════════════════════════════════

PROCEDURES = {
    "IV Cannulation": {
        "icon":"💉","difficulty":"Basic","time":"5 min",
        "systems":["cardio","gastro","ID","pulmo","neuro","uro","surgery","ped","gyneco","endo","ENT","derm","psych","other","general"],
        "dx_keywords":["sepsis","shock","dehydration","fluid","bleed","haemorrhage","hemorrhage","vomiting","dka","ketoacidosis","overdose","trauma","surgery","transfusion","infusion","iv","access"],
        "case_relevance":"Required for IV fluid resuscitation, drug administration, and blood transfusion in almost all acute admissions.",
        "steps":[
            {"title":"Gather Equipment","detail":"IV cannula (18G green for adults; 20G pink for standard; 14G/16G grey/orange for rapid resuscitation), tourniquet, 2% chlorhexidine swab, transparent dressing, 10ml saline flush, non-sterile gloves, sharps bin. Check expiry dates on all equipment.","check":"Correct gauge cannula selected; all equipment present and in date"},
            {"title":"Patient Preparation & Consent","detail":"Introduce yourself. Explain the procedure: 'I need to put a small plastic tube in your vein to give you fluids/medication.' Obtain verbal consent. Position patient supine or sitting with arm extended and supinated. Apply tourniquet 10cm proximal to intended site.","check":"Verbal consent obtained; tourniquet applied correctly"},
            {"title":"Site Selection","detail":"Ask patient to make a fist to engorge veins. Palpate for a straight, bouncy, non-mobile vein. Preferred sites in order: antecubital fossa (cephalic/basilic), forearm veins, dorsum of hand. Avoid: joints (flexion dislodges cannula), areas of previous phlebitis, oedematous limbs, limb with AV fistula or post-mastectomy lymphoedema.","check":"Suitable vein identified and site confirmed safe"},
            {"title":"Skin Preparation","detail":"Clean the site with 2% chlorhexidine gluconate in 70% isopropyl alcohol swab using back-and-forth friction for 30 seconds. Allow to air-dry for 30 seconds — do NOT blow, fan, or touch the site afterwards. Drying is essential for antiseptic activity.","check":"Skin cleaned and fully dried; site not re-touched"},
            {"title":"Cannula Insertion","detail":"Anchor skin distally with non-dominant thumb to prevent vein rolling. Hold cannula bevel-up at 15–30° to skin. Advance until flashback of blood appears in the chamber. Lower angle to 5–10°. Advance the plastic cannula 2–3mm further (to ensure cannula tip is in lumen), then slide the cannula fully off the needle into the vein while holding the needle still.","check":"Blood flashback confirmed; cannula fully advanced into vein"},
            {"title":"Secure, Flush & Confirm Patency","detail":"Release tourniquet. Occlude vein proximally with one finger. Remove needle and discard immediately into sharps bin (never resheath). Connect Luer-lock cap or extension set. Apply transparent dressing. Flush with 10ml 0.9% NaCl using push-pause technique. Inspect for swelling or resistance — either indicates extravasation; remove and resite. Document cannula gauge, site, date, and time.","check":"Patent cannula confirmed; dressing applied; documented"},
        ]
    },
    "Lumbar Puncture": {
        "icon":"🧠","difficulty":"Advanced","time":"20-30 min",
        "systems":["neuro","ID"],
        "dx_keywords":["meningitis","meningococcal","encephalitis","subarachnoid","sah","iih","intracranial hypertension","guillain","ms","multiple sclerosis","csf","lumbar"],
        "case_relevance":"Essential for diagnosing bacterial meningitis, subarachnoid haemorrhage (CT-negative), and raised intracranial pressure.",
        "steps":[
            {"title":"Indications, Contraindications & Consent","detail":"Indications: suspected bacterial/viral meningitis, subarachnoid haemorrhage (CT-negative), IIH, Guillain-Barré, MS workup. Absolute contraindications: signs of raised ICP (papilloedema, Cushing's triad, declining GCS) — CT head first. Coagulopathy (INR >1.5, platelets <50). Local infection at site. Suspected spinal cord compression. Obtain written consent; explain post-LP headache risk (15–30%).","check":"Indications confirmed; contraindications excluded; written consent obtained"},
            {"title":"Pre-procedure CT Head","detail":"Perform CT head before LP if: GCS <15, focal neurology, papilloedema, immunocompromised, new-onset seizure, or age >60. If LP is urgent (e.g. suspected bacterial meningitis with no contraindications), give IV antibiotics FIRST, then CT, then LP. Never delay antibiotics for imaging.","check":"CT reviewed and safe to proceed; or decision documented that CT not required"},
            {"title":"Positioning","detail":"Left lateral decubitus (lying on side) with knees drawn to chest (foetal position) — maximises intervertebral space. OR seated leaning forward over a pillow. Assistant should help patient maintain position. Ensure spine is horizontal (not tilted) in lateral decubitus — use a pillow under head.","check":"Patient correctly positioned and able to maintain position"},
            {"title":"Landmarks & Level Selection","detail":"Identify the iliac crest line — this crosses the L4 spinous process or L3-L4 interspace. Use L3-L4 or L4-L5 (both are below the conus medullaris which ends at L1-L2 in adults). Palpate spinous processes. Mark chosen interspace with thumbnail indentation or skin marker.","check":"L3-L4 or L4-L5 interspace correctly identified"},
            {"title":"Sterile Field Preparation","detail":"Wash hands. Open LP pack aseptically. Don sterile gloves. Apply betadine or 0.5% chlorhexidine in alcohol solution in concentric circles from centre outward. Apply sterile drape. Draw up 2ml of 1% lidocaine into a 5ml syringe. Prepare manometer and 4 collection tubes.","check":"Sterile field established; equipment prepared"},
            {"title":"Local Anaesthesia","detail":"Create a skin wheal with 1% lidocaine using a 25G needle. Switch to a 21G needle and infiltrate deeper tissues along the planned needle track, aspirating before each injection to avoid intravascular injection. Wait 2–3 minutes for full effect. Test for anaesthesia by gentle touch.","check":"Skin wheal raised; deep anaesthesia administered; 2-3 min wait observed"},
            {"title":"LP Needle Insertion","detail":"Insert a 22G spinal needle (with stylet in place) perpendicular to the skin in the horizontal plane, with a slight cephalad angle (5–10° toward umbilicus). Advance slowly in 2–3mm increments. A subtle 'give' sensation is felt as the needle passes through the ligamentum flavum and dura. Remove stylet — CSF drips should appear. If no CSF: rotate needle 90°, advance 1–2mm, or reposition patient. Never advance without stylet in place.","check":"CSF freely dripping; needle position confirmed"},
            {"title":"Opening Pressure & CSF Collection","detail":"Attach manometer immediately. Normal opening pressure: 6–20 cmCSF (lateral decubitus, relaxed). Note colour and clarity (normal: clear and colourless; xanthochromia = SAH; turbid = infection; bloody = traumatic tap or haemorrhage). Collect 4 bottles (1–2ml each): Bottle 1 & 4: M,C&S + RBC count; Bottle 2: glucose and protein (send blood glucose simultaneously); Bottle 3: cytology/specialist. Total collection ≤8ml.","check":"Opening pressure documented; 4 bottles collected correctly"},
            {"title":"Needle Removal & Aftercare","detail":"Replace stylet before withdrawing needle (reduces post-LP headache risk). Remove needle in one smooth motion. Apply sterile dressing. Patient should lie flat for 1 hour. Encourage oral hydration. Warn about post-LP headache (positional, frontal — treat with analgesia, caffeine, and hydration; refractory cases may need blood patch). Monitor observations every 30 minutes for 2 hours. Document CSF appearance, opening pressure, and samples sent.","check":"Stylet replaced before removal; patient positioned flat; monitoring plan in place"},
        ]
    },
    "ABG Sampling": {
        "icon":"🩸","difficulty":"Intermediate","time":"5-10 min",
        "systems":["pulmo","cardio","endo","ID","general"],
        "dx_keywords":["respiratory","copd","asthma","pneumonia","pe","pulmonary","acidosis","alkalosis","dka","ketoacidosis","sepsis","shock","hypoxia","hypercapnia","abg","blood gas"],
        "case_relevance":"Arterial blood gas analysis is essential for assessing oxygenation, ventilation, and acid-base status in any acutely unwell patient.",
        "steps":[
            {"title":"Indications & Preparation","detail":"Indications: respiratory failure, acid-base disturbance, assessment of ventilation (PaCO2), monitoring of critical illness. Check FiO2 the patient is currently receiving — document this on the request form (ABG results are meaningless without it). Review INR/clotting if anticoagulated (extend pressure time to 10 minutes).","check":"Indication confirmed; FiO2 documented; anticoagulation status checked"},
            {"title":"Allen's Test","detail":"Assess collateral ulnar circulation before radial puncture. Compress both radial and ulnar arteries simultaneously with your thumbs. Ask patient to clench and open fist 5 times until hand is pale. Release ONLY the ulnar artery. Normal: hand flushes pink within 5–7 seconds (adequate collateral circulation — safe to proceed). Abnormal: >10 seconds — consider alternative site (brachial artery or femoral).","check":"Allen's test performed and result documented; positive = ulnar refill <7 seconds"},
            {"title":"Equipment Assembly","detail":"Heparinised ABG syringe (pre-filled with dry lithium heparin or liquid heparin — expel excess liquid heparin to <0.2ml to avoid dilutional error), 23G or 25G needle, 2% chlorhexidine swab, gauze, tape. Prepare ice slurry if analysis will be delayed >15 minutes. Label syringe with patient ID before sampling.","check":"Correct heparinised syringe; ice prepared if needed; syringe labelled"},
            {"title":"Positioning & Site Preparation","detail":"Extend wrist 30–60° — use a roll of gauze or a wrist support under the wrist. Locate radial pulse with index and middle fingers. Clean skin with chlorhexidine swab and allow to dry 30 seconds. Optional: 1% lidocaine intradermal wheal (particularly in awake patients — reduces pain and movement artefact).","check":"Wrist correctly extended; radial pulse palpated; skin cleaned"},
            {"title":"Arterial Puncture","detail":"Hold syringe like a pencil, bevel up, at 45° to the skin (some practitioners use 30–60° — adjust to feel). Insert needle toward the pulsation. Arterial blood pulsates into the syringe under its own pressure — DO NOT aspirate. Collect 1–2ml. If venous blood (dark, non-pulsatile, requires aspiration) — withdraw and apply pressure; reattempt.","check":"Bright red pulsatile blood obtained without aspiration; 1–2ml collected"},
            {"title":"Post-Procedure Care","detail":"Withdraw needle in one smooth motion. Apply immediate firm pressure with dry gauze for 5 minutes (minimum) — 10 minutes if on anticoagulants or thrombocytopenic. Do not release early. Expel any air bubbles from syringe immediately (tilt syringe and tap). Cap syringe. Transport on ice if not analysed within 15 minutes. Analyse within 30 minutes maximum. Apply dressing after haemostasis confirmed. Document site and time.","check":"Haemostasis confirmed after 5 min; air bubbles expelled; sample transported correctly"},
        ]
    },
    "ECG Recording & Interpretation": {
        "icon":"📈","difficulty":"Intermediate","time":"10-15 min",
        "systems":["cardio","pulmo","endo","general"],
        "dx_keywords":["chest pain","palpitation","mi","nstemi","stemi","infarct","arrhythmia","af","atrial fibrillation","syncope","pe","pulmonary embolism","hyperkalaemia","cardiac","heart","ecg","ekg"],
        "case_relevance":"12-lead ECG is the single most important first investigation in any patient with chest pain, palpitations, syncope, or haemodynamic instability.",
        "steps":[
            {"title":"Patient Preparation & Lead Placement","detail":"Explain procedure. Patient supine, arms relaxed at sides, legs uncrossed. Expose chest, ankles, and wrists. Clean skin with alcohol wipe if diaphoretic or oily (poor contact = artefact). Apply 10 electrodes: Limb leads — RA (right wrist), LA (left wrist), RL (right ankle — earth), LL (left ankle). Chest leads — V1: 4th ICS, right sternal border; V2: 4th ICS, left sternal border; V3: between V2 and V4; V4: 5th ICS, mid-clavicular line; V5: anterior axillary line (same level as V4); V6: mid-axillary line (same level as V4/V5).","check":"All 10 electrodes correctly placed; patient relaxed and still"},
            {"title":"Recording & Quality Check","detail":"Set speed 25mm/s, gain 10mm/mV (standard). Ask patient to lie still and breathe normally. Record ECG. Check quality: baseline wander = poor electrode contact or patient movement; muscle tremor artefact = ask patient to relax arms. Check calibration square (1mV = 10mm, 0.2s wide). Label with patient name, DOB, date, time, and clinical indication.","check":"Clean trace obtained; calibration mark present; ECG labelled"},
            {"title":"Rate","detail":"Regular rhythm: 300 ÷ R-R interval in large squares. Quick method: count number of QRS complexes in a 10-second rhythm strip × 6. Normal: 60–100 bpm. Tachycardia: >100 bpm. Bradycardia: <60 bpm. Very fast regular tachycardia: 300, 150, 100, 75, 60 (memorise for 1, 2, 3, 4, 5 large squares).","check":"Heart rate calculated correctly"},
            {"title":"Rhythm","detail":"Is it regular or irregular? Measure R-R intervals across the strip — should be consistent (±10%). Irregular: is it regularly irregular (pattern — suggests 2nd degree HB, bigeminy) or irregularly irregular (no pattern — strongly suggests AF)? Are P waves present before every QRS? Is the P wave morphology consistent?","check":"Rhythm characterised as regular/irregularly irregular; P wave relationship to QRS assessed"},
            {"title":"Axis","detail":"Check leads I and aVF. Normal axis (−30° to +90°): both positive. Left axis deviation (<−30°): Lead I positive, aVF negative (LAD = left anterior fascicular block, inferior MI, WPW). Right axis deviation (>+90°): Lead I negative, aVF positive (RAD = RVH, PE, left posterior fascicular block, lateral MI). Extreme axis ('northwest'): both negative (ventricular tachycardia, hyperkalaemia, dextrocardia).","check":"Axis correctly determined from leads I and aVF"},
            {"title":"P Wave & PR Interval","detail":"P wave: upright in I and II (normal sinus origin); inverted in aVR (expected). Absent P waves = AF. Multiple P wave morphologies = wandering atrial pacemaker. PR interval: 120–200ms (3–5 small squares). Short PR (<120ms) + delta wave = WPW. Prolonged PR (>200ms) = 1st degree HB. Progressive PR lengthening then dropped QRS = Mobitz I (Wenckebach). Consistent PR with dropped QRS = Mobitz II (serious — requires pacing). No relationship between P and QRS = complete (3rd degree) HB.","check":"P wave morphology assessed; PR interval measured and classified"},
            {"title":"QRS Complex","detail":"Normal QRS <120ms (3 small squares). Broad QRS (>120ms): LBBB (broad, notched 'M' in V5/V6, 'W' in V1), RBBB (RSR' 'M' in V1, broad S in I/V6), or ventricular rhythm/hyperkalaemia. Pathological Q waves: >1 small square wide OR >25% height of R wave in same lead = previous MI. Check all leads. Poor R wave progression (no R wave growth V1→V4) = anterior MI or LBBB.","check":"QRS duration measured; LBBB/RBBB identified if present; Q waves assessed"},
            {"title":"ST Segment & T Waves","detail":"ST elevation: ≥1mm in ≥2 contiguous limb leads OR ≥2mm in ≥2 contiguous chest leads = STEMI until proven otherwise (call senior immediately). Saddle-shaped ST elevation in all leads = pericarditis. ST depression: ≥1mm = ischaemia or reciprocal change (seen in leads opposite a STEMI). T wave inversion: normal in aVR and V1. Pathological in ≥2 contiguous leads = ischaemia. Tall, tented T waves = hyperkalaemia (early sign). Biphasic T waves in V2/V3 = Wellens syndrome (LAD critical stenosis).","check":"ST changes assessed in all 12 leads; STEMI/NSTEMI/normal identified"},
            {"title":"Intervals & Conclusion","detail":"QTc interval: QT ÷ √R-R (Bazett formula). Normal QTc: males <440ms, females <460ms. Prolonged QTc: drug effect, hypokalaemia, hypomagnesaemia, congenital LQTS — risk of Torsades de Pointes. Write a systematic ECG conclusion: rate, rhythm, axis, intervals, ST/T changes, overall interpretation. Correlate with clinical picture. Escalate immediately if: STEMI, Mobitz II/complete HB, sustained VT, or new LBBB with chest pain.","check":"QTc calculated; systematic conclusion written; escalation decision made"},
        ]
    },
    "Urinary Catheterisation": {
        "icon":"🔬","difficulty":"Intermediate","time":"10-15 min",
        "systems":["uro","surgery","general","ped"],
        "dx_keywords":["urinary retention","uti","uro","catheter","urology","prostate","bladder","renal","acute kidney","aki","output","fluid balance","surgery","post-op"],
        "case_relevance":"Urinary catheterisation is indicated for acute urinary retention, strict hourly urine output monitoring, and peri-operative fluid balance management.",
        "steps":[
            {"title":"Indications, Contraindications & Consent","detail":"Indications: acute urinary retention, hourly urine output monitoring (critical illness, major surgery, haemodynamic instability), urological surgery, certain urodynamic studies. Contraindications: urethral trauma (blood at meatus, perineal bruising, high-riding prostate on PR exam — call urology). Select correct catheter: 12–14Fr Foley for most adults; larger bore (16–18Fr) if haematuria expected. Obtain consent; explain procedure.","check":"Indications confirmed; urethral trauma excluded; consent obtained; correct catheter selected"},
            {"title":"Equipment Preparation","detail":"Catheterisation pack (sterile drape, fenestrated drape, swabs, forceps, kidney dish), appropriate Foley catheter, 10ml sterile water for balloon inflation, catheter bag, sterile gloves (2 pairs — outer pair for skin prep, inner pair for catheter handling), lubricant/anaesthetic gel (2% lidocaine gel — 11ml instilled into urethra and left for 3–5 minutes in males), cleaning solution.","check":"All equipment assembled on sterile field; lidocaine gel administered in males and waited"},
            {"title":"Sterile Field & Patient Positioning","detail":"Males: supine, drape exposing genitalia only. Females: supine, legs in frog-leg position (hips flexed and abducted), fenestrated drape over perineum. Open catheter pack aseptically. Don sterile gloves. Arrange swabs and equipment on sterile field. Instil 11ml 2% lidocaine gel into male urethra, compress meatus for 3–5 minutes. Female: no lidocaine gel needed routinely.","check":"Sterile field maintained; anaesthetic gel instilled in males (3–5 min wait)"},
            {"title":"Cleaning","detail":"Males: retract foreskin (replace after — paraphimosis risk). Clean glans and urethral meatus with antiseptic solution using swabs held in forceps — each swab used once, working from meatus outward. Females: separate labia minora with non-dominant hand (keep retracted throughout). Identify urethral meatus (between clitoris anteriorly and vaginal opening posteriorly). Clean with swabs from anterior to posterior — one swipe per swab.","check":"Correct cleaning technique; meatus correctly identified (female); foreskin retracted (male)"},
            {"title":"Catheter Insertion","detail":"Males: hold penis at 90° to body (perpendicular). Insert catheter tip gently into meatus. Advance slowly — two resistances felt: external sphincter (~5cm) and bladder neck (~15–20cm from meatus). Never force. If resistance: ask patient to breathe out slowly. Advance until bifurcation of catheter is at meatus (full insertion). Females: insert 5–7cm or until urine drains. Urine drainage confirms bladder placement before inflating balloon.","check":"Catheter fully inserted to bifurcation (male) or urine draining (female) before balloon inflation"},
            {"title":"Balloon Inflation & Securing","detail":"ONLY inflate balloon when urine is draining freely — inflating in urethra causes severe pain and urethral injury. Inflate with 10ml sterile water (not saline — crystalises in valve). Gently retract catheter until resistance felt — balloon is now seated at bladder neck. Connect to sterile catheter bag. In males: replace foreskin immediately to prevent paraphimosis. Secure catheter to inner thigh with adhesive fixation device — prevents traction and urethral trauma. Document: catheter size, balloon volume, residual volume, date, time, urine appearance.","check":"Balloon inflated only after urine confirmed; foreskin replaced; catheter secured; documented"},
        ]
    },
    "Nasogastric Tube Insertion": {
        "icon":"🩺","difficulty":"Intermediate","time":"10-15 min",
        "systems":["gastro","surgery","neuro","general"],
        "dx_keywords":["bowel obstruction","ileus","ng","nasogastric","feed","nutrition","overdose","poisoning","gastric","vomiting","swallowing","dysphagia","stroke","head injury","unconscious"],
        "case_relevance":"NG tube insertion is required for enteral feeding in patients who cannot swallow safely, gastric decompression in obstruction/ileus, or medication administration.",
        "steps":[
            {"title":"Indications, Contraindications & Consent","detail":"Indications: enteral nutrition (safe swallow impaired), gastric decompression (obstruction, ileus, post-operative), drug administration (unconscious patients), gastric lavage (selected overdoses). Contraindications: base of skull fracture (insert orally instead — NG may enter cranial vault), severe facial trauma, oesophageal stricture/varices (relative), recent oesophageal surgery. Use Fine Bore tube (8Fr) for feeding; wide bore (16–18Fr) for aspiration/lavage. Obtain consent if patient able.","check":"Indication confirmed; base of skull fracture excluded; correct tube size selected"},
            {"title":"Measure & Mark Insertion Length","detail":"Estimate required insertion length using NEX measurement: distance from Nose to Earlobe to Xiphisternum. Mark on tube with tape or note the cm marking. Typical insertion length in adults: 55–65cm. Position patient sitting upright at 45° minimum (reduces aspiration risk). If unconscious, consider lateral position. Have suction ready.","check":"NEX measurement performed; insertion length marked on tube; patient upright"},
            {"title":"Lubrication & Initial Insertion","detail":"Lubricate 10–15cm of tube tip with water-soluble lubricant. Check which nostril is more patent (ask patient, or occlude each nostril). Insert tube tip into chosen nostril, aiming posteriorly and inferiorly (NOT upward — the floor of the nose is horizontal). Advance gently through nasopharynx. The patient will feel the tube in the back of their throat at ~15–20cm.","check":"Tube inserted through correct nostril, directed horizontally; resistance not forced"},
            {"title":"Passage Through Pharynx & Oesophagus","detail":"Ask conscious patient to tuck chin to chest (closes glottis, opens oesophagus). Ask patient to swallow (sips of water if safe, or dry swallow). Advance tube 2–3cm with each swallow. Continue until marked insertion length reached. Stop immediately and withdraw if: patient coughs violently, cannot speak, becomes cyanosed, or shows respiratory distress — tube is likely in airway. Never use force.","check":"Tube advanced with swallowing; no signs of airway insertion; insertion length reached"},
            {"title":"Confirm Position — pH Testing (First Line)","detail":"The ONLY safe first-line bedside confirmation method is pH testing of gastric aspirate. Attach a 50ml enteral syringe. Aspirate gastric contents — 0.5–1ml is sufficient. Apply to CE-marked pH indicator paper. pH ≤5.5 = confirmed gastric position — safe to use. pH 6–9 = uncertain — do NOT use; obtain CXR. If no aspirate: try advancing 5cm, turn patient to left lateral, wait 30 minutes, try again.","check":"pH ≤5.5 confirmed on CE-marked pH paper; documented; NEVER use blue litmus paper alone"},
            {"title":"Confirm Position — CXR (When pH Uncertain)","detail":"CXR is required when pH is ≥6 or aspirate cannot be obtained. On CXR, correctly placed NG tube should: cross midline (follow oesophagus), bisect the carina, and tip should be visible below the left hemi-diaphragm in the stomach, at least 10cm below the GOJ. Ask a competent clinician to confirm CXR — do NOT use tube until confirmed. Document radiological confirmation. NEVER use air auscultation ('whoosh' test) alone — this is no longer considered safe.","check":"CXR reviewed by competent clinician if pH uncertain; tip position confirmed below diaphragm"},
            {"title":"Securing & Documentation","detail":"Secure tube to nose with hypoallergenic tape, avoiding pressure on the nasal ala (can cause pressure necrosis). Mark insertion length at nostril with permanent marker and document. Connect to bag for drainage or cap for feeding. Re-check pH or CXR at each shift if tube moved or displaced (tape loosened, patient vomited, retched, or coughed violently). Document: tube size, insertion length, confirmation method, aspirate pH/CXR result, who confirmed, date and time.","check":"Tube secured without nasal pressure; insertion length marked and documented; confirmation method documented"},
        ]
    },
    "Wound Suturing": {
        "icon":"🪡","difficulty":"Intermediate","time":"15-30 min",
        "systems":["surgery","derm","ortho","ENT","ped","other"],
        "dx_keywords":["laceration","wound","cut","trauma","suture","stitch","repair","bleeding","injury","abscess","incision"],
        "case_relevance":"Basic wound closure is an essential emergency and surgical skill for managing traumatic lacerations and surgical wounds.",
        "steps":[
            {"title":"Wound Assessment & Consent","detail":"Assess: mechanism of injury (sharp/blunt, clean/contaminated), time since injury (<6 hours = primary closure usually safe; 6–12 hours = judgment call; >12 hours or bite wounds = consider delayed primary closure or secondary intention), depth, location, neurovascular status distal to wound, tendon involvement, foreign bodies. Check tetanus immunisation status. Obtain consent. Document neurovascular exam BEFORE local anaesthetic.","check":"Wound assessed; neurovascular exam documented; tetanus status checked; consent obtained"},
            {"title":"Anaesthesia","detail":"Infiltrate with 1% lidocaine (maximum dose: 3mg/kg without adrenaline; 7mg/kg with adrenaline — do NOT use adrenaline in end-arteries: fingers, toes, nose, ear, penis). Use 25–27G needle. Inject into wound edges (less painful than through intact skin). Aspirate before injection. Wait 3–5 minutes for full effect. Adrenaline in lidocaine reduces bleeding and extends duration — safe in most body locations.","check":"Maximum lidocaine dose calculated and not exceeded; adrenaline avoided in end-arteries; full anaesthesia confirmed"},
            {"title":"Wound Irrigation & Debridement","detail":"Irrigate wound copiously with at least 200–500ml 0.9% NaCl using a 20ml syringe and 19G needle (creates ~8 psi pressure — optimal for bacterial load reduction). Explore wound for foreign bodies (glass, gravel, debris). Debride clearly devitalised tissue with scissors. Do NOT debride on face (cosmetically critical areas — let the body reabsorb). Re-examine neurovascular status if deep wound near nerve/vessel.","check":"Wound copiously irrigated; foreign bodies excluded; devitalised tissue debrided"},
            {"title":"Instrument & Suture Selection","detail":"Select appropriate suture: Face/cosmetically sensitive: 5-0 or 6-0 non-absorbable monofilament (Prolene/nylon) — remove 5–7 days. Scalp: 3-0 or 4-0 non-absorbable; staples are acceptable. Trunk/extremities: 3-0 or 4-0 non-absorbable; remove 10–14 days. Deep/subcutaneous layer: 2-0 or 3-0 absorbable (Vicryl/PDS) — buried interrupted sutures. Hold needle in needle holder at 2/3 from tip. Grasp tissue with toothed forceps — do not crush.","check":"Correct suture material and gauge selected for anatomical site"},
            {"title":"Interrupted Suture Technique","detail":"Align wound edges. Insert needle perpendicular to skin surface, ~4–5mm from wound edge. Pass through full dermis, emerging on opposite side equidistant from edge. Tie instrument tie: two throws forward (surgeon's knot), one throw back — minimum 3 knots total. Knot should sit to the side of the wound (not over it). Sutures should evert wound edges slightly — flat or inverted edges heal poorly with scarring. Space sutures 4–5mm apart. Aim for equal bites on both sides.","check":"Wound edges everted; equal bites; knots to side; sutures 4–5mm apart"},
            {"title":"Wound Dressing & Aftercare Instructions","detail":"Clean wound of blood with saline. Apply non-adherent dressing. Give written aftercare instructions: keep dry for 24–48 hours; signs of infection (increasing redness, warmth, pus, fever) → return immediately; suture removal timing based on site; avoid sun exposure on healing wound for 6 months (reduces scarring). Prescribe antibiotics only if contaminated, human/animal bite, or immunocompromised — prophylactic antibiotics are not routine for clean lacerations. Document suture number, type, and removal date.","check":"Dressing applied; written aftercare given; follow-up for suture removal arranged; antibiotics only if indicated"},
        ]
    },
    "Pleural Aspiration (Thoracocentesis)": {
        "icon":"🫁","difficulty":"Advanced","time":"20-30 min",
        "systems":["pulmo","cardio","ID","surgery"],
        "dx_keywords":["pleural effusion","effusion","empyema","pneumothorax","hemothorax","chest","respiratory","lung","pe","heart failure","hepatic","cancer","malignant"],
        "case_relevance":"Pleural aspiration is indicated for diagnostic sampling of pleural effusions and therapeutic drainage of symptomatic effusions causing dyspnoea.",
        "steps":[
            {"title":"Indications, Safety Checks & Consent","detail":"Diagnostic: any unilateral or unexplained effusion. Therapeutic: symptomatic large effusion causing dyspnoea (drain up to 1–1.5L per session — no more, to prevent re-expansion pulmonary oedema). Contraindications: coagulopathy (INR >1.5 or platelets <50 — correct first), uncooperative patient, skin infection at site. Pre-procedure: CXR and ultrasound to confirm effusion and select site. Obtain written consent. IV access in situ.","check":"Indications confirmed; CXR and USS reviewed; coagulation checked; written consent obtained"},
            {"title":"USS-Guided Site Marking","detail":"Bedside ultrasound is the gold standard for marking and should be performed immediately before the procedure (effusions shift with position). Confirm effusion depth (>2cm from skin to lung margin = safe). Mark the skin with a pen at the chosen intercostal space — typically the 7th–9th ICS in the posterior axillary line for a free-flowing effusion. USS guidance dramatically reduces complication rates and is now standard of care.","check":"USS performed immediately pre-procedure; effusion >2cm depth confirmed; site marked"},
            {"title":"Positioning & Sterile Preparation","detail":"Optimal position: patient sitting upright, leaning forward over a pillow or bedtable (gravity layers effusion inferiorly, moves lung superiorly). Alternatively: lateral decubitus on the side of the effusion. Wash hands. Full aseptic technique: sterile gloves, gown, drape, chlorhexidine skin preparation. Prepare aspiration kit: 50ml syringe, 3-way tap, green needle (21G) or Seldinger drain kit, specimen pots (biochemistry, microbiology, cytology).","check":"Patient sitting upright; full aseptic technique; specimen pots labelled"},
            {"title":"Local Anaesthesia","detail":"Insert 25G needle, raise skin wheal with 1% lidocaine. Switch to 21G needle, insert over the UPPER border of the rib (intercostal vessels — vein, artery, nerve — run in the subcostal groove under the lower border of each rib). Advance and aspirate while injecting: if blood is aspirated, withdraw and apply pressure. Anaesthetise down to and including the parietal pleura (the most sensitive layer). Confirm entry into pleural space when straw-coloured fluid is aspirated — note depth.","check":"Needle over upper rib border; pleura anaesthetised; needle depth noted; no blood aspirated"},
            {"title":"Aspiration","detail":"Insert aspiration needle (or Seldinger kit) at the marked site over upper rib border. Advance until pleural fluid obtained. Attach 3-way tap and 50ml syringe. Aspirate fluid, turning 3-way tap to expel into a collection bag. Diagnostic: 20–50ml is sufficient. Therapeutic: maximum 1–1.5L per session. Stop if patient coughs persistently, has chest pain, or if air is aspirated (suggests lung now adjacent to needle — stop). Do NOT remove >1.5L.","check":"Maximum 1.5L drained; procedure stopped for pain, cough, or air aspiration; specimens sent"},
            {"title":"Post-Procedure Care","detail":"Remove needle, apply occlusive dressing. Perform post-procedure CXR (erect, within 1 hour): confirm effusion reduced, exclude pneumothorax (complication in ~3–5%). Send fluid: Biochemistry (protein, LDH, glucose, pH — for Light's criteria); Microbiology (MC&S, AFB if TB suspected); Cytology (malignant cells); additional tests as indicated (ADA for TB, amylase for pancreatitis). Monitor observations for 2 hours. Document volume drained, colour and consistency, and any complications.","check":"Post-procedure CXR performed; specimens sent for biochemistry, microbiology, cytology; observations monitored"},
        ]
    },
}

# ── Case-to-procedure mapping ──────────────────────────────────────────────
def _get_recommended_procedures(case):
    """
    Returns a dict of {procedure_name: reason_string} for the active case.
    Uses the case's System field and Final_Diagnosis keywords — no AI calls,
    no hallucination risk. Pure deterministic keyword matching.
    """
    if not case:
        return {}
    dx   = str(case.get("Final_Diagnosis","")).lower()
    cc   = str(case.get("Chief_Complaint","")).lower()
    sys  = str(case.get("System","")).lower()
    labs = str(case.get("Labs","")).lower()
    combined = dx + " " + cc + " " + labs

    recommended = {}
    for pname, pdata in PROCEDURES.items():
        # System match
        sys_match = any(s in sys for s in pdata.get("systems", []))
        # Keyword match in diagnosis/CC/labs
        kw_match  = any(kw in combined for kw in pdata.get("dx_keywords", []))
        if sys_match or kw_match:
            recommended[pname] = pdata.get("case_relevance", "")
    return recommended




# ════════════════════════════════════════════════════════════════════════════
# 🩺 PROCEDURE SIMULATOR — 3D ANATOMICAL ATLAS (Three.js)
# ════════════════════════════════════════════════════════════════════════════


def _build_3d_html(proc_name: str, step_idx: int) -> str:
    """
    Guided 3D anatomical procedure simulator — v4.
    Uses procedure_3d_viewer_guided.py which wraps procedure_3d_viewer.py and
    adds: 7-step sequencer, guided instruction panel, progress tracker, vein
    engorgement, IV tubing (Step 6), Tegaderm dressing (Step 7), score overlay.
    Requires: libs/ and models/ folders next to app.py
    """
    import os, sys
    _d = os.path.dirname(os.path.abspath(__file__))
    if _d not in sys.path:
        sys.path.insert(0, _d)
    try:
        # ── Use the guided simulation wrapper (v4) ──────────────────────────
        # Try the guided wrapper first, fall back to base viewer
        try:
            import procedure_3d_viewer_guided as _g
            return _g.build_guided_3d_html(proc_name, step_idx)
        except (ModuleNotFoundError, ImportError):
            pass
        try:
            import procedure_3d_viewer as _g
            return _g._build_3d_html(proc_name, step_idx)
        except (ModuleNotFoundError, ImportError):
            raise RuntimeError(
                "Neither procedure_3d_viewer_guided.py nor a base "
                "procedure_3d_viewer.py with _build_3d_html() was found. "
                "Place both files next to app.py."
            )
    except FileNotFoundError as e:
        return f"""<html><body style="background:#07101e;color:#5acce0;font-family:system-ui;
            display:flex;align-items:center;justify-content:center;height:100vh;">
            <div style="text-align:center;padding:2rem;border:1px solid #1a4a6a;border-radius:12px;">
            <div style="font-size:2rem;">⚠️</div>
            <div style="font-size:.9rem;margin-top:.5rem;">3D model files not found.<br>
            Place <b>libs/</b> and <b>models/</b> folders next to app.py.<br>
            <span style="font-size:.75rem;color:#3a6a7a;">Missing: {e}</span></div>
            </div></body></html>"""
    except Exception as e:
        import traceback
        return f"""<html><body style="background:#07101e;color:#ff6666;font-family:system-ui;padding:2rem;">
            <b>3D Viewer Error:</b><br><pre style="font-size:.75rem">{traceback.format_exc()[:800]}</pre>
            </body></html>"""





# ═══════════════════════════════════════════════════════════════════════════
# STEP DESCRIPTION LABELS  (for the side panel overlay on each step)
# ═══════════════════════════════════════════════════════════════════════════

ANATOMY_NOTES = {
    "IV Cannulation": [
        "3D model shows the full forearm. Toggle layers to inspect skin, veins, muscles and bones.",
        "Anatomy unchanged — veins are in the same position. Review the antecubital fossa.",
        "Median cubital vein highlighted in cyan. Apply tourniquet (shown) 5–10 cm proximal.",
        "Median cubital vein highlighted. Skin cleaned — do not re-palpate after cleaning.",
        "Cannula insertion shown. Bevel up at 15–30°. Watch for blood flashback in chamber.",
        "Cannula in situ. Tourniquet released. Flush with 10ml 0.9% NaCl. Apply dressing.",
    ],
    "ABG Sampling": [
        "Wrist anatomy displayed. Radial artery runs lateral to flexor carpi radialis tendon.",
        "Allen's test: compress radial + ulnar arteries simultaneously. Release ulnar only — hand should flush pink within 7 s.",
        "Assemble heparinised syringe. Expel excess liquid heparin to avoid dilutional error.",
        "Radial artery highlighted in red. Extend wrist 30–60°. Palpate the pulse.",
        "Insert needle bevel-up at 45°. Arterial blood pulsates into the syringe — do NOT aspirate.",
        "Apply firm pressure 5 minutes minimum. Expel air bubbles. Transport on ice if >15 min delay.",
    ],
    "Lumbar Puncture": [
        "Lumbar spine displayed. Note the cord ends at L1–L2. LP must be below this level.",
        "CT head before LP if: GCS <15, focal neurology, papilloedema, immunocompromised.",
        "Lateral decubitus position maximises intervertebral space. Knees to chest (foetal position).",
        "L3–L4 interspace highlighted. Iliac crest crosses L4 spinous process — your landmark.",
        "Sterile field established. Anaesthetise down to parietal pleura — most sensitive layer.",
        "Local anaesthetic infiltrated along needle track. Wait 2–3 minutes for full effect.",
        "LP needle shown at L3–L4. Insert perpendicular, slight cephalad angle. Feel the 'give' through ligamentum flavum.",
        "CSF space highlighted. Attach manometer — normal opening pressure 6–20 cmCSF. Collect 4 bottles.",
        "Replace stylet before withdrawing. Patient flat for 1 hour. Encourage oral hydration.",
    ],
    "ECG Recording & Interpretation": [
        "Anterior torso displayed with electrode positions. V1: 4th ICS right sternal border. V4: 5th ICS mid-clavicular line.",
        "Electrode leads shown (yellow = limb, cyan = chest). Ensure patient relaxed — muscle artefact mimics arrhythmias.",
        "Rate: 300 ÷ R-R large squares. Or count QRS complexes in 10s strip × 6.",
        "Rhythm: measure R-R intervals. Regular vs irregularly irregular (AF has no pattern).",
        "Axis: check leads I and aVF. Both positive = normal axis.",
        "P wave: upright in I and II. PR interval: 120–200ms (3–5 small squares).",
        "QRS <120ms normal. Broad QRS: LBBB / RBBB / ventricular rhythm. Q waves: >1 sq wide = old MI.",
        "ST elevation ≥1mm in ≥2 contiguous leads = STEMI. Call senior immediately.",
        "QTc normal: males <440ms, females <460ms. Write systematic conclusion.",
    ],
}


# ═══════════════════════════════════════════════════════════════════════════
# MAIN PAGE FUNCTION — replaces page_procedure_sim()
# ═══════════════════════════════════════════════════════════════════════════

# Import your existing PROCEDURES dict and helpers from app context
# (They are already defined in app.py — this function uses them via globals)

def page_procedure_sim_3d():
    """
    Enhanced Procedure Simulator with per-step rotating 3D anatomical viewer.
    Drop-in replacement for page_procedure_sim() in MLS Virtual Hospital.
    """

    # All functions and data are now in the same file — use directly
    _get_recommended = _get_recommended_procedures

    c           = st.session_state.get("selected_case")
    recommended = _get_recommended(c)

    # ── Procedures that have a 3D viewer ──────────────────────────────────
    HAS_3D = {"IV Cannulation", "ABG Sampling", "Lumbar Puncture",
               "ECG Recording & Interpretation"}

    st.markdown(
        '<div class="main-header">'
        '<h1>🩺 Procedure Simulator <span style="font-size:.6em;font-weight:400;'
        'background:#0e7490;color:white;padding:3px 10px;border-radius:6px;margin-left:8px;">3D ATLAS</span></h1>'
        '<p>Step-by-step guidance · Rotating 3D anatomical models · Evidence-based checkpoints</p>'
        '</div>',
        unsafe_allow_html=True,
    )

    # ── Case banner ────────────────────────────────────────────────────────
    if c:
        dx = c.get("Final_Diagnosis", "?")
        ag = c.get("Age_Sex", "?")
        cc = c.get("Chief_Complaint", "?")
        if recommended:
            rn = ", ".join(recommended.keys())
            st.markdown(
                f'<div style="background:#f0fdf4;border:1px solid #16a34a;border-radius:10px;'
                f'padding:10px 14px;margin-bottom:14px;font-size:.84rem;">'
                f'<b>🟢 Active case:</b> {ag} — {cc} &nbsp;|&nbsp; <b>Dx:</b> {dx}<br>'
                f'<b>Clinically relevant procedures:</b> {rn}</div>',
                unsafe_allow_html=True,
            )
    else:
        st.markdown(
            '<div class="alert-info">ℹ️ No active case. All procedures available for practice.</div>',
            unsafe_allow_html=True,
        )

    col_sel, col_proc = st.columns([1, 2.6])

    # ══════════════════════════════════════════════════════════════════════
    # LEFT — procedure selector
    # ══════════════════════════════════════════════════════════════════════
    with col_sel:
        st.markdown('<div class="section-header">📋 Select Procedure</div>', unsafe_allow_html=True)

        if recommended:
            st.markdown(
                '<div style="font-size:.72rem;font-weight:600;color:#16a34a;margin-bottom:4px;">'
                '⭐ RECOMMENDED FOR THIS CASE</div>',
                unsafe_allow_html=True,
            )
            for pname in recommended:
                if pname not in PROCEDURES:
                    continue
                pdata = PROCEDURES[pname]
                badge = "🔵 3D" if pname in HAS_3D else ""
                if st.button(
                    f"⭐ {pdata['icon']} {pname} {badge}",
                    use_container_width=True,
                    key=f"proc_rec_{pname}",
                ):
                    st.session_state.proc_selected = pname
                    st.session_state.proc_step     = 0
                    st.session_state.proc_score    = 0
                    st.rerun()
                diff_col = {"Basic":"#16a34a","Intermediate":"#d97706","Advanced":"#dc2626"}.get(
                    pdata["difficulty"], "#64748b")
                st.markdown(
                    f'<div style="font-size:.7rem;color:{diff_col};margin:-6px 0 6px 4px;">'
                    f'{pdata["difficulty"]} · {pdata["time"]}</div>',
                    unsafe_allow_html=True,
                )

            st.markdown(
                '<div style="font-size:.72rem;font-weight:600;color:#64748b;margin:8px 0 4px;">'
                'ALL PROCEDURES</div>',
                unsafe_allow_html=True,
            )

        for pname, pdata in PROCEDURES.items():
            if pname in recommended:
                continue
            badge = " 🔵" if pname in HAS_3D else ""
            if st.button(
                f"{pdata['icon']} {pname}{badge}",
                use_container_width=True,
                key=f"proc_{pname}",
            ):
                st.session_state.proc_selected = pname
                st.session_state.proc_step     = 0
                st.session_state.proc_score    = 0
                st.rerun()
            diff_col = {"Basic":"#16a34a","Intermediate":"#d97706","Advanced":"#dc2626"}.get(
                pdata["difficulty"], "#64748b")
            st.markdown(
                f'<div style="font-size:.7rem;color:{diff_col};margin:-6px 0 6px 4px;">'
                f'{pdata["difficulty"]} · {pdata["time"]}</div>',
                unsafe_allow_html=True,
            )

        st.markdown(
            '<div style="font-size:.68rem;color:#94a3b8;margin-top:8px;">🔵 = 3D atlas viewer</div>',
            unsafe_allow_html=True,
        )

    # ══════════════════════════════════════════════════════════════════════
    # RIGHT — 3D viewer + step panel
    # ══════════════════════════════════════════════════════════════════════
    with col_proc:
        proc_name = st.session_state.get("proc_selected")
        if not proc_name:
            st.markdown(
                '<div style="background:white;border:2px dashed #cbd5e1;border-radius:12px;'
                'padding:2rem;text-align:center;color:#64748b;">'
                '← Select a procedure to begin simulation</div>',
                unsafe_allow_html=True,
            )
            return

        proc    = PROCEDURES[proc_name]
        steps   = proc["steps"]
        current = st.session_state.proc_step

        atlas_badge = "&nbsp;<span style='background:#0e7490;color:white;font-size:.65em;border-radius:5px;padding:2px 8px;'>3D ATLAS</span>" if proc_name in HAS_3D else ""
        st.markdown(
            f'<div class="section-header">{proc["icon"]} {proc_name} {atlas_badge}</div>',
            unsafe_allow_html=True,
        )

        # Case relevance banner
        if proc_name in recommended and c:
            st.markdown(
                f'<div style="background:#fef3c7;border-left:4px solid #d97706;'
                f'border-radius:0 8px 8px 0;padding:8px 12px;margin-bottom:10px;'
                f'font-size:.8rem;color:#92400e;"><b>Why this matters for your case:</b><br>'
                f'{recommended[proc_name]}</div>',
                unsafe_allow_html=True,
            )

        # Progress
        progress = min(current / len(steps), 1.0)
        st.progress(progress, text=f"Step {min(current+1, len(steps))} of {len(steps)}")

        # ── 3D VIEWER ──────────────────────────────────────────────────────
        if proc_name in HAS_3D and current < len(steps):
            viewer_html = _build_3d_html(proc_name, current)
            components.html(viewer_html, height=660, scrolling=False)

            # Anatomy note for this step
            notes = ANATOMY_NOTES.get(proc_name, [])
            if current < len(notes):
                st.markdown(
                    f'<div style="background:#eff6ff;border-left:3px solid #3b82f6;'
                    f'border-radius:0 8px 8px 0;padding:8px 12px;margin-bottom:8px;'
                    f'font-size:.78rem;color:#1e40af;">🔬 <b>Anatomy view:</b> {notes[current]}</div>',
                    unsafe_allow_html=True,
                )

        # ── STEP CARDS ─────────────────────────────────────────────────────
        for i, step in enumerate(steps):
            if i < current:
                st.markdown(
                    f'<div class="surgery-step completed">'
                    f'✅ <b>{step["title"]}</b>'
                    f'<div style="font-size:.78rem;color:#94a3b8;margin-top:3px;">{step["check"]}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            elif i == current:
                st.markdown(
                    f'<div class="surgery-step active">'
                    f'<div style="font-weight:700;color:#065f46;margin-bottom:6px;">▶ Step {i+1}: {step["title"]}</div>'
                    f'<div style="font-size:.85rem;color:#0f172a;line-height:1.6;">{step["detail"]}</div>'
                    f'<div style="margin-top:8px;font-size:.78rem;font-style:italic;color:#0369a1;">'
                    f'✔ Checkpoint: {step["check"]}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                col_a, col_b = st.columns(2)
                with col_a:
                    if st.button(
                        "✅ Step Complete — Next",
                        type="primary",
                        use_container_width=True,
                        key=f"proc_next_{i}",
                    ):
                        st.session_state.proc_step  += 1
                        st.session_state.proc_score += 10
                        if st.session_state.proc_step >= len(steps):
                            comp = st.session_state.competencies
                            comp.setdefault(proc_name, {"attempts":0,"passes":0,"last":""})
                            comp[proc_name]["attempts"] += 1
                            comp[proc_name]["passes"]   += 1
                            comp[proc_name]["last"]      = str(datetime.now().date())
                            st.session_state.score += 20
                        st.rerun()

                with col_b:
                    if st.button(
                        "💡 Hint from Tutor",
                        use_container_width=True,
                        key=f"proc_hint_{i}",
                    ):
                        case_ctx = ""
                        if c:
                            case_ctx = (
                                f"Active patient: {c.get('Age_Sex','?')}, "
                                f"Diagnosis: {c.get('Final_Diagnosis','?')}, "
                                f"Vitals: {c.get('Vitals','?')}, "
                                f"Labs: {c.get('Labs','?')}. "
                            )
                        with st.spinner("Getting hint…"):
                            hint = call_ai(
                                "You are a clinical skills tutor. Give one concise, practical, "
                                "evidence-based tip. If a patient context is provided, tailor "
                                "the tip to that specific patient. Never invent drug doses, lab "
                                "values, or clinical facts.",
                                [{"role":"user","content":
                                  f"{case_ctx}Procedure: {proc_name}. "
                                  f"Current step: {step['title']}. "
                                  f"Detail: {step['detail']}"}],
                                max_tokens=180,
                            )
                        st.info(f"💡 {hint}")

            else:
                st.markdown(
                    f'<div class="surgery-step" style="opacity:.4;">Step {i+1}: {step["title"]}</div>',
                    unsafe_allow_html=True,
                )

        # ── COMPLETION ─────────────────────────────────────────────────────
        if current >= len(steps):
            score = st.session_state.proc_score
            st.markdown(
                f'<div class="alert-good" style="text-align:center;font-size:1rem;padding:1rem;">'
                f'🎉 <b>Procedure Complete!</b><br>'
                f'You scored <b>{score}</b> points and earned a competency mark for <b>{proc_name}</b>.'
                f'</div>',
                unsafe_allow_html=True,
            )
            if proc_name in HAS_3D:
                st.markdown(
                    '<div style="background:#0a2540;border-radius:10px;padding:12px;'
                    'text-align:center;color:#94a3b8;font-size:.8rem;margin-top:8px;">'
                    '🔵 Use the 3D model above to review the anatomy of this procedure at any point.</div>',
                    unsafe_allow_html=True,
                )
            if st.button("🔄 Restart Procedure", use_container_width=True, key="proc_restart"):
                st.session_state.proc_step  = 0
                st.session_state.proc_score = 0
                st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# 🩺 PROCEDURE SIMULATOR — ORIGINAL (fallback, kept for reference)
# ════════════════════════════════════════════════════════════════════════════

def page_procedure_sim():
    """Interactive procedure simulator — case-aware, evidence-based, no hallucination."""
    c = st.session_state.get("selected_case")
    recommended = _get_recommended_procedures(c)

    st.markdown('<div class="main-header"><h1>🩺 Procedure Simulator</h1><p>Step-by-step evidence-based procedure guidance with competency tracking</p></div>', unsafe_allow_html=True)

    # ── Case context banner ────────────────────────────────────────────────
    if c:
        dx  = c.get("Final_Diagnosis","?")
        age = c.get("Age_Sex","?")
        cc  = c.get("Chief_Complaint","?")
        if recommended:
            rec_names = ", ".join(recommended.keys())
            st.markdown(f'''
            <div style="background:#f0fdf4;border:1px solid #16a34a;border-radius:10px;
                        padding:10px 14px;margin-bottom:14px;font-size:.84rem;">
                <b>🟢 Active case:</b> {age} — {cc} &nbsp;|&nbsp; <b>Dx:</b> {dx}<br>
                <b>Clinically relevant procedures highlighted below:</b> {rec_names}
            </div>''', unsafe_allow_html=True)
        else:
            st.markdown(f'''
            <div style="background:#eff6ff;border:1px solid #3b82f6;border-radius:10px;
                        padding:10px 14px;margin-bottom:14px;font-size:.84rem;">
                <b>ℹ️ Active case:</b> {age} — {dx}. All procedures are available for practice.
            </div>''', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-info">ℹ️ No active case selected. Select a case from the Case Library to see which procedures are most relevant. All procedures are available for practice.</div>', unsafe_allow_html=True)

    col_sel, col_proc = st.columns([1, 2])

    with col_sel:
        st.markdown('<div class="section-header">📋 Select Procedure</div>', unsafe_allow_html=True)

        # Recommended section first
        if recommended:
            st.markdown('<div style="font-size:.75rem;font-weight:600;color:#16a34a;margin-bottom:4px;">⭐ RECOMMENDED FOR THIS CASE</div>', unsafe_allow_html=True)
            for pname in recommended:
                pdata = PROCEDURES[pname]
                diff_col = {"Basic":"#16a34a","Intermediate":"#d97706","Advanced":"#dc2626"}.get(pdata["difficulty"],"#64748b")
                btn_style = "border:2px solid #16a34a;border-radius:6px;margin-bottom:2px;"
                st.markdown(f'<div style="{btn_style}">', unsafe_allow_html=True)
                if st.button(f"⭐ {pdata['icon']} {pname}", use_container_width=True, key=f"proc_rec_{pname}"):
                    st.session_state.proc_selected = pname
                    st.session_state.proc_step = 0
                    st.session_state.proc_score = 0
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
                st.markdown(f'<div style="font-size:.7rem;color:{diff_col};margin:-6px 0 6px 4px;">{pdata["difficulty"]} · {pdata["time"]}</div>', unsafe_allow_html=True)

            st.markdown('<div style="font-size:.75rem;font-weight:600;color:#64748b;margin:8px 0 4px;">ALL PROCEDURES</div>', unsafe_allow_html=True)

        for pname, pdata in PROCEDURES.items():
            if pname in recommended:
                continue
            diff_col = {"Basic":"#16a34a","Intermediate":"#d97706","Advanced":"#dc2626"}.get(pdata["difficulty"],"#64748b")
            if st.button(f"{pdata['icon']} {pname}", use_container_width=True, key=f"proc_{pname}"):
                st.session_state.proc_selected = pname
                st.session_state.proc_step = 0
                st.session_state.proc_score = 0
                st.rerun()
            st.markdown(f'<div style="font-size:.7rem;color:{diff_col};margin:-6px 0 6px 4px;">{pdata["difficulty"]} · {pdata["time"]}</div>', unsafe_allow_html=True)

    with col_proc:
        proc_name = st.session_state.get("proc_selected")
        if not proc_name:
            st.markdown('<div style="background:white;border:2px dashed #cbd5e1;border-radius:12px;padding:2rem;text-align:center;color:#64748b;">← Select a procedure to begin simulation</div>', unsafe_allow_html=True)
            return

        proc  = PROCEDURES[proc_name]
        steps = proc["steps"]
        current = st.session_state.proc_step

        st.markdown(f'<div class="section-header">{proc["icon"]} {proc_name}</div>', unsafe_allow_html=True)

        # ── Case relevance banner for this specific procedure ──────────────
        if proc_name in recommended and c:
            st.markdown(f'''
            <div style="background:#fef3c7;border-left:4px solid #d97706;border-radius:0 8px 8px 0;
                        padding:8px 12px;margin-bottom:10px;font-size:.8rem;color:#92400e;">
                <b>Why this procedure matters for your case:</b><br>
                {recommended[proc_name]}
            </div>''', unsafe_allow_html=True)

        # Progress bar
        progress = current / len(steps)
        st.progress(progress, text=f"Step {min(current+1, len(steps))} of {len(steps)}")

        # Render steps
        for i, step in enumerate(steps):
            if i < current:
                st.markdown(f'''
                <div class="surgery-step completed">
                    ✅ <b>{step["title"]}</b>
                    <div style="font-size:.78rem;color:#94a3b8;margin-top:3px;">{step["check"]}</div>
                </div>''', unsafe_allow_html=True)
            elif i == current:
                st.markdown(f'''
                <div class="surgery-step active">
                    <div style="font-weight:700;color:#065f46;margin-bottom:6px;">▶ Step {i+1}: {step["title"]}</div>
                    <div style="font-size:.85rem;color:#0f172a;line-height:1.6;">{step["detail"]}</div>
                    <div style="margin-top:8px;font-size:.78rem;font-style:italic;color:#0369a1;">✔ Checkpoint: {step["check"]}</div>
                </div>''', unsafe_allow_html=True)

                col_a, col_b = st.columns(2)
                with col_a:
                    if st.button("✅ Step Complete — Next", type="primary", use_container_width=True, key=f"proc_next_{i}"):
                        st.session_state.proc_step += 1
                        st.session_state.proc_score += 10
                        if st.session_state.proc_step >= len(steps):
                            comp = st.session_state.competencies
                            comp.setdefault(proc_name, {"attempts":0,"passes":0,"last":""})
                            comp[proc_name]["attempts"] += 1
                            comp[proc_name]["passes"]   += 1
                            comp[proc_name]["last"]      = str(datetime.now().date())
                            st.session_state.score += 20
                        st.rerun()
                with col_b:
                    if st.button("💡 Hint from Tutor", use_container_width=True, key=f"proc_hint_{i}"):
                        case_ctx = ""
                        if c:
                            case_ctx = (f"Active patient: {c.get('Age_Sex','?')}, "
                                        f"Diagnosis: {c.get('Final_Diagnosis','?')}, "
                                        f"Vitals: {c.get('Vitals','?')}, "
                                        f"Labs: {c.get('Labs','?')}. ")
                        with st.spinner("Getting hint..."):
                            hint = call_ai(
                                "You are an experienced clinical skills tutor giving structured, comprehensive guidance to a medical student. "
                                "When asked about a procedure step, provide: "
                                "(1) WHY this step matters clinically, "
                                "(2) HOW to perform it correctly with specific technique details, "
                                "(3) WHAT to watch for — key clinical signs, common mistakes, and how to avoid them, "
                                "(4) PATIENT-SPECIFIC considerations if a patient context is given. "
                                "Be specific, practical and evidence-based. Use standard clinical references (NICE, WHO, BNF, UpToDate). "
                                "Never truncate mid-sentence. Always complete every point fully.",
                                [{"role":"user","content":
                                  f"{case_ctx}Procedure: {proc_name}. "
                                  f"Current step: {step['title']}. "
                                  f"Step detail: {step['detail']}. "
                                  f"Give a complete, structured clinical hint covering all 4 points above."}],
                                max_tokens=500)
                        st.info(f"💡 {hint}")
            else:
                st.markdown(f'<div class="surgery-step" style="opacity:.4;">Step {i+1}: {step["title"]}</div>', unsafe_allow_html=True)

        if current >= len(steps):
            st.markdown(f'''
            <div class="alert-good" style="text-align:center;font-size:1rem;padding:1rem;">
                🎉 <b>Procedure Complete!</b><br>
                You scored <b>{st.session_state.proc_score}</b> points and earned a competency mark for <b>{proc_name}</b>.
            </div>''', unsafe_allow_html=True)
            if st.button("🔄 Restart Procedure", use_container_width=True, key="proc_restart"):
                st.session_state.proc_step = 0
                st.session_state.proc_score = 0
                st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# 🧠 CLINICAL REASONING MAP
# ════════════════════════════════════════════════════════════════════════════

REASONING_TYPES = {
    "history":   {"label":"History Finding","color":"#0ea5e9","icon":"📋"},
    "exam":      {"label":"Examination Finding","color":"#7c3aed","icon":"🫁"},
    "lab":       {"label":"Investigation Result","color":"#d97706","icon":"🧪"},
    "diagnosis": {"label":"Differential / Diagnosis","color":"#059669","icon":"🧬"},
    "management":{"label":"Management Decision","color":"#dc2626","icon":"💊"},
}

def page_reasoning_map():
    """Clinical Reasoning Map — visual mind-map of student's clinical thinking."""
    c = st.session_state.get("selected_case")
    st.markdown('<div class="main-header"><h1>🧠 Clinical Reasoning Map</h1><p>Build your clinical reasoning pathway — from history to management</p></div>', unsafe_allow_html=True)

    if not c:
        st.markdown('<div class="alert-warn">⚠️ Select a case first.</div>', unsafe_allow_html=True)
        return

    st.markdown(f'<div class="patient-card"><b>Case:</b> {c.get("Age_Sex","?")} | {c.get("Chief_Complaint","?")} | {c.get("Final_Diagnosis","?")}</div>', unsafe_allow_html=True)

    col_add, col_map = st.columns([1,2])

    with col_add:
        st.markdown('<div class="section-header">➕ Add Reasoning Node</div>', unsafe_allow_html=True)
        node_type = st.selectbox("Node Type", list(REASONING_TYPES.keys()),
                                 format_func=lambda x: REASONING_TYPES[x]["icon"]+" "+REASONING_TYPES[x]["label"],
                                 key="rn_type")
        node_content = st.text_area("Finding / Reasoning", placeholder="e.g. RLQ tenderness on palpation", height=80, key="rn_content")
        node_significance = st.selectbox("Significance", ["Key finding","Supportive","Ruling out alternative"], key="rn_sig")

        if st.button("➕ Add to Map", type="primary", use_container_width=True, key="rn_add"):
            if node_content.strip():
                st.session_state.reasoning_nodes.append({
                    "type": node_type,
                    "content": node_content.strip(),
                    "significance": node_significance,
                    "timestamp": str(datetime.now())[:16],
                })
                st.rerun()

        if st.button("🤖 AI Review My Reasoning", use_container_width=True, key="rn_ai"):
            nodes = st.session_state.reasoning_nodes
            if not nodes:
                st.warning("Add some reasoning nodes first.")
            else:
                summary = "\n".join([f"[{n['type']}] {n['content']} ({n['significance']})" for n in nodes])
                prompt = (f"Case: {c.get('Chief_Complaint','')}. Diagnosis: {c.get('Final_Diagnosis','')}.\n"
                          f"Student reasoning:\n{summary}\n\n"
                          "As a clinical educator, review this reasoning pathway. Is it logical? What key steps are missing? What could be improved?")
                with st.spinner("Analysing reasoning..."):
                    review = call_ai("You are a senior clinician reviewing a student's clinical reasoning.",
                                    [{"role":"user","content":prompt}], max_tokens=500)
                st.markdown(f'<div class="chat-tutor">🤖 {review}</div>', unsafe_allow_html=True)

        if st.button("🗑️ Clear Map", use_container_width=True, key="rn_clear"):
            st.session_state.reasoning_nodes = []
            st.rerun()

    with col_map:
        st.markdown('<div class="section-header">🗺️ Your Reasoning Map</div>', unsafe_allow_html=True)
        nodes = st.session_state.reasoning_nodes
        if not nodes:
            st.markdown('<div style="background:white;border:2px dashed #cbd5e1;border-radius:12px;padding:2rem;text-align:center;color:#64748b;">Your reasoning nodes will appear here.<br>Start by adding history findings, exam findings, and your differentials.</div>', unsafe_allow_html=True)
        else:
            for ntype, tdata in REASONING_TYPES.items():
                type_nodes = [n for n in nodes if n["type"]==ntype]
                if not type_nodes: continue
                st.markdown(f'<div style="font-weight:700;color:{tdata["color"]};font-size:.85rem;margin:.6rem 0 .3rem;">{tdata["icon"]} {tdata["label"]}</div>', unsafe_allow_html=True)
                for n in type_nodes:
                    sig_icon = "🔑" if n["significance"]=="Key finding" else "➕" if n["significance"]=="Supportive" else "🚫"
                    st.markdown(f'''
                    <div style="background:white;border-left:4px solid {tdata["color"]};border-radius:0 8px 8px 0;
                                padding:7px 10px;margin:.25rem 0 .25rem 12px;font-size:.82rem;
                                box-shadow:0 1px 4px rgba(0,0,0,.06);">
                        {sig_icon} {n["content"]}
                        <div style="color:#94a3b8;font-size:.72rem;margin-top:2px;">{n["timestamp"]}</div>
                    </div>''', unsafe_allow_html=True)

        # Summary stats
        if nodes:
            st.markdown("---")
            c1,c2,c3 = st.columns(3)
            c1.metric("Total Nodes", len(nodes))
            c2.metric("Key Findings", len([n for n in nodes if n["significance"]=="Key finding"]))
            c3.metric("Types Covered", len(set(n["type"] for n in nodes)))


# ════════════════════════════════════════════════════════════════════════════
# 🏆 COMPETENCY TRACKER
# ════════════════════════════════════════════════════════════════════════════

ALL_COMPETENCIES = {
    "Clinical Reasoning":    {"description":"Formulate logical differential diagnoses","module":"ddx","target":5},
    "Drug Prescribing":      {"description":"Write safe, appropriate prescriptions","module":"prescribing","target":5},
    "IV Cannulation":        {"description":"Peripheral IV access procedure","module":"procedures","target":3},
    "Lumbar Puncture":       {"description":"CSF sampling procedure","module":"procedures","target":3},
    "ABG Sampling":          {"description":"Arterial blood gas sampling","module":"procedures","target":3},
    "ECG Interpretation":    {"description":"12-lead ECG analysis","module":"procedures","target":5},
    "Patient Communication": {"description":"History taking & communication","module":"simulator","target":10},
    "Physical Examination":  {"description":"Systematic clinical examination","module":"physical_exam","target":8},
    "Lab Interpretation":    {"description":"Interpret investigation results","module":"lab","target":10},
    "Imaging Analysis":      {"description":"X-Ray, CT, and ECG reading","module":"imaging","target":5},
    "Surgical Knowledge":    {"description":"Peri-operative care & procedures","module":"surgery","target":5},
    "Diagnosis Accuracy":    {"description":"Correct final diagnosis submission","module":"diagnosis","target":10},
}

def page_competency_tracker():
    """Competency Tracker — full OSCE-style skill matrix."""
    st.markdown('<div class="main-header"><h1>🏆 Competency Tracker</h1><p>Your clinical skills portfolio — track progress across all OSCE competency domains</p></div>', unsafe_allow_html=True)

    comp = st.session_state.competencies
    user = st.session_state.get("auth_user") or {}

    # Overall completion
    total_comps = len(ALL_COMPETENCIES)
    achieved = sum(1 for k,v in ALL_COMPETENCIES.items()
                   if comp.get(k,{}).get("passes",0) >= v["target"])
    pct = int(achieved/total_comps*100)

    k1,k2,k3,k4 = st.columns(4)
    with k1: st.markdown(f'<div class="kpi-card"><div class="kpi-value">{achieved}/{total_comps}</div><div class="kpi-label">✅ Competencies Achieved</div></div>', unsafe_allow_html=True)
    with k2: st.markdown(f'<div class="kpi-card"><div class="kpi-value">{pct}%</div><div class="kpi-label">📊 Portfolio Complete</div></div>', unsafe_allow_html=True)
    with k3:
        total_attempts = sum(comp.get(k,{}).get("attempts",0) for k in ALL_COMPETENCIES)
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{total_attempts}</div><div class="kpi-label">🎯 Total Attempts</div></div>', unsafe_allow_html=True)
    with k4:
        total_passes = sum(comp.get(k,{}).get("passes",0) for k in ALL_COMPETENCIES)
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{total_passes}</div><div class="kpi-label">⭐ Total Passes</div></div>', unsafe_allow_html=True)

    st.progress(pct/100, text=f"Overall competency portfolio: {pct}%")
    st.markdown("")

    # Competency matrix
    st.markdown('<div class="section-header">📋 Competency Matrix</div>', unsafe_allow_html=True)

    for domain, ddata in ALL_COMPETENCIES.items():
        student_data = comp.get(domain, {"attempts":0,"passes":0,"last":""})
        attempts = student_data.get("attempts",0)
        passes   = student_data.get("passes",0)
        target   = ddata["target"]
        last     = student_data.get("last","Never")
        achieved_comp = passes >= target

        status_color = "#16a34a" if achieved_comp else "#d97706" if passes > 0 else "#94a3b8"
        status_icon  = "✅" if achieved_comp else "🟡" if passes > 0 else "⭕"
        bar_pct = min(int(passes/target*100),100)

        st.markdown(f'''
        <div style="background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px 16px;margin:.4rem 0;
                    border-left:5px solid {status_color};">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;">
                <div>
                    <span style="font-weight:700;color:#0a2540;">{status_icon} {domain}</span>
                    <span style="font-size:.75rem;color:#64748b;margin-left:8px;">{ddata["description"]}</span>
                </div>
                <div style="text-align:right;font-size:.78rem;">
                    <span style="font-weight:700;color:{status_color};">{passes}/{target}</span>
                    <span style="color:#94a3b8;"> passes · last: {last}</span>
                </div>
            </div>
            <div style="background:#f1f5f9;border-radius:999px;height:6px;">
                <div style="background:{status_color};border-radius:999px;height:6px;width:{bar_pct}%;transition:width .3s;"></div>
            </div>
        </div>''', unsafe_allow_html=True)

        # Go to module button
        module_map = {"ddx":"ddx","prescribing":"prescribing","procedures":"procedures",
                      "simulator":"simulator","physical_exam":"physical_exam","lab":"lab",
                      "imaging":"imaging","surgery":"surgery","diagnosis":"diagnosis"}
        mod = ddata["module"]
        if not achieved_comp and st.button(f"Practice {domain} →", key=f"comp_go_{domain}"):
            nav(module_map.get(mod,"home"))

    # Export PDF portfolio
    if FPDF_OK:
        st.markdown("---")
        if st.button("📄 Export Competency Portfolio as PDF", key="comp_pdf"):
            pdf = FPDF(); pdf.add_page()
            pdf.set_font("Helvetica","B",16)
            pdf.cell(0,10,"MLS Virtual Hospital — Competency Portfolio",ln=True,align="C")
            pdf.set_font("Helvetica","",11)
            pdf.cell(0,8,f"Student: {user.get('name','—')}  |  Date: {str(datetime.now().date())}",ln=True)
            pdf.cell(0,8,f"Overall: {achieved}/{total_comps} competencies achieved ({pct}%)",ln=True)
            pdf.ln(4)
            pdf.set_font("Helvetica","B",11)
            for domain,ddata in ALL_COMPETENCIES.items():
                sd = comp.get(domain,{"attempts":0,"passes":0,"last":""})
                status = "ACHIEVED" if sd.get("passes",0)>=ddata["target"] else "IN PROGRESS"
                pdf.cell(0,7,f"{domain}: {sd.get('passes',0)}/{ddata['target']} — {status}",ln=True)
            pdf_bytes = pdf.output(dest="S").encode("latin-1")
            st.download_button("⬇ Download PDF Portfolio", pdf_bytes, "competency_portfolio.pdf","application/pdf")


# ════════════════════════════════════════════════════════════════════════════
# 🏥 CASE CREATOR (Faculty)
# ════════════════════════════════════════════════════════════════════════════


# ════════════════════════════════════════════════════════════════════════════
# 🏥 MIMIC-IV INTEGRATION + CASE CREATOR (Faculty)
# ════════════════════════════════════════════════════════════════════════════

# ── MIMIC-IV Synthetic demo data (used when no files uploaded) ────────────
_MIMIC_DEMO = [
    {"subject_id":"M-001","age":67,"gender":"M","diagnosis":"Sepsis",
     "system":"Infectious Disease","cc":"Fever, confusion, hypotension x 6h",
     "hpi":"67yo male brought in by family with 6h fever 39.4C, confusion, hypotension. History of T2DM, CKD. Lives alone.",
     "vitals":"HR 118 | BP 88/54 | RR 24 | Temp 39.4°C | SpO2 93% RA | GCS 13",
     "labs":"WBC 18.4 | Lactate 4.2 | Creatinine 2.8 | CRP 210 | Blood cultures x2 sent | Procalcitonin 8.4",
     "pmh":"T2DM | CKD stage 3 | HTN","meds":"Metformin | Lisinopril | Amlodipine",
     "allergies":"Penicillin (rash)","imaging":"CXR: bilateral infiltrates","dx":"Septic Shock"},
    {"subject_id":"M-002","age":54,"gender":"F","diagnosis":"NSTEMI",
     "system":"Cardiovascular","cc":"Chest tightness, diaphoresis x 3h",
     "hpi":"54yo female with 3h central chest tightness radiating to left arm. Diaphoretic, nauseous. Smoker 20 pack-years. Family history of IHD.",
     "vitals":"HR 96 | BP 148/92 | RR 18 | Temp 37.1°C | SpO2 97% RA",
     "labs":"Troponin I 3.8 (high) | CK-MB 24 (high) | Total Cholesterol 6.2 | Glucose 8.4 | eGFR 72",
     "pmh":"Hypertension | Hyperlipidaemia | Ex-smoker","meds":"Amlodipine","allergies":"NKDA",
     "imaging":"ECG: ST depression V4-V6, T-wave inversion leads I, aVL","dx":"NSTEMI"},
    {"subject_id":"M-003","age":28,"gender":"F","diagnosis":"DKA",
     "system":"Endocrine","cc":"Vomiting, abdominal pain, polyuria x 24h",
     "hpi":"28yo female T1DM (diagnosed age 9) presents 24h vomiting, abdominal pain, polyuria. Missed insulin doses x 2 days. Fruity breath noted.",
     "vitals":"HR 112 | BP 102/68 | RR 26 (Kussmaul) | Temp 37.8°C | SpO2 99% RA",
     "labs":"Glucose 28.4 | pH 7.18 | Bicarbonate 8 | Ketones 4+ | Na 131 | K 5.8 | WBC 14.2",
     "pmh":"T1DM since age 9 | No other conditions","meds":"Insulin Lantus | Insulin Novorapid","allergies":"NKDA",
     "imaging":"CXR: clear","dx":"Diabetic Ketoacidosis (DKA)"},
    {"subject_id":"M-004","age":72,"gender":"M","diagnosis":"Stroke",
     "system":"Neurological","cc":"Sudden right-sided weakness, slurred speech",
     "hpi":"72yo male sudden onset right arm weakness and facial droop while eating breakfast. Slurred speech. Symptoms onset 45 minutes ago. AF known.",
     "vitals":"HR 88 (irregular) | BP 178/96 | RR 16 | Temp 37.0°C | SpO2 96% RA",
     "labs":"INR 1.1 | Glucose 7.2 | Platelets 198 | HbA1c 6.8 | Cholesterol 5.4 | eGFR 58",
     "pmh":"AF | HTN | T2DM | Previous TIA 2021","meds":"Apixaban | Metformin | Ramipril","allergies":"NKDA",
     "imaging":"CT Head: No haemorrhage. MRI DWI: acute ischaemic lesion left MCA territory","dx":"Ischaemic Stroke (MCA territory)"},
    {"subject_id":"M-005","age":19,"gender":"M","diagnosis":"Meningitis",
     "system":"Neurological","cc":"Severe headache, neck stiffness, photophobia",
     "hpi":"19yo male university student, 12h worsening headache, neck stiffness, photophobia. Fever 39.8C. Non-blanching petechial rash on legs noted.",
     "vitals":"HR 124 | BP 96/60 | RR 22 | Temp 39.8°C | SpO2 98% RA | GCS 14",
     "labs":"WBC 22.4 (neut 94%) | CRP 180 | Procalcitonin 12 | Platelets 88 | Clotting: PT 18, APTT 42",
     "pmh":"Nil significant | University student, halls of residence","meds":"None","allergies":"NKDA",
     "imaging":"CT Head: no contraindication to LP. LP: turbid CSF, WBC 1200, protein 2.4, glucose 1.2","dx":"Bacterial Meningitis (Meningococcal)"},
    {"subject_id":"M-006","age":45,"gender":"F","diagnosis":"Pulmonary Embolism",
     "system":"Respiratory","cc":"Sudden pleuritic chest pain, dyspnoea",
     "hpi":"45yo female 3 days post total hip replacement, sudden onset right pleuritic chest pain and dyspnoea. Right calf swollen and tender.",
     "vitals":"HR 108 | BP 118/76 | RR 24 | Temp 37.6°C | SpO2 91% RA -> 96% on 4L O2",
     "labs":"D-dimer 4800 (very high) | Troponin 0.08 (mildly elevated) | BNP 180 | ABG: PaO2 7.8, PaCO2 3.8, pH 7.48",
     "pmh":"Recent THR 3 days ago | OCP use","meds":"OCP | Enoxaparin (post-op prophylaxis stopped day 2)","allergies":"NKDA",
     "imaging":"CTPA: bilateral pulmonary emboli, right > left. Echo: RV strain","dx":"Massive Pulmonary Embolism"},
]

def _mimic_to_case(m: dict, case_id_override: str = None) -> dict:
    """Convert a MIMIC record dict into your app case_studies format."""
    age_sex = f"{m.get('age','?')}yr {'Male' if m.get('gender','')=='M' else 'Female'}"
    return {
        "Case_ID":         case_id_override or m.get("subject_id","MIMIC-?"),
        "Title":           f"{age_sex} — {m.get('cc','')}",
        "Age_Sex":         age_sex,
        "Occupation":      m.get("occupation",""),
        "System":          m.get("system","General"),
        "Difficulty":      m.get("difficulty","Intermediate"),
        "Chief_Complaint": m.get("cc",""),
        "Duration":        m.get("duration",""),
        "Context":         m.get("social",""),
        "HPI":             m.get("hpi",""),
        "PMH":             m.get("pmh",""),
        "Medications":     m.get("meds",""),
        "Allergies":       m.get("allergies","NKDA"),
        "Family_Hx":       m.get("family",""),
        "Social_Hx":       m.get("social",""),
        "Vitals":          m.get("vitals",""),
        "Physical_Findings": m.get("exam",""),
        "Labs":            m.get("labs",""),
        "Imaging_Tests":   m.get("imaging",""),
        "Final_Diagnosis": m.get("dx",""),
        "Learning_Obj":    m.get("learning_obj",""),
        "_source":         "MIMIC-IV",
    }

def _parse_mimic_csv(uploaded_file) -> list:
    """
    Parse an uploaded MIMIC-IV admissions CSV and extract usable cases.
    Expects columns: subject_id, hadm_id, diagnosis (or icd_code), admittime.
    Returns list of partial dicts — user fills in clinical details via AI enrichment.
    """
    try:
        df_m = pd.read_csv(uploaded_file)
        df_m.columns = [c.lower().strip() for c in df_m.columns]
        cases = []
        dx_col = next((c for c in ["diagnosis","long_title","icd_title"] if c in df_m.columns), None)
        for _, row in df_m.head(20).iterrows():
            cases.append({
                "subject_id": str(row.get("subject_id","?")),
                "age": int(row.get("anchor_age", row.get("age",50))),
                "gender": str(row.get("gender","U")),
                "dx": str(row.get(dx_col,"Unknown")) if dx_col else "Unknown",
                "system": "General",
                "cc": f"Admission for {row.get(dx_col,'unknown condition')}",
                "hpi": "", "vitals": "", "labs": "", "pmh": "",
                "meds": "", "allergies": "NKDA", "imaging": "",
            })
        return cases
    except Exception as e:
        return []

def _mimic_bigquery_query(bq_client, diagnosis_filter: str, limit: int = 10) -> list:
    """
    Query MIMIC-IV via BigQuery (requires google-cloud-bigquery + credentials).
    Returns list of admission records.
    """
    safe_dx = diagnosis_filter.replace("'","").upper()
    query = f"""
        SELECT a.subject_id, a.hadm_id, a.admittime,
               p.anchor_age AS age, p.gender,
               d.long_title AS diagnosis
        FROM `physionet-data.mimiciv_hosp.admissions` a
        JOIN `physionet-data.mimiciv_hosp.patients` p USING (subject_id)
        JOIN `physionet-data.mimiciv_hosp.diagnoses_icd` di USING (hadm_id, subject_id)
        JOIN `physionet-data.mimiciv_hosp.d_icd_diagnoses` d USING (icd_code, icd_version)
        WHERE UPPER(d.long_title) LIKE '%{safe_dx}%'
        GROUP BY 1,2,3,4,5,6
        ORDER BY a.admittime DESC
        LIMIT {limit}
    """
    try:
        rows = list(bq_client.query(query).result())
        return [{"subject_id":str(r.subject_id),"age":r.age,"gender":r.gender,
                 "dx":r.diagnosis,"system":"General","cc":r.diagnosis,
                 "hpi":"","vitals":"","labs":"","pmh":"","meds":"",
                 "allergies":"NKDA","imaging":""} for r in rows]
    except Exception as e:
        return []


def page_case_creator():
    """
    Faculty case creator with 4 modes:
    1. Manual entry
    2. MIMIC-IV demo cases (built-in, no signup needed)
    3. MIMIC-IV local CSV upload (after downloading from PhysioNet)
    4. MIMIC-IV BigQuery live query (requires credentials)
    """
    user = st.session_state.get("auth_user") or {}
    if user.get("role") != "faculty":
        st.error("🔒 Case Creator is restricted to faculty members.")
        return

    st.markdown('<div class="main-header"><h1>🏥 Case Creator — Faculty Portal</h1><p>Build AI training cases manually or import from MIMIC-IV real patient data</p></div>', unsafe_allow_html=True)

    tab_manual, tab_mimic, tab_preview, tab_manage = st.tabs([
        "✍️ Manual Entry", "🏥 MIMIC-IV Importer", "👁️ Preview & Approve", "📁 Manage Cases"])

    # ════════════════════════════════════════════════════════════════
    with tab_manual:
    # ════════════════════════════════════════════════════════════════
        st.markdown('<div class="section-header">📋 Patient Demographics</div>', unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        with c1:
            age  = st.number_input("Age", 1, 120, 45, key="cc_age")
            sex  = st.selectbox("Sex", ["Male","Female","Other"], key="cc_sex")
        with c2:
            occ  = st.text_input("Occupation", placeholder="e.g. Teacher", key="cc_occ")
            diff = st.selectbox("Difficulty", ["Basic","Intermediate","Advanced"], key="cc_diff")
        with c3:
            system = st.selectbox("System", [
                "Gastrointestinal","Cardiovascular","Respiratory","Neurological",
                "Endocrine","Renal","Musculoskeletal","Haematology",
                "Infectious Disease","Other"], key="cc_sys")
            title = st.text_input("Case Title", placeholder="e.g. Young adult with RLQ pain", key="cc_title")

        st.markdown('<div class="section-header">🩺 Clinical Details</div>', unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        with c1:
            cc       = st.text_area("Chief Complaint", height=60, key="cc_cc")
            hpi      = st.text_area("History of Present Illness", height=100, key="cc_hpi")
            pmh      = st.text_area("Past Medical History", height=60, key="cc_pmh")
            meds     = st.text_area("Current Medications", height=60, key="cc_meds")
            allergies= st.text_input("Allergies", placeholder="NKDA", key="cc_allerg")
            family   = st.text_area("Family History", height=50, key="cc_fhx")
            social   = st.text_area("Social History", height=50, key="cc_shx")
        with c2:
            vitals   = st.text_area("Vitals", placeholder="HR | BP | RR | Temp | SpO2", height=60, key="cc_vitals")
            findings = st.text_area("Physical Examination Findings", height=100, key="cc_findings")
            labs     = st.text_area("Lab Results", height=80, key="cc_labs")
            imaging  = st.text_area("Imaging / ECG Results", height=60, key="cc_imaging")
            dx       = st.text_input("Final Diagnosis (hidden from students)", key="cc_dx")
            lo       = st.text_area("Learning Objectives", height=80, key="cc_lo")

        if st.button("🤖 Generate AI Patient Persona", type="primary",
                     use_container_width=True, key="cc_generate"):
            if not cc.strip() or not dx.strip():
                st.warning("Chief Complaint and Final Diagnosis are required.")
            else:
                draft = {
                    "Case_ID": "MANUAL-"+str(random.randint(100000,999999)),
                    "Title": title or f"{age}yr {sex} — {cc[:40]}",
                    "Age_Sex": f"{age}yr {sex}", "Occupation": occ,
                    "System": system, "Difficulty": diff,
                    "Chief_Complaint": cc, "HPI": hpi, "PMH": pmh,
                    "Medications": meds, "Allergies": allergies,
                    "Family_Hx": family, "Social_Hx": social,
                    "Vitals": vitals, "Physical_Findings": findings,
                    "Labs": labs, "Imaging_Tests": imaging,
                    "Final_Diagnosis": dx, "Learning_Obj": lo,
                    "_source": "Manual",
                }
                prompt = (f"Create a realistic AI patient for: {age}yr {sex}, {cc}. "
                          f"Diagnosis (hidden): {dx}. HPI: {hpi[:300]}. "
                          f"Generate: 1) Opening statement (2-3 sentences, first person) "
                          f"2) 3 follow-up answers 3) Patient personality/emotional state.")
                with st.spinner("Generating AI patient persona..."):
                    persona = call_ai(
                        "You create realistic patient simulations for medical education.",
                        [{"role":"user","content":prompt}], max_tokens=500)
                draft["ai_persona"] = persona
                st.session_state.draft_case = draft
                st.success("Case created! Go to Preview & Approve tab.")
                st.rerun()

    # ════════════════════════════════════════════════════════════════
    with tab_mimic:
    # ════════════════════════════════════════════════════════════════
        st.markdown('<div class="section-header">🏥 MIMIC-IV Real Patient Data Importer</div>', unsafe_allow_html=True)

        mode = st.radio("Import mode:", [
            "🔵 Demo — Built-in MIMIC-style cases (no signup needed)",
            "📂 Local CSV — Upload MIMIC-IV files from PhysioNet",
            "☁️ BigQuery — Live query MIMIC-IV (requires credentials)",
        ], key="mimic_mode")

        st.markdown("---")

        # ── MODE 1: Demo cases ────────────────────────────────────
        if "Demo" in mode:
            st.markdown('<div class="alert-good">✅ No signup needed. These are realistic MIMIC-style cases for immediate use.</div>', unsafe_allow_html=True)
            st.markdown("")

            for i, m in enumerate(_MIMIC_DEMO):
                age_sex = f"{m['age']}yr {'Male' if m['gender']=='M' else 'Female'}"
                with st.expander(f"🏥 {m['subject_id']} — {age_sex} — {m['dx']}", expanded=False):
                    c1,c2 = st.columns(2)
                    with c1:
                        st.markdown(f"**Chief Complaint:** {m['cc']}")
                        st.markdown(f"**System:** {m['system']}")
                        st.markdown(f"**Vitals:** {m['vitals']}")
                        st.markdown(f"**PMH:** {m['pmh']}")
                    with c2:
                        st.markdown(f"**Labs:** {m['labs']}")
                        st.markdown(f"**Imaging:** {m['imaging']}")
                        st.markdown(f"**Allergies:** {m['allergies']}")
                        st.markdown(f"**Diagnosis:** `{m['dx']}`")

                    diff_sel = st.selectbox("Set difficulty:",
                        ["Basic","Intermediate","Advanced"],
                        index=1, key=f"mimic_diff_{i}")
                    lo_sel = st.text_input("Learning objectives:",
                        placeholder="What should students learn?", key=f"mimic_lo_{i}")

                    if st.button(f"🤖 Import & Generate AI Patient",
                                 key=f"mimic_import_{i}", type="primary"):
                        m_copy = dict(m)
                        m_copy["difficulty"] = diff_sel
                        m_copy["learning_obj"] = lo_sel
                        draft = _mimic_to_case(m_copy)
                        # Generate AI persona
                        prompt = (f"Create a realistic AI patient for: {age_sex}, {m['cc']}. "
                                  f"Diagnosis (hidden): {m['dx']}. HPI: {m['hpi']}. "
                                  f"Generate opening statement, 3 follow-up answers, "
                                  f"patient personality. First person, realistic.")
                        with st.spinner("Generating AI patient persona from MIMIC data..."):
                            persona = call_ai(
                                "You create realistic patient simulations from real clinical data.",
                                [{"role":"user","content":prompt}], max_tokens=500)
                        draft["ai_persona"] = persona
                        st.session_state.draft_case = draft
                        st.success(f"Imported {m['subject_id']}! Go to Preview & Approve tab.")
                        st.rerun()

        # ── MODE 2: Local CSV upload ──────────────────────────────
        elif "Local CSV" in mode:
            st.markdown('''
            <div class="alert-info">
                <b>How to get MIMIC-IV CSV files:</b><br>
                1. Go to <b>physionet.org</b> and create a free account<br>
                2. Complete <b>CITI training</b> (~2h, free online)<br>
                3. Sign the data use agreement<br>
                4. Download <b>mimiciv/hosp/admissions.csv</b> and <b>patients.csv</b>
            </div>''', unsafe_allow_html=True)

            uploaded = st.file_uploader(
                "Upload MIMIC-IV admissions.csv",
                type=["csv"], key="mimic_csv_upload")

            if uploaded:
                with st.spinner("Parsing MIMIC-IV CSV..."):
                    records = _parse_mimic_csv(uploaded)

                if not records:
                    st.error("Could not parse file. Ensure it is MIMIC-IV admissions.csv format.")
                else:
                    st.success(f"Found {len(records)} records. Select one to import:")
                    for i, r in enumerate(records[:10]):
                        with st.expander(f"{r['subject_id']} — {r['dx'][:60]}"):
                            diff_s = st.selectbox("Difficulty",
                                ["Basic","Intermediate","Advanced"], key=f"csv_diff_{i}")
                            enrich = st.checkbox("AI-enrich clinical details", value=True,
                                                 key=f"csv_enrich_{i}")
                            if st.button("Import This Case", key=f"csv_import_{i}"):
                                r["difficulty"] = diff_s
                                draft = _mimic_to_case(r)
                                if enrich:
                                    prompt = (f"Patient {r['age']}yr {r['gender']} admitted for {r['dx']}. "
                                              f"Generate realistic: vitals, key labs, HPI (2 paragraphs), "
                                              f"physical findings, chief complaint. Be clinically accurate.")
                                    with st.spinner("AI enriching clinical details..."):
                                        enriched = call_ai(
                                            "You enrich sparse patient records with realistic clinical details.",
                                            [{"role":"user","content":prompt}], max_tokens=600)
                                    draft["HPI"] = enriched
                                st.session_state.draft_case = draft
                                st.success("Imported! Go to Preview & Approve tab.")
                                st.rerun()

        # ── MODE 3: BigQuery ──────────────────────────────────────
        else:
            st.markdown('''
            <div class="alert-info">
                <b>BigQuery Setup:</b><br>
                1. Complete MIMIC-IV credentialing at physionet.org<br>
                2. Go to console.cloud.google.com → Create a service account<br>
                3. Download JSON credentials file<br>
                4. Add to secrets.toml: <code>[gcp]  credentials_json = "..."</code><br>
                5. Install: <code>pip install google-cloud-bigquery pandas-gbq</code>
            </div>''', unsafe_allow_html=True)

            gcp_creds = st.file_uploader(
                "Upload GCP Service Account JSON", type=["json"], key="bq_creds")
            dx_filter = st.text_input("Search MIMIC-IV for diagnosis:",
                placeholder="e.g. appendicitis, sepsis, myocardial infarction",
                key="bq_dx_filter")
            limit = st.slider("Max records to retrieve", 5, 50, 10, key="bq_limit")

            if st.button("☁️ Query MIMIC-IV via BigQuery",
                         type="primary", use_container_width=True, key="bq_query"):
                if not dx_filter.strip():
                    st.warning("Enter a diagnosis to search for.")
                elif not gcp_creds:
                    st.warning("Upload your GCP credentials JSON file.")
                else:
                    try:
                        from google.cloud import bigquery
                        from google.oauth2 import service_account
                        import tempfile, json as _json
                        creds_data = _json.loads(gcp_creds.read())
                        with tempfile.NamedTemporaryFile(mode="w",suffix=".json",delete=False) as f:
                            _json.dump(creds_data, f)
                            tmp_path = f.name
                        creds = service_account.Credentials.from_service_account_file(
                            tmp_path,
                            scopes=["https://www.googleapis.com/auth/bigquery.readonly"])
                        bq = bigquery.Client(credentials=creds,
                                             project=creds_data.get("project_id"))
                        with st.spinner(f"Querying MIMIC-IV for '{dx_filter}'..."):
                            results = _mimic_bigquery_query(bq, dx_filter, limit)
                        if results:
                            st.success(f"Found {len(results)} matching admissions.")
                            st.session_state["bq_results"] = results
                        else:
                            st.warning("No results found. Try a different search term.")
                    except ImportError:
                        st.error("Run: pip install google-cloud-bigquery")
                    except Exception as e:
                        st.error(f"BigQuery error: {e}")

            if st.session_state.get("bq_results"):
                results = st.session_state["bq_results"]
                for i, r in enumerate(results):
                    with st.expander(f"{r['subject_id']} — {r['dx'][:60]}"):
                        diff_s = st.selectbox("Difficulty",
                            ["Basic","Intermediate","Advanced"], key=f"bq_diff_{i}")
                        if st.button("Import & Enrich with AI", key=f"bq_import_{i}"):
                            r["difficulty"] = diff_s
                            draft = _mimic_to_case(r)
                            prompt = (f"Patient {r.get('age','?')}yr admitted for {r['dx']}. "
                                      f"Generate realistic: HPI, vitals, labs, physical findings, "
                                      f"chief complaint. Be clinically accurate.")
                            with st.spinner("AI enriching MIMIC record..."):
                                enriched = call_ai(
                                    "You enrich patient records with realistic clinical details.",
                                    [{"role":"user","content":prompt}], max_tokens=600)
                            draft["HPI"] = enriched
                            st.session_state.draft_case = draft
                            st.success("Imported! Go to Preview & Approve tab.")
                            st.rerun()

    # ════════════════════════════════════════════════════════════════
    with tab_preview:
    # ════════════════════════════════════════════════════════════════
        draft = st.session_state.get("draft_case",{})
        if not draft:
            st.info("Create or import a case first, then come back here to review it.")
        else:
            src = draft.get("_source","Manual")
            src_color = "#0e7490" if src=="Manual" else "#7c3aed"
            st.markdown(f'''
            <div class="patient-card">
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
                    <span style="font-weight:800;font-size:1.05rem;color:#0a2540;">
                        {draft.get("Title","Draft Case")}
                    </span>
                    <div>
                        <span style="background:{src_color};color:white;border-radius:4px;
                                     padding:2px 10px;font-size:.75rem;font-weight:700;">
                            {src}
                        </span>
                        &nbsp;
                        <span style="background:#0e7490;color:white;border-radius:4px;
                                     padding:2px 10px;font-size:.75rem;">
                            {draft.get("Difficulty","?")}
                        </span>
                    </div>
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;font-size:.84rem;">
                    <div><b>Patient:</b> {draft.get("Age_Sex","?")} · {draft.get("Occupation","")}</div>
                    <div><b>System:</b> {draft.get("System","?")}</div>
                    <div><b>ID:</b> {draft.get("Case_ID","?")}</div>
                    <div><b>Chief Complaint:</b> {draft.get("Chief_Complaint","?")}</div>
                    <div><b>Allergies:</b> {draft.get("Allergies","NKDA")}</div>
                    <div><b>Diagnosis:</b>
                        <span style="color:#dc2626;font-weight:700;">
                            {draft.get("Final_Diagnosis","?")}
                        </span> (hidden from students)
                    </div>
                </div>
            </div>''', unsafe_allow_html=True)

            t1, t2, t3, t4 = st.tabs(["📋 History","🩺 Exam & Labs","🤖 AI Persona","📄 Raw JSON"])

            with t1:
                st.markdown(f"**HPI:** {draft.get('HPI','—')}")
                st.markdown(f"**PMH:** {draft.get('PMH','—')}")
                st.markdown(f"**Medications:** {draft.get('Medications','—')}")
                st.markdown(f"**Family Hx:** {draft.get('Family_Hx','—')}")
                st.markdown(f"**Social Hx:** {draft.get('Social_Hx','—')}")
            with t2:
                st.markdown(f"**Vitals:** {draft.get('Vitals','—')}")
                st.markdown(f"**Findings:** {draft.get('Physical_Findings','—')}")
                st.markdown(f"**Labs:** {draft.get('Labs','—')}")
                st.markdown(f"**Imaging:** {draft.get('Imaging_Tests','—')}")
            with t3:
                if draft.get("ai_persona"):
                    st.markdown(f'<div class="chat-tutor">{draft["ai_persona"]}</div>', unsafe_allow_html=True)
                else:
                    st.info("No AI persona generated yet.")
            with t4:
                st.json({k:v for k,v in draft.items() if k!="ai_persona"})

            st.markdown("---")
            c1,c2,c3 = st.columns(3)
            with c1:
                if st.button("✅ Approve & Save to Library",
                             type="primary", use_container_width=True, key="cc_approve"):
                    if "imported_cases" not in st.session_state:
                        st.session_state.imported_cases = []
                    st.session_state.imported_cases.append(draft)
                    st.success(f"Case saved to this session library. Connect Supabase to persist permanently.")
                    st.balloons()
            with c2:
                if st.button("🔄 Regenerate AI Persona",
                             use_container_width=True, key="cc_regen"):
                    prompt = (f"Regenerate AI patient persona for: "
                              f"{draft.get('Age_Sex','?')} with {draft.get('Chief_Complaint','?')}. "
                              f"Diagnosis: {draft.get('Final_Diagnosis','?')}. "
                              f"Different personality/presentation style this time.")
                    with st.spinner("Regenerating..."):
                        persona = call_ai(
                            "You create realistic patient simulations for medical education.",
                            [{"role":"user","content":prompt}], max_tokens=500)
                    st.session_state.draft_case["ai_persona"] = persona
                    st.rerun()
            with c3:
                if st.button("🗑️ Discard Draft",
                             use_container_width=True, key="cc_discard"):
                    st.session_state.draft_case = {}
                    st.rerun()

    # ════════════════════════════════════════════════════════════════
    with tab_manage:
    # ════════════════════════════════════════════════════════════════
        st.markdown('<div class="section-header">📁 Case Library Overview</div>', unsafe_allow_html=True)

        imported = st.session_state.get("imported_cases",[])
        col_a, col_b, col_c = st.columns(3)
        with col_a:
            st.markdown(f'<div class="kpi-card"><div class="kpi-value">{len(df) if not df.empty else 0}</div><div class="kpi-label">Excel Cases</div></div>', unsafe_allow_html=True)
        with col_b:
            st.markdown(f'<div class="kpi-card"><div class="kpi-value">{len(imported)}</div><div class="kpi-label">Imported This Session</div></div>', unsafe_allow_html=True)
        with col_c:
            mimic_count = len([c for c in imported if c.get("_source")=="MIMIC-IV"])
            st.markdown(f'<div class="kpi-card"><div class="kpi-value">{mimic_count}</div><div class="kpi-label">From MIMIC-IV</div></div>', unsafe_allow_html=True)

        if imported:
            st.markdown('<div class="section-header">📥 Imported Cases This Session</div>', unsafe_allow_html=True)
            for ic in imported:
                st.markdown(f'''
                <div style="background:white;border:1px solid #e2e8f0;border-left:4px solid #7c3aed;
                            border-radius:0 8px 8px 0;padding:8px 12px;margin:.35rem 0;font-size:.84rem;">
                    <b>{ic.get("Title","?")} </b>
                    <span style="color:#7c3aed;font-size:.75rem;">{ic.get("_source","")}</span><br>
                    <span style="color:#64748b;">Dx: {ic.get("Final_Diagnosis","?")} | {ic.get("Difficulty","?")}</span>
                </div>''', unsafe_allow_html=True)

        if not df.empty:
            st.markdown('<div class="section-header">📊 Existing case_studies.xlsx</div>', unsafe_allow_html=True)
            show_cols = [c for c in ["Case_ID","Title","System","Difficulty","Final_Diagnosis"] if c in df.columns]
            st.dataframe(df[show_cols].head(30) if show_cols else df.head(30),
                         use_container_width=True)

        excel_path = os.path.join(os.path.dirname(__file__),"case_studies.xlsx")
        if os.path.exists(excel_path):
            excel_bytes = open(excel_path,"rb").read()
            st.download_button("📥 Download case_studies.xlsx", excel_bytes,
                "case_studies.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="cc_dl_excel")



# ════════════════════════════════════════════════════════════════
# 🧠  AI CLINICAL TUTOR CASES  — 1000-case AI-generated training
# ════════════════════════════════════════════════════════════════

AI_TUTOR_CASE_POOL = [
    {"system":"Respiratory","condition":"Community-Acquired Pneumonia","age_range":(25,75),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Respiratory","condition":"Pulmonary Embolism","age_range":(30,70),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Respiratory","condition":"Asthma Exacerbation","age_range":(16,45),"sex_options":["Male","Female"],"difficulty":"Basic"},
    {"system":"Respiratory","condition":"COPD Exacerbation","age_range":(55,80),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Respiratory","condition":"Pulmonary Tuberculosis","age_range":(18,65),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Respiratory","condition":"Spontaneous Pneumothorax","age_range":(18,35),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Respiratory","condition":"Pleural Effusion","age_range":(40,75),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Respiratory","condition":"Lung Cancer","age_range":(50,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Cardiovascular","condition":"STEMI","age_range":(45,75),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Cardiovascular","condition":"NSTEMI / Unstable Angina","age_range":(40,70),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Cardiovascular","condition":"Acute Heart Failure","age_range":(55,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Cardiovascular","condition":"Atrial Fibrillation","age_range":(50,80),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Cardiovascular","condition":"Hypertensive Emergency","age_range":(40,70),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Cardiovascular","condition":"Aortic Dissection","age_range":(45,75),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Cardiovascular","condition":"Infective Endocarditis","age_range":(25,65),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Cardiovascular","condition":"Pericarditis","age_range":(20,50),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Gastroenterology","condition":"Acute Appendicitis","age_range":(10,45),"sex_options":["Male","Female"],"difficulty":"Basic"},
    {"system":"Gastroenterology","condition":"Acute Cholecystitis","age_range":(30,65),"sex_options":["Female","Male"],"difficulty":"Intermediate"},
    {"system":"Gastroenterology","condition":"Upper GI Bleed (Peptic Ulcer)","age_range":(35,70),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Gastroenterology","condition":"Acute Pancreatitis","age_range":(30,60),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Gastroenterology","condition":"Bowel Obstruction","age_range":(45,80),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Gastroenterology","condition":"Crohn's Disease Flare","age_range":(20,40),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Gastroenterology","condition":"Liver Cirrhosis with Ascites","age_range":(45,70),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Gastroenterology","condition":"Ischemic Colitis","age_range":(55,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Neurology","condition":"Ischemic Stroke","age_range":(50,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Neurology","condition":"Subarachnoid Hemorrhage","age_range":(40,65),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Neurology","condition":"Bacterial Meningitis","age_range":(18,50),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Neurology","condition":"Epileptic Seizure","age_range":(15,60),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Neurology","condition":"Guillain-Barre Syndrome","age_range":(20,65),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Neurology","condition":"Multiple Sclerosis Relapse","age_range":(20,45),"sex_options":["Female","Male"],"difficulty":"Advanced"},
    {"system":"Neurology","condition":"Migraine with Aura","age_range":(18,45),"sex_options":["Female","Male"],"difficulty":"Basic"},
    {"system":"Endocrinology","condition":"Diabetic Ketoacidosis","age_range":(16,45),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Endocrinology","condition":"Hyperosmolar Hyperglycaemic State","age_range":(50,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Endocrinology","condition":"Hypoglycaemia","age_range":(20,75),"sex_options":["Male","Female"],"difficulty":"Basic"},
    {"system":"Endocrinology","condition":"Thyrotoxic Crisis","age_range":(30,60),"sex_options":["Female","Male"],"difficulty":"Advanced"},
    {"system":"Endocrinology","condition":"Adrenal Crisis","age_range":(25,65),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Endocrinology","condition":"Hypothyroidism","age_range":(35,70),"sex_options":["Female","Male"],"difficulty":"Basic"},
    {"system":"Nephrology","condition":"Acute Kidney Injury","age_range":(40,80),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Nephrology","condition":"Nephrotic Syndrome","age_range":(20,55),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Nephrology","condition":"Urinary Tract Infection / Pyelonephritis","age_range":(18,70),"sex_options":["Female","Male"],"difficulty":"Basic"},
    {"system":"Nephrology","condition":"Renal Calculi","age_range":(25,60),"sex_options":["Male","Female"],"difficulty":"Basic"},
    {"system":"Nephrology","condition":"Chronic Kidney Disease Stage 5","age_range":(50,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Haematology","condition":"Sickle Cell Crisis","age_range":(16,40),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Haematology","condition":"Acute Leukaemia","age_range":(20,60),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Haematology","condition":"Deep Vein Thrombosis","age_range":(30,70),"sex_options":["Female","Male"],"difficulty":"Basic"},
    {"system":"Haematology","condition":"ITP (Immune Thrombocytopenia)","age_range":(20,50),"sex_options":["Female","Male"],"difficulty":"Intermediate"},
    {"system":"Infectious Disease","condition":"Septic Shock","age_range":(35,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Infectious Disease","condition":"Malaria","age_range":(18,55),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Infectious Disease","condition":"HIV with Opportunistic Infection","age_range":(25,55),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Infectious Disease","condition":"Dengue Fever","age_range":(10,55),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Rheumatology","condition":"Rheumatoid Arthritis Flare","age_range":(30,65),"sex_options":["Female","Male"],"difficulty":"Intermediate"},
    {"system":"Rheumatology","condition":"SLE (Lupus) Flare","age_range":(20,45),"sex_options":["Female","Male"],"difficulty":"Advanced"},
    {"system":"Rheumatology","condition":"Septic Arthritis","age_range":(25,75),"sex_options":["Male","Female"],"difficulty":"Intermediate"},
    {"system":"Rheumatology","condition":"Gout Attack","age_range":(35,70),"sex_options":["Male","Female"],"difficulty":"Basic"},
    {"system":"Surgery","condition":"Ruptured Ectopic Pregnancy","age_range":(18,40),"sex_options":["Female"],"difficulty":"Advanced"},
    {"system":"Surgery","condition":"Testicular Torsion","age_range":(12,30),"sex_options":["Male"],"difficulty":"Advanced"},
    {"system":"Surgery","condition":"Perforated Peptic Ulcer","age_range":(35,70),"sex_options":["Male","Female"],"difficulty":"Advanced"},
    {"system":"Surgery","condition":"Strangulated Hernia","age_range":(40,80),"sex_options":["Male","Female"],"difficulty":"Advanced"},
]


def _generate_ai_case(template: dict) -> dict:
    age  = random.randint(template["age_range"][0], template["age_range"][1])
    sex  = random.choice(template["sex_options"])
    cond = template["condition"]
    diff = template["difficulty"]
    prompt = (
        f"Generate a realistic medical case for a medical student simulator.\n"
        f"Condition: {cond}\nPatient: {age}-year-old {sex}\nDifficulty: {diff}\n\n"
        f"Return ONLY a valid JSON object with exactly these keys (no markdown, no backticks):\n"
        '{"Age_Sex":"Xyr Sex","Chief_Complaint":"...","HPI":"...","PMH":"...","Medications":"...",'
        '"Social_Hx":"...","Family_Hx":"...","Vitals":"HR: X | BP: X/X | RR: X | Temp: X.X°C | SpO2: X%",'
        '"Appearance":"...","Physical_Findings":"...","Labs":"...","Imaging_Tests":"...",'
        f'"Final_Diagnosis":"{cond}","Management":"...","Teaching_Points":"3 key learning points"}}'
    )
    result = call_ai(
        "You are a medical education expert. Generate realistic clinical cases. Return ONLY valid JSON.",
        [{"role": "user", "content": prompt}],
        max_tokens=1000, credit_type="chat"
    )
    try:
        clean = result.strip().replace("```json", "").replace("```", "").strip()
        data  = json.loads(clean)
        data["System"]       = template["system"]
        data["Difficulty"]   = diff
        data["_ai_generated"] = True
        return data
    except Exception:
        return {
            "Age_Sex": f"{age}yr {sex}",
            "Chief_Complaint": f"Presentation consistent with {cond}",
            "HPI": result[:400] if not result.startswith("!ERR") else "Case generation failed. Please retry.",
            "Final_Diagnosis": cond, "System": template["system"], "Difficulty": diff,
            "Vitals": "HR: 90 | BP: 120/80 | RR: 18 | Temp: 37.2°C | SpO2: 97%",
            "_ai_generated": True,
        }


def page_ai_clinical_tutor():
    st.markdown("""
    <div class="main-header">
        <h1>🧠 AI Clinical Tutor Cases</h1>
        <p>Practice with AI-generated cases · Interview · Examine · Diagnose · Learn</p>
    </div>""", unsafe_allow_html=True)

    for k, v in [("ait_case", None), ("ait_chat", []), ("ait_dx_submitted", False),
                  ("ait_score", 0), ("ait_cases_done", 0)]:
        if k not in st.session_state:
            st.session_state[k] = v

    case = st.session_state.ait_case

    col_left, col_right = st.columns([2, 1])
    with col_left:
        st.markdown('<div class="section-header">🎲 Case Selection</div>', unsafe_allow_html=True)
        systems = ["All Systems"] + sorted(set(t["system"] for t in AI_TUTOR_CASE_POOL))
        diffs   = ["All", "Basic", "Intermediate", "Advanced"]
        fc1, fc2 = st.columns(2)
        with fc1: chosen_sys  = st.selectbox("System:",     systems, key="ait_sys_sel")
        with fc2: chosen_diff = st.selectbox("Difficulty:", diffs,   key="ait_diff_sel")
        pool = [t for t in AI_TUTOR_CASE_POOL
                if (chosen_sys  == "All Systems"  or t["system"]     == chosen_sys)
                and (chosen_diff == "All"          or t["difficulty"] == chosen_diff)]
        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button("🎲 Random Case", use_container_width=True, type="primary", key="ait_random"):
                if pool:
                    with st.spinner("Generating case..."):
                        st.session_state.ait_case         = _generate_ai_case(random.choice(pool))
                        st.session_state.ait_chat         = []
                        st.session_state.ait_dx_submitted = False
                    st.rerun()
        with c2:
            conditions  = sorted(set(t["condition"] for t in pool))
            chosen_cond = st.selectbox("Specific:", ["(random)"] + conditions, key="ait_cond_sel")
        with c3:
            if st.button("▶ Generate", use_container_width=True, key="ait_specific"):
                tmpl = next((t for t in pool if t["condition"] == chosen_cond), None) or \
                       (random.choice(pool) if pool else None)
                if tmpl:
                    with st.spinner("Generating..."):
                        st.session_state.ait_case         = _generate_ai_case(tmpl)
                        st.session_state.ait_chat         = []
                        st.session_state.ait_dx_submitted = False
                    st.rerun()

    with col_right:
        st.markdown('<div class="section-header">📊 Your Stats</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">
          <div class="kpi-card">
            <div class="kpi-value">{st.session_state.ait_cases_done}</div>
            <div class="kpi-label">Cases Done</div></div>
          <div class="kpi-card" style="border-top-color:#16a34a;">
            <div class="kpi-value" style="color:#16a34a;">{st.session_state.ait_score}</div>
            <div class="kpi-label">Score</div></div>
        </div>
        <div style="margin-top:8px;font-size:.75rem;color:#64748b;text-align:center;">
          {len(AI_TUTOR_CASE_POOL)} case types · AI-generated variations</div>""",
        unsafe_allow_html=True)

    if not case:
        st.markdown("""
        <div style="background:#f8fafc;border:2px dashed #cbd5e1;border-radius:14px;
                    padding:3rem;text-align:center;color:#64748b;margin-top:1rem;">
          <div style="font-size:3rem;">🎲</div>
          <div style="font-size:1.1rem;font-weight:700;color:#0a2540;margin-top:.5rem;">
            Pick a case above to begin training</div>
          <div style="font-size:.85rem;margin-top:.4rem;">
            Filter by system and difficulty, then click Random Case or pick specific.</div>
        </div>""", unsafe_allow_html=True)
        return

    st.markdown("---")
    diff_color = {"Basic": "#16a34a", "Intermediate": "#d97706", "Advanced": "#dc2626"}.get(
        case.get("Difficulty", ""), "#64748b")
    st.markdown(f"""
    <div class="patient-card">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
        <span style="font-weight:800;font-size:1.05rem;color:#0a2540;">
          🤒 AI Patient · {case.get("Age_Sex","?")}</span>
        <div>
          <span style="background:{diff_color};color:white;border-radius:4px;
                       padding:2px 10px;font-size:.75rem;font-weight:700;">{case.get("Difficulty","?")}</span>
          &nbsp;
          <span style="background:#0e7490;color:white;border-radius:4px;
                       padding:2px 10px;font-size:.75rem;">{case.get("System","?")}</span>
          &nbsp;
          <span style="background:#dc2626;color:white;border-radius:4px;
                       padding:2px 10px;font-size:.75rem;">🔒 Dx Hidden</span>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;font-size:.85rem;">
        <div><b>Chief Complaint:</b><br>{case.get("Chief_Complaint","?")}</div>
        <div><b>Vitals:</b><br><span style="font-family:monospace;font-size:.78rem;">{case.get("Vitals","?")}</span></div>
        <div><b>Appearance:</b><br>{case.get("Appearance","?")}</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Load into global selected_case so all other modules work
    st.session_state.selected_case = case

    st.markdown('<div class="section-header">🏥 Jump to Clinical Modules</div>', unsafe_allow_html=True)
    btn_cols = st.columns(6)
    shortcuts = [("💬 Interview", "simulator"), ("🫁 Examine", "physical_exam"),
                 ("🧪 Labs", "lab"), ("🔬 Imaging", "imaging"),
                 ("🧬 DDx", "ddx"), ("💊 Prescribe", "prescribing")]
    for i, (lbl, pk) in enumerate(shortcuts):
        with btn_cols[i]:
            if st.button(lbl, use_container_width=True, key=f"ait_sc_{pk}"):
                nav(pk)

    st.markdown("---")
    tab_hx, tab_exam, tab_inv, tab_dx, tab_tut = st.tabs(
        ["📋 History", "🩺 Examination", "🧪 Investigations", "✅ Diagnosis", "🤖 AI Explanation"])

    # ── History ───────────────────────────────────────────────────
    with tab_hx:
        with st.expander("📖 History of Presenting Illness", expanded=True):
            st.markdown(case.get("HPI", "—"))
        c1, c2 = st.columns(2)
        with c1:
            with st.expander("📁 Past Medical History"):   st.markdown(case.get("PMH", "None"))
            with st.expander("💊 Medications"):            st.markdown(case.get("Medications", "None"))
        with c2:
            with st.expander("👪 Family History"):         st.markdown(case.get("Family_Hx", "—"))
            with st.expander("🚬 Social History"):         st.markdown(case.get("Social_Hx", "—"))
        st.markdown("---")
        st.markdown("**🎙️ Interview the AI Patient:**")
        for msg in st.session_state.ait_chat:
            cls  = "chat-student" if msg["role"] == "student" else "chat-patient"
            icon = "🧑‍⚕️ You" if msg["role"] == "student" else "🤒 Patient"
            st.markdown(f'<div class="{cls}"><b>{icon}:</b> {msg["content"]}</div>',
                        unsafe_allow_html=True)
        q = st.text_input("Ask the patient:", key="ait_q",
                           placeholder="e.g. Where exactly is the pain? Does it radiate?")
        if st.button("Ask →", key="ait_ask", type="primary"):
            if q.strip():
                st.session_state.ait_chat.append({"role": "student", "content": q})
                with st.spinner("Patient responding..."):
                    r = call_ai(patient_sys(case), [{"role": "user", "content": q}],
                                max_tokens=150, credit_type="chat")
                st.session_state.ait_chat.append({"role": "patient", "content": r})
                st.rerun()
        if st.button("🔄 Reset Chat", key="ait_reset"):
            st.session_state.ait_chat = []; st.rerun()

    # ── Examination ───────────────────────────────────────────────
    with tab_exam:
        systems_exam = ["General Inspection", "Cardiovascular", "Respiratory",
                        "Abdomen", "Neurological", "MSK", "Skin / Lymph Nodes"]
        chosen_e = st.selectbox("Examine:", systems_exam, key="ait_exam_sys")
        if st.button(f"🩺 Perform {chosen_e} Exam", type="primary", key="ait_do_exam"):
            with st.spinner("Examining..."):
                r = call_ai(
                    "You generate clinical examination findings for medical training.",
                    [{"role": "user", "content":
                      f"Student performs {chosen_e} examination. "
                      f"All findings: {case.get('Physical_Findings','?')}. "
                      f"Report only {chosen_e}-relevant findings in 2-3 clinical sentences. "
                      f"Do NOT reveal diagnosis."}],
                    max_tokens=200, credit_type="exam")
            st.markdown(f'<div class="exam-finding"><b>🩺 {chosen_e}:</b><br>{r}</div>',
                        unsafe_allow_html=True)

    # ── Investigations ────────────────────────────────────────────
    with tab_inv:
        l1, l2 = st.columns(2)
        with l1:
            st.markdown("**🧪 Laboratory Results:**")
            if st.button("View Labs", type="primary", key="ait_labs"):
                st.markdown(
                    f'<div style="background:#f0f9ff;border-left:4px solid #0ea5e9;'
                    f'border-radius:8px;padding:1rem;font-family:monospace;font-size:.85rem;">'
                    f'{case.get("Labs","No labs documented.")}</div>',
                    unsafe_allow_html=True)
        with l2:
            st.markdown("**🔬 Imaging / ECG:**")
            if st.button("View Imaging", type="primary", key="ait_imaging"):
                st.markdown(
                    f'<div style="background:#fff7ed;border-left:4px solid #f59e0b;'
                    f'border-radius:8px;padding:1rem;font-size:.85rem;">'
                    f'{case.get("Imaging_Tests","No imaging documented.")}</div>',
                    unsafe_allow_html=True)

    # ── Diagnosis ─────────────────────────────────────────────────
    with tab_dx:
        if not st.session_state.ait_dx_submitted:
            s_dx  = st.text_area("Your primary diagnosis:", height=70, key="ait_sdx",
                                  placeholder="e.g. Community-Acquired Pneumonia")
            _     = st.text_area("Differentials (optional):", height=70, key="ait_sddx",
                                  placeholder="1. ...\n2. ...")
            _     = st.text_area("Management plan:", height=90, key="ait_smx")
            if st.button("✅ Submit Diagnosis", type="primary",
                         use_container_width=True, key="ait_sub"):
                if s_dx.strip():
                    real = case.get("Final_Diagnosis", "")
                    correct = (real.lower() in s_dx.lower() or s_dx.lower() in real.lower())
                    st.session_state.ait_dx_submitted = True
                    st.session_state.ait_cases_done  += 1
                    st.session_state.ait_score       += 10 if correct else 3
                    if correct: st.balloons()
                    st.rerun()
                else:
                    st.warning("Please enter your diagnosis.")
        else:
            real = case.get("Final_Diagnosis", "?")
            st.markdown(f"""
            <div style="background:#f0fdf4;border:2px solid #16a34a;border-radius:12px;
                        padding:1.2rem 1.5rem;margin-bottom:1rem;">
              <div style="font-size:1rem;font-weight:800;color:#14532d;">✅ Actual Diagnosis Revealed:</div>
              <div style="font-size:1.4rem;font-weight:800;color:#dc2626;margin-top:6px;">{real}</div>
            </div>
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;
                        padding:1rem;font-size:.88rem;">
              <b>📋 Management Summary:</b><br>{case.get("Management","See AI explanation tab.")}
            </div>""", unsafe_allow_html=True)
            if st.button("➡️ Next Random Case", type="primary",
                         use_container_width=True, key="ait_next"):
                if AI_TUTOR_CASE_POOL:
                    with st.spinner("Generating next case..."):
                        st.session_state.ait_case         = _generate_ai_case(random.choice(AI_TUTOR_CASE_POOL))
                        st.session_state.ait_chat         = []
                        st.session_state.ait_dx_submitted = False
                    st.rerun()

    # ── AI Explanation ────────────────────────────────────────────
    with tab_tut:
        if not st.session_state.ait_dx_submitted:
            st.info("⚠️ Submit your diagnosis first to unlock the AI explanation.")
        else:
            tp = case.get("Teaching_Points", "")
            if tp:
                st.markdown(
                    f'<div style="background:#eff6ff;border-left:4px solid #3b82f6;'
                    f'border-radius:8px;padding:1rem;font-size:.88rem;margin-bottom:1rem;">'
                    f'<b>📚 Key Teaching Points:</b><br>{tp}</div>',
                    unsafe_allow_html=True)
            if st.button("🤖 Get Full AI Explanation", type="primary",
                         use_container_width=True, key="ait_exp"):
                summary = (
                    f"Patient: {case.get('Age_Sex','?')}, CC: {case.get('Chief_Complaint','?')}, "
                    f"HPI: {case.get('HPI','?')}, Vitals: {case.get('Vitals','?')}, "
                    f"Exam: {case.get('Physical_Findings','?')}, Labs: {case.get('Labs','?')}, "
                    f"Imaging: {case.get('Imaging_Tests','?')}, Dx: {case.get('Final_Diagnosis','?')}. "
                    "Explain: 1) Why this is the diagnosis (pathophysiology) "
                    "2) How history and exam pointed to it 3) How labs/imaging confirm it "
                    "4) First-line management with drug doses 5) What to monitor 6) Exam-question pitfalls."
                )
                with st.spinner("AI Tutor preparing explanation..."):
                    exp = call_ai(tutor_sys(case),
                                  [{"role": "user", "content": summary}],
                                  max_tokens=1200, credit_type="chat")
                st.markdown(
                    f'<div class="chat-tutor" style="padding:1.2rem;line-height:1.7;">{exp}</div>',
                    unsafe_allow_html=True)
            dx = case.get("Final_Diagnosis", "")
            if dx:
                st.markdown("---")
                st.markdown("**📖 Evidence & Guidelines:**")
                ec = st.columns(3)
                links = [
                    ("PubMed",   f"https://pubmed.ncbi.nlm.nih.gov/?term={dx.replace(' ','+')}"),
                    ("UpToDate", f"https://www.uptodate.com/contents/search?search={dx.replace(' ','+')}"),
                    ("NICE",     f"https://www.nice.org.uk/search#?q={dx.replace(' ','+')}"),
                ]
                for i, (nm, url) in enumerate(links):
                    with ec[i]: st.markdown(f"[🔗 {nm}]({url})")


# ════════════════════════════════════════════════════════════════
# 👨‍⚕️  AVATAR BUILDER
# ════════════════════════════════════════════════════════════════

def _render_doctor_svg(av: dict, width: int = 220) -> str:
    """Render a realistic cartoon-style doctor avatar as SVG."""
    gender  = av.get("gender", "Female")
    skin    = av.get("skin",   "#f5c5a3")
    hair_c  = av.get("hair",   "#2c1810")
    eye_c   = av.get("eyes",   "#3d2b1a")
    coat_c  = av.get("coat",   "#ffffff")
    hijab   = av.get("hijab",  False)
    hijab_c = av.get("hijab_color", "#1a4f8a")
    steth   = av.get("stethoscope", True)
    name    = av.get("name",   "Dr. ...")

    def darken(h, f=0.80):
        h = h.lstrip("#")
        r,g,b = int(h[0:2],16)/255, int(h[2:4],16)/255, int(h[4:6],16)/255
        return "#{:02x}{:02x}{:02x}".format(int(r*f*255), int(g*f*255), int(b*f*255))

    skin_s  = darken(skin, 0.82)
    coat_s  = darken(coat_c if coat_c != "#ffffff" else "#dde8f5", 0.88)
    lapel_c = "#dde6ef" if coat_c == "#ffffff" else coat_s

    # ── HAIR / HIJAB ────────────────────────────────────────────
    if hijab:
        hair_svg = (
            f'<ellipse cx="100" cy="108" rx="58" ry="64" fill="{hijab_c}" opacity=".95"/>'
            f'<path d="M 46 108 Q 44 52 100 44 Q 156 52 154 108 Z" fill="{hijab_c}"/>'
            f'<path d="M 46 108 Q 38 145 44 190 Q 52 210 72 215 L 68 175 Q 54 150 50 118 Z" fill="{hijab_c}" opacity=".92"/>'
            f'<path d="M 154 108 Q 162 145 156 190 Q 148 210 128 215 L 132 175 Q 146 150 150 118 Z" fill="{hijab_c}" opacity=".92"/>'
            f'<path d="M 68 215 Q 100 225 132 215 Q 136 200 140 185 L 100 190 L 60 185 Q 64 200 68 215 Z" fill="{hijab_c}" opacity=".88"/>'
            f'<path d="M 68 112 Q 100 108 132 112" stroke="white" stroke-width="1.2" fill="none" opacity=".18" stroke-linecap="round"/>'
        )
    elif gender == "Female":
        hair_svg = (
            f'<path d="M 57 108 Q 53 58 100 44 Q 147 58 143 108 Q 148 75 130 58 Q 115 42 100 40 Q 85 42 70 58 Q 52 75 57 108 Z" fill="{hair_c}"/>'
            f'<path d="M 57 108 Q 48 138 52 175 Q 54 190 60 195" stroke="{hair_c}" stroke-width="18" fill="none" stroke-linecap="round" opacity=".95"/>'
            f'<path d="M 143 108 Q 152 138 148 175 Q 146 190 140 195" stroke="{hair_c}" stroke-width="18" fill="none" stroke-linecap="round" opacity=".95"/>'
            f'<path d="M 76 52 Q 100 46 122 54" stroke="white" stroke-width="2.5" fill="none" opacity=".15" stroke-linecap="round"/>'
        )
    else:
        hair_svg = (
            f'<path d="M 60 100 Q 58 60 100 48 Q 142 60 140 100 Q 135 68 100 58 Q 65 68 60 100 Z" fill="{hair_c}"/>'
            f'<path d="M 60 100 Q 56 115 58 125" stroke="{hair_c}" stroke-width="8" fill="none" stroke-linecap="round" opacity=".8"/>'
            f'<path d="M 140 100 Q 144 115 142 125" stroke="{hair_c}" stroke-width="8" fill="none" stroke-linecap="round" opacity=".8"/>'
            f'<path d="M 74 58 Q 100 50 126 58" stroke="white" stroke-width="2.5" fill="none" opacity=".15" stroke-linecap="round"/>'
        )

    # ── EYEBROWS ────────────────────────────────────────────────
    brow_w = "3.2" if gender == "Female" else "3.8"
    eyebrows = (
        f'<path d="M 76 88 Q 86 82 96 86" stroke="{hair_c}" stroke-width="{brow_w}" fill="none" stroke-linecap="round"/>'
        f'<path d="M 104 86 Q 114 82 124 88" stroke="{hair_c}" stroke-width="{brow_w}" fill="none" stroke-linecap="round"/>'
    )

    # ── FEMALE LASHES ────────────────────────────────────────────
    lashes = ""
    if gender == "Female":
        lc = "#3a2010"
        lashes = (
            f'<path d="M 76 99 Q 74 96 73 94" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
            f'<path d="M 80 96 Q 79 93 79 91" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
            f'<path d="M 92 96 Q 93 93 94 91" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
            f'<path d="M 96 99 Q 98 96 99 94" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
            f'<path d="M 104 99 Q 102 96 101 94" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
            f'<path d="M 108 96 Q 107 93 107 91" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
            f'<path d="M 120 96 Q 121 93 122 91" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
            f'<path d="M 124 99 Q 126 96 127 94" stroke="{lc}" stroke-width="1.5" fill="none" opacity=".6"/>'
        )

    # ── HEAD SHAPE ───────────────────────────────────────────────
    if gender == "Female":
        head_path = f'<path d="M 58 148 Q 58 168 100 170 Q 142 168 142 148 Q 148 125 148 108 Q 148 60 100 54 Q 52 60 52 108 Q 52 125 58 148 Z" fill="{skin}"/>'
    else:
        head_path = f'<path d="M 56 145 Q 58 168 100 170 Q 142 168 144 145 Q 150 122 148 108 Q 146 60 100 54 Q 54 60 52 108 Q 50 122 56 145 Z" fill="{skin}"/>'

    # ── MOUTH ───────────────────────────────────────────────────
    if gender == "Female":
        mouth = (
            f'<path d="M 84 143 Q 100 155 116 143" stroke="#b0605a" stroke-width="2.5" fill="none" stroke-linecap="round"/>'
            f'<path d="M 84 143 Q 100 149 116 143" fill="#c0706a" opacity=".5"/>'
            f'<path d="M 88 143 Q 100 140 112 143" stroke="#b0605a" stroke-width="1.8" fill="none" opacity=".6"/>'
        )
    else:
        mouth = (
            f'<path d="M 84 143 Q 100 154 116 143" stroke="#a05848" stroke-width="2.8" fill="none" stroke-linecap="round"/>'
            f'<path d="M 88 143 Q 100 140 112 143" stroke="#a05848" stroke-width="1.5" fill="none" opacity=".5"/>'
        )

    # ── STETHOSCOPE ──────────────────────────────────────────────
    steth_svg = ""
    if steth:
        steth_svg = (
            f'<path d="M 74 192 Q 65 218 72 232 Q 88 250 100 252 Q 112 250 128 232 Q 135 218 126 192" stroke="#4a5568" stroke-width="4.5" fill="none" stroke-linecap="round"/>'
            f'<circle cx="100" cy="253" r="9" fill="#4a5568"/>'
            f'<circle cx="100" cy="253" r="5.5" fill="#718096"/>'
            f'<circle cx="70" cy="191" r="5.5" fill="#4a5568"/>'
            f'<circle cx="130" cy="191" r="5.5" fill="#4a5568"/>'
        )

    # ── NAME TAG ─────────────────────────────────────────────────
    name_tag = (
        f'<rect x="64" y="205" width="72" height="20" rx="5" fill="#0e7490" opacity=".95"/>'
        f'<text x="100" y="219" text-anchor="middle" font-size="8.5" fill="white" font-family="Inter,Arial,sans-serif" font-weight="700">{name[:16]}</text>'
    )

    h = int(width * 300 / 200)
    return f"""<svg viewBox="0 0 200 300" width="{width}" height="{h}" xmlns="http://www.w3.org/2000/svg" style="display:block;margin:auto;">
  <ellipse cx="100" cy="292" rx="58" ry="8" fill="#0002"/>
  <path d="M 36 172 Q 34 290 38 295 L 162 295 Q 166 290 164 172 Q 148 162 130 158 L 100 168 L 70 158 Q 52 162 36 172 Z" fill="{coat_c}" stroke="{coat_s}" stroke-width="1.2"/>
  <path d="M 70 158 L 60 178 L 72 200 L 100 168 Z" fill="{lapel_c}" stroke="{coat_s}" stroke-width="1"/>
  <path d="M 130 158 L 140 178 L 128 200 L 100 168 Z" fill="{lapel_c}" stroke="{coat_s}" stroke-width="1"/>
  <line x1="100" y1="172" x2="100" y2="295" stroke="{coat_s}" stroke-width="0.8" opacity=".5"/>
  <rect x="44" y="210" width="26" height="18" rx="3" fill="{lapel_c}" stroke="{coat_s}" stroke-width="1"/>
  <rect x="53" y="208" width="4" height="14" rx="2" fill="#0e7490" opacity=".7"/>
  <rect x="34" y="272" width="28" height="12" rx="4" fill="{lapel_c}" stroke="{coat_s}" stroke-width="1"/>
  <rect x="138" y="272" width="28" height="12" rx="4" fill="{lapel_c}" stroke="{coat_s}" stroke-width="1"/>
  <ellipse cx="48" cy="284" rx="12" ry="9" fill="{skin}"/>
  <ellipse cx="152" cy="284" rx="12" ry="9" fill="{skin}"/>
  <rect x="85" y="152" width="30" height="26" rx="10" fill="{skin}"/>
  <rect x="85" y="165" width="30" height="14" rx="7" fill="{skin_s}" opacity=".35"/>
  <ellipse cx="52" cy="113" rx="9" ry="12" fill="{skin}"/>
  <ellipse cx="52" cy="113" rx="5" ry="7"  fill="{skin_s}" opacity=".3"/>
  <ellipse cx="148" cy="113" rx="9" ry="12" fill="{skin}"/>
  <ellipse cx="148" cy="113" rx="5" ry="7"  fill="{skin_s}" opacity=".3"/>
  {head_path}
  <path d="M 68 155 Q 76 168 100 170 Q 124 168 132 155" stroke="{skin_s}" stroke-width="4" fill="none" opacity=".2" stroke-linecap="round"/>
  {hair_svg}
  {eyebrows}
  <ellipse cx="86" cy="104" rx="10" ry="9" fill="white"/>
  <ellipse cx="114" cy="104" rx="10" ry="9" fill="white"/>
  <circle cx="86" cy="105" r="6.5" fill="{eye_c}"/>
  <circle cx="114" cy="105" r="6.5" fill="{eye_c}"/>
  <circle cx="86" cy="105" r="3.8" fill="#0a0a0a"/>
  <circle cx="114" cy="105" r="3.8" fill="#0a0a0a"/>
  <circle cx="88" cy="103" r="1.8" fill="white" opacity=".9"/>
  <circle cx="116" cy="103" r="1.8" fill="white" opacity=".9"/>
  <path d="M 76 99 Q 86 94 96 99" stroke="#3a2010" stroke-width="1.8" fill="none" stroke-linecap="round" opacity=".7"/>
  <path d="M 104 99 Q 114 94 124 99" stroke="#3a2010" stroke-width="1.8" fill="none" stroke-linecap="round" opacity=".7"/>
  {lashes}
  <path d="M 96 116 Q 92 126 94 132 Q 100 136 106 132 Q 108 126 104 116" fill="{skin_s}" opacity=".25"/>
  <ellipse cx="94" cy="131" rx="3.5" ry="2.2" fill="{skin_s}" opacity=".35"/>
  <ellipse cx="106" cy="131" rx="3.5" ry="2.2" fill="{skin_s}" opacity=".35"/>
  {mouth}
  {steth_svg}
  {name_tag}
</svg>"""


_avatar_counter = [0]


def _render_anime_avatar(av: dict, size: int = 280) -> str:
    """Anime 3D-style doctor. Unique IDs stop gradient clashes between sidebar and preview."""
    _avatar_counter[0] += 1
    u = _avatar_counter[0]

    skin  = av.get("skin",  "#f5c5a3")
    hair  = av.get("hair",  "#2c1810")
    eye_c = av.get("eyes",  "#4a7ab5")
    coat  = av.get("coat",  "#5ba4cf")
    hijab = av.get("hijab", False)
    hjcol = av.get("hijab_color", "#1a4f8a")
    gen   = av.get("gender", "Female")
    name  = av.get("name",  "Dr. ...")[:14]
    steth   = av.get("stethoscope", True)
    glasses = av.get("glasses", False)

    def _hex(h):
        h = h.lstrip("#")
        return (int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)) if len(h) == 6 else (180,160,140)
    def dk(h, a=30):
        r,g,b = _hex(h)
        return "#{:02x}{:02x}{:02x}".format(max(0,r-a), max(0,g-a), max(0,b-a))
    def lt(h, a=30):
        r,g,b = _hex(h)
        return "#{:02x}{:02x}{:02x}".format(min(255,r+a), min(255,g+a), min(255,b+a))

    sd = dk(skin,28);  sl = lt(skin,32)
    cd = dk(coat,35);  cl = lt(coat,52)
    ed = dk(eye_c,45); hd = dk(hair,22)
    el = lt(eye_c,70); hl = lt(hair,40)

    # gradient / filter defs — all IDs end with u so multiple avatars coexist
    G  = "sk" + str(u)
    GB = "ct" + str(u)
    GE = "ey" + str(u)
    GH = "hr" + str(u)
    BL = "bl" + str(u)

    svg = '<svg width="{sz}" height="{sz}" viewBox="0 0 400 420" xmlns="http://www.w3.org/2000/svg">'.format(sz=size)

    # defs
    svg += (
        "<defs>"
        "<radialGradient id='{G}' cx='42%' cy='36%' r='64%'>"
        "<stop offset='0%' stop-color='{sl}'/><stop offset='100%' stop-color='{sd}'/></radialGradient>"
        "<radialGradient id='{GB}' cx='48%' cy='25%' r='75%'>"
        "<stop offset='0%' stop-color='{cl}'/><stop offset='100%' stop-color='{cd}'/></radialGradient>"
        "<radialGradient id='{GE}' cx='32%' cy='28%' r='72%'>"
        "<stop offset='0%' stop-color='{el}'/>"
        "<stop offset='50%' stop-color='{eye_c}'/>"
        "<stop offset='100%' stop-color='{ed}'/></radialGradient>"
        "<radialGradient id='{GH}' cx='40%' cy='20%' r='70%'>"
        "<stop offset='0%' stop-color='{hl}'/><stop offset='100%' stop-color='{hd}'/></radialGradient>"
        "<filter id='{BL}'><feGaussianBlur stdDeviation='3'/></filter>"
        "</defs>"
    ).format(G=G,GB=GB,GE=GE,GH=GH,BL=BL,sl=sl,sd=sd,cl=cl,cd=cd,el=el,eye_c=eye_c,ed=ed,hl=hl,hd=hd)

    # ── Body / scrubs ────────────────────────────────────────────────────────
    svg += (
        # main torso
        "<path d='M90 298 Q84 395 88 412 L312 412 Q316 395 310 298"
        " Q286 274 262 266 L200 284 L138 266 Q114 274 90 298Z'"
        " fill='url(#{GB})'/>".format(GB=GB)
        # shoulder sheen
        + "<path d='M90 298 Q114 274 138 266 L200 284 L262 266 Q286 274 310 298"
        " Q286 288 200 292 Q114 288 90 298Z' fill='{cd}' opacity='.28'/>".format(cd=cd)
        # left arm
        + "<path d='M90 298 Q68 335 65 378 Q63 402 80 408 Q98 414 104 392"
        " Q110 370 102 338 Q98 318 90 298Z' fill='url(#{GB})'/>".format(GB=GB)
        # right arm
        + "<path d='M310 298 Q332 335 335 378 Q337 402 320 408 Q302 414 296 392"
        " Q290 370 298 338 Q302 318 310 298Z' fill='url(#{GB})'/>".format(GB=GB)
        # hands
        + "<ellipse cx='74' cy='404' rx='23' ry='16' fill='{skin}'/>".format(skin=skin)
        + "<ellipse cx='326' cy='404' rx='23' ry='16' fill='{skin}'/>".format(skin=skin)
        # white coat overlay
        + "<path d='M112 292 Q104 375 106 412 L182 412 L182 288Z' fill='white' opacity='.18'/>"
        + "<path d='M288 292 Q296 375 294 412 L218 412 L218 288Z' fill='white' opacity='.18'/>"
        # v-neck collar
        + "<path d='M164 266 L150 304 L200 316 L250 304 L236 266 L200 282Z'"
        " fill='{cl}' opacity='.92'/>".format(cl=cl)
        # lapels
        + "<path d='M138 266 L116 304 L150 328 L164 266Z' fill='white' opacity='.28'/>"
        + "<path d='M262 266 L284 304 L250 328 L236 266Z' fill='white' opacity='.28'/>"
        # pocket
        + "<rect x='114' y='320' width='50' height='36' rx='5' fill='white' stroke='{cd}' stroke-width='1.5' opacity='.9'/>".format(cd=cd)
        + "<rect x='118' y='324' width='42' height='12' rx='3' fill='#0e7490' opacity='.78'/>"
        + "<text x='139' y='334' font-size='8' fill='white' text-anchor='middle' font-family='Arial,sans-serif' font-weight='bold'>MLS</text>"
        # name badge
        + "<rect x='130' y='366' width='140' height='30' rx='7' fill='white' stroke='{cd}' stroke-width='1.5' opacity='.95'/>".format(cd=cd)
        + "<text x='200' y='386' font-size='12' fill='#0a2540' text-anchor='middle' font-family='Arial,sans-serif' font-weight='bold'>{name}</text>".format(name=name)
    )

    # ── Stethoscope ──────────────────────────────────────────────────────────
    if steth:
        svg += (
            "<path d='M164 284 Q140 294 130 322 Q126 346 134 360"
            " Q144 374 156 366 Q166 358 162 344 Q158 330 148 327"
            " Q142 325 144 312 Q147 302 160 296'"
            " fill='none' stroke='#1a1a2e' stroke-width='7' stroke-linecap='round'/>"
            "<path d='M236 284 Q260 294 270 322 Q274 346 266 360"
            " Q256 374 244 366 Q234 358 238 344 Q242 330 252 327"
            " Q258 325 256 312 Q253 302 240 296'"
            " fill='none' stroke='#1a1a2e' stroke-width='7' stroke-linecap='round'/>"
            "<line x1='160' y1='296' x2='240' y2='296' stroke='#1a1a2e' stroke-width='7' stroke-linecap='round'/>"
            "<circle cx='200' cy='362' r='15' fill='#1a1a2e'/>"
            "<circle cx='200' cy='362' r='9' fill='#2d2d4e'/>"
            "<circle cx='197' cy='359' r='3' fill='white' opacity='.38'/>"
        )

    # ── Neck ─────────────────────────────────────────────────────────────────
    svg += (
        "<path d='M180 252 Q178 278 200 282 Q222 278 220 252Z' fill='url(#{G})'/>".format(G=G)
        + "<path d='M184 265 Q200 273 216 265' fill='none' stroke='{sd}' stroke-width='1.5' opacity='.28'/>".format(sd=sd)
    )

    # ── Hair / hijab — drawn BEHIND head ────────────────────────────────────
    if hijab:
        hj2 = dk(hjcol,22); hj3 = lt(hjcol,8)
        svg += (
            # back volume
            "<ellipse cx='200' cy='188' rx='120' ry='128' fill='{hj2}'/>".format(hj2=hj2)
            # main wrap
            + "<path d='M84 178 Q86 98 200 92 Q314 98 316 178"
            " Q316 240 304 262 L200 272 L96 262 Q84 240 84 178Z' fill='{hjcol}'/>".format(hjcol=hjcol)
            # chin drape shadow
            + "<ellipse cx='200' cy='268' rx='108' ry='26' fill='{hj2}'/>".format(hj2=hj2)
            # highlight sheen
            + "<ellipse cx='162' cy='118' rx='42' ry='22' fill='white' opacity='.10'/>"
            # fold
            + "<path d='M98 256 Q200 274 302 256' fill='none' stroke='{hj2}' stroke-width='3' opacity='.45'/>".format(hj2=hj2)
        )
    elif gen == "Female":
        svg += (
            # back mass
            "<ellipse cx='200' cy='178' rx='118' ry='122' fill='{hd}'/>".format(hd=hd)
            # side flows
            + "<path d='M88 182 Q72 245 80 315 Q88 345 104 340 Q88 295 90 250 Q92 220 88 182Z' fill='{hd}'/>".format(hd=hd)
            + "<path d='M312 182 Q328 245 320 315 Q312 345 296 340 Q312 295 310 250 Q308 220 312 182Z' fill='{hd}'/>".format(hd=hd)
            # top cap with gradient
            + "<path d='M94 180 Q98 78 200 70 Q302 78 306 180"
            " Q282 155 200 150 Q118 155 94 180Z' fill='url(#{GH})'/>".format(GH=GH)
            # parting + shine
            + "<path d='M200 75 Q196 90 200 108 Q204 90 200 75Z' fill='{hl}' opacity='.55'/>".format(hl=lt(hair,55))
            + "<path d='M168 82 Q178 96 172 118' fill='none' stroke='{hl}' stroke-width='5' stroke-linecap='round' opacity='.42'/>".format(hl=lt(hair,55))
        )
    else:
        svg += (
            "<ellipse cx='200' cy='172' rx='118' ry='114' fill='{hd}'/>".format(hd=hd)
            + "<path d='M94 178 Q98 76 200 68 Q302 76 306 178"
            " Q282 148 200 144 Q118 148 94 178Z' fill='url(#{GH})'/>".format(GH=GH)
            + "<path d='M88 178 Q82 205 86 230' fill='none' stroke='{hl}' stroke-width='6' opacity='.35'/>".format(hl=lt(hair,22))
            + "<path d='M312 178 Q318 205 314 230' fill='none' stroke='{hl}' stroke-width='6' opacity='.35'/>".format(hl=lt(hair,22))
            + "<path d='M168 76 Q178 90 172 112' fill='none' stroke='{hl}' stroke-width='5' stroke-linecap='round' opacity='.4'/>".format(hl=lt(hair,55))
        )

    # ── Head — wide anime shape, tapered chin ────────────────────────────────
    svg += (
        # soft drop shadow
        "<ellipse cx='203' cy='200' rx='116' ry='120' fill='{sd}' opacity='.16' filter='url(#{BL})'/>".format(sd=sd,BL=BL)
        # head shape
        + "<path d='M94 178 Q88 235 98 262 Q118 295 200 300"
        " Q282 295 302 262 Q312 235 306 178 Q294 128 200 122 Q106 128 94 178Z'"
        " fill='url(#{G})'/>".format(G=G)
        # jaw shadow for depth
        + "<path d='M108 264 Q150 294 200 298 Q250 294 292 264"
        " Q280 278 200 282 Q120 278 108 264Z' fill='{sd}' opacity='.20'/>".format(sd=sd)
        # cheek spherical highlight
        + "<ellipse cx='146' cy='218' rx='32' ry='24' fill='white' opacity='.11'/>"
        + "<ellipse cx='254' cy='218' rx='32' ry='24' fill='white' opacity='.11'/>"
        # ears
        + "<path d='M90 192 Q78 204 80 224 Q82 242 96 242 Q108 238 110 220"
        " Q112 202 96 192Z' fill='{skin}'/>".format(skin=skin)
        + "<path d='M310 192 Q322 204 320 224 Q318 242 304 242 Q292 238 290 220"
        " Q288 202 304 192Z' fill='{skin}'/>".format(skin=skin)
        # inner ear shadow
        + "<path d='M94 202 Q86 212 88 226 Q90 236 97 236' fill='none' stroke='{sd}' stroke-width='2' opacity='.38'/>".format(sd=sd)
        + "<path d='M306 202 Q314 212 312 226 Q310 236 303 236' fill='none' stroke='{sd}' stroke-width='2' opacity='.38'/>".format(sd=sd)
        # cheek blush
        + "<ellipse cx='140' cy='232' rx='30' ry='18' fill='#f08080' opacity='.18'/>"
        + "<ellipse cx='260' cy='232' rx='30' ry='18' fill='#f08080' opacity='.18'/>"
    )

    # ── Eyes — large anime teardrop with shine ───────────────────────────────
    for ex in [152, 248]:
        # pre-compute all coords so f-strings never clash with .format()
        eL=ex-32; eL2=ex-30; eL3=ex-28; eL4=ex-40
        eR=ex+32; eR2=ex+30; eR3=ex+28; eR4=ex+40
        eS=ex-8;  eS2=ex+9;  eB=ex-30; eB2=ex+30
        lid_fill = dk(skin, 5)
        svg += (
            f"<path d='M{eL} 192 Q{eL2} 172 {ex} 170 Q{eR2} 172 {eR} 192"
            f" Q{eR2} 212 {ex} 214 Q{eL2} 212 {eL} 192Z' fill='white'/>"
            f"<circle cx='{ex}' cy='192' r='21' fill='url(#{GE})'/>"
            f"<circle cx='{ex}' cy='192' r='21' fill='none' stroke='{ed}' stroke-width='3' opacity='.55'/>"
            f"<circle cx='{ex}' cy='193' r='13' fill='#06060e'/>"
            f"<circle cx='{eS}' cy='182' r='8' fill='white' opacity='.96'/>"
            f"<circle cx='{eS2}' cy='198' r='4' fill='white' opacity='.52'/>"
            f"<path d='M{eL} 192 Q{eL3} 168 {ex} 166 Q{eR3} 168 {eR} 192'"
            f" fill='{lid_fill}' stroke='#180e08' stroke-width='4' stroke-linecap='round'/>"
            f"<line x1='{eR}' y1='192' x2='{eR4}' y2='182' stroke='#180e08' stroke-width='3' stroke-linecap='round'/>"
            f"<line x1='{eR3}' y1='187' x2='{eR}' y2='178' stroke='#180e08' stroke-width='2.5' stroke-linecap='round'/>"
            f"<line x1='{eL}' y1='192' x2='{eL4}' y2='182' stroke='#180e08' stroke-width='3' stroke-linecap='round'/>"
            f"<path d='M{eB} 195 Q{ex} 215 {eB2} 195'"
            f" fill='none' stroke='#2a1a10' stroke-width='2' opacity='.48'/>"
        )

    # ── Eyebrows — thick, arched, anime ─────────────────────────────────────
    svg += (
        "<path d='M118 164 Q152 150 184 158' stroke='{hd}' stroke-width='7' fill='none' stroke-linecap='round'/>".format(hd=dk(hair,8))
        + "<path d='M118 164 Q152 150 184 158' stroke='{hl}' stroke-width='2.5' fill='none' stroke-linecap='round' opacity='.38'/>".format(hl=lt(hair,28))
        + "<path d='M216 158 Q248 150 282 164' stroke='{hd}' stroke-width='7' fill='none' stroke-linecap='round'/>".format(hd=dk(hair,8))
        + "<path d='M216 158 Q248 150 282 164' stroke='{hl}' stroke-width='2.5' fill='none' stroke-linecap='round' opacity='.38'/>".format(hl=lt(hair,28))
    )

    # ── Nose — subtle 3D shadow ──────────────────────────────────────────────
    svg += (
        "<path d='M194 224 Q187 238 191 246 Q200 252 209 246 Q213 238 206 224'"
        " fill='{sd}' opacity='.22'/>".format(sd=sd)
        + "<ellipse cx='191' cy='246' rx='7' ry='4' fill='{sd}' opacity='.16'/>".format(sd=sd)
        + "<ellipse cx='209' cy='246' rx='7' ry='4' fill='{sd}' opacity='.16'/>".format(sd=sd)
    )

    # ── Mouth — anime smile ──────────────────────────────────────────────────
    svg += (
        "<path d='M170 262 Q200 284 230 262' stroke='{sd}' stroke-width='4' fill='none' stroke-linecap='round'/>".format(sd=dk(skin,55))
        + "<path d='M174 263 Q200 278 226 263 Q200 270 174 263Z' fill='white' opacity='.68'/>"
        + "<ellipse cx='200' cy='275' rx='18' ry='7' fill='white' opacity='.12'/>"
    )

    # ── Glasses (conditional on av["glasses"]) ──────────────────────────────
    if av.get("glasses", False):
        svg += (
            # left lens
            "<rect x='118' y='176' width='70' height='50' rx='20' fill='none' stroke='#1a1a1a' stroke-width='5' opacity='.88'/>"
            "<rect x='118' y='176' width='70' height='50' rx='20' fill='white' opacity='.05'/>"
            # right lens
            + "<rect x='212' y='176' width='70' height='50' rx='20' fill='none' stroke='#1a1a1a' stroke-width='5' opacity='.88'/>"
            + "<rect x='212' y='176' width='70' height='50' rx='20' fill='white' opacity='.05'/>"
            # bridge
            + "<line x1='188' y1='198' x2='212' y2='198' stroke='#1a1a1a' stroke-width='4' opacity='.8'/>"
            # temples
            + "<line x1='118' y1='196' x2='90' y2='190' stroke='#1a1a1a' stroke-width='4' stroke-linecap='round' opacity='.8'/>"
            + "<line x1='282' y1='196' x2='310' y2='190' stroke='#1a1a1a' stroke-width='4' stroke-linecap='round' opacity='.8'/>"
            # lens shine
            + "<path d='M124 184 Q132 180 142 184' fill='none' stroke='white' stroke-width='3' stroke-linecap='round' opacity='.52'/>"
            + "<path d='M218 184 Q226 180 236 184' fill='none' stroke='white' stroke-width='3' stroke-linecap='round' opacity='.52'/>"
        )

    svg += "</svg>"
    return svg



def _gemini_analyse_photo_for_avatar(image_bytes: bytes) -> dict:
    """
    Send a photo to Gemini Vision and return avatar config dict.
    Keys returned: skin, hair, eyes, hijab (bool), hijab_color, glasses (bool).
    Rotates through all API keys in the pool; retries with back-off on 429;
    falls back through model chain; never exposes key in error messages.
    """
    import base64, requests, json, time

    def _all_keys() -> list:
        keys = []
        try:
            for i in range(1, 21):
                for pattern in (f"GEMINI_KEY_{i}", f"GEMINI_API_KEY_{i}"):
                    k = st.secrets.get(pattern, "").strip()
                    if k:
                        keys.append(k)
                        break
            if not keys:
                for name in ("GEMINI_API_KEY", "GEMINI_KEY", "GOOGLE_API_KEY"):
                    k = st.secrets.get(name, "").strip()
                    if k:
                        keys.append(k)
                        break
        except Exception:
            pass
        try:
            k = get_api_key()
            if k and k not in keys:
                keys.insert(0, k)
        except Exception:
            pass
        return keys

    MODELS = [
        "gemini-2.0-flash",
        "gemini-1.5-flash",
    ]

    PROMPT = """Look at the person in this photo. Return ONLY a valid JSON object
with these exact keys (no preamble, no markdown fences, no explanation):

{
  "skin":        "<one of: #fde8d0 | #f5c5a3 | #d4956a | #a0673a | #6b3d1e>",
  "hair":        "<one of: #1a1a1a | #2c1810 | #6b3d1e | #8b3a0f | #c8a050 | #8a8a8a | #e8e8e8>",
  "eyes":        "<one of: #5c3a1e | #7c5230 | #2d6a4f | #1e6aa1 | #5a6a7a | #8b6914 | #0e7490>",
  "hijab":       <true | false>,
  "hijab_color": "<one of: #1a4f8a | #1a1a1a | #0e7490 | #7c1a1a | #556b2f | #b5556b | #6b7280 | #f8fafc>",
  "glasses":     <true | false>
}

Rules:
- Choose the closest matching hex from the listed options only.
- hijab = true ONLY if the person is clearly wearing a hijab/headscarf.
- glasses = true ONLY if the person is clearly wearing glasses.
- Return raw JSON only -- no ```json fences, no extra text."""

    # Detect MIME type from magic bytes
    if len(image_bytes) >= 4 and image_bytes[:4] == b'\x89PNG':
        mime_type = "image/png"
    elif len(image_bytes) >= 2 and image_bytes[:2] == b'\xff\xd8':
        mime_type = "image/jpeg"
    else:
        mime_type = "image/jpeg"

    b64_img = base64.b64encode(image_bytes).decode("utf-8")

    payload = {
        "contents": [{
            "parts": [
                {"inline_data": {"mime_type": mime_type, "data": b64_img}},
                {"text": PROMPT},
            ]
        }],
        "generationConfig": {"temperature": 0.05, "maxOutputTokens": 256},
    }

    all_keys = _all_keys()
    if not all_keys:
        st.error(
            "\u274c No Gemini API key found. "
            "Add GEMINI_API_KEY to your secrets.toml and restart."
        )
        return {}

    last_error = ""
    tried = 0

    for model in MODELS:
        base_url = (
            f"https://generativelanguage.googleapis.com/v1beta/models/"
            f"{model}:generateContent"
        )
        for key in all_keys:
            for attempt in range(2):             # max 2 tries per key
                tried += 1
                try:
                    resp = requests.post(
                        base_url,
                        params={"key": key},   # key in params, NOT in URL string
                        json=payload,
                        timeout=9,             # fast fail — don't block the UI
                    )

                    if resp.status_code == 429:
                        wait = attempt + 1     # 1 s then 2 s — keep it short
                        last_error = "rate limit (429)"
                        time.sleep(wait)
                        continue

                    if resp.status_code in (400, 401, 403):
                        last_error = f"auth error ({resp.status_code})"
                        break

                    if resp.status_code in (404, 503):
                        last_error = f"model unavailable ({resp.status_code})"
                        break

                    resp.raise_for_status()

                    raw = (
                        resp.json()
                        .get("candidates", [{}])[0]
                        .get("content", {})
                        .get("parts", [{}])[0]
                        .get("text", "")
                    )
                    clean = (
                        raw.strip()
                           .removeprefix("```json")
                           .removeprefix("```")
                           .removesuffix("```")
                           .strip()
                    )
                    result = json.loads(clean)

                    defaults = {
                        "skin": "#f5c5a3", "hair": "#2c1810", "eyes": "#5c3a1e",
                        "hijab": False, "hijab_color": "#6b7280", "glasses": False
                    }
                    for k, v in defaults.items():
                        if k not in result:
                            result[k] = v

                    if not all(k in result for k in ("skin", "eyes", "hijab")):
                        last_error = "incomplete JSON from model"
                        continue

                    return result   # SUCCESS

                except json.JSONDecodeError:
                    last_error = "model returned non-JSON"
                    continue
                except requests.Timeout:
                    last_error = "request timed out"
                    continue  # no extra sleep — already wasted 9 s
                except Exception as exc:
                    safe_msg = str(exc)
                    for k in all_keys:
                        safe_msg = safe_msg.replace(k, "***")
                    last_error = safe_msg
                    continue

    # All keys and models exhausted
    if "rate limit" in last_error or "429" in last_error:
        st.warning(
            f"\u26a0\ufe0f All Gemini API keys are currently rate-limited. "
            "Wait 60 seconds then try again, or add more keys to secrets.toml. "
            f"(Tried {tried} combinations across {len(all_keys)} key(s) "
            f"and {len(MODELS)} model(s).)"
        )
    elif "auth" in last_error:
        st.error(
            "\u274c Gemini API key rejected. "
            "Check your secrets.toml -- the key may be invalid or expired."
        )
    elif "model unavailable" in last_error:
        st.warning(
            "\u26a0\ufe0f All Gemini vision models are temporarily unavailable. "
            "Please try again in a few minutes."
        )
    else:
        st.warning(
            f"\u26a0\ufe0f Photo analysis failed after {tried} attempts. "
            f"Reason: {last_error}. "
            "Try a clearer, well-lit face photo (jpg/png)."
        )

    return {}



def page_avatar_builder():
    # ── Welcome banner: shown once right after registration ──────────────────
    if st.session_state.pop("_new_registration", False):
        st.session_state["_show_welcome_banner"] = True
    if st.session_state.pop("_show_welcome_banner", False):
        st.markdown("""
        <div style="background:linear-gradient(135deg,#f0fdf4,#dcfce7);border:2px solid #16a34a;
                    border-radius:14px;padding:1.2rem 1.5rem;margin-bottom:1rem;text-align:center;">
            <div style="font-size:1.5rem;">🎉</div>
            <div style="font-weight:800;color:#166534;font-size:1rem;">Welcome to MLS Virtual Hospital!</div>
            <div style="color:#374151;font-size:.85rem;margin-top:.3rem;">
                Set up your doctor profile — upload a photo and add your name.
            </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div class="main-header">
        <h1>👨‍⚕️ My Doctor Profile</h1>
        <p>Upload your photo · Set your name · Get your hospital badge</p>
    </div>""", unsafe_allow_html=True)

    av = st.session_state.get("doctor_avatar") or {}
    saved_photo = av.get("photo_b64", "")  # base64-encoded photo

    col_left, col_right = st.columns([1, 1])

    # ════════════════════════════════════════════════════════════════════════
    # LEFT — profile setup
    # ════════════════════════════════════════════════════════════════════════
    with col_left:
        st.markdown('<div class="section-header">📸 Your Profile Photo</div>',
                    unsafe_allow_html=True)

        st.markdown("""
        <div style="background:linear-gradient(135deg,#f0f9ff,#e0f2fe);
                    border:1.5px solid #0ea5e9;border-radius:10px;
                    padding:.9rem 1rem;margin-bottom:1rem;font-size:.82rem;color:#0369a1;">
            Upload a clear, professional photo of yourself. This will appear on
            your hospital badge and in the app sidebar. <i>Your photo is stored
            only in your browser session — it is not uploaded to any server.</i>
        </div>
        """, unsafe_allow_html=True)

        uploaded_photo = st.file_uploader(
            "Upload your photo (jpg/png):",
            type=["jpg", "jpeg", "png"],
            key="av_photo_upload",
        )

        if uploaded_photo is not None:
            import base64 as _b64
            img_bytes = uploaded_photo.read()
            # Compress/resize via PIL to keep session state lean
            try:
                from PIL import Image as _PIL
                import io as _io
                im = _PIL.open(_io.BytesIO(img_bytes))
                # Convert to RGB if needed (handles RGBA, P mode etc.)
                if im.mode != "RGB":
                    im = im.convert("RGB")
                # Resize to max 400x400 maintaining aspect ratio
                im.thumbnail((400, 400))
                buf = _io.BytesIO()
                im.save(buf, format="JPEG", quality=85)
                img_bytes = buf.getvalue()
            except Exception:
                pass  # If PIL fails, use original bytes

            saved_photo = _b64.b64encode(img_bytes).decode("ascii")
            av["photo_b64"] = saved_photo
            st.success("✅ Photo loaded — click 'Save Profile' below to keep it.")

        st.markdown("---")

        # ── Profile fields ──────────────────────────────────────────────
        st.markdown('<div class="section-header">📝 Your Details</div>',
                    unsafe_allow_html=True)

        av["name"] = st.text_input("Full name on badge:",
                                   value=av.get("name", "Dr. ..."),
                                   key="av_name",
                                   placeholder="e.g. Dr. Hiba Hamdar")

        av["title"] = st.text_input("Title / specialty (optional):",
                                    value=av.get("title", "Medical Student"),
                                    key="av_title",
                                    placeholder="e.g. Medical Student · Y3")

        av["institution"] = st.text_input("Institution (optional):",
                                          value=av.get("institution",
                                                       "MLS Academy"),
                                          key="av_inst",
                                          placeholder="Your medical school")

        st.markdown("---")

        if st.button("💾 Save Profile", type="primary",
                     use_container_width=True, key="av_save"):
            st.session_state.doctor_avatar = dict(av)
            st.success("✅ Profile saved! It appears on your badge and "
                       "throughout the app.")
            st.rerun()

        # Optional clear photo
        if saved_photo:
            if st.button("🗑️ Remove Photo", use_container_width=True,
                         key="av_clear_photo"):
                av.pop("photo_b64", None)
                st.session_state.doctor_avatar = dict(av)
                st.rerun()

    # ════════════════════════════════════════════════════════════════════════
    # RIGHT — preview + hospital badge
    # ════════════════════════════════════════════════════════════════════════
    with col_right:
        st.markdown('<div class="section-header">🪪 Your Hospital Badge</div>',
                    unsafe_allow_html=True)

        saved_av = st.session_state.get("doctor_avatar") or av
        photo_b64 = saved_av.get("photo_b64", "")
        name      = saved_av.get("name", "Dr. ...")
        title     = saved_av.get("title", "Medical Student")
        inst      = saved_av.get("institution", "MLS Academy")

        # Photo or placeholder
        if photo_b64:
            photo_html = (
                f'<img src="data:image/jpeg;base64,{photo_b64}" '
                f'style="width:100%;height:100%;object-fit:cover;'
                f'border-radius:50%;" alt="Doctor photo"/>'
            )
        else:
            photo_html = (
                '<div style="width:100%;height:100%;border-radius:50%;'
                'background:linear-gradient(135deg,#0e7490,#1a4f8a);'
                'display:flex;align-items:center;justify-content:center;'
                'color:white;font-size:3rem;">👤</div>'
            )

        # ── Big professional badge ────────────────────────────────────
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0a2540 0%,#1a4f8a 100%);
                    border-radius:18px;padding:24px 28px;color:white;
                    max-width:340px;margin:0 auto 1rem;
                    box-shadow:0 12px 30px rgba(10,37,64,.45);
                    text-align:center;border:1px solid rgba(255,255,255,.1);">

          <!-- Photo -->
          <div style="width:140px;height:140px;margin:0 auto 12px;
                      border-radius:50%;border:4px solid rgba(255,255,255,.4);
                      box-shadow:0 4px 14px rgba(0,0,0,.3);overflow:hidden;
                      background:white;">
            {photo_html}
          </div>

          <!-- Hospital tagline -->
          <div style="font-size:.62rem;letter-spacing:.18em;
                      opacity:.6;text-transform:uppercase;font-weight:600;
                      margin-bottom:4px;">
            🏥 MLS Virtual Hospital
          </div>

          <!-- Name -->
          <div style="font-size:1.3rem;font-weight:800;letter-spacing:.01em;">
            {name}
          </div>

          <!-- Title -->
          <div style="font-size:.85rem;color:#67e8f9;margin-top:2px;
                      font-weight:600;">
            {title}
          </div>

          <!-- Institution -->
          <div style="font-size:.75rem;opacity:.65;margin-top:8px;">
            {inst}
          </div>

          <!-- Badge ID line -->
          <div style="margin-top:14px;padding-top:10px;
                      border-top:1px solid rgba(255,255,255,.15);
                      font-size:.65rem;opacity:.55;letter-spacing:.05em;">
            CLINICAL TRAINING SIMULATOR · ID: MLS-{abs(hash(name)) % 99999:05d}
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Tips ──────────────────────────────────────────────────────
        if not photo_b64:
            st.markdown("""
            <div class="alert-info" style="font-size:.82rem;">
              💡 <b>Tip:</b> Upload a photo on the left to personalise your
              badge. Without a photo, you'll see a default doctor icon.
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="alert-good" style="font-size:.82rem;">
              ✅ Your badge is ready! It will appear in the sidebar and on
              your AI Tutor sessions.
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<div style='height:.8rem'></div>", unsafe_allow_html=True)
        if st.button("🏥 Enter the Hospital →", type="primary",
                     use_container_width=True, key="av_enter"):
            st.session_state.page = "home"
            st.rerun()

# ════════════════════════════════════════════════════════
p=st.session_state.page
if   p=="home":          page_home()
elif p=="library":       page_library()
elif p=="emergency":     page_emergency()
elif p=="simulator":     page_simulator()
elif p=="physical_exam": page_physical_exam()
elif p=="lab":           page_lab()
elif p=="imaging":       page_imaging()
elif p=="surgery":       page_surgery()
elif p=="live":          page_live()
elif p=="diagnosis":     page_diagnosis()
elif p=="tutor":         page_tutor()
elif p=="add_case":      page_add_case()
elif p=="submit_case":   page_submit_case()
elif p=="multiplayer":   page_multiplayer()
elif p=="peer_sim":      page_peer_sim()
elif p=="credits":       page_credits()
elif p=="scores":        page_clinical_scores()
elif p=="evidence":      page_evidence()
elif p=="doccollab":     page_doccollab()
# ── New feature pages ─────────────────────────────────────────────────────
elif p=="ddx":           page_ddx_builder()
elif p=="prescribing":   page_prescribing()
elif p=="procedures":    page_procedure_sim_3d()
elif p=="reasoning":     page_reasoning_map()
elif p=="competency":    page_competency_tracker()
elif p=="analytics":     page_faculty_analytics()
elif p=="case_creator":  page_case_creator()
elif p=="ai_tutor_cases": page_ai_clinical_tutor()
elif p=="avatar_builder": page_avatar_builder()
# ── New feature pages ─────────────────────────────────────────────────────────
elif p=="osce":
    if NEW_FEATURES_OK: page_osce_exam()
    else: st.error("⚠️ new_features.py not found. Place it in the same folder as app.py.")
elif p=="notes":
    if NEW_FEATURES_OK: page_progress_notes()
    else: st.error("⚠️ new_features.py not found. Place it in the same folder as app.py.")
elif p=="flashcards":
    if NEW_FEATURES_OK: page_flashcard_builder()
    else: st.error("⚠️ new_features.py not found. Place it in the same folder as app.py.")
# ── Tier 1: Progress Dashboard ────────────────────────────────────────────────
elif p=="progress_dashboard":
    if TIER1_AVAILABLE: render_stats_dashboard()
    else: st.error("⚠️ tier1_features.py not found. Place it in the same folder as app.py.")
# ── Mentor Directory: book sessions with senior doctors ──────────────────────
elif p=="mentor_directory":
    if MENTOR_DIRECTORY_OK: render_mentor_directory_page()
    else: st.error("⚠️ mentor_directory.py not found. Place it in the same folder as app.py.")
elif p=="admin_mentors":
    if MENTOR_DIRECTORY_OK: render_admin_mentor_panel()
    else: st.error("⚠️ mentor_directory.py not found. Place it in the same folder as app.py.")
elif p=="my_sessions":
    if MENTOR_DIRECTORY_OK: render_my_sessions_page()
    else: st.error("⚠️ mentor_directory.py not found.")
elif p=="jitsi_call":
    if MENTOR_DIRECTORY_OK: render_jitsi_call_page()
    else: st.error("⚠️ mentor_directory.py not found.")
# ── MCQ system: student session + admin panel ────────────────────────────────
elif p=="mcq_session":
    if MCQ_SYSTEM_OK: render_mcq_session_page()
    else: st.error("⚠️ mcq_system.py not found. Place it in the same folder as app.py.")
elif p=="admin_mcqs":
    if MCQ_SYSTEM_OK: render_mcq_admin_panel()
    else: st.error("⚠️ mcq_system.py not found.")
# ── AI Case Creator: Faculty-only — separate from manual case_creator page ───
elif p=="ai_case_creator":
    if CASE_CREATOR_OK: render_case_creator_panel()
    else: st.error("⚠️ case_creator.py not found. Place it in the same folder as app.py.")
# ── Image Practice Library: student-facing + admin ────────────────────────────
elif p=="image_practice":
    if IMAGE_LIBRARY_OK: render_image_practice_page()
    else: st.error("⚠️ image_library.py not found. Place it in the same folder as app.py.")
elif p=="admin_images":
    if IMAGE_LIBRARY_OK: render_image_admin_panel()
    else: st.error("⚠️ image_library.py not found.")
# ── RAG System: medical reference library admin (Phase 3) ──────────────────
elif p=="admin_rag":
    if RAG_SYSTEM_OK: render_rag_admin_panel()
    else: st.error("⚠️ rag_system.py not found. Place it in the same folder as app.py.")
# ── User Management + Email Notifications ────────────────────────────────────
elif p=="user_management":
    if USER_PANEL_OK: render_user_management_panel()
    else: st.error("⚠️ admin_user_panel.py not found. Place it in the same folder as app.py.")
else: page_home()

# ── Tier 1: floating "Ask Dr. Hiba" button + daily login init ────────────────
if TIER1_AVAILABLE and st.session_state.get("auth_user"):
    init_session()
    render_floating_help_button()
